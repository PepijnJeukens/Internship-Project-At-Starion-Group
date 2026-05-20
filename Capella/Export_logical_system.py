'''
Created on 24 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import *

#Path names
aird_path = "/DVS/DVS/DVS.aird"

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering
se = model.get_system_engineering()
lc_pkg = se.get_logical_architecture().get_logical_component_pkg()

# create  a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + "/" + se.get_name() + "_Logical_System.xlsx"

print("writing " + xlsx_file_name)

def get_max_depth(java_object, current_depth=1):
    """Recursively find the maximum depth of the component hierarchy using Java objects."""
    max_depth = current_depth
    try:
        children = java_object.getOwnedLogicalComponents()
        for child in children:
            child_depth = get_max_depth(child, current_depth + 1)
            if child_depth > max_depth:
                max_depth = child_depth
    except Exception as e:
        print(f"Error getting children for {java_object.getName()}: {e}")
    return max_depth

def get_component_id(component):
    """Get the ID of a component, handling both LogicalComponent and LogicalActor."""
    try:
        return component.get_id()
    except AttributeError:
        return component.get_java_object().getId()

def get_component_name(component):
    """Get the name of a component, handling both LogicalComponent and LogicalActor."""
    try:
        return component.get_name()
    except AttributeError:
        return component.get_java_object().getName()

# Determine the maximum depth of the hierarchy
max_depth = 0
# Use Java object directly to get all owned logical components
java_lc_pkg = lc_pkg.get_java_object()
children = java_lc_pkg.getOwnedLogicalComponents()
for child in children:
    depth = get_max_depth(child)
    if depth > max_depth:
        max_depth = depth

print(f"Maximum depth of the hierarchy: {max_depth}")

# create a workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active
ws.title = 'Logical system'

# Create headers based on the maximum depth
for i in range(max_depth):
    col = i * 3 + 1
    ws.cell(row=1, column=col, value=f"{'Sub' * i}System ID")
    ws.cell(row=1, column=col + 1, value=f"{'Sub' * i}System Name")
    ws.cell(row=1, column=col + 2, value=f"{'Sub' * i}System Type")

def export_component(component, row, level=0):
    """Recursively export a component and its children."""
    global ws
    col = level * 3 + 1
    component_id = get_component_id(component)
    component_name = get_component_name(component)

    ws.cell(row=row, column=col, value=component_id)
    ws.cell(row=row, column=col + 1, value=component_name)
    ws.cell(row=row, column=col + 2, value="Actor" if component.get_java_object().isActor() else "Component")

    row += 1  # Move to the next row for children

    try:
        java_object = component.get_java_object()
        children = java_object.getOwnedLogicalComponents()
        if children.size() > 0:
            for child in children:
                if child.isActor():
                    child_component = LogicalActor(child)
                else:
                    child_component = LogicalComponent(child)
                row = export_component(child_component, row, level + 1)
    except Exception as e:
        print(f"Error exporting children for {get_component_name(component)}: {e}")

    return row

# Export components
row = 2
for lc in lc_pkg.get_owned_logical_components():
    # print(f"Exporting system level component: {lc.get_name() if not lc.get_java_object().isActor() else lc.get_java_object().getName()}, Type: {'Actor' if lc.get_java_object().isActor() else 'Component'}")
    if lc.get_java_object().isActor():
        component = LogicalActor(lc.get_java_object())
    else:
        component = lc
    row = export_component(component, row)
    
try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

CapellaPlatform.refresh(folder)




