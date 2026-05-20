'''
Created on 02 Apr 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import Workbook
import os

# Path names
aird_path = "/DVS/DVS/DVS.aird"
# aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering and Logical Architecture
se = model.get_system_engineering()
la = se.get_logical_architecture()

# Get the capability realization package
capability_realization_pkg = la.get_capability_realization_pkg()
if not capability_realization_pkg:
    print("No Capability Realization Package found in Logical Architecture")
    exit()

# Create a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
folder_path = CapellaPlatform.getAbsolutePath(folder)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Set the correct file name format
xlsx_file_name = folder_path + "/" + se.get_name() + "_LA_Capabilities.xlsx"

print("Writing " + xlsx_file_name)

def get_logical_functions(capability_realization):
    """Get logical functions directly involved with a capability realization"""
    functions = set()

    try:
        # Get the Java object of the capability realization
        java_cr = capability_realization.get_java_object()

        # Get all owned AbstractFunctionAbstractCapabilityInvolvements
        involvements = java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements()

        # Iterate through all involvements to find functions
        for involvement in involvements:
            try:
                # Get the involved element (function)
                involved_element = involvement.getInvolved()
                if involved_element and involved_element.eClass().getName() == "LogicalFunction":
                    functions.add(LogicalFunction(involved_element))
            except Exception as e:
                print(f"Error processing involvement: {e}")
                continue

    except Exception as e:
        print(f"Error getting logical functions: {e}")

    return list(functions)

def get_involved_components_with_ids(capability_realization):
    """Get involved components (actors and components) for a capability realization with involvement IDs"""
    involved_components = []

    try:
        # Get the Java object of the capability realization
        java_cr = capability_realization.get_java_object()

        # Try to use getInvolvedComponents() method
        try:
            # Get all involved components directly
            involved_elements = java_cr.getInvolvedComponents()

            # Get all owned involvements
            involvements = java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements()

            # Create a mapping of element ID to involvement ID
            involvement_map = {}
            for involvement in involvements:
                try:
                    involved_element = involvement.getInvolved()
                    if involved_element:
                        involvement_map[involved_element.getId()] = involvement.getId()
                except Exception as e:
                    print(f"Error processing involvement: {e}")
                    continue

            # Iterate through all involved components
            for element in involved_elements:
                try:
                    element_id = element.getId()
                    involvement_id = involvement_map.get(element_id, "")

                    # Create the appropriate wrapper object based on isActor()
                    if element.isActor():
                        involved_components.append(("actor", LogicalActor(element), involvement_id))
                    else:
                        involved_components.append(("component", LogicalComponent(element), involvement_id))
                except Exception as e:
                    print(f"Error processing involved component: {e}")
                    continue
        except Exception as e:
            print(f"Error using getInvolvedComponents(): {e}")

            # Fallback method: Check CapabilityRealizationInvolvement relationships
            try:
                # Get all owned involvements
                involvements = java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements()

                # Iterate through all involvements to find components
                for involvement in involvements:
                    try:
                        involved_element = involvement.getInvolved()
                        if involved_element:
                            involvement_id = involvement.getId()

                            # Create the appropriate wrapper object based on isActor()
                            if involved_element.isActor():
                                involved_components.append(("actor", LogicalActor(involved_element), involvement_id))
                            else:
                                involved_components.append(("component", LogicalComponent(involved_element), involvement_id))
                    except Exception as e:
                        print(f"Error processing involvement: {e}")
                        continue
            except Exception as e:
                print(f"Error using fallback method: {e}")

    except Exception as e:
        print(f"Error getting involved components: {e}")

    return involved_components

def get_component_id(component):
    """Get the ID of a component, handling both LogicalComponent and LogicalActor."""
    try:
        if hasattr(component, 'get_id'):
            return component.get_id()
        elif hasattr(component, 'getId'):
            return component.getId()
        elif hasattr(component, 'get_java_object'):
            java_obj = component.get_java_object()
            if hasattr(java_obj, 'getId'):
                return java_obj.getId()
    except Exception as e:
        print(f"Error getting component ID: {e}")
    return None

def get_component_name(component):
    """Get the name of a component, handling both LogicalComponent and LogicalActor."""
    try:
        if hasattr(component, 'get_name'):
            return component.get_name()
        elif hasattr(component, 'getName'):
            return component.getName()
        elif hasattr(component, 'get_java_object'):
            java_obj = component.get_java_object()
            if hasattr(java_obj, 'getName'):
                return java_obj.getName()
    except Exception as e:
        print(f"Error getting component name: {e}")
    return None

def get_functional_chains(capability_realization):
    """Get functional chains owned or involved by a capability realization"""
    functional_chains = []

    try:
        # Get the Java object
        java_cr = capability_realization.get_java_object()

        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)

        for chain in all_chains:
            try:
                # Check if owned by capability realization
                if chain.get_java_object().eContainer() == java_cr:
                    functional_chains.append(chain)
                # Check if capability realization is involved
                else:
                    for involvement in chain.get_java_object().getOwnedFunctionalChainInvolvements():
                        if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                            involved = involvement.getInvolvedElement()
                            if involved and involved.getId() == java_cr.getId():
                                functional_chains.append(chain)
                                break
            except Exception as e:
                print(f"Error checking chain: {e}")
                continue
    except Exception as e:
        print(f"Error getting functional chains: {e}")

    return functional_chains

def get_involvement_id(element, chain):
    """Get the involvement ID for an element in a chain"""
    try:
        for involvement in chain.get_java_object().getOwnedFunctionalChainInvolvements():
            if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                involved = involvement.getInvolvedElement()
                if involved and involved.getId() == element.get_id():
                    return involvement.getId()
            elif involvement.eClass().getName() == "FunctionalChainInvolvementLink":
                involved = involvement.getInvolvedElement()
                if involved and involved.getId() == element.get_id():
                    return involvement.getId()
    except Exception as e:
        print(f"Error getting involvement ID: {e}")
    return ""

# First, find the maximum number of functional chains, logical functions, and involved components
max_functional_chains = 0
max_logical_functions = 0
max_involved_components = 0

# Temporary analysis to determine maximum sizes
for cr in capability_realization_pkg.get_owned_capability_realizations():
    try:
        # Count functional chains
        functional_chains = []
        for chain in se.get_all_contents_by_type(FunctionalChain):
            try:
                # Check if owned by capability realization
                if chain.get_java_object().eContainer() == cr.get_java_object():
                    functional_chains.append(chain)
                # Check if capability realization is involved
                else:
                    for involvement in chain.get_java_object().getOwnedFunctionalChainInvolvements():
                        if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                            involved = involvement.getInvolvedElement()
                            if involved and involved.getId() == cr.get_id():
                                functional_chains.append(chain)
                                break
            except:
                continue

        if len(functional_chains) > max_functional_chains:
            max_functional_chains = len(functional_chains)

        # Count logical functions
        functions = get_logical_functions(cr)
        if len(functions) > max_logical_functions:
            max_logical_functions = len(functions)

        # Count involved components
        involved_components = get_involved_components_with_ids(cr)
        if len(involved_components) > max_involved_components:
            max_involved_components = len(involved_components)

    except Exception as e:
        print(f"Error analyzing capability realization: {e}")
        continue

print(f"Found max: {max_functional_chains} functional chains, {max_logical_functions} logical functions, and {max_involved_components} involved components")

# Create a workbook
wb = Workbook()
ws = wb.active
ws.title = 'LA Capabilities'

# Create headers
ws.cell(row=1, column=1, value="Capability ID")
ws.cell(row=1, column=2, value="Capability Name")

# Add headers for functional chains
function_start_col = 3
for i in range(max_functional_chains):
    ws.cell(row=1, column=function_start_col + i*3, value=f"Functional Chain {i+1} ID")
    ws.cell(row=1, column=function_start_col + i*3 + 1, value=f"Functional Chain {i+1} Name")
    ws.cell(row=1, column=function_start_col + i*3 + 2, value=f"Functional Chain {i+1} Involvement ID")

# Add headers for logical functions
function_end_col = function_start_col + max_functional_chains*3
for i in range(max_logical_functions):
    ws.cell(row=1, column=function_end_col + i*3, value=f"Logical Function {i+1} ID")
    ws.cell(row=1, column=function_end_col + i*3 + 1, value=f"Logical Function {i+1} Name")
    ws.cell(row=1, column=function_end_col + i*3 + 2, value=f"Logical Function {i+1} Involvement ID")

# Add headers for involved components
component_start_col = function_end_col + max_logical_functions*3
for i in range(max_involved_components):
    ws.cell(row=1, column=component_start_col + i*4, value=f"Involved Component {i+1} ID")
    ws.cell(row=1, column=component_start_col + i*4 + 1, value=f"Involved Component {i+1} Name")
    ws.cell(row=1, column=component_start_col + i*4 + 2, value=f"Involved Component {i+1} Type")
    ws.cell(row=1, column=component_start_col + i*4 + 3, value=f"Involved Component {i+1} Involvement ID")

# Track the next available row
next_row = 2

# Export all capability realizations
for cr in capability_realization_pkg.get_owned_capability_realizations():
    try:
        # Get basic info
        cr_id = get_component_id(cr)
        cr_name = get_component_name(cr)

        # Write basic capability realization info
        ws.cell(row=next_row, column=1, value=cr_id)
        ws.cell(row=next_row, column=2, value=cr_name)

        # Get and write functional chains
        functional_chains = get_functional_chains(cr)
        for i, chain in enumerate(functional_chains):
            try:
                chain_id = get_component_id(chain)
                chain_name = get_component_name(chain)
                involvement_id = get_involvement_id(cr, chain)

                ws.cell(row=next_row, column=function_start_col + i*3, value=chain_id)
                ws.cell(row=next_row, column=function_start_col + i*3 + 1, value=chain_name)
                ws.cell(row=next_row, column=function_start_col + i*3 + 2, value=involvement_id)
            except Exception as e:
                print(f"Error writing functional chain {i+1}: {e}")

        # Get and write logical functions
        logical_functions = get_logical_functions(cr)
        for i, func in enumerate(logical_functions):
            try:
                func_id = get_component_id(func)
                func_name = get_component_name(func)
                involvement_id = ""

                # Try to find involvement ID
                for involvement in cr.get_java_object().getOwnedAbstractFunctionAbstractCapabilityInvolvements():
                    if involvement.getInvolved() and involvement.getInvolved().getId() == func.get_id():
                        involvement_id = involvement.getId()
                        break

                ws.cell(row=next_row, column=function_end_col + i*3, value=func_id)
                ws.cell(row=next_row, column=function_end_col + i*3 + 1, value=func_name)
                ws.cell(row=next_row, column=function_end_col + i*3 + 2, value=involvement_id)
            except Exception as e:
                print(f"Error writing logical function {i+1}: {e}")

        # Get and write involved components
        involved_components = get_involved_components_with_ids(cr)
        for i, (comp_type, comp, involvement_id) in enumerate(involved_components):
            try:
                comp_id = get_component_id(comp)
                comp_name = get_component_name(comp)

                ws.cell(row=next_row, column=component_start_col + i*4, value=comp_id)
                ws.cell(row=next_row, column=component_start_col + i*4 + 1, value=comp_name)
                ws.cell(row=next_row, column=component_start_col + i*4 + 2, value=comp_type)
                ws.cell(row=next_row, column=component_start_col + i*4 + 3, value=involvement_id)
            except Exception as e:
                print(f"Error writing involved component {i+1}: {e}")

        next_row += 1

    except Exception as e:
        print(f"Error processing capability realization {get_component_name(cr)}: {e}")
        import traceback
        traceback.print_exc()
        continue

# Save and refresh
try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

CapellaPlatform.refresh(folder)
print(f"Export completed. Exported data to {xlsx_file_name}")