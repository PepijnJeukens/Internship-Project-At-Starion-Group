'''
Created on 24 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import *

#Path names
aird_path = "/Import_test/Import_test.aird"
xlsx_path = "/DVS/results/DVS_Logical_System.xlsx"
# aird_path = "/Import_test_IFES/Import_test_IFES.aird"
# xlsx_path = "/In-Flight Entertainment System/results/In-Flight Entertainment System_Logical_System.xlsx"


model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering
se = model.get_system_engineering()
lc_pkg = se.get_logical_architecture().get_logical_component_pkg()

# create  a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Read " + xlsx_file_name)

# load the workbook
wb = load_workbook(xlsx_file_name)

# grab the active worksheet
ws = wb.active

def find_component_by_id(components, component_id):
    """Find a component by its ID in a list of components."""
    for component in components:
        try:
            if component.get_sid() == component_id:
                return component
        except AttributeError:
            if component.get_java_object().getSid() == component_id:
                return component
    return None

def get_component_by_id(java_object, component_id):
    """Find a component by its ID in the children of a Java object."""
    for child in java_object.getOwnedLogicalComponents():
        if child.getSid() == component_id:
            return child
    return None

#Start the import
model.start_transaction()

try:
    # Determine the maximum depth from the Excel file
    max_col = ws.max_column
    max_depth = max_col // 3
    print(f"Maximum depth detected from Excel: {max_depth}")

    # Create a list to keep track of parent components at each level
    parent_components = [None] * max_depth

    for row in ws.iter_rows(min_row=2):       
        # Create a list of component data for each level
        components = []
        for level in range(max_depth):
            col = level * 3
            component_id = row[col].value if row[col].value else None
            component_name = row[col + 1].value if col + 1 < len(row) and row[col + 1].value else None
            component_type = row[col + 2].value if col + 2 < len(row) and row[col + 2].value else None
            components.append((component_id, component_name, component_type))

        # Import the hierarchy
        for level in range(max_depth):
            component_id, component_name, component_type = components[level]

            if not component_id:
                continue  # Skip empty entries

            if level == 0:
                # Top level component
                existing_component = find_component_by_id(lc_pkg.get_owned_logical_components(), component_id)
                if existing_component is None:
                    print(f"Creating new top-level component: {component_name} with ID: {component_id} as {component_type}")
                    if component_type == "Actor":
                        component = LogicalActor()
                    else:
                        component = LogicalComponent()
                    lc_pkg.get_owned_logical_components().add(component)
                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(component.get_java_object())
                    component.set_name(component_name)
                    component.set_sid(component_id)
                else:
                    component = existing_component
                    print(f"Top-level component: {component_name} with ID: {component_id} already exists in the model.")
            else:
                # Child component
                parent_component = parent_components[level - 1]
                if parent_component is None:
                    print(f"Parent component not found for {component_name} at level {level}")
                    continue

                java_parent = parent_component.get_java_object()
                existing_component = get_component_by_id(java_parent, component_id)

                if existing_component is None:
                    print(f"Creating new component: {component_name} with ID: {component_id} under {parent_component.get_name()} as {component_type}")
                    if component_type == "Actor":
                        component = LogicalActor()
                    else:
                        component = LogicalComponent()
                    java_parent.getOwnedLogicalComponents().add(component.get_java_object())
                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(component.get_java_object())
                    component.set_name(component_name)
                    component.set_sid(component_id)
                else:
                    if component_type == "Actor":
                        component = LogicalActor(existing_component)
                    else:
                        component = LogicalComponent(existing_component)
                    print(f"Component {component_name} with ID: {component_id} already exists under {parent_component.get_name()}")

            parent_components[level] = component
                       
except Exception as e:
    print("Error: " + e)
    model.rollback_transaction()
    raise 

else:
    model.commit_transaction()
    
# At the end of your script, after the transaction is committed
model.save()
print("Model saved")           