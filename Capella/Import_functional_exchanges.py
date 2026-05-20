'''
Created on 25 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
aird_path = "/Import_test/Import_test.aird"
xlsx_path = "/DVS/results/DVS_Logical_Functional_Exchanges.xlsx"
# aird_path = "/Import_test_IFES/Import_test_IFES.aird"
# xlsx_path = "/In-Flight Entertainment System/results/In-Flight Entertainment System_Logical_Functional_Exchanges.xlsx"


model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()
lf_pkg = la.get_logical_function_pkg()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Reading " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

def find_function_by_id(parent, function_id):
    """Find a function by its ID in the hierarchy."""
    if parent.get_sid() == function_id:
        return parent

    for function in parent.get_owned_functions():
        if function.get_sid() == function_id:
            return function
        found = find_function_by_id(function, function_id)
        if found:
            return found
    return None

def find_functional_exchange_by_id(exchange_id):
    """Find a functional exchange by its ID in the model"""
    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for fe in all_exchanges:
            if fe.get_sid() == exchange_id:
                return fe
        return None
    except Exception as e:
        print(f"Error finding functional exchange by ID {exchange_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_port(function, is_input, port_name, port_id):
    """Create a function port (input or output) for a function."""
    try:
        if is_input:
            # Create input port using the correct method
            port = FunctionInputPort()
            function.get_inputs().add(port)
            port.set_name(port_name)
            port.get_java_object().setSid(port_id)
            print(f"Created input port {port_name} for function {function.get_name()}")
            return port
        else:
            # Create output port using the correct method
            port = FunctionOutputPort()
            function.get_outputs().add(port)
            port.set_name(port_name)
            port.get_java_object().setSid(port_id)
            print(f"Created output port {port_name} for function {function.get_name()}")
            return port
    except Exception as e:
        print(f"Error creating port {port_name} for function {function.get_name()}: {e}")
        return None

def find_or_create_port(function, is_input, port_name, port_id):
    """Find an existing port or create a new one."""
    try:
        if is_input:
            # Check existing input ports
            for port in function.get_inputs():
                if port.get_sid() == port_id:
                    print(f"Found existing input port {port_name} for function {function.get_name()}")
                    return port
            # Create new input port if not found
            return create_port(function, True, port_name, port_id)
        else:
            # Check existing output ports
            for port in function.get_outputs():
                if port.get_sid() == port_id:
                    print(f"Found existing output port {port_name} for function {function.get_name()}")
                    return port
            # Create new output port if not found
            return create_port(function, False, port_name, port_id)
    except Exception as e:
        print(f"Error finding or creating port {port_name} for function {function.get_name()}: {e}")
        return None

def check_functional_exchange_exists(exchange_id):
    """Check if a functional exchange with the given ID already exists"""
    try:
        # Find the functional exchange by ID
        existing_fe = find_functional_exchange_by_id(exchange_id)
        if existing_fe:
            print(f"Functional exchange with ID {exchange_id} already exists")
            return True
        return False
    except Exception as e:
        print(f"Error checking if functional exchange exists: {e}")
        return False

# Start the import
model.start_transaction()

try:
    # Find the Root Logical Function
    rlf = None
    for lf in lf_pkg.get_owned_logical_functions():
        if lf.get_name() == "Root Logical Function":
            rlf = lf
            break

    if rlf is None:
        print("Root Logical Function not found!")
        exit()

    # Process each row in the Excel file
    for row in ws.iter_rows(min_row=2):
        # Extract data from the row
        source_func_id = row[0].value
        source_func_name = row[1].value
        source_port_id = row[2].value
        source_port_name = row[3].value
        fe_id = row[4].value
        fe_name = row[5].value
        target_func_id = row[6].value
        target_func_name = row[7].value
        target_port_id = row[8].value
        target_port_name = row[9].value

        print(f"\nProcessing exchange: {fe_name} between {source_func_name} and {target_func_name}")

        # Check if the functional exchange already exists by ID
        if check_functional_exchange_exists(fe_id):
            print(f"Skipping - Functional exchange with ID {fe_id} already exists")
            continue

        # Find source and target functions
        source_function = find_function_by_id(rlf, source_func_id)
        target_function = find_function_by_id(rlf, target_func_id)

        if not source_function or not target_function:
            print(f"Could not find source or target function for exchange {fe_name}")
            continue

        print(f"Found source function: {source_function.get_name()} with ID: {source_function.get_id()}")
        print(f"Found target function: {target_function.get_name()} with ID: {target_function.get_id()}")

        # Find or create source port
        source_port = find_or_create_port(source_function, False, source_port_name, source_port_id)
        if not source_port:
            print(f"Failed to create/get source port {source_port_name} for {source_func_name}")
            continue

        print(f"Source port: {source_port.get_name()} with ID: {source_port.get_id()}")

        # Find or create target port
        target_port = find_or_create_port(target_function, True, target_port_name, target_port_id)
        if not target_port:
            print(f"Failed to create/get target port {target_port_name} for {target_func_name}")
            continue

        print(f"Target port: {target_port.get_name()} with ID: {target_port.get_id()}")

        # Create the functional exchange
        fe = FunctionalExchange()

        # Add the exchange to the source function
        fe.set_source_port(source_port)
        fe.set_target_port(target_port)

        source_function.get_owned_functional_exchanges().add(fe)
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fe.get_java_object())

        fe.set_name(fe_name)
        fe.get_java_object().setSid(fe_id)

        print(f"Successfully created exchange {fe_name} with ID {fe_id}")

except Exception as e:
    print("Error: " + str(e))
    import traceback
    traceback.print_exc()
    model.rollback_transaction()
    raise

# else:
model.commit_transaction()

model.save()
print("Functional exchanges import completed successfully")