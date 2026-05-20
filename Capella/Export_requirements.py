'''
Created on 06 May 2026

@author: p.jeukens
'''

#----------------------------------- 1. Input, Imports and Includes
#==================================================================================
# Input, Library Imports and other Inclusions (e.g.Capella API) go here
#==================================================================================
import xml.etree.ElementTree as ET
from xml.dom import minidom
import html
from html.parser import HTMLParser
import os
import warnings
import random
import string
from datetime import datetime
import pytz
import re
from openpyxl import Workbook

warnings.simplefilter(action='ignore', category=UserWarning)

# include dependencies
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *
include('workspace://Python4Capella/simplified_api/requirement.py')
if False:
    from simplified_api.requirement import *
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

#----------------------------------- 2. Functions
#==================================================================================
# This section contains any required function definitions
#==================================================================================

def generate_random_id(length=23):
    """Generate a random ReqIF identifier."""
    prefix = "_"
    characters = string.ascii_letters + string.digits
    random_part = ''.join(random.choices(characters, k=length))
    return prefix + random_part

def get_current_time():
    """Get current time in ReqIF format."""
    tz = pytz.timezone('Europe/Amsterdam')
    now = datetime.now(tz)
    formatted_time = now.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + now.strftime('%z')
    formatted_time = formatted_time[:-2] + ':' + formatted_time[-2:]
    return formatted_time

def set_ReqIF_identifier(ModelObject, model):
    """Set ReqIF identifier for model objects that don't have one."""
    try:
        ModelObject = ModelObject.get_java_object()
    except:
        pass
    if ModelObject.getReqIFIdentifier():
        return ModelObject.getReqIFIdentifier()
    else:
        ReqIFId = generate_random_id(length=23)
        try:
            model.start_transaction()
            ModelObject.setReqIFIdentifier(ReqIFId)
        except:
            print('Writing to model failed')
            model.rollback_transaction()
        else:
            print('set ReqIF.Id for model object:', ModelObject.getReqIFLongName())
            model.commit_transaction()
        model.save()
    return ReqIFId

XHTML_NS = "http://www.w3.org/1999/xhtml"

def collapse_xhtml_blocks(s, prefix="xhtml"):
    """Collapse XHTML blocks for pretty printing."""
    out = []
    pos = 0
    open_re = re.compile(r'<%s:([A-Za-z0-9_\-]+)\b' % re.escape(prefix))
    while True:
        m = open_re.search(s, pos)
        if not m:
            out.append(s[pos:])
            break
        start = m.start()
        tag = m.group(1)
        out.append(s[pos:start])
        pattern = re.compile(
            r'(</{p}:{t}\s*>)|(<{p}:{t}\b[^>]*?/>)|(<{p}:{t}\b[^>]*>)'.format(p=re.escape(prefix), t=re.escape(tag)),
            flags=re.S,
        )
        depth = 0
        found = False
        for mm in pattern.finditer(s, start):
            tok = mm.group(0)
            if tok.startswith('</'):
                depth -= 1
                if depth == 0:
                    end = mm.end()
                    block = s[start:end]
                    out.append(_collapse_block(block))
                    pos = end
                    found = True
                    break
            elif tok.endswith('/>'):
                if depth == 0:
                    end = mm.end()
                    block = s[start:end]
                    out.append(_collapse_block(block))
                    pos = end
                    found = True
                    break
            else:
                depth += 1
        if not found:
            out.append(s[start:])
            break
    return "".join(out)

def _collapse_block(block):
    """Collapse whitespace in a block."""
    b = re.sub(r'\r?\n', ' ', block)
    b = re.sub(r'\s+', ' ', b)
    b = re.sub(r'>\s+<', '><', b)
    b = re.sub(r'>\s+([^<\s])', r'>\1', b)
    b = re.sub(r'(\S)\s+<', r'\1<', b)
    return b.strip()

def wrap_in_div_p(text: str) -> str:
    """Wrap text in div/p tags."""
    return f'<div><p style="">{html.escape(text)}</p></div>'

class HTMLToXHTMLParser(HTMLParser):
    """Parser for converting HTML to XHTML."""
    def __init__(self):
        super().__init__()
        self.fragments = []
        self.stack = []

    def handle_starttag(self, tag, attrs):
        attrib = dict(attrs)
        if tag == "p" and "style" not in attrib:
            attrib["style"] = ""
        el = ET.Element(f"{{{XHTML_NS}}}{tag}", attrib)
        if self.stack:
            self.stack[-1].append(el)
        else:
            self.fragments.append(el)
        self.stack.append(el)

    def handle_endtag(self, tag):
        if self.stack:
            self.stack.pop()

    def handle_data(self, data):
        if self.stack:
            if self.stack[-1].text:
                self.stack[-1].text += data
            else:
                self.stack[-1].text = data

    def handle_startendtag(self, tag, attrs):
        attrib = dict(attrs)
        el = ET.Element(f"{{{XHTML_NS}}}{tag}", attrib)
        if self.stack:
            self.stack[-1].append(el)
        else:
            self.fragments.append(el)

def convert_html_to_xhtml(html_string):
    """Convert HTML string to XHTML."""
    parser = HTMLToXHTMLParser()
    parser.feed(html_string)
    wrapper = ET.Element(f"{{{XHTML_NS}}}div")
    for frag in parser.fragments:
        if frag.tag == f"{{{XHTML_NS}}}div":
            for child in list(frag):
                wrapper.append(child)
        else:
            wrapper.append(frag)
    return ET.tostring(wrapper, encoding="unicode")

#----------------------------------- 3. Run Main Script
#==================================================================================
# This section runs the main script
#==================================================================================

# Path names - ONLY THIS NEEDS TO BE SPECIFIED
aird_path = "/DVS/DVS/DVS.aird"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()

# Create a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + "/" + se.get_name() + "_Requirements.xlsx"

print("Writing " + xlsx_file_name)

# Create workbook
wb = Workbook()
del wb['Sheet']  # Remove default sheet

# Create all worksheets
ws_header = wb.create_sheet("HEADER")
ws_datatypes = wb.create_sheet("DATATYPES")
ws_enum_values = wb.create_sheet("ENUM-VALUES")
ws_spec_types = wb.create_sheet("SPEC-TYPES")
ws_spec_attributes = wb.create_sheet("SPEC-ATTRIBUTES")
ws_spec_objects = wb.create_sheet("SPEC-OBJECTS")
ws_attribute_values = wb.create_sheet("ATTRIBUTE-VALUES")
ws_spec_relations = wb.create_sheet("SPEC-RELATIONS")
ws_specifications = wb.create_sheet("SPECIFICATIONS")
ws_spec_hierarchy = wb.create_sheet("SPEC-HIERARCHY")

# Populate HEADER worksheet
print("Writing HEADER to Excel...")
headers = ["IDENTIFIER", "TITLE", "xmlns=", "version=", "encoding=", "comment", "req-if-tool-id", "source-tool-id", "xmlns:xhtml="]
for item in se.get_all_contents_by_type(TypesFolder):
    header_id = set_ReqIF_identifier(item, model)
header_title = se.get_name() + " Requirements Specification"
values = [header_id, header_title, "http://www.omg.org/spec/ReqIF/20110401/reqif.xsd", "1.0", "UTF-8",
    "Created by: Starion Group", "ReqIF-Export for Capella (https://github.com/STARIONGROUP/...)", "Capella Version 7 (https://mbse-capella.org/)", "http://www.w3.org/1999/xhtml"]
for col, (header, value) in enumerate(zip(headers, values), start=1):
    ws_header.cell(row=1, column=col, value=header)
    ws_header.cell(row=2, column=col, value=value)

# Populate DATATYPES worksheet
print("Writing DATATYPES to Excel...")
headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "DATATYPE-DEFINITION-KIND", "ALTERNATIVE-ID"]
for col, header in enumerate(headers, start=1):
    ws_datatypes.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(DataTypeDefinition):
    ws_datatypes[f"A{i}"] = set_ReqIF_identifier(item, model)
    ws_datatypes[f"B{i}"] = get_current_time()
    ws_datatypes[f"C{i}"] = item.get_java_object().getReqIFLongName()
    ws_datatypes[f"D{i}"] = 'DATATYPE-DEFINITION-XHTML' if "HTML" in item.get_java_object().getReqIFLongName() else 'DATATYPE-DEFINITION-STRING'
    ws_datatypes[f"E{i}"] = item.get_java_object().getId()
    i += 1
for item in se.get_all_contents_by_type(EnumerationDataTypeDefinition):
    ws_datatypes[f"A{i}"] = set_ReqIF_identifier(item, model)
    ws_datatypes[f"B{i}"] = get_current_time()
    ws_datatypes[f"C{i}"] = item.get_java_object().getReqIFLongName()
    ws_datatypes[f"D{i}"] = 'DATATYPE-DEFINITION-ENUMERATION'
    ws_datatypes[f"E{i}"] = item.get_java_object().getId()
    i += 1

# Populate ENUM-VALUES worksheet
print("Writing ENUM-VALUES to Excel...")
headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "KEY", "DATATYPE-DEFINITION-ENUMERATION-REF", "ALTERNATIVE-ID"]
for col, header in enumerate(headers, start=1):
    ws_enum_values.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(EnumerationDataTypeDefinition):
    for value in item.get_java_object().getSpecifiedValues():
        ws_enum_values[f"A{i}"] = set_ReqIF_identifier(value, model)
        ws_enum_values[f"B{i}"] = get_current_time()
        ws_enum_values[f"C{i}"] = value.getReqIFLongName()
        ws_enum_values[f"D{i}"] = str(i-1)
        ws_enum_values[f"E{i}"] = set_ReqIF_identifier(item, model)
        ws_enum_values[f"F{i}"] = value.getId()
        i += 1

# Populate SPEC-TYPES worksheet
print("Writing SPEC-TYPES to Excel...")
headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "KIND", "ALTERNATIVE-ID"]
for col, header in enumerate(headers, start=1):
    ws_spec_types.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(RequirementType):
    ws_spec_types[f"A{i}"] = set_ReqIF_identifier(item, model)
    ws_spec_types[f"B{i}"] = get_current_time()
    ws_spec_types[f"C{i}"] = item.get_java_object().getReqIFLongName()
    ws_spec_types[f"D{i}"] = 'SPEC-OBJECT-TYPE'
    ws_spec_types[f"E{i}"] = item.get_java_object().getId()
    i += 1
for item in se.get_all_contents_by_type(ModuleType):
    ws_spec_types[f"A{i}"] = set_ReqIF_identifier(item, model)
    ws_spec_types[f"B{i}"] = get_current_time()
    ws_spec_types[f"C{i}"] = item.get_java_object().getReqIFLongName()
    ws_spec_types[f"D{i}"] = 'SPECIFICATION-TYPE'
    ws_spec_types[f"E{i}"] = item.get_java_object().getId()
    i += 1
for item in se.get_all_contents_by_type(RelationType):
    ws_spec_types[f"A{i}"] = set_ReqIF_identifier(item, model)
    ws_spec_types[f"B{i}"] = get_current_time()
    ws_spec_types[f"C{i}"] = item.get_java_object().getReqIFLongName()
    ws_spec_types[f"D{i}"] = 'SPEC-RELATION-TYPE'
    ws_spec_types[f"E{i}"] = item.get_java_object().getId()
    i += 1

# Populate SPEC-ATTRIBUTES worksheet
print("Writing SPEC-ATTRIBUTES to Excel...")
headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "ATTRIBUTE-DEF-KIND", "SPEC-TYPE-REF", "DATATYPE-DEFINITION-STRING-REF", "ALTERNATIVE-ID"]
for col, header in enumerate(headers, start=1):
    ws_spec_attributes.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(AttributeDefinition):
    ws_spec_attributes[f"A{i}"] = set_ReqIF_identifier(item, model)
    ws_spec_attributes[f"B{i}"] = get_current_time()
    ws_spec_attributes[f"C{i}"] = item.get_java_object().getReqIFLongName()
    def_type = item.get_java_object().getDefinitionType()
    if def_type and def_type.eClass().getName() == "DataTypeDefinition":
        if "HTML" in def_type.getReqIFLongName() or item.get_java_object().getReqIFLongName() == "ReqIF.Text":
            ws_spec_attributes[f"D{i}"] = 'ATTRIBUTE-DEFINITION-XHTML'
        else:
            ws_spec_attributes[f"D{i}"] = 'ATTRIBUTE-DEFINITION-STRING'
    elif def_type and def_type.eClass().getName() == "EnumerationDataTypeDefinition":
        ws_spec_attributes[f"D{i}"] = 'ATTRIBUTE-DEFINITION-ENUMERATION'
    ws_spec_attributes[f"E{i}"] = item.get_java_object().eContainer().getReqIFIdentifier()
    if item.get_java_object().getReqIFLongName() == "ReqIF.Text":
        for type in se.get_all_contents_by_type(DataTypeDefinition):
            if type.get_java_object().getReqIFLongName() == "DataTypeDefinitionXHTML":
                ws_spec_attributes[f"F{i}"] = type.get_java_object().getReqIFIdentifier()
    else:
        ws_spec_attributes[f"F{i}"] = def_type.getReqIFIdentifier() if def_type else ""
    ws_spec_attributes[f"G{i}"] = item.get_java_object().getId()
    i += 1

# Populate SPEC-OBJECTS worksheet
print("Writing SPEC-OBJECTS to Excel...")
headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "SPEC-TYPE-REF", "ALTERNATIVE-ID", "Capella BridgeTraces Type"]
for col, header in enumerate(headers, start=1):
    ws_spec_objects.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(Requirement):
    ws_spec_objects[f"A{i}"] = set_ReqIF_identifier(item, model)
    ws_spec_objects[f"B{i}"] = get_current_time()
    ws_spec_objects[f"C{i}"] = item.get_java_object().getReqIFLongName()
    ws_spec_objects[f"D{i}"] = item.get_java_object().getRequirementType().getReqIFIdentifier()
    if item.get_java_object().getRequirementTypeProxy() != item.get_java_object().getRequirementType().getReqIFIdentifier():
        try:
            model.start_transaction()
            item.get_java_object().setRequirementTypeProxy(item.get_java_object().getRequirementType().getReqIFIdentifier())
        except:
            print('Writing to model failed')
            model.rollback_transaction()
        else:
            print('set RequirementTypeProxy for model object:', item.get_java_object().getReqIFLongName())
            model.commit_transaction()
        model.save()
    ws_spec_objects[f"E{i}"] = item.get_java_object().getId()
    if item.get_java_object().eClass().getName() == 'Folder':
        if len(item.get_java_object().getOwnedRequirements()) == 0:
            ws_spec_objects[f"F{i}"] = 'Requirement'
        else:
            ws_spec_objects[f"F{i}"] = item.get_java_object().eClass().getName()
    else:
        ws_spec_objects[f"F{i}"] = item.get_java_object().eClass().getName()
    i += 1

# Populate ATTRIBUTE-VALUES worksheet
print("Writing ATTRIBUTE-VALUES to Excel...")
headers = ["VALUE / ENUM-VALUE-REF", "ATTRIBUTE-VALUE-KIND", "ATTRIBUTE-DEFINITION-REF", "SPEC-OBJECT-REF", "ALTERNATIVE-ID"]
for col, header in enumerate(headers, start=1):
    ws_attribute_values.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(Requirement):
    attr_list = item.get_java_object().getRequirementType().getOwnedAttributes()
    for attribute in item.get_java_object().getOwnedAttributes():
        try:
            if attribute.eClass().getName() == "EnumerationValueAttribute":
                ws_attribute_values[f"A{i}"] = attribute.getValues()[0].getReqIFIdentifier()
                ws_attribute_values[f"B{i}"] = "ATTRIBUTE-VALUE-ENUMERATION"
            elif attribute.eClass().getName() == "StringValueAttribute":
                ws_attribute_values[f"A{i}"] = attribute.getValue()
                ws_attribute_values[f"B{i}"] = "ATTRIBUTE-VALUE-STRING"
            ws_attribute_values[f"C{i}"] = attribute.getDefinition().getReqIFIdentifier()
            ws_attribute_values[f"D{i}"] = attribute.eContainer().getReqIFIdentifier()
            ws_attribute_values[f"E{i}"] = attribute.getId()
            i += 1
        except:
            print("Attribute Skipped: attribute value or definition missing for SPEC-OBJECT:", get_name(item))
    if item.get_java_object().getReqIFName():
        ws_attribute_values[f"A{i}"] = item.get_java_object().getReqIFName()
        ws_attribute_values[f"B{i}"] = "ATTRIBUTE-VALUE-STRING"
        for attr in attr_list:
            if attr.getReqIFLongName() == "ReqIF.Name":
                ws_attribute_values[f"C{i}"] = attr.getReqIFIdentifier()
        ws_attribute_values[f"D{i}"] = set_ReqIF_identifier(item, model)
        i += 1
    if item.get_java_object().getReqIFChapterName():
        ws_attribute_values[f"A{i}"] = item.get_java_object().getReqIFChapterName()
        ws_attribute_values[f"B{i}"] = "ATTRIBUTE-VALUE-STRING"
        for attr in attr_list:
            if attr.getReqIFLongName() == "ReqIF.ChapterName":
                ws_attribute_values[f"C{i}"] = attr.getReqIFIdentifier()
        ws_attribute_values[f"D{i}"] = set_ReqIF_identifier(item, model)
        i += 1
    if item.get_java_object().getReqIFPrefix():
        ws_attribute_values[f"A{i}"] = item.get_java_object().getReqIFPrefix()
        ws_attribute_values[f"B{i}"] = "ATTRIBUTE-VALUE-STRING"
        for attr in attr_list:
            if attr.getReqIFLongName() == "ReqIF.Prefix":
                ws_attribute_values[f"C{i}"] = attr.getReqIFIdentifier()
        ws_attribute_values[f"D{i}"] = set_ReqIF_identifier(item, model)
        i += 1
    if item.get_java_object().getReqIFText():
        ws_attribute_values[f"A{i}"] = item.get_java_object().getReqIFText()
        ws_attribute_values[f"B{i}"] = "ATTRIBUTE-VALUE-XHTML"
        for attr in attr_list:
            if attr.getReqIFLongName() == "ReqIF.Text":
                ws_attribute_values[f"C{i}"] = attr.getReqIFIdentifier()
        ws_attribute_values[f"D{i}"] = set_ReqIF_identifier(item, model)
        i += 1

# Populate SPEC-RELATIONS worksheet
print("Writing SPEC-RELATIONS to Excel...")
headers = ["IDENTIFIER", "LAST-CHANGE", "SOURCE-SPEC-OBJECT-REF", "TARGET-SPEC-OBJECT-REF", "SPEC-RELATION-TYPE-REF", "ALTERNATIVE-ID"]
for col, header in enumerate(headers, start=1):
    ws_spec_relations.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(AbstractRelation):
    try:
        ws_spec_relations[f"A{i}"] = set_ReqIF_identifier(item, model)
        ws_spec_relations[f"B{i}"] = get_current_time()
        ws_spec_relations[f"C{i}"] = item.get_java_object().getSource().getReqIFIdentifier()
        ws_spec_relations[f"D{i}"] = item.get_java_object().getTarget().getReqIFIdentifier()
        ws_spec_relations[f"E{i}"] = item.get_java_object().getRelationType().getReqIFIdentifier()
        ws_spec_relations[f"F{i}"] = item.get_java_object().getId()
        i += 1
    except:
        pass

# Populate SPECIFICATIONS worksheet
print("Writing SPECIFICATIONS to Excel...")
headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "SPEC-OBJECT-TYPE-REF", "ALTERNATIVE-ID", "Capella BridgeTraces Type"]
for col, header in enumerate(headers, start=1):
    ws_specifications.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(CapellaModule):
    ws_specifications[f"A{i}"] = set_ReqIF_identifier(item, model)
    ws_specifications[f"B{i}"] = get_current_time()
    ws_specifications[f"C{i}"] = item.get_java_object().getReqIFLongName()
    ws_specifications[f"D{i}"] = item.get_java_object().getModuleType().getReqIFIdentifier()
    ws_specifications[f"E{i}"] = item.get_java_object().getId()
    ws_specifications[f"F{i}"] = "Module"
    i += 1

# Populate SPEC-HIERARCHY worksheet
print("Writing SPEC-HIERARCHY to Excel...")
headers = ["IDENTIFIER", "LAST-CHANGE", "SPEC-OBJECT-REF", "PARENT-SPEC-OBJECT-REF"]
for col, header in enumerate(headers, start=1):
    ws_spec_hierarchy.cell(row=1, column=col, value=header)
i = 2
for item in se.get_all_contents_by_type(Requirement):
    ws_spec_hierarchy[f"A{i}"] = generate_random_id()
    ws_spec_hierarchy[f"B{i}"] = get_current_time()
    ws_spec_hierarchy[f"C{i}"] = set_ReqIF_identifier(item, model)
    ws_spec_hierarchy[f"D{i}"] = item.get_java_object().eContainer().getReqIFIdentifier()
    i += 1

# Save the workbook
try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

# Refresh the folder in Capella
CapellaPlatform.refresh(folder)

print("Done with everything")