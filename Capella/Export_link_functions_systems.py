'''
Created on 24 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import *
import os

#Path names
aird_path = "/DVS/DVS/DVS.aird"
# aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()
lf_pkg = la.get_logical_function_pkg()

# Find the Root Logical Function
rlf = None
for lf in lf_pkg.get_owned_logical_functions():
    if lf.get_name() == "Root Logical Function":
        rlf = lf
        break

if rlf is None:
    print("Root Logical Function not found!")
    exit()

# Get the logical component package
lc_pkg = la.get_logical_component_pkg()

# Create a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
folder_path = CapellaPlatform.getAbsolutePath(folder)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)

xlsx_file_name = folder_path + "/" + se.get_name() + "_Function_System_Links.xlsx"

print("Writing " + xlsx_file_name)

def get_system_max_depth(java_object, current_depth=1):
    """Recursively find the maximum depth of the component hierarchy using Java objects."""
    max_depth = current_depth
    try:
        children = java_object.getOwnedLogicalComponents()
        for child in children:
            child_depth = get_system_max_depth(child, current_depth + 1)
            if child_depth > max_depth:
                max_depth = child_depth
    except Exception as e:
        print(f"Error getting children for {java_object.getName()}: {e}")
    return max_depth

# Determine the maximum depth of the system hierarchy
system_max_depth = 0
# Use Java object directly to get all owned logical components
java_lc_pkg = lc_pkg.get_java_object()
children = java_lc_pkg.getOwnedLogicalComponents()
for child in children:
    depth = get_system_max_depth(child)
    if depth > system_max_depth:
        system_max_depth = depth

def get_function_max_depth(function, current_depth=1):
    """Recursively find the maximum depth of the function hierarchy."""
    max_depth = current_depth
    try:
        children = function.get_java_object().getOwnedFunctions()
        for child in children:
            child_function = LogicalFunction(child)
            child_depth = get_function_max_depth(child_function, current_depth + 1)
            if child_depth > max_depth:
                max_depth = child_depth
    except Exception as e:
        print(f"Error getting children for {function.get_name()}: {e}")
    return max_depth

# Determine the maximum depth of the hierarchy under RLF
function_max_depth = 0
children = rlf.get_java_object().getOwnedFunctions()
for child in children:
    child_function = LogicalFunction(child)
    depth = get_function_max_depth(child_function, 1)  # Start from 1 since we're skipping RLF
    if depth > function_max_depth:
        function_max_depth = depth

print(f"Maximum system depth: {system_max_depth}")
print(f"Maximum function depth: {function_max_depth}")

# Create a workbook
wb = Workbook()

# Grab the active worksheet
ws = wb.active
ws.title = 'Function-System Links'

# Create headers for systems
for i in range(system_max_depth):
    col = i * 2 + 1  # 2 columns per level (ID and Name)
    ws.cell(row=1, column=col, value=f"{'Sub' * i}System ID")
    ws.cell(row=1, column=col + 1, value=f"{'Sub' * i}System Name")

# Create headers for functions
function_start_col = system_max_depth * 2 + 1
for i in range(function_max_depth):
    col = function_start_col + i * 2
    ws.cell(row=1, column=col, value=f"{'Sub' * i}Function ID")
    ws.cell(row=1, column=col + 1, value=f"{'Sub' * i}Function Name")

def export_system_functions(system, row, system_level=0):
    """Export all functions allocated to a system and its children."""
    global ws

    try:
        # Write system info at current level
        col = system_level * 2 + 1
        ws.cell(row=row, column=col, value=system.get_id())
        ws.cell(row=row, column=col + 1, value=system.get_name())
        row += 1
        
        # Get allocated functions for this system
        java_object = system.get_java_object()
        allocated_functions = java_object.getAllocatedLogicalFunctions()
        
        # Write each allocated function at function level 0
        for func in allocated_functions:
            function = LogicalFunction(func)
            func_col = function_start_col
            ws.cell(row=row, column=func_col, value=function.get_id())
            ws.cell(row=row, column=func_col + 1, value=function.get_name())
            row += 1
        
        # Process children at next system level
        children = java_object.getOwnedLogicalComponents()
        for child in children:
            if child.isActor():
                child_system = LogicalActor(child)
            else:
                child_system = LogicalComponent(child)
            
            row = export_system_functions(child_system, row, system_level + 1)

    except Exception as e:
        print(f"Error processing system {system.get_name()}: {e}")

    return row

# Export all system-function links
row = 2
for lc in lc_pkg.get_owned_logical_components():
    row = export_system_functions(lc, row)

try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

CapellaPlatform.refresh(folder)