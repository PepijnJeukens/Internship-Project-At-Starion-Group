'''
Created on 13 Apr 2026

@author: p.jeukens
'''

""" 
This script contains all functions necessary to import the Logical Architecture layer from Capella to Excel
It can import systems (components/actors), functions, the link between systems and functions, functional exchanges,
component exchanges, the link between functional and component exchanges, functional chains and capabilities.

Additionally it checks the file that is being exported to for all already existing parts. In case an item is already in the
Capella Project but not in the imported file, the item will be deleted from the project. This allows a project to be updated
in case the orginal project which was exported is updated as well. It also creates a deletion log with the id's and names of
the deleted items. In case you desire that you only add items and not delete them the Master_import.py file should use
Import_functions.py instead of Import_and_delete_functions.py
"""

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import os

def get_model_and_workbook(aird_path, xlsx_path):
    """Open the model and load the workbook with project info"""
    try:
        model = CapellaModel()
        model.open(aird_path)

        # Get system engineering
        se = model.get_system_engineering()
        la = se.get_logical_architecture()
        lc_pkg = la.get_logical_component_pkg()
        lf_pkg = la.get_logical_function_pkg()

        # Find the Root Logical Function
        rlf = None
        for lf in lf_pkg.get_owned_logical_functions():
            if lf.get_name() == "Root Logical Function":
                rlf = lf
                break

        # Load the workbook
        xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
        xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)
        print("Read " + xlsx_file_name)
        wb = load_workbook(xlsx_file_name)

        # Store project info in workbook
        if not hasattr(wb, '_custom_doc_props'):
            wb._custom_doc_props = {}

        # Get project info - only get what we need without trying to access folders
        try:
            model_path = CapellaPlatform.getModelPath(se)
            if model_path and "/" in model_path:
                project_name = model_path.split("/")[1]
                wb._custom_doc_props["project_name"] = project_name
                wb._custom_doc_props["se"] = se
            else:
                project_name = os.path.basename(aird_path).split('.')[0]
                wb._custom_doc_props["project_name"] = project_name
                wb._custom_doc_props["se"] = se
        except Exception as e:
            print(f"Error getting project info: {e}")
            project_name = os.path.basename(aird_path).split('.')[0]
            wb._custom_doc_props = {
                "se": se,
                "project_name": project_name
            }

        # Store the rest of the model information
        wb._custom_doc_props["la"] = la
        wb._custom_doc_props["lc_pkg"] = lc_pkg
        wb._custom_doc_props["lf_pkg"] = lf_pkg
        wb._custom_doc_props["rlf"] = rlf

        return model, se, la, lc_pkg, lf_pkg, rlf, wb, xlsx_file_name

    except Exception as e:
        print(f"Error in get_model_and_workbook: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, None, None, None, None, None, None, None

def find_component_by_id(component_input, component_id):
    """
    Find a component by its ID.
    Works with either a Java object or a list of components.
    """
    try:
        # Case 1: component_input is a Java object (for recursive search)
        if hasattr(component_input, 'getSid') or hasattr(component_input, 'getOwnedLogicalComponents'):
            java_obj = component_input
            # Check if this object matches the ID
            if hasattr(java_obj, 'getSid') and java_obj.getSid() == component_id:
                return java_obj

            # Check children recursively
            if hasattr(java_obj, 'getOwnedLogicalComponents'):
                children = java_obj.getOwnedLogicalComponents()
                for child in children:
                    found = find_component_by_id(child, component_id)
                    if found:
                        return found
            return None

        # Case 2: component_input is a list of components
        elif hasattr(component_input, '__iter__'):
            for component in component_input:
                try:
                    if component.get_sid() == component_id:
                        return component
                except AttributeError:
                    if hasattr(component, 'get_java_object') and component.get_java_object().getSid() == component_id:
                        return component
            return None

        # Unknown input type
        return None
    except Exception as e:
        print(f"Error finding component by ID {component_id}: {e}")
        return None
def get_component_by_id(java_object, component_id):
    """Find a component by its ID in the children of a Java object."""
    for child in java_object.getOwnedLogicalComponents():
        if child.getSid() == component_id:
            return child
    return None

def find_component_recursive(java_object, component_id):
    """Recursively find a component by its ID in the hierarchy."""
    # First check direct children
    component = get_component_by_id(java_object, component_id)
    if component:
        return component

    # If not found, recursively check children
    children = java_object.getOwnedLogicalComponents()
    for child in children:
        found = find_component_recursive(child, component_id)
        if found:
            return found

    return None

def find_function_by_id(parent, function_id):
    """Find a function by ID in the hierarchy"""
    if parent.get_sid() == function_id:
        return parent

    for function in parent.get_owned_functions():
        if function.get_sid() == function_id:
            return function
        found = find_function_by_id(function, function_id)
        if found:
            return found
    return None

def set_function_kind(function, kind):
    """Set the kind of function using the proper Function.set_kind() method"""
    try:
        # Convert to uppercase to match the expected values
        kind = kind.upper()

        # Use the set_kind method from the Function class
        function.set_kind(kind)
        return True
    except Exception as e:
        print(f"Error setting function kind for {function.get_name()}: {e}")
        return False

def get_or_create_rlf(lf_pkg):
    """Get or create the Root Logical Function"""
    # Check if the root logical function exists
    rlf = None
    for lf in lf_pkg.get_owned_logical_functions():
        if lf.get_name() == "Root Logical Function":
            rlf = lf
            break

    # If RLF doesn't exist, create it
    if rlf is None:
        print("Creating Root Logical Function")
        rlf = LogicalFunction()
        lf_pkg.get_owned_logical_functions().add(rlf)
        rlf.set_name("Root Logical Function")
        # Set an ID for RLF if needed
        rlf.get_java_object().setSid("RLF_ID")
    return rlf

# Global variable to store the log path
deletion_log_path = None

def create_deletion_log_workbook():
    """Create a new workbook with all required deletion log sheets"""
    wb = Workbook()

    # Remove default sheet if it exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create worksheets with headers
    systems_sheet = wb.create_sheet("Systems")
    systems_sheet.append(["Timestamp", "Element Type", "Element ID", "Element Name", "Parent ID", "Parent Name"])

    functions_sheet = wb.create_sheet("Functions")
    functions_sheet.append(["Timestamp", "Element Type", "Element ID", "Element Name", "Parent ID", "Parent Name", "Function Kind"])

    allocations_sheet = wb.create_sheet("Allocations")
    allocations_sheet.append(["Timestamp", "Element Type", "System ID", "System Name", "Function ID", "Function Name"])

    functional_exchanges_sheet = wb.create_sheet("FunctionalExchanges")
    functional_exchanges_sheet.append(["Timestamp", "Element Type", "Exchange ID", "Exchange Name",
                                      "Source Function ID", "Source Function Name",
                                      "Target Function ID", "Target Function Name"])

    component_exchanges_sheet = wb.create_sheet("ComponentExchanges")
    component_exchanges_sheet.append(["Timestamp", "Element Type", "Exchange ID", "Exchange Name",
                                     "Source Component ID", "Source Component Name",
                                     "Target Component ID", "Target Component Name"])

    link_exchanges_sheet = wb.create_sheet("LinkExchanges")
    link_exchanges_sheet.append(["Timestamp", "Element Type", "Component Exchange ID", "Component Exchange Name",
                                "Functional Exchange ID", "Functional Exchange Name", "Allocation ID"])

    functional_chains_sheet = wb.create_sheet("FunctionalChains")
    functional_chains_sheet.append(["Timestamp", "Element Type", "Chain ID", "Chain Name"])

    capabilities_sheet = wb.create_sheet("Capabilities")
    capabilities_sheet.append(["Timestamp", "Element Type", "Capability ID", "Capability Name"])

    return wb

def setup_deletion_log(aird_path):
    """
    Set up a deletion log file using the exact approach you specified
    """
    try:
        # Create timestamp for filename
        now = datetime.now()
        date_str = now.strftime("%d_%m_%Y")
        time_str = now.strftime("%H_%M_%S")

        # Open model to get project information
        model = CapellaModel()
        model.open(aird_path)
        se = model.get_system_engineering()

        # Get project information using your exact code
        model_path = CapellaPlatform.getModelPath(se)
        project_name = model_path[0:(model_path.index("/", 1) + 1)]
        project_name = project_name.replace("/", "")
        project = CapellaPlatform.getProject(project_name)

        # Create the Error Log workbook
        log_wb = create_deletion_log_workbook()
        log_filename = f"{project_name}_Deleted_Items_{date_str}_{time_str}.xlsx"

        # Try to get the Error Log folder
        try:
            folder = CapellaPlatform.getFolder(project, "Deleted Log")
            log_path = CapellaPlatform.getAbsolutePath(folder) + "/" + log_filename
        except:
            # If Error Log folder doesn't exist, create it using CapellaPlatform
            try:
                # Create the folder through CapellaPlatform
                error_log_folder = project.getFolder("Error Log")
                if not error_log_folder.exists():
                    error_log_folder.create(True, True, None)
                log_path = CapellaPlatform.getAbsolutePath(error_log_folder) + "/" + log_filename
            except Exception as e:
                print(f"Could not create Error Log folder: {e}")
                # Fall back to project root
                log_path = CapellaPlatform.getAbsolutePath(project) + "/" + log_filename

        # Save the workbook
        log_wb.save(log_path)

        # Verify the file was created
        if not os.path.exists(log_path):
            print(f"ERROR: File was not created at {log_path}")
            return None

        # Refresh the project to make the folder visible in Capella
        try:
            CapellaPlatform.refresh(project)
        except Exception as e:
            print(f"Could not refresh project: {e}")
            print("You may need to manually refresh the project to see the Error Log folder")

        # Return the log workbook and path
        return {
            "path": log_path,
            "wb": log_wb
        }, project

    except Exception as e:
        print(f"Error setting up deletion log: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_deletion_log(wb):
    """Get the deletion log workbook and path"""
    if hasattr(wb, '_deletion_log'):
        return wb._deletion_log.get("wb"), wb._deletion_log.get("path")
    return None, None

def log_deletion(wb, sheet_name, element_type, **kwargs):
    """
    Generic function to log any deletion
    kwargs should contain the column values (element_id, element_name, etc.)
    """
    try:
        log_wb, log_path = get_deletion_log(wb)
        if not log_wb or not log_path:
            return False

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if sheet_name not in log_wb.sheetnames:
            print(f"Warning: Sheet {sheet_name} not found in deletion log")
            return False

        ws = log_wb[sheet_name]

        # Build the row data in the correct order
        # The order should match the headers defined in setup_import_deletion_log
        row_data = [timestamp, element_type]

        # Add the rest of the data based on the sheet type
        if sheet_name == "Systems":
            row_data.extend([
                kwargs.get("element_id"),
                kwargs.get("element_name"),
                kwargs.get("parent_id"),
                kwargs.get("parent_name")
            ])
        elif sheet_name == "Functions":
            row_data.extend([
                kwargs.get("element_id"),
                kwargs.get("element_name"),
                kwargs.get("parent_id"),
                kwargs.get("parent_name"),
                kwargs.get("function_kind", "N/A")
            ])
        elif sheet_name == "Allocations":
            row_data.extend([
                kwargs.get("system_id"),
                kwargs.get("system_name"),
                kwargs.get("function_id"),
                kwargs.get("function_name")            
            ])
        elif sheet_name == "Functional Exchanges":
            row_data.extend([
                kwargs.get("exchange_id"),
                kwargs.get("fe_name"),
                kwargs.get("source_func_id"),
                kwargs.get("source_func_name"),
                kwargs.get("target_func_id"),
                kwargs.get("target_func_name"),
            ])
        elif sheet_name == "Component Exchanges":
            row_data.extend([
                kwargs.get("exchange_id"),
                kwargs.get("fe_name"),
                kwargs.get("source_func_id"),
                kwargs.get("source_func_name"),
                kwargs.get("target_func_id"),
                kwargs.get("target_func_name"),
            ])
        elif sheet_name == "LinkExchanges":
            row_data.extend([
                kwargs.get("ce_id"),
                kwargs.get("ce_name"),
                kwargs.get("fe_id"),
                kwargs.get("fe_name"),
                kwargs.get("allocation_id"),
            ])
        elif sheet_name == "FunctionalChains":
            row_data.extend([
                kwargs.get("chain_id"),
                kwargs.get("chain_name")
            ])
        elif sheet_name == "Capabilities":
            row_data.extend([
                kwargs.get("cr_id"),
                kwargs.get("cr_name")                
            ])
            
        ws.append(row_data)
        log_wb.save(log_path)
        return True

    except Exception as e:
        print(f"Error logging deletion: {e}")
        return False

def collect_all_component_ids(component):
    """Recursively collect all component IDs from a component and its children"""
    component_ids = set()

    try:
        # Add the current component's ID
        component_ids.add(get_component_id(component))

        # Recursively collect IDs from children
        java_obj = component.get_java_object()
        for child in java_obj.getOwnedLogicalComponents():
            child_component = LogicalComponent(child) if not child.isActor() else LogicalActor(child)
            component_ids.update(collect_all_component_ids(child_component))

    except Exception as e:
        print(f"Error collecting component IDs: {e}")

    return component_ids

def delete_component_and_children(wb, component):
    """Delete a component and all its children recursively, logging deletions"""
    try:
        # Get the Python wrapper for the component 
        if component.isActor():
            comp_to_delete = LogicalActor(component)
        else:
            comp_to_delete = LogicalComponent(component)

        # Collect all children to log
        all_to_delete = [comp_to_delete]
        stack = [component]

        # Collect all children recursively
        while stack:
            current = stack.pop()
            try:
                for child in current.getOwnedLogicalComponents():
                    child_comp = LogicalComponent(child) if not child.isActor() else LogicalActor(child)
                    all_to_delete.append(child_comp)
                    stack.append(child)
            except:
                continue

        # Log and delete all components
        for comp in reversed(all_to_delete):  # Delete in reverse order (children first)
            try:
                # Get component info for logging
                component_id = comp.get_sid()
                component_name = comp.get_name()

                # Get parent info if available
                parent = None
                parent_id = None
                parent_name = None
                if comp != comp_to_delete:  # If it's not the root component being deleted
                    java_obj = comp.get_java_object()
                    parent_java = java_obj.eContainer()
                    if parent_java:
                        parent_id = parent_java.getSid()
                        try:
                            parent = LogicalComponent(parent_java) if not parent_java.isActor() else LogicalActor(parent_java)
                            parent_name = parent.get_name()
                        except:
                            parent_name = "Unknown"

                # Delete the component
                EObject.delete_e_object(comp)

                # Log the deletion using the generic log_deletion function
                log_deletion(wb, "Systems", "System",
                           element_id=component_id,
                           element_name=component_name,
                           parent_id=parent_id,
                           parent_name=parent_name)

            except Exception as e:
                print(f"Error deleting component {get_component_id(comp) if hasattr(comp, 'get_sid') else 'unknown'}: {e}")

    except Exception as e:
        print(f"Error in delete_component_and_children: {e}")

def delete_unused_components(wb, parent_component, imported_ids):
    """Delete components that exist in the model but were not imported"""
    try:
        # Get all components from the parent
        java_parent = parent_component.get_java_object()
        children = list(java_parent.getOwnedLogicalComponents())

        # Create a list of components to delete
        to_delete = []

        for child in children:
            child_id = child.getSid()
            if child_id not in imported_ids:
                to_delete.append(child)

        # Delete the components (in reverse order to avoid index issues)
        for component in reversed(to_delete):
            try:
                delete_component_and_children(wb, component)
            except Exception as e:
                print(f"Error deleting component {component.getSid()}: {e}")

        # Recursively check children of imported components
        for child in children:
            child_id = child.getSid()
            if child_id in imported_ids:
                child_component = LogicalComponent(child) if not child.isActor() else LogicalActor(child)
                delete_unused_components(wb, child_component, imported_ids)

    except Exception as e:
        print(f"Error in delete_unused_components: {e}")

def import_systems(wb, model, lc_pkg):
    """Import systems from the Systems worksheet with deletion functionality"""
    try:
        ws = wb["Systems"]

        # Determine the maximum depth from the Excel file
        max_col = ws.max_column
        max_depth = max_col // 3

        # First, collect all existing component IDs in the model
        existing_component_ids = set()
        for component in lc_pkg.get_owned_logical_components():
            existing_component_ids.update(collect_all_component_ids(component))

        # Set to track imported component IDs (only those explicitly in the import file)
        imported_component_ids = set()

        # Create a list to keep track of parent components at each level
        parent_components = [None] * max_depth

        for row in ws.iter_rows(min_row=2):
            # Create a list of component data for each level
            components = []
            for level in range(max_depth):
                col = level * 3
                component_id = row[col].value if col < len(row) and row[col].value else None
                component_name = row[col + 1].value if col + 1 < len(row) and row[col + 1].value else None
                component_type = row[col + 2].value if col + 2 < len(row) and row[col + 2].value else None
                components.append((component_id, component_name, component_type))

            # Import the hierarchy
            for level in range(max_depth):
                component_id, component_name, component_type = components[level]

                if not component_id:
                    continue  # Skip empty entries

                if level == 0:
                    # Top level component
                    existing_component = find_component_by_id(lc_pkg.get_owned_logical_components(), component_id)
                    if existing_component is None:
                        if component_type == "Actor":
                            component = LogicalActor()
                        else:
                            component = LogicalComponent()
                        lc_pkg.get_owned_logical_components().add(component)
                        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(component.get_java_object())
                        component.set_name(component_name)
                        component.set_sid(component_id)
                    else:
                        component = existing_component

                    # Add ONLY this component's ID to imported IDs (not its children)
                    imported_component_ids.add(component_id)

                else:
                    # Child component
                    parent_component = parent_components[level - 1]
                    if parent_component is None:
                        print(f"Parent component not found for {component_name} at level {level}")
                        continue

                    java_parent = parent_component.get_java_object()
                    existing_component = get_component_by_id(java_parent, component_id)

                    if existing_component is None:
                        if component_type == "Actor":
                            component = LogicalActor()
                        else:
                            component = LogicalComponent()
                        java_parent.getOwnedLogicalComponents().add(component.get_java_object())
                        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(component.get_java_object())
                        component.set_name(component_name)
                        component.set_sid(component_id)
                    else:
                        if component_type == "Actor":
                            component = LogicalActor(existing_component)
                        else:
                            component = LogicalComponent(existing_component)

                    # Add ONLY this component's ID to imported IDs (not its children)
                    imported_component_ids.add(component_id)

                parent_components[level] = component

        # After import, check for components that need to be deleted
        unused_ids = existing_component_ids - imported_component_ids
        deleted_systems_count = len(unused_ids)

        if unused_ids:
            print(f"Found {deleted_systems_count} components that were not imported and will be deleted")
            delete_unused_components(wb, lc_pkg, imported_component_ids)

        print(f"Imported Systems worksheet successfully")
        return True, deleted_systems_count
    except Exception as e:
        print(f"Error importing systems: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 0
    
def collect_all_function_ids(function):
    """Recursively collect all function IDs from a function and its children"""
    function_ids = set()

    try:
        # Add the current function's ID
        function_ids.add(function.get_sid())

        # Recursively collect IDs from children
        for child in function.get_owned_functions():
            function_ids.update(collect_all_function_ids(child))

    except Exception as e:
        print(f"Error collecting function IDs: {e}")

    return function_ids

def delete_function_and_children(wb, function):
    """Delete a function and all its children recursively, logging deletions"""
    try:
        # Collect all children to log and delete
        all_to_delete = [function]
        stack = [function]

        # Collect all children recursively
        while stack:
            current = stack.pop()
            try:
                for child in current.get_owned_functions():
                    all_to_delete.append(child)
                    stack.append(child)
            except:
                continue

        # Log and delete all functions in reverse order (children first)
        for func in reversed(all_to_delete):
            try:
                # Get function info for logging
                function_id = func.get_sid()
                function_name = func.get_name()

                # Get function kind
                function_kind = "N/A"
                try:
                    function_kind = func.get_kind()
                except:
                    pass

                # Get parent info if available
                parent = None
                parent_id = None
                parent_name = None
                if func != function:  # If it's not the root function being deleted
                    try:
                        parent = func.eContainer()
                        if hasattr(parent, 'get_java_object'):
                            parent = parent.get_java_object()
                        if hasattr(parent, 'getSid'):
                            parent_id = parent.getSid()
                            parent_func = LogicalFunction(parent)
                            parent_name = parent_func.get_name()
                    except:
                        pass

                # Log the deletion using the generic log_deletion function
                log_deletion(wb, "Functions", "Function",
                           element_id=function_id,
                           element_name=function_name,
                           parent_id=parent_id,
                           parent_name=parent_name,
                           function_kind=function_kind)

                # Delete the function
                EObject.delete_e_object(func)

            except Exception as e:
                print(f"Error deleting function {func.get_sid() if hasattr(func, 'get_sid') else 'unknown'}: {e}")

    except Exception as e:
        print(f"Error in delete_function_and_children: {e}")

def delete_unused_functions(wb, parent_function, imported_ids):
    """Delete functions that exist in the model but were not imported"""
    try:
        # Get all functions from the parent
        children = list(parent_function.get_owned_functions())

        # Create a list of functions to delete
        to_delete = []

        for child in children:
            child_id = child.get_sid()
            if child_id not in imported_ids:
                to_delete.append(child)

        # Delete the functions (in reverse order to avoid index issues)
        for function in reversed(to_delete):
            try:
                delete_function_and_children(wb, function)
            except Exception as e:
                print(f"Error deleting function {function.get_sid()}: {e}")

        # Recursively check children of imported functions
        for child in children:
            child_id = child.get_sid()
            if child_id in imported_ids:
                delete_unused_functions(wb, child, imported_ids)

    except Exception as e:
        print(f"Error in delete_unused_functions: {e}")

def import_functions(wb, model, lf_pkg):
    """Import functions from the Functions worksheet with deletion functionality"""
    try:
        ws = wb["Functions"]

        # Determine the maximum depth from the Excel file
        max_col = ws.max_column
        max_depth = max_col // 3  # 3 columns per level (ID, Name, Kind)

        # Get or create RLF
        rlf = get_or_create_rlf(lf_pkg)

        # First, collect all existing function IDs in the model
        existing_function_ids = set()
        if rlf:
            existing_function_ids.update(collect_all_function_ids(rlf))

        # Set to track imported function IDs
        imported_function_ids = set()
        imported_function_ids.add(rlf.get_sid())  # Add RLF to imported IDs

        # Create a list to keep track of parent functions at each level
        parent_functions = [None] * max_depth

        for row in ws.iter_rows(min_row=2):
            # Create a list of function data for each level
            functions = []
            for level in range(max_depth):
                col = level * 3  # 3 columns per level (ID, Name, Kind)
                function_id = row[col].value if col < len(row) and row[col].value else None
                function_name = row[col + 1].value if col + 1 < len(row) and row[col + 1].value else None
                function_kind = row[col + 2].value if col + 2 < len(row) and row[col + 2].value else None
                functions.append((function_id, function_name, function_kind))

            # Import the hierarchy
            for level in range(max_depth):
                function_id, function_name, function_kind = functions[level]

                if not function_id or not function_name:
                    continue  # Skip empty entries

                # For level 0, parent is RLF
                if level == 0:
                    parent_function = rlf
                else:
                    parent_function = parent_functions[level - 1]

                if parent_function is None:
                    print(f"Parent function not found for {function_name} at level {level+1}")
                    continue

                # Check if the function already exists
                existing_function = find_function_by_id(parent_function, function_id)

                if existing_function is None:
                    function = LogicalFunction()
                    parent_function.get_owned_functions().add(function)
                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(function.get_java_object())
                    function.set_name(function_name)
                    function.get_java_object().setSid(function_id)

                    # Set the function kind if provided
                    if function_kind:
                        success = set_function_kind(function, function_kind)
                        if not success:
                            print(f"Failed to set function kind to {function_kind} for {function_name}")
                    else:
                        print(f"No function kind specified for {function_name}, defaulting to FUNCTION")
                else:
                    function = existing_function

                    # Update the function kind if provided
                    if function_kind:
                        success = set_function_kind(function, function_kind)
                        if not success:
                            print(f"Failed to update function kind to {function_kind} for {function_name}")

                # Add to imported IDs
                imported_function_ids.add(function_id)

                # Only set parent for next level if this is not the last level
                if level < max_depth - 1:
                    parent_functions[level] = function

        # After import, check for functions that need to be deleted
        unused_ids = existing_function_ids - imported_function_ids
        deleted_functions_count = len(unused_ids)
        
        if deleted_functions_count > 0:
            print(f"Found {deleted_functions_count} functions that were not imported and will be deleted")
    
            # Delete unused functions from the root logical function
            if rlf:
                delete_unused_functions(wb, rlf, imported_function_ids)

        print(f"Imported Functions worksheet successfully")
        return True, deleted_functions_count
    except Exception as e:
        print(f"Error importing functions: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 0

def collect_all_system_allocations(lc_pkg):
    """Collect all existing system-function allocations in the model"""
    allocations = {}

    try:
        # Get all components
        java_lc_pkg = lc_pkg.get_java_object()
        components = [java_lc_pkg.getOwnedLogicalComponents()]

        while components:
            current_level = components.pop()
            for component in current_level:
                # Check for allocations
                allocated_functions = component.getAllocatedLogicalFunctions()
                for func in allocated_functions:
                    system_id = component.getSid()
                    function_id = func.getSid()
                    if system_id not in allocations:
                        allocations[system_id] = []
                    allocations[system_id].append(function_id)

                # Add children to the processing queue
                children = component.getOwnedLogicalComponents()
                if children:
                    components.append(children)

        return allocations
    except Exception as e:
        print(f"Error collecting allocations: {e}")
        return {}

def delete_unused_system_allocations(wb, imported_allocations, lc_pkg):
    """Delete allocations that exist in the model but were not imported"""
    try:
        # Get all components
        java_lc_pkg = lc_pkg.get_java_object()
        components = [java_lc_pkg.getOwnedLogicalComponents()]

        deleted_count = 0

        while components:
            current_level = components.pop()
            for component in current_level:
                # Get the system name for logging
                system_name = "Unknown"
                try:
                    if component.isActor():
                        system = LogicalActor(component)
                    else:
                        system = LogicalComponent(component)
                    system_name = system.get_name()
                except:
                    pass

                # Check for allocations to delete
                allocated_functions = list(component.getAllocatedLogicalFunctions())
                for func in allocated_functions:
                    system_id = component.getSid()
                    function_id = func.getSid()

                    # Check if this allocation should be kept
                    if system_id in imported_allocations and function_id in imported_allocations[system_id]:
                        continue

                    # This allocation is not in the import file, delete it
                    try:
                        # Get function name for logging
                        function_name = "Unknown"
                        try:
                            function = LogicalFunction(func)
                            function_name = function.get_name()
                        except:
                            pass

                        # Log the deletion                        
                        log_deletion(wb, "Allocations", "Function-System Allocation",
                            system_id=system_id,
                            system_name=system_name,
                            function_id=function_id,
                            function_name=function_name)

                        # Find the allocation object
                        allocations = component.getOwnedFunctionalAllocation()
                        for allocation in allocations:
                            if allocation.getTargetElement().getSid() == function_id:
                                component.getOwnedFunctionalAllocation().remove(allocation)
                                deleted_count += 1
                                break
                    except Exception as e:
                        print(f"Error deleting allocation between {system_name} and function {function_id}: {str(e)}")

                # Add children to the processing queue
                children = component.getOwnedLogicalComponents()
                if children:
                    components.append(children)

        return deleted_count

    except Exception as e:
        print(f"Error in delete_unused_system_allocations: {str(e)}")
        return 0

def import_link_functions_systems(wb, model, lc_pkg, lf_pkg, rlf):
    """Import links between systems and functions from the Link Systems and Functions worksheet with deletion functionality"""
    try:
        ws = wb["Link Systems and Functions"]

        # Find the column where "Function ID" appears to determine system depth
        function_id_col = None
        for cell in ws[1]:
            if cell.value == "Function ID":
                function_id_col = cell.column
                break

        if function_id_col is None:
            print("Could not find 'Function ID' column in the Excel file.")
            return False

        # Calculate system depth
        system_max_depth = (function_id_col - 1) // 2

        # First collect all existing allocations
        existing_allocations = collect_all_system_allocations(lc_pkg)

        # Dictionary to store imported allocations: {system_id: [function_ids]}
        imported_allocations = {}

        # Store the current system path for each row
        current_system_java = None
        current_system_id = None
        current_system_level = 0

        for row in ws.iter_rows(min_row=2):
            # Check if this row contains a system (has value in any system column)
            system_found = False
            system_id = None
            system_col = -1

            # Find which system column has a value
            for level in range(system_max_depth):
                col = level * 2
                if col < len(row) and row[col].value:
                    system_id = row[col].value
                    system_col = col
                    system_found = True
                    break  # Use the first system found (highest level)

            # If we found a system, determine its level and find it
            if system_found:
                # Level is determined by column position: col 0-1 = level 0, col 2-3 = level 1, etc.
                current_system_level = system_col // 2
                current_system_id = system_id

                # Find the system using recursive search
                java_lc_pkg = lc_pkg.get_java_object()
                current_system_java = find_component_recursive(java_lc_pkg, current_system_id)

                if not current_system_java:
                    print(f"System with ID {current_system_id} not found at level {current_system_level + 1}")
                    continue

                # Create a Python wrapper for display purposes
                if current_system_java.isActor():
                    current_system = LogicalActor(current_system_java)
                else:
                    current_system = LogicalComponent(current_system_java)

            # If this row contains a function (has value in function ID column)
            if function_id_col <= len(row) and row[function_id_col-1].value:
                function_id = row[function_id_col-1].value
                function_name = row[function_id_col].value if function_id_col < len(row) else ""

                if current_system_java:
                    # Find the function
                    function = find_function_by_id(rlf, function_id)

                    if function:
                        try:
                            # Check if the allocation already exists
                            allocated_functions = current_system_java.getAllocatedLogicalFunctions()
                            already_allocated = False
                            for allocated_func in allocated_functions:
                                if allocated_func.getSid() == function_id:
                                    already_allocated = True
                                    break

                            if not already_allocated:
                                # Create the allocation using the exact same method as the working script
                                try:
                                    # Use the Capella Java API directly
                                    cfa = org.polarsys.capella.core.data.fa.FaFactory.eINSTANCE.createComponentFunctionalAllocation()
                                    cfa.setSourceElement(current_system_java)
                                    cfa.setTargetElement(function.get_java_object())

                                    # Add to the correct containment feature
                                    current_system_java.getOwnedFunctionalAllocation().add(cfa)

                                    # Create the service
                                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(cfa)

                                except Exception as e:
                                    print(f"Failed to create allocation: {str(e)}")
                                    print(f"Failed to allocate function '{function.get_name()}' to system '{current_system.get_name()}': {str(e)}")

                            # Add to imported allocations
                            if current_system_id not in imported_allocations:
                                imported_allocations[current_system_id] = []
                            imported_allocations[current_system_id].append(function_id)
                        except Exception as e:
                            print(f"Failed to check existing allocations for function '{function.get_name()}' to system '{current_system.get_name()}': {str(e)}")
                    else:
                        print(f"Function with ID {function_id} and name {function_name} not found.")
                else:
                    print("No current system defined for function allocation.")

        # After import, delete unused allocations
        deleted_count = delete_unused_system_allocations(wb, imported_allocations, lc_pkg)
        if deleted_count > 0:
            print(f"Deleted {deleted_count} allocations that were not imported and will be deleted")

        print(f"Imported Link Systems and Functions worksheet successfully")
        return True, deleted_count
    except Exception as e:
        print(f"Error importing link functions systems: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 0

def find_functional_exchange_by_id_existance(exchange_id, se):
    """Find a functional exchange by its ID in the model"""
    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for fe in all_exchanges:
            if fe.get_sid() == exchange_id:
                return fe
        return None
    except Exception as e:
        print(f"Error finding functional exchange by ID {exchange_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def find_or_create_function_port(function, is_input, port_name, port_id):
    """Find an existing function port or create a new one"""
    try:
        if is_input:
            # Check existing input ports
            for port in function.get_inputs():
                if port.get_sid() == port_id:
                    return port
            # Create new input port if not found
            port = FunctionInputPort()
            function.get_inputs().add(port)
            port.set_name(port_name)
            port.get_java_object().setSid(port_id)
            return port
        else:
            # Check existing output ports
            for port in function.get_outputs():
                if port.get_sid() == port_id:
                    return port
            # Create new output port if not found
            port = FunctionOutputPort()
            function.get_outputs().add(port)
            port.set_name(port_name)
            port.get_java_object().setSid(port_id)
            return port
    except Exception as e:
        print(f"Error finding or creating function port {port_name}: {e}")
        return None

def check_functional_exchange_exists(exchange_id, se):
    """Check if a functional exchange with the given ID already exists"""
    try:
        # Find the functional exchange by ID
        existing_fe = find_functional_exchange_by_id_existance(exchange_id, se)
        if existing_fe:
            return True
        return False
    except Exception as e:
        print(f"Error checking if functional exchange exists: {e}")
        return False
    
def log_exchange_deletion(exchange_id, exchange_name, source_comp_id, source_comp_name, target_comp_id, target_comp_name):
    """Log a component exchange deletion event"""
    try:
        log_path = get_deletion_log_path()
        if not log_path or not os.path.exists(log_path):
            return

        wb = load_workbook(log_path)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Get or create the component exchanges sheet
        if "FunctionalExchanges" not in wb.sheetnames:
            ws = wb.create_sheet("FunctionalExchanges")
            ws.append(["Timestamp", "Element Type", "Exchange ID", "Exchange Name",
                      "Source Function ID", "Source Function Name",
                      "Target Function ID", "Target Function Name"])
        else:
            ws = wb["FunctionalExchanges"]

        # Add the deletion record
        ws.append([timestamp, "FunctionalExchange", exchange_id, exchange_name,
                  source_func_id, source_func_name, target_func_id, target_func_name])
        wb.save(log_path)
    except Exception as e:
        print(f"Error logging exchange deletion: {e}")

def collect_all_exchanges(se):
    """Collect all existing functional exchanges in the model"""
    exchanges = {}

    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for fe in all_exchanges:
            source_port = fe.get_source_port()
            target_port = fe.get_target_port()

            if source_port and target_port:
                source_func = source_port.get_java_object().eContainer()
                target_func = target_port.get_java_object().eContainer()

                if source_func and target_func:
                    exchange_id = fe.get_sid()
                    source_func_id = source_func.getSid()
                    target_func_id = target_func.getSid()

                    if source_func_id not in exchanges:
                        exchanges[source_func_id] = {}
                    if target_func_id not in exchanges[source_func_id]:
                        exchanges[source_func_id][target_func_id] = []

                    exchanges[source_func_id][target_func_id].append(exchange_id)

        return exchanges
    except Exception as e:
        print(f"Error collecting existing exchanges: {e}")
        return {}

def delete_unused_exchanges(wb, imported_exchanges, se):
    """Delete exchanges that exist in the model but were not imported"""
    try:
        deleted_count = 0

        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)

        for fe in all_exchanges:
            source_port = fe.get_source_port()
            target_port = fe.get_target_port()

            if not source_port or not target_port:
                continue

            try:
                source_func = source_port.get_java_object().eContainer()
                target_func = target_port.get_java_object().eContainer()

                if not source_func or not target_func:
                    continue

                source_func_id = source_func.getSid()
                target_func_id = target_func.getSid()
                exchange_id = fe.get_sid()

                # Check if this exchange should be kept
                if (source_func_id in imported_exchanges and
                    target_func_id in imported_exchanges[source_func_id] and
                    exchange_id in imported_exchanges[source_func_id][target_func_id]):
                    continue  # This exchange is in the import file

                # Get function names for logging
                source_func_name = "Unknown"
                target_func_name = "Unknown"
                fe_name = fe.get_name()

                try:
                    source_func_obj = LogicalFunction(source_func)
                    source_func_name = source_func_obj.get_name()
                except:
                    pass

                try:
                    target_func_obj = LogicalFunction(target_func)
                    target_func_name = target_func_obj.get_name()
                except:
                    pass

                # Log the deletion if log_path is provided
                log_deletion(wb, "FunctionalExchanges", "Functional Exchange",
                    exchange_id=exchange_id, fe_name=fe_name,
                    source_func_id=source_func_id, source_func_name=source_func_name,
                    target_func_id=target_func_id, target_func_name=target_func_name)
                
                EObject.delete_e_object(fe)
                deleted_count += 1

            except Exception as e:
                print(f"Error processing exchange {fe.get_sid()}: {e}")

        return deleted_count

    except Exception as e:
        print(f"Error in delete_unused_exchanges: {e}")
        return 0    

def import_functional_exchanges(wb, model, se, lf_pkg):
    """Import functional exchanges from the Functional Exchanges worksheet with deletion functionality"""
    try:
        ws = wb["Functional Exchanges"]

        # Find the Root Logical Function
        rlf = None
        for lf in lf_pkg.get_owned_logical_functions():
            if lf.get_name() == "Root Logical Function":
                rlf = lf
                break

        if rlf is None:
            print("Root Logical Function not found!")
            return False

        # First collect all existing exchanges
        existing_exchanges = collect_all_exchanges(se)

        # Dictionary to store imported exchanges: {source_func_id: {target_func_id: [exchange_ids]}}
        imported_exchanges = {}

        # Process each row in the Excel file
        for row in ws.iter_rows(min_row=2):
            # Extract data from the row
            source_func_id = row[0].value
            source_func_name = row[1].value
            source_port_id = row[2].value
            source_port_name = row[3].value
            fe_id = row[4].value
            fe_name = row[5].value
            target_func_id = row[6].value
            target_func_name = row[7].value
            target_port_id = row[8].value
            target_port_name = row[9].value

            # Check if the functional exchange already exists by ID
            existing_fe = find_functional_exchange_by_id_existance(fe_id, se)
            if existing_fe:
                # Add to imported exchanges
                if source_func_id not in imported_exchanges:
                    imported_exchanges[source_func_id] = {}
                if target_func_id not in imported_exchanges[source_func_id]:
                    imported_exchanges[source_func_id][target_func_id] = []
                imported_exchanges[source_func_id][target_func_id].append(fe_id)
                continue

            # Find source and target functions
            source_function = find_function_by_id(rlf, source_func_id)
            target_function = find_function_by_id(rlf, target_func_id)

            if not source_function or not target_function:
                print(f"Could not find source or target function for exchange {fe_name}")
                continue

            # Find or create source port (output port for source function)
            source_port = find_or_create_function_port(source_function, False, source_port_name, source_port_id)
            if not source_port:
                print(f"Failed to create/get source port {source_port_name}")
                continue

            # Find or create target port (input port for target function)
            target_port = find_or_create_function_port(target_function, True, target_port_name, target_port_id)
            if not target_port:
                print(f"Failed to create/get target port {target_port_name}")
                continue

            # Create the functional exchange
            fe = FunctionalExchange()

            # Set source and target ports
            fe.set_source_port(source_port)
            fe.set_target_port(target_port)

            source_function.get_owned_functional_exchanges().add(fe)
            org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fe.get_java_object())

            fe.set_name(fe_name)
            fe.get_java_object().setSid(fe_id)

            # Add to imported exchanges
            if source_func_id not in imported_exchanges:
                imported_exchanges[source_func_id] = {}
            if target_func_id not in imported_exchanges[source_func_id]:
                imported_exchanges[source_func_id][target_func_id] = []
            imported_exchanges[source_func_id][target_func_id].append(fe_id)

        # After import, delete unused exchanges
        deleted_count = delete_unused_exchanges(wb, imported_exchanges, se)
        if deleted_count > 0:
            print(f"Deleted {deleted_count} exchanges that were not imported and will be deleted")

        print(f"Imported Functional Exchanges worksheet successfully")
        return True, deleted_count
    except Exception as e:
        print(f"Error importing functional exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 0 

def create_component_from_java(java_obj):
    """Create the appropriate component type from a Java object"""
    try:
        if hasattr(java_obj, 'isActor') and java_obj.isActor():
            return LogicalActor(java_obj)
        else:
            return LogicalComponent(java_obj)
    except Exception as e:
        print(f"Error creating component: {e}")
        return None

def find_or_create_component_port(component, port_id, port_name, port_direction, port_kind):
    """Find an existing component port or create a new one with the specified direction and kind"""
    try:
        java_obj = component.get_java_object()

        # Check existing ports
        for port in java_obj.getOwnedFeatures():
            if hasattr(port, 'getSid') and port.getSid() == port_id:
                # Found existing port, update its direction and kind
                set_port_direction(ComponentPort(port), port_direction)
                set_port_kind(ComponentPort(port), port_kind)
                return ComponentPort(port)

        # Create new port if not found
        port = ComponentPort()
        java_obj.getOwnedFeatures().add(port.get_java_object())
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(port.get_java_object())

        port.set_name(port_name)
        port.get_java_object().setSid(port_id)

        # Set the port direction
        set_port_direction(port, port_direction)

        # Set the port kind
        set_port_kind(port, port_kind)

        return port
    except Exception as e:
        print(f"Error finding or creating component port {port_name}: {e}")
        return None

def set_port_direction(port, direction):
    """Set the direction of a port using the correct method"""
    try:
        # Use the correct method to set orientation
        if direction == "IN":
            port.set_orientation("IN")
        elif direction == "OUT":
            port.set_orientation("OUT")
        elif direction == "INOUT":
            port.set_orientation("INOUT")
        elif direction == "UNSET":
            port.set_orientation("UNSET")
        else:
            print(f"Unknown port direction: {direction}")
    except Exception as e:
        print(f"Error setting port direction: {e}")

def set_port_kind(port, kind):
    """Set the kind of a port (STANDARD or FLOW)"""
    try:
        # Set the kind based on the input
        if kind == "STANDARD":
            port.set_kind("STANDARD")
        elif kind == "FLOW":
            port.set_kind("FLOW")
        else:
            # Default to STANDARD if unknown kind
            port.set_kind("STANDARD")
    except Exception as e:
        print(f"Error setting port kind: {e}")

def get_ancestor_path(component):
    """Get the path from a component to the root as a list of component IDs"""
    path = []
    current = component.get_java_object()
    while current:
        try:
            # Add the current component's ID to the path
            path.append(current.getSid())
            container = current.eContainer()
            if container and hasattr(container, 'getSid'):
                current = container
            else:
                current = None
        except:
            current = None
    return path

def find_common_ancestor(component1, component2, lc_pkg):
    """Find the common ancestor of two components by comparing IDs"""
    try:
        # Get the paths from each component to the root as lists of IDs
        path1 = get_ancestor_path(component1)
        path2 = get_ancestor_path(component2)

        # Find all common ancestor IDs
        common_ancestors = []
        for i in range(len(path1)):
            for j in range(len(path2)):
                if path1[i] == path2[j]:
                    common_ancestors.append(path1[i])

        # If no common ancestors found, use the logical component package
        if not common_ancestors:
            return lc_pkg

        # Use the first common ancestor as you specified
        common_ancestor_id = common_ancestors[0]

        # Find the component with the common ancestor ID
        def find_component_by_id_recursive(java_object, target_id):
            try:
                # Check if this object matches the ID
                if java_object.getSid() == target_id:
                    return java_object

                # Check children
                try:
                    children = java_object.getOwnedLogicalComponents()
                    for child in children:
                        found = find_component_by_id_recursive(child, target_id)
                        if found:
                            return found
                except:
                    pass
            except:
                pass
            return None

        # Search for the component with the common ancestor ID
        common_ancestor_java = find_component_by_id_recursive(lc_pkg.get_java_object(), common_ancestor_id)

        if common_ancestor_java:
            # Create a component object from the Java object
            try:
                # Try to create a LogicalComponent first
                return LogicalComponent(common_ancestor_java)
            except:
                try:
                    # If that fails, try to create a LogicalActor
                    return LogicalActor(common_ancestor_java)
                except:
                    # If both fail, return the logical component package as fallback
                    return lc_pkg
        else:
            # Fallback to logical component package if we can't find the component
            return lc_pkg

    except Exception as e:
        print(f"Error finding common ancestor: {e}")
        return lc_pkg

def log_component_exchange_deletion(exchange_id, exchange_name, source_comp_id, source_comp_name, target_comp_id, target_comp_name):
    """Log a component exchange deletion event"""
    try:
        log_path = get_deletion_log_path()
        if not log_path or not os.path.exists(log_path):
            return

        wb = load_workbook(log_path)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Get or create the component exchanges sheet
        if "ComponentExchanges" not in wb.sheetnames:
            ws = wb.create_sheet("ComponentExchanges")
            ws.append(["Timestamp", "Element Type", "Exchange ID", "Exchange Name",
                      "Source Component ID", "Source Component Name",
                      "Target Component ID", "Target Component Name"])
        else:
            ws = wb["ComponentExchanges"]

        # Add the deletion record
        ws.append([timestamp, "ComponentExchange", exchange_id, exchange_name,
                  source_comp_id, source_comp_name, target_comp_id, target_comp_name])
        wb.save(log_path)
    except Exception as e:
        print(f"Error logging component exchange deletion: {e}")

def collect_all_component_exchanges(lc_pkg):
    """Collect all existing component exchanges in the model"""
    exchanges = {}

    try:
        # Start from the logical component package
        def search_recursive(java_object):
            try:
                # Check component exchanges owned by this object
                for ce in java_object.getOwnedComponentExchanges():
                    source_port = ce.getSource()
                    target_port = ce.getTarget()

                    if source_port and target_port:
                        source_comp = source_port.eContainer()
                        target_comp = target_port.eContainer()

                        if source_comp and target_comp:
                            exchange_id = ce.getSid()
                            source_comp_id = source_comp.getSid()
                            target_comp_id = target_comp.getSid()

                            if source_comp_id not in exchanges:
                                exchanges[source_comp_id] = {}
                            if target_comp_id not in exchanges[source_comp_id]:
                                exchanges[source_comp_id][target_comp_id] = []

                            exchanges[source_comp_id][target_comp_id].append(exchange_id)

                # Recursively check children
                children = java_object.getOwnedLogicalComponents()
                for child in children:
                    search_recursive(child)
            except:
                pass

        lc_pkg_java = lc_pkg.get_java_object()
        search_recursive(lc_pkg_java)

        return exchanges
    except Exception as e:
        print(f"Error collecting existing component exchanges: {e}")
        return {}

def delete_unused_component_exchanges(wb, imported_exchanges, lc_pkg):
    """Delete component exchanges that exist in the model but were not imported"""
    try:
        deleted_count = 0

        # Start from the logical component package
        def search_and_delete_recursive(java_object):
            nonlocal deleted_count
            try:
                # Get all component exchanges owned by this object
                exchanges = list(java_object.getOwnedComponentExchanges())
                for ce in exchanges:
                    try:
                        source_port = ce.getSource()
                        target_port = ce.getTarget()

                        if source_port and target_port:
                            source_comp = source_port.eContainer()
                            target_comp = target_port.eContainer()

                            if source_comp and target_comp:
                                exchange_id = ce.getSid()
                                source_comp_id = source_comp.getSid()
                                target_comp_id = target_comp.getSid()

                                # Check if this exchange should be kept
                                if (source_comp_id in imported_exchanges and
                                    target_comp_id in imported_exchanges[source_comp_id] and
                                    exchange_id in imported_exchanges[source_comp_id][target_comp_id]):
                                    continue  # This exchange is in the import file

                                # Get component names for logging
                                source_comp_name = "Unknown"
                                target_comp_name = "Unknown"
                                ce_name = ce.getName()

                                try:
                                    source_comp_obj = create_component_from_java(source_comp)
                                    if source_comp_obj:
                                        source_comp_name = source_comp_obj.get_name()
                                except:
                                    pass

                                try:
                                    target_comp_obj = create_component_from_java(target_comp)
                                    if target_comp_obj:
                                        target_comp_name = target_comp_obj.get_name()
                                except:
                                    pass

                                # Log the deletion if log_path is provided
                                log_deletion(wb, "ComponentExchanges", "Component Exchange",
                                    exchange_id=exchange_id, ce_name=ce_name,
                                    source_comp_id=source_comp_id, source_comp_name=source_comp_name,
                                    target_comp_id=target_comp_id, target_comp_name=target_comp_name)
                                
                                EObject.delete_e_object(ComponentExchange(ce))
                                deleted_count += 1

                    except Exception as e:
                        print(f"Error processing component exchange {ce.getSid()}: {e}")

                # Recursively check children
                children = java_object.getOwnedLogicalComponents()
                for child in children:
                    search_and_delete_recursive(child)
            except:
                pass

        lc_pkg_java = lc_pkg.get_java_object()
        search_and_delete_recursive(lc_pkg_java)

        return deleted_count

    except Exception as e:
        print(f"Error in delete_unused_component_exchanges: {e}")
        return 0

def import_component_exchanges(wb, model, lc_pkg):
    """Import component exchanges from the Component Exchanges worksheet with deletion functionality"""
    try:
        ws = wb["Component Exchanges"]

        # First collect all existing component exchanges
        existing_exchanges = collect_all_component_exchanges(lc_pkg)

        # Dictionary to store imported exchanges: {source_comp_id: {target_comp_id: [exchange_ids]}}
        imported_exchanges = {}

        # Process each row in the Excel file
        for row in ws.iter_rows(min_row=2):
            # Extract data from the row
            source_comp_id = row[0].value
            source_comp_name = row[1].value
            source_port_id = row[2].value
            source_port_name = row[3].value
            source_port_dir = row[4].value
            source_port_kind = row[5].value
            ce_id = row[6].value
            ce_name = row[7].value
            ce_kind = row[8].value
            target_comp_id = row[9].value
            target_comp_name = row[10].value
            target_port_id = row[11].value
            target_port_name = row[12].value
            target_port_dir = row[13].value
            target_port_kind = row[14].value

            # Find source and target components
            source_java = find_component_by_id(lc_pkg.get_java_object(), source_comp_id)
            target_java = find_component_by_id(lc_pkg.get_java_object(), target_comp_id)

            if not source_java or not target_java:
                print(f"Could not find source or target component for exchange {ce_name}")
                continue

            # Create component objects
            source_component = create_component_from_java(source_java)
            target_component = create_component_from_java(target_java)

            if not source_component or not target_component:
                print(f"Could not create component objects for exchange {ce_name}")
                continue

            # Find or create source port with direction and kind
            source_port = find_or_create_component_port(source_component, source_port_id, source_port_name, source_port_dir, source_port_kind)
            if not source_port:
                print(f"Failed to create/get source port {source_port_name}")
                continue

            # Find or create target port with direction and kind
            target_port = find_or_create_component_port(target_component, target_port_id, target_port_name, target_port_dir, target_port_kind)
            if not target_port:
                print(f"Failed to create/get target port {target_port_name}")
                continue

            # Find the common ancestor to store the component exchange
            common_ancestor = find_common_ancestor(source_component, target_component, lc_pkg)

            # Check if the component exchange already exists in the common ancestor
            existing_ce = None
            common_ancestor_java = common_ancestor.get_java_object()
            for ce in common_ancestor_java.getOwnedComponentExchanges():
                if ce.getSid() == ce_id:
                    existing_ce = ComponentExchange(ce)
                    break

            if existing_ce:
                # Add to imported exchanges
                if source_comp_id not in imported_exchanges:
                    imported_exchanges[source_comp_id] = {}
                if target_comp_id not in imported_exchanges[source_comp_id]:
                    imported_exchanges[source_comp_id][target_comp_id] = []
                imported_exchanges[source_comp_id][target_comp_id].append(ce_id)
                continue

            # Create the component exchange
            ce = ComponentExchange()

            # Set the exchange kind
            if ce_kind:
                ce.set_kind(ce_kind)

            # Set source and target ports
            ce.get_java_object().setSource(source_port.get_java_object())
            ce.get_java_object().setTarget(target_port.get_java_object())

            # Add the exchange to the common ancestor
            common_ancestor.get_java_object().getOwnedComponentExchanges().add(ce.get_java_object())
            org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(ce.get_java_object())

            # Set basic properties
            ce.set_name(ce_name)
            ce.get_java_object().setSid(ce_id)

            # Add to imported exchanges
            if source_comp_id not in imported_exchanges:
                imported_exchanges[source_comp_id] = {}
            if target_comp_id not in imported_exchanges[source_comp_id]:
                imported_exchanges[source_comp_id][target_comp_id] = []
            imported_exchanges[source_comp_id][target_comp_id].append(ce_id)

        # After import, delete unused component exchanges
        deleted_count = delete_unused_component_exchanges(wb, imported_exchanges, lc_pkg)
        if deleted_count > 0:
            print(f"Deleted {deleted_count} component exchanges that were not imported and will be deleted")

        print(f"Imported Component Exchanges worksheet successfully")
        return True, deleted_count
    except Exception as e:
        print(f"Error importing component exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 0
   
def find_component_exchange_by_id(ce_id, lc_pkg):
    """Find a component exchange by its ID in the model using recursive search"""
    try:
        # Start from the logical component package
        lc_pkg_java = lc_pkg.get_java_object()

        # Use a stack for iterative depth-first search
        stack = [lc_pkg_java]

        while stack:
            current = stack.pop()

            try:
                # Check exchanges owned by this object
                if hasattr(current, 'getOwnedComponentExchanges'):
                    exchanges = current.getOwnedComponentExchanges()
                    for ce in exchanges:
                        if ce.getSid() == ce_id:
                            return ComponentExchange(ce)

                # Add children to stack for processing
                if hasattr(current, 'getOwnedLogicalComponents'):
                    children = current.getOwnedLogicalComponents()
                    for child in reversed(children):
                        stack.append(child)
            except:
                continue

        print(f"Could not find component exchange with ID {ce_id}")
        return None
    except Exception as e:
        print(f"Error finding component exchange by ID {ce_id}: {e}")
        return None

def find_functional_exchange_by_id(exchange_id, se):
    """Find a functional exchange by its ID in the model"""
    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for fe in all_exchanges:
            if fe.get_sid() == exchange_id:
                return fe

        print(f"Could not find functional exchange with ID {exchange_id}")
        return None
    except Exception as e:
        print(f"Error finding functional exchange by ID {exchange_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def check_allocation_exists(ce, fe):
    """Check if an allocation between component exchange and functional exchange already exists"""
    try:
        # Check all allocations owned by the component exchange
        allocations = ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations()
        for allocation in allocations:
            target = allocation.getTargetElement()
            if target and target.getSid() == fe.get_sid():
                return True
        return False
    except Exception as e:
        print(f"Error checking if allocation exists: {e}")
        return False

def create_allocation(ce, fe, allocation_id):
    """Create an allocation between component exchange and functional exchange"""
    try:
        # Check if allocation already exists
        if check_allocation_exists(ce, fe):
            return None

        # Create the allocation
        allocation = ComponentExchangeFunctionalExchangeAllocation()

        # Add the allocation to the component exchange's owned allocations
        ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations().add(allocation.get_java_object())
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(allocation.get_java_object())

        # Set the ID for the allocation
        allocation.get_java_object().setSid(allocation_id)

        # Set source as component exchange and target as functional exchange
        allocation.get_java_object().setSourceElement(ce.get_java_object())
        allocation.get_java_object().setTargetElement(fe.get_java_object())

        return allocation
    except Exception as e:
        print(f"Error creating allocation: {e}")
        import traceback
        traceback.print_exc()
        return None

def synchronize_port_allocations(ce, fe):
    """Synchronize port allocations between component exchange and functional exchange"""
    try:
        # Use the Capella API method to synchronize port allocations
        org.polarsys.capella.core.model.helpers.ComponentExchangeExt.synchronizePortAllocations(
            ce.get_java_object(),
            fe.get_java_object()
        )
        return True
    except Exception as e:
        print(f"Error synchronizing port allocations: {e}")
        import traceback
        traceback.print_exc()
        return False
    
def collect_all_allocations(lc_pkg):
    """Collect all existing allocations in the model"""
    allocations = {}

    try:

        # Start from the logical component package
        def search_recursive(java_object):
            try:
                # Check if this object has component exchanges
                if hasattr(java_object, 'getOwnedComponentExchanges'):
                    # Get all component exchanges owned by this object
                    exchanges = java_object.getOwnedComponentExchanges()
                    for ce in exchanges:
                        try:
                            # Get allocations for this component exchange
                            allocations_list = ce.getOwnedComponentExchangeFunctionalExchangeAllocations()
                            for allocation in allocations_list:
                                try:
                                    fe = allocation.getTargetElement()
                                    if fe:
                                        ce_id = ce.getSid()
                                        fe_id = fe.getSid()
                                        allocation_id = allocation.getSid()

                                        # print(f"Found allocation: CE={ce_id}, FE={fe_id}, Alloc={allocation_id}")

                                        if ce_id not in allocations:
                                            allocations[ce_id] = {}
                                        if fe_id not in allocations[ce_id]:
                                            allocations[ce_id][fe_id] = []

                                        allocations[ce_id][fe_id].append(allocation_id)
                                except Exception as e:
                                    print(f"Error processing allocation: {e}")
                        except Exception as e:
                            print(f"Error getting allocations for component exchange: {e}")

                # Recursively check children
                if hasattr(java_object, 'getOwnedLogicalComponents'):
                    children = java_object.getOwnedLogicalComponents()
                    for child in children:
                        search_recursive(child)
            except Exception as e:
                print(f"Error in search_recursive: {e}")

        # Start the recursive search from the logical component package
        lc_pkg_java = lc_pkg.get_java_object()
        search_recursive(lc_pkg_java)

        total_count = sum(len(v) for d in allocations.values() for v in d.values())
        # print(f"Found {total_count} existing allocations")
        return allocations
    except Exception as e:
        print(f"Error in collect_all_allocations: {e}")
        return {}

def delete_unused_allocations(wb, imported_allocations, lc_pkg):
    """Delete allocations that exist in the model but were not imported"""
    try:
        deleted_count = 0

        # Start from the logical component package
        def search_and_delete_recursive(java_object):
            nonlocal deleted_count
            try:
                # Check if this object has component exchanges
                if hasattr(java_object, 'getOwnedComponentExchanges'):
                    exchanges = java_object.getOwnedComponentExchanges()
                    for ce_java in exchanges:
                        try:
                            ce = ComponentExchange(ce_java)

                            # Get allocations for this component exchange
                            allocations_list = list(ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations())
                            for allocation in allocations_list:
                                try:
                                    fe_java = allocation.getTargetElement()
                                    if fe_java:
                                        fe = FunctionalExchange(fe_java)
                                        ce_id = ce.get_sid()
                                        fe_id = fe.get_sid()
                                        allocation_id = allocation.getSid()

                                        # Check if this allocation should be kept
                                        if (ce_id in imported_allocations and
                                            fe_id in imported_allocations[ce_id] and
                                            allocation_id in imported_allocations[ce_id][fe_id]):
                                            continue  # This allocation is in the import file

                                        # Get names for logging
                                        ce_name = ce.get_name()
                                        fe_name = fe.get_name()

                                        # Log the deletion if log_path is provided
                                        log_deletion(wb, "LinkExchanges", "Functional Exchange to Component Exchange Allocation",
                                            ce_id=ce_id, ce_name=ce_name, fe_id=fe_id, fe_name=fe_name, allocation_id=allocation_id)
                                        
                                        EObject.delete_e_object(ComponentExchangeFunctionalExchangeAllocation(allocation))
                                        deleted_count += 1

                                except Exception as e:
                                    print(f"Error processing allocation: {e}")
                        except Exception as e:
                            print(f"Error getting allocations for component exchange: {e}")

                # Recursively check children
                if hasattr(java_object, 'getOwnedLogicalComponents'):
                    children = java_object.getOwnedLogicalComponents()
                    for child in children:
                        search_and_delete_recursive(child)
            except Exception as e:
                print(f"Error in search_and_delete_recursive: {e}")

        # Start the recursive search from the logical component package
        lc_pkg_java = lc_pkg.get_java_object()
        search_and_delete_recursive(lc_pkg_java)

        return deleted_count

    except Exception as e:
        print(f"Error in delete_unused_allocations: {e}")
        return 0   
   
def import_link_exchanges(wb, model, se, lc_pkg, lf_pkg):
    """Import link exchanges from the Link Exchanges worksheet with deletion functionality"""
    try:
        ws = wb["Link Exchanges"]

        # First collect all existing allocations
        existing_allocations = collect_all_allocations(lc_pkg)

        # Dictionary to store imported allocations: {ce_id: {fe_id: [allocation_ids]}}
        imported_allocations = {}

        # Process each row in the Excel file
        for row in ws.iter_rows(min_row=2):
            # Extract data from the row
            ce_id = row[0].value
            ce_name = row[1].value
            ce_source_port_id = row[2].value
            ce_source_port_name = row[3].value
            source_comp_id = row[4].value
            source_comp_name = row[5].value
            ce_target_port_id = row[6].value
            ce_target_port_name = row[7].value
            target_comp_id = row[8].value
            target_comp_name = row[9].value
            fe_id = row[10].value
            fe_name = row[11].value
            fe_source_port_id = row[12].value
            fe_source_port_name = row[13].value
            fe_target_port_id = row[14].value
            fe_target_port_name = row[15].value
            allocation_id = row[16].value
            source_func_id = row[17].value
            target_func_id = row[18].value

            # Find component exchange
            ce = find_component_exchange_by_id(ce_id, lc_pkg)
            if not ce:
                print(f"Skipping - Could not find component exchange with ID {ce_id}")
                continue

            # Find functional exchange
            fe = find_functional_exchange_by_id(fe_id, se)
            if not fe:
                print(f"Skipping - Could not find functional exchange with ID {fe_id}")
                continue

            # Check if allocation already exists
            if check_allocation_exists(ce, fe):
                # Find the existing allocation to get its ID
                existing_allocation = None
                for alloc in ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations():
                    if alloc.getTargetElement().getSid() == fe.get_sid():
                        existing_allocation = alloc
                        break

                if existing_allocation:
                    allocation_id = existing_allocation.getSid()
                else:
                    print(f"Could not find existing allocation ID")
                    continue
            else:
                # Create the allocation between component exchange and functional exchange
                allocation = create_allocation(ce, fe, allocation_id)
                if not allocation:
                    print(f"Skipping - Could not create allocation between {ce_name} and {fe_name}")
                    continue

            # Add to imported allocations
            if ce_id not in imported_allocations:
                imported_allocations[ce_id] = {}
            if fe_id not in imported_allocations[ce_id]:
                imported_allocations[ce_id][fe_id] = []
            imported_allocations[ce_id][fe_id].append(allocation_id)

            # Synchronize port allocations
            synchronize_port_allocations(ce, fe)

        # After ALL imports are complete, delete unused allocations
        should_delete_count = 0
        for ce_id, fe_dict in existing_allocations.items():
            for fe_id, alloc_ids in fe_dict.items():
                if ce_id in imported_allocations and fe_id in imported_allocations[ce_id]:
                    # Check which allocations are not in imported list
                    imported_ids = set(imported_allocations[ce_id][fe_id])
                    for alloc_id in alloc_ids:
                        if alloc_id not in imported_ids:
                            should_delete_count += 1
                else:
                    # All allocations for this CE-FE pair should be deleted
                    should_delete_count += len(alloc_ids)

        # Now actually delete unused allocations
        deleted_count = delete_unused_allocations(wb, imported_allocations, lc_pkg)
        if deleted_count > 0:
            print(f"Deleted {deleted_count} allocations that were not imported and will be deleted")

        if should_delete_count != deleted_count:
            print(f"Warning: Expected to delete {should_delete_count} allocations but deleted {deleted_count}")

        print(f"Imported Link Exchanges worksheet successfully")
        return True, deleted_count
    except Exception as e:
        print(f"Error importing link exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 0

def find_function_by_id_global(function_id, se):
    """Find a function by its ID by searching through all functions in the model"""
    try:
        # First try the hierarchical search starting from RLF
        rlf = None
        for lf in se.get_logical_architecture().get_logical_function_pkg().get_owned_logical_functions():
            if lf.get_name() == "Root Logical Function":
                rlf = lf
                break

        if rlf:
            found = find_function_by_id(rlf, function_id)
            if found:
                return found

        # If not found in hierarchy, search through all functions
        all_functions = se.get_all_contents_by_type(Function)
        for func in all_functions:
            if func.get_sid() == function_id:
                return func
        return None
    except Exception as e:
        print(f"Error finding function by ID {function_id}: {e}")
        return None
    
def functional_chain_exists(fc_id, se):
    """Check if a functional chain already exists"""
    try:
        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)
        for chain in all_chains:
            if chain.get_sid() == fc_id:
                return True
        return False
    except Exception as e:
        print(f"Error checking if functional chain exists: {e}")
        return False

def find_function_involvement(function_involvements, function_id):
    """Find a function involvement by function ID"""
    for fcif in function_involvements:
        try:
            involved = fcif.getInvolvedElement()
            if involved and involved.getSid() == function_id:
                return fcif
        except Exception as e:
            print(f"Error checking function involvement: {e}")
    return None

def create_function_involvement(fc, func, involvement_id=None):
    """Create a functional chain involvement function"""
    try:
        # Use the Capella factory to create the involvement
        factory = org.polarsys.capella.core.data.fa.FaFactory.eINSTANCE
        fcif = factory.createFunctionalChainInvolvementFunction()
        fcif.setInvolved(func.get_java_object())
        fc.get_java_object().getOwnedFunctionalChainInvolvements().add(fcif)
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fcif)
        if involvement_id:
            fcif.setSid(involvement_id)
        return fcif
    except Exception as e:
        print(f"Error creating function involvement: {e}")
        return None

def create_link_involvement(fc, exchange, source_fcif, target_fcif, involvement_id=None):
    """Create a functional chain involvement link"""
    try:
        # Use the Capella factory to create the involvement
        factory = org.polarsys.capella.core.data.fa.FaFactory.eINSTANCE
        fcil = factory.createFunctionalChainInvolvementLink()
        fcil.setInvolved(exchange.get_java_object())
        fcil.setSource(source_fcif)
        fcil.setTarget(target_fcif)
        fc.get_java_object().getOwnedFunctionalChainInvolvements().add(fcil)
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fcil)
        if involvement_id:
            fcil.setSid(involvement_id)
        return fcil
    except Exception as e:
        print(f"Error creating link involvement: {e}")
        return None
    
def collect_all_functional_chains(se):
    """Collect all existing functional chains in the model"""
    chains = {}

    try:
        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)
        for chain in all_chains:
            chain_id = chain.get_sid()
            chains[chain_id] = chain

        return chains
    except Exception as e:
        print(f"Error collecting existing functional chains: {e}")
        return {}

def delete_unused_functional_chains(wb, imported_chains, se):
    """Delete functional chains that exist in the model but were not imported"""
    try:
        deleted_count = 0

        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)
        for chain in all_chains:
            chain_id = chain.get_sid()

            # Check if this chain should be kept
            if chain_id in imported_chains:
                continue  # This chain is in the import file

            # Get name for logging
            chain_name = chain.get_name()

            # Log the deletion if log_path is provided
            log_deletion(wb, "FunctionalChains", "Functional Chains",
                chain_id=chain_id, chain_name=chain_name)
            
            EObject.delete_e_object(chain)
            deleted_count += 1

        return deleted_count
    except Exception as e:
        print(f"Error in delete_unused_functional_chains: {e}")
        return 0

def import_functional_chains(wb, model, se, lf_pkg):
    """Import functional chains from the Functional Chains worksheet with explicit source/target for links"""
    try:
        ws = wb["Functional Chains"]

        # Find the Root Logical Function
        rlf = None
        for lf in lf_pkg.get_owned_logical_functions():
            if lf.get_name() == "Root Logical Function":
                rlf = lf
                break

        if not rlf:
            print("Root Logical Function not found!")
            return False, 0

        # Find the exchange start column
        header_row = ws[1]
        function_start_col = 3
        exchange_start_col = None
        for cell in header_row:
            if cell.value and "Exchange 1 ID" in str(cell.value):
                exchange_start_col = cell.column
                break

        if exchange_start_col is None:
            print("Error: Could not find exchange start column in header")
            return False, 0

        # First collect all existing functional chains
        existing_chains = collect_all_functional_chains(se)
        imported_chains = {}

        # Process each row in the Excel file (skip header row)
        for row in ws.iter_rows(min_row=2):
            fc_id = row[0].value
            fc_name = row[1].value

            if not fc_id:
                continue

            if fc_id in existing_chains:
                imported_chains[fc_id] = existing_chains[fc_id]
                continue

            # Create a new functional chain
            fc = FunctionalChain()
            fc.get_java_object().setSid(fc_id)
            fc.set_name(fc_name)
            imported_chains[fc_id] = fc

            # Process functions: Store in a dict for lookup by ID
            function_col = function_start_col - 1  # 0-based index
            function_involvements = {}  # {func_id: fcif_java}
            involved_functions = []

            while function_col < exchange_start_col - 1:
                try:
                    func_id = row[function_col].value
                    func_name = row[function_col + 1].value
                    func_involvement_id = row[function_col + 2].value

                    if not func_id:
                        function_col += 3
                        continue

                    func = find_function_by_id_global(func_id, se)
                    if not func:
                        print(f"Could not find function with ID {func_id} and name {func_name}")
                        function_col += 3
                        continue

                    fcif = create_function_involvement(fc, func, func_involvement_id)
                    if fcif:
                        function_involvements[func_id] = fcif  # Store in dict for lookup
                        involved_functions.append(func)

                    function_col += 3
                except Exception as e:
                    print(f"Error processing function: {e}")
                    function_col += 3
                    continue

            # Process exchanges: Use explicit source/target from Excel columns
            exchange_col = exchange_start_col - 1  # 0-based index
            while exchange_col < len(row):
                try:
                    exchange_id = row[exchange_col].value
                    exchange_name = row[exchange_col + 1].value
                    exchange_involvement_id = row[exchange_col + 2].value

                    if not exchange_id:
                        break

                    exchange = find_functional_exchange_by_id(exchange_id, se)
                    if not exchange:
                        print(f"Could not find functional exchange with ID {exchange_id} and name {exchange_name}")
                        exchange_col += 7  # 7 columns per exchange
                        continue

                    # --- KEY CHANGE: Read source/target IDs from Excel ---
                    source_func_id = row[exchange_col + 3].value  # Source Function ID (column +3)
                    target_func_id = row[exchange_col + 5].value  # Target Function ID (column +5)

                    if not source_func_id or not target_func_id:
                        print(f"Warning: Exchange {exchange_id} has missing source/target function IDs")
                        exchange_col += 7
                        continue

                    # Look up the function involvements by ID
                    source_fcif = function_involvements.get(source_func_id)
                    target_fcif = function_involvements.get(target_func_id)

                    if not source_fcif or not target_fcif:
                        print(f"Warning: Could not find function involvements for source {source_func_id} or target {target_func_id}")
                        exchange_col += 7
                        continue

                    # Create link involvement with EXPLICIT source/target
                    create_link_involvement(fc, exchange, source_fcif, target_fcif, exchange_involvement_id)

                    exchange_col += 7  # Move to next exchange (7 columns per exchange)

                except Exception as e:
                    print(f"Error processing exchange: {e}")
                    exchange_col += 7
                    continue

            # Add the functional chain to the root logical function
            try:
                rlf.get_java_object().getOwnedFunctionalChains().add(fc.get_java_object())
                org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fc.get_java_object())
                fc.get_java_object().setSid(fc_id)
                fc.set_name(fc_name)
            except Exception as e:
                print(f"Error adding chain to root: {e}")

            # Add the functional chain to all involved functions
            for func in involved_functions:
                try:
                    func.get_java_object().getOwnedFunctionalChains().add(fc.get_java_object())
                except Exception as e:
                    print(f"Error adding chain to function {get_name(func)}: {e}")

        # Delete unused chains
        deleted_count = delete_unused_functional_chains(wb, imported_chains, se)
        if deleted_count > 0:
            print(f"Deleted {deleted_count} unused functional chains")

        print("Imported Functional Chains worksheet successfully")
        return True, deleted_count

    except Exception as e:
        print(f"Error importing functional chains: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 0
    
def get_java_object(obj):
    """Get the Java object from a Python wrapper or return the object itself"""
    try:
        if hasattr(obj, 'get_java_object'):
            return obj.get_java_object()
        return obj
    except Exception as e:
        print(f"Error getting Java object: {str(e)}")
        return None    

def get_component_id(component):
    """Get the ID of a component, handling both LogicalComponent and LogicalActor."""
    try:
        if hasattr(component, 'get_sid'):
            return component.get_sid()
        elif hasattr(component, 'getSid'):
            return component.getSid()
        elif hasattr(component, 'get_java_object'):
            java_obj = component.get_java_object()
            if hasattr(java_obj, 'getSid'):
                return java_obj.getSid()
    except Exception as e:
        print(f"Error getting component ID: {e}")
    return None

def find_capability_realization_by_id(capability_realization_pkg, cr_id):
    """Find a capability realization by its ID"""
    try:
        java_pkg = get_java_object(capability_realization_pkg)
        if not java_pkg:
            return None

        for cr in java_pkg.getOwnedCapabilityRealizations():
            if cr.getSid() == cr_id:
                return CapabilityRealization(cr)
        return None
    except Exception as e:
        print(f"Error finding capability realization by ID: {e}")
        return None

def find_functional_chain_by_id(fc_id, se):
    """Find a functional chain by its ID by searching through all functions"""
    try:
        # Get all functions in the model
        all_functions = se.get_all_contents_by_type(Function)

        # Search through all functions for owned functional chains
        for func in all_functions:
            java_func = get_java_object(func)
            for chain in java_func.getOwnedFunctionalChains():
                if chain.getSid() == fc_id:
                    return FunctionalChain(chain)

        return None
    except Exception as e:
        print(f"Error finding functional chain by ID: {e}")
        return None

def find_logical_function_by_id(func_id, se):
    """Find a logical function by its ID"""
    try:
        # Get all logical functions from the model
        all_functions = se.get_all_contents_by_type(LogicalFunction)
        for func in all_functions:
            if get_component_id(func) == func_id:
                return func
        return None
    except Exception as e:
        print(f"Error finding logical function by ID: {e}")
        return None

def find_logical_component_by_id(comp_id, se):
    """Find a logical component by its ID"""
    try:
        # Get all logical components from the model
        all_components = se.get_all_contents_by_type(LogicalComponent)
        for comp in all_components:
            if get_component_id(comp) == comp_id:
                return comp
        return None
    except Exception as e:
        print(f"Error finding logical component by ID: {e}")
        return None

def find_logical_actor_by_id(actor_id, se):
    """Find a logical actor by its ID"""
    try:
        # Get all logical actors from the model
        all_actors = se.get_all_contents_by_type(LogicalActor)
        for actor in all_actors:
            if get_component_id(actor) == actor_id:
                return actor
        return None
    except Exception as e:
        print(f"Error finding logical actor by ID: {e}")
        return None

def create_function_capability_involvement(cr, element, involvement_id=None):
    """Create an involvement between a function and a capability realization"""
    try:
        # Create the involvement object using the factory
        factory = org.polarsys.capella.core.data.interaction.InteractionFactory.eINSTANCE
        involvement = factory.createAbstractFunctionAbstractCapabilityInvolvement()

        # Get the Java objects
        java_cr = get_java_object(cr)
        java_element = get_java_object(element)

        # Add the involvement to the capability realization
        java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements().add(involvement)

        # Set the involved element
        involvement.setInvolved(java_element)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)

        if involvement_id:
            involvement.setSid(involvement_id)

        return involvement
    except Exception as e:
        print(f"Error creating function/capability involvement: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_functional_chain_capability_involvement(cr, fc, involvement_id=None):
    """Create an involvement between a functional chain and a capability realization"""
    try:
        # Create the involvement object using the factory
        factory = org.polarsys.capella.core.data.interaction.InteractionFactory.eINSTANCE
        involvement = factory.createFunctionalChainAbstractCapabilityInvolvement()

        # Get the Java objects
        java_cr = get_java_object(cr)
        java_element = get_java_object(fc)

        # Add the involvement to the capability realization
        java_cr.getOwnedFunctionalChainAbstractCapabilityInvolvements().add(involvement)

        # Set the involved element
        involvement.setInvolved(java_element)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)

        if involvement_id:
            involvement.setSid(involvement_id)

        return involvement
    except Exception as e:
        print(f"Error creating function/capability involvement: {e}")
        import traceback
        traceback.print_exc()
        return None

def add_component_to_capability_realization(cr, component, involvement_id=None):
    """Add a component or actor to a capability realization"""
    try:
        # Get the Java objects
        java_cr = get_java_object(cr)
        java_component = get_java_object(component)

        # Check if the component is already involved
        already_involved = False
        for involved in java_cr.getInvolvedComponents():
            if involved.getSid() == java_component.getSid():
                already_involved = True
                break

        if already_involved:
            return True

        # Create the involvement using the factory
        factory = org.polarsys.capella.core.data.capellacommon.CapellacommonFactory.eINSTANCE
        involvement = factory.createCapabilityRealizationInvolvement()

        # Set the involved element
        involvement.setInvolved(java_component)

        # Add the involvement to the capability realization
        java_cr.getOwnedCapabilityRealizationInvolvements().add(involvement)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)

        # Set the involvement ID if provided
        if involvement_id:
            involvement.setSid(involvement_id)

        return True
    except Exception as e:
        print(f"Error adding component to capability realization: {e}")
        import traceback
        traceback.print_exc()
        return False
    
def collect_all_capability_realizations(capability_realization_pkg):
    """Collect all existing capability realizations in the model"""
    realizations = {}

    try:
        if not capability_realization_pkg:
            print("No Capability Realization Package found in Logical Architecture")
            return {}

        # Get all capability realizations in the package
        java_pkg = get_java_object(capability_realization_pkg)
        for cr in java_pkg.getOwnedCapabilityRealizations():
            cr_id = cr.getSid()
            realizations[cr_id] = cr

        total_count = len(realizations)
        return realizations
    except Exception as e:
        print(f"Error in collect_all_capability_realizations: {e}")
        return {}

def collect_capability_involvements(cr):
    """Collect all involvements for a capability realization"""
    involvements = {
        'functional_chains': {},
        'logical_functions': {},
        'components': {}
    }

    try:
        java_cr = get_java_object(cr)

        # Collect functional chain involvements
        for involvement in java_cr.getOwnedFunctionalChainAbstractCapabilityInvolvements():
            fc = involvement.getInvolved()
            if fc:
                fc_id = fc.getSid()
                involvement_id = involvement.getSid()
                involvements['functional_chains'][fc_id] = involvement_id
                print(f"Found functional chain involvement: {fc_id} (Involvement ID: {involvement_id})")

        # Collect function involvements
        for involvement in java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements():
            func = involvement.getInvolved()
            if func:
                func_id = func.getSid()
                involvement_id = involvement.getSid()
                involvements['logical_functions'][func_id] = involvement_id
                print(f"Found function involvement: {func_id} (Involvement ID: {involvement_id})")

        # Collect component involvements
        for involvement in java_cr.getOwnedCapabilityRealizationInvolvements():
            component = involvement.getInvolved()
            if component:
                component_id = component.getSid()
                involvement_id = involvement.getSid()
                involvements['components'][component_id] = involvement_id
                print(f"Found component involvement: {component_id} (Involvement ID: {involvement_id})")

        return involvements
    except Exception as e:
        print(f"Error collecting capability involvements: {e}")
        return {}

def delete_unused_capability_realizations(wb, imported_realizations, capability_realization_pkg):
    """Delete capability realizations that exist in the model but were not imported"""
    try:
        deleted_count = 0

        # Get all capability realizations in the package
        java_pkg = get_java_object(capability_realization_pkg)
        existing_realizations = java_pkg.getOwnedCapabilityRealizations()

        for cr_java in existing_realizations:
            try:
                cr = CapabilityRealization(cr_java)
                cr_id = cr.get_sid()
                cr_name = cr.get_name()

                # Check if this capability realization should be kept
                if cr_id in imported_realizations:
                    continue  # This capability realization is in the import file

                # Log the deletion if log_path is provided
                log_deletion(wb, "Capabilities", "Capability",
                    cr_id=cr_id, cr_name=cr_name)

                EObject.delete_e_object(cr)
                deleted_count += 1

            except Exception as e:
                print(f"Error processing capability realization {cr_java.getSid()}: {e}")

        return deleted_count

    except Exception as e:
        print(f"Error in delete_unused_capability_realizations: {e}")
        return 0

def import_capabilities(wb, model, se, la):
    """Import capabilities from the Capabilities worksheet with deletion functionality"""
    try:
        ws = wb["Capabilities"]

        # Find max functional chains, logical functions, and involved components from headers
        max_functional_chains = 0
        max_logical_functions = 0
        max_involved_components = 0

        function_start_col = 3
        for col in range(function_start_col, ws.max_column + 1, 3):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "Functional Chain" in cell_value:
                max_functional_chains += 1

        function_end_col = function_start_col + max_functional_chains * 3
        for col in range(function_end_col, ws.max_column + 1, 3):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "Logical Function" in cell_value:
                max_logical_functions += 1

        component_start_col = function_end_col + max_logical_functions * 3
        for col in range(component_start_col, ws.max_column + 1, 4):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "Involved Component" in cell_value:
                max_involved_components += 1

        # Get the capability realization package
        capability_realization_pkg = la.get_capability_realization_pkg()
        if not capability_realization_pkg:
            print("No Capability Realization Package found in Logical Architecture")
            return False

        # First collect all existing capability realizations
        existing_realizations = collect_all_capability_realizations(capability_realization_pkg)

        # Dictionary to store imported capability realizations
        imported_realizations = {}

        # Process each row in the Excel file (skip header row)
        for row in ws.iter_rows(min_row=2):
            try:
                # Extract basic capability realization info
                cr_id = row[0].value
                cr_name = row[1].value

                if not cr_id:
                    continue

                # Check if the capability realization already exists
                existing_cr = find_capability_realization_by_id(capability_realization_pkg, cr_id)
                if existing_cr:
                    imported_realizations[cr_id] = existing_cr
                    continue

                # Create a new capability realization
                cr = CapabilityRealization()
                java_cr = get_java_object(cr)

                # Add to the capability realization package first
                java_pkg = get_java_object(capability_realization_pkg)
                java_pkg.getOwnedCapabilityRealizations().add(java_cr)
                org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(java_cr)

                # Now set the ID and name
                java_cr.setSid(cr_id)
                cr.set_name(cr_name)

                # Add to imported realizations
                imported_realizations[cr_id] = cr

                # Process functional chains
                for i in range(max_functional_chains):
                    fc_id_col = function_start_col + i*3 - 1  # Convert to 0-based index
                    fc_name_col = function_start_col + i*3
                    fc_involvement_col = function_start_col + i*3 + 1

                    if fc_id_col >= len(row) or not row[fc_id_col].value:
                        continue

                    fc_id = row[fc_id_col].value
                    fc_name = row[fc_name_col].value if fc_name_col < len(row) else ""
                    fc_involvement_id = row[fc_involvement_col].value if fc_involvement_col < len(row) else ""

                    if not fc_id:
                        continue

                    # Find the functional chain
                    fc = find_functional_chain_by_id(fc_id, se)
                    if not fc:
                        print(f"WARNING: Could not find functional chain with ID {fc_id} and name '{fc_name}'. Skipping this chain.")
                        continue

                    # Create involvement between functional chain and capability realization
                    involvement = create_functional_chain_capability_involvement(cr, fc, fc_involvement_id)
                    if not involvement:
                        print(f"Failed to create involvement for functional chain {fc_name}")

                # Process logical functions
                for i in range(max_logical_functions):
                    func_id_col = function_end_col + i*3 - 1  # Convert to 0-based index
                    func_name_col = function_end_col + i*3
                    func_involvement_col = function_end_col + i*3 + 1

                    if func_id_col >= len(row) or not row[func_id_col].value:
                        continue

                    func_id = row[func_id_col].value
                    func_name = row[func_name_col].value if func_name_col < len(row) else ""
                    func_involvement_id = row[func_involvement_col].value if func_involvement_col < len(row) else ""

                    if not func_id:
                        continue

                    # Find the logical function
                    func = find_logical_function_by_id(func_id, se)
                    if not func:
                        print(f"WARNING: Could not find logical function with ID {func_id} and name '{func_name}'. Skipping this function.")
                        continue

                    # Create involvement between logical function and capability realization
                    involvement = create_function_capability_involvement(cr, func, func_involvement_id)
                    if not involvement:
                        print(f"Failed to create involvement for logical function {func_name}")

                # Process involved components
                for i in range(max_involved_components):
                    comp_id_col = component_start_col + i*4 - 1  # Convert to 0-based index
                    comp_name_col = component_start_col + i*4
                    comp_type_col = component_start_col + i*4 + 1
                    comp_involvement_col = component_start_col + i*4 + 2

                    if comp_id_col >= len(row) or not row[comp_id_col].value:
                        continue

                    comp_id = row[comp_id_col].value
                    comp_name = row[comp_name_col].value if comp_name_col < len(row) else ""
                    comp_type = row[comp_type_col].value if comp_type_col < len(row) else ""
                    comp_involvement_id = row[comp_involvement_col].value if comp_involvement_col < len(row) else ""

                    if not comp_id:
                        continue

                    # Find the component or actor based on type
                    component = None
                    if comp_type == "actor":
                        component = find_logical_actor_by_id(comp_id, se)
                    else:
                        component = find_logical_component_by_id(comp_id, se)

                    if not component:
                        print(f"WARNING: Could not find {comp_type} with ID {comp_id} and name '{comp_name}'. Skipping this component.")
                        continue

                    # Add component/actor to capability realization
                    success = add_component_to_capability_realization(cr, component, comp_involvement_id)
                    if not success:
                        print(f"Failed to add {comp_type} {comp_name} to capability realization {cr_name}")

            except Exception as e:
                print(f"Error processing row: {str(e)}")
                import traceback
                traceback.print_exc()
                continue

        # After import, delete unused capability realizations
        deleted_count = delete_unused_capability_realizations(wb, imported_realizations, capability_realization_pkg)
        if deleted_count > 0:
            print(f"Deleted {deleted_count} capability realizations that were not imported and will be deleted")

        print(f"Imported Capabilities worksheet successfully")
        return True, deleted_count
    except Exception as e:
        print(f"Error importing capabilities: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, 0 
    
def verify_deletion_log(wb):
    """Verify that the deletion log was created and is accessible"""
    try:
        if hasattr(wb, '_deletion_log') and wb._deletion_log.get("path"):
            log_path = wb._deletion_log["path"]
            print(f"\nVerifying deletion log at: {log_path}")
            print(f"File exists: {os.path.exists(log_path)}")
            if os.path.exists(log_path):
                print(f"File size: {os.path.getsize(log_path)} bytes")
                print(f"File is readable: {os.access(log_path, os.R_OK)}")
                print(f"File is writable: {os.access(log_path, os.W_OK)}")

                # Try to open the file to verify it's not locked
                try:
                    with open(log_path, 'rb') as f:
                        print("File is not locked and can be read")
                except IOError:
                    print("WARNING: File appears to be locked by another process")
            else:
                print("ERROR: Deletion log file does not exist at the expected path")
        else:
            print("No deletion log information found in workbook")
    except Exception as e:
        print(f"Error verifying deletion log: {e}")

def full_import(aird_path, xlsx_path):
    """
    Perform a full import of all model elements from a single Excel workbook
    with multiple worksheets.
    """
    try:
        # Get model and workbook
        model, se, la, lc_pkg, lf_pkg, rlf, wb, xlsx_file_name = get_model_and_workbook(aird_path, xlsx_path)
        if not wb:
            print("Failed to setup workbook")
            return False

        # Start the import transaction
        model.start_transaction()

        # Initialize deletion counter
        total_deletions = 0

        try:
            # Set up deletion log using the aird path
            log_info, project = setup_deletion_log(aird_path)
            if not log_info:
                print("Warning: Could not create deletion log. Continuing without logging...")
                log_path = None
                log_wb = None
            else:
                log_path = log_info["path"]
                log_wb = log_info["wb"]

                # Store the log path in the main workbook for reference
                if not hasattr(wb, '_deletion_log'):
                    wb._deletion_log = {}
                wb._deletion_log["path"] = log_path
                wb._deletion_log["wb"] = log_wb

            # Import all elements - modify each call to capture deletion counts
            success = True

            # Modify each import function call to capture both success and deletion count
            systems_success, systems_deletions = import_systems(wb, model, lc_pkg)
            success &= systems_success
            total_deletions += systems_deletions

            functions_success, functions_deletions = import_functions(wb, model, lf_pkg)
            success &= functions_success
            total_deletions += functions_deletions

            links_success, links_deletions = import_link_functions_systems(wb, model, lc_pkg, lf_pkg, rlf)
            success &= links_success
            total_deletions += links_deletions

            functional_exchanges_success, functional_exchanges_deletions = import_functional_exchanges(wb, model, se, lf_pkg)
            success &= functional_exchanges_success
            total_deletions += functional_exchanges_deletions

            component_exchanges_success, component_exchanges_deletions = import_component_exchanges(wb, model, lc_pkg)
            success &= component_exchanges_success
            total_deletions += component_exchanges_deletions

            link_exchanges_success, link_exchanges_deletions = import_link_exchanges(wb, model, se, lc_pkg, lf_pkg)
            success &= link_exchanges_success
            total_deletions += link_exchanges_deletions

            functional_chains_success, functional_chains_deletions = import_functional_chains(wb, model, se, lf_pkg)
            success &= functional_chains_success
            total_deletions += functional_chains_deletions

            capabilities_success, capabilities_deletions = import_capabilities(wb, model, se, la)
            success &= capabilities_success
            total_deletions += capabilities_deletions
            
            if total_deletions > 0:
                print(f"Total items deleted: {total_deletions}")

            if success:
                model.commit_transaction()
                model.save()

                # Check if we should delete the empty log file
                if log_path and total_deletions == 0:
                    try:
                        os.remove(log_path)
                        
                        try:
                            CapellaPlatform.refresh(project)
                        except Exception as e:
                            print(f"Could not refresh the deleted log folder: {e}")
                        
                    except Exception as e:
                        print(f"Could not remove empty deletion log file: {e}")

                print("Full import completed successfully!")
                return True
            else:
                model.rollback_transaction()
                print("Some imports failed. Check error messages above.")
                return False

        except Exception as e:
            print(f"Error during import: {str(e)}")
            model.rollback_transaction()
            import traceback
            traceback.print_exc()
            return False

    except Exception as e:
        print(f"Error in full import: {str(e)}")
        import traceback
        traceback.print_exc()
        return False