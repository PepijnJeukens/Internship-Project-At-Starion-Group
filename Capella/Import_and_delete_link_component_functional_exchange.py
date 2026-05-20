'''
Created on 13 Apr 2026

@author: p.jeukens
'''
# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
aird_path = "/delete_test/delete_test.aird"
xlsx_path = "/DVS/results/DVS_Exchange_Allocations2.xlsx"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()
lc_pkg = la.get_logical_component_pkg()
lf_pkg = la.get_logical_function_pkg()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Reading " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

def find_component_exchange_by_id(component_id):
    """Find a component exchange by its ID in the model using recursive search"""
    try:
        # Start from the logical component package
        lc_pkg_java = lc_pkg.get_java_object()

        # Use a stack for iterative depth-first search
        stack = [lc_pkg_java]

        while stack:
            current = stack.pop()

            try:
                # Check exchanges owned by this object
                if hasattr(current, 'getOwnedComponentExchanges'):
                    exchanges = current.getOwnedComponentExchanges()
                    for ce in exchanges:
                        if ce.getSid() == component_id:
                            return ComponentExchange(ce)

                # Add children to stack for processing
                if hasattr(current, 'getOwnedLogicalComponents'):
                    children = current.getOwnedLogicalComponents()
                    for child in reversed(children):
                        stack.append(child)
            except:
                continue

        print(f"Could not find component exchange with ID {component_id}")
        return None
    except Exception as e:
        print(f"Error finding component exchange by ID {component_id}: {e}")
        return None

def find_functional_exchange_by_id(exchange_id):
    """Find a functional exchange by its ID in the model"""
    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for fe in all_exchanges:
            if fe.get_sid() == exchange_id:
                return fe

        print(f"Could not find functional exchange with ID {exchange_id}")
        return None
    except Exception as e:
        print(f"Error finding functional exchange by ID {exchange_id}: {e}")
        return None

def find_functional_exchange_port(fe, expected_port_id):
    """Find a functional exchange port by ID from the functional exchange"""
    try:
        # Get source and target ports from the functional exchange
        source_port = fe.get_source_port()
        target_port = fe.get_target_port()

        # Check if either port matches the expected ID
        if source_port and source_port.get_sid() == expected_port_id:
            return source_port
        if target_port and target_port.get_sid() == expected_port_id:
            return target_port

        print(f"Could not find functional exchange port with ID {expected_port_id} in exchange {fe.get_name()}")
        return None
    except Exception as e:
        print(f"Error finding functional exchange port: {e}")
        return None

def check_allocation_exists(ce, fe):
    """Check if an allocation between component exchange and functional exchange already exists"""
    try:
        # Check all allocations owned by the component exchange
        allocations = ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations()
        for allocation in allocations:
            target = allocation.getTargetElement()
            if target and target.getSid() == fe.get_sid():
                return True
        return False
    except Exception as e:
        print(f"Error checking if allocation exists: {e}")
        return False

def check_port_allocation_exists(fe_port, ce_port):
    """Check if a port allocation already exists between functional and component ports"""
    try:
        if fe_port and ce_port:
            # Check if the functional port already has this component port as allocator
            allocator = fe_port.get_allocator_component_port()
            if allocator and allocator.getSid() == ce_port.get_sid():
                return True
        return False
    except Exception as e:
        print(f"Error checking if port allocation exists: {e}")
        return False

def create_allocation(ce, fe, allocation_id):
    """Create an allocation between component exchange and functional exchange"""
    try:
        # Check if allocation already exists
        if check_allocation_exists(ce, fe):
            # print(f"Allocation already exists between {ce.get_name()} and {fe.get_name()}")
            return None

        # Create the allocation
        allocation = ComponentExchangeFunctionalExchangeAllocation()

        # Add the allocation to the component exchange's owned allocations
        ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations().add(allocation.get_java_object())
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(allocation.get_java_object())

        # Set the ID for the allocation
        allocation.get_java_object().setSid(allocation_id)

        # Set source as component exchange and target as functional exchange
        allocation.get_java_object().setSourceElement(ce.get_java_object())
        allocation.get_java_object().setTargetElement(fe.get_java_object())

        # print(f"Successfully created allocation between {ce.get_name()} and {fe.get_name()}")
        return allocation
    except Exception as e:
        print(f"Error creating allocation: {e}")
        return None

def synchronize_port_allocations(ce, fe):
    """Synchronize port allocations between component exchange and functional exchange"""
    try:
        # Use the Capella API method to synchronize port allocations
        org.polarsys.capella.core.model.helpers.ComponentExchangeExt.synchronizePortAllocations(
            ce.get_java_object(),
            fe.get_java_object()
        )
        # print(f"Successfully synchronized port allocations between {ce.get_name()} and {fe.get_name()}")
        return True
    except Exception as e:
        print(f"Error synchronizing port allocations: {e}")
        return False

def collect_all_allocations():
    """Collect all existing allocations in the model"""
    allocations = {}

    try:
        print("Starting to collect all existing allocations...")

        # Start from the logical component package
        def search_recursive(java_object):
            try:
                # Check if this object has component exchanges
                if hasattr(java_object, 'getOwnedComponentExchanges'):
                    # Get all component exchanges owned by this object
                    exchanges = java_object.getOwnedComponentExchanges()
                    for ce in exchanges:
                        try:
                            # Get allocations for this component exchange
                            allocations_list = ce.getOwnedComponentExchangeFunctionalExchangeAllocations()
                            for allocation in allocations_list:
                                try:
                                    fe = allocation.getTargetElement()
                                    if fe:
                                        ce_id = ce.getSid()
                                        fe_id = fe.getSid()
                                        allocation_id = allocation.getSid()

                                        # print(f"Found allocation: CE={ce_id}, FE={fe_id}, Alloc={allocation_id}")

                                        if ce_id not in allocations:
                                            allocations[ce_id] = {}
                                        if fe_id not in allocations[ce_id]:
                                            allocations[ce_id][fe_id] = []

                                        allocations[ce_id][fe_id].append(allocation_id)
                                except Exception as e:
                                    print(f"Error processing allocation: {e}")
                        except Exception as e:
                            print(f"Error getting allocations for component exchange: {e}")

                # Recursively check children
                if hasattr(java_object, 'getOwnedLogicalComponents'):
                    children = java_object.getOwnedLogicalComponents()
                    for child in children:
                        search_recursive(child)
            except Exception as e:
                print(f"Error in search_recursive: {e}")

        # Start the recursive search from the logical component package
        lc_pkg_java = lc_pkg.get_java_object()
        search_recursive(lc_pkg_java)

        total_count = sum(len(v) for d in allocations.values() for v in d.values())
        # print(f"Found {total_count} existing allocations")
        return allocations
    except Exception as e:
        print(f"Error in collect_all_allocations: {e}")
        return {}

def delete_unused_allocations(imported_allocations):
    """Delete allocations that exist in the model but were not imported"""
    try:
        deleted_count = 0

        print("Starting to delete unused allocations...")

        # Start from the logical component package
        def search_and_delete_recursive(java_object):
            nonlocal deleted_count
            try:
                # Check if this object has component exchanges
                if hasattr(java_object, 'getOwnedComponentExchanges'):
                    exchanges = java_object.getOwnedComponentExchanges()
                    for ce_java in exchanges:
                        try:
                            ce = ComponentExchange(ce_java)

                            # Get allocations for this component exchange
                            allocations_list = list(ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations())
                            for allocation in allocations_list:
                                try:
                                    fe_java = allocation.getTargetElement()
                                    if fe_java:
                                        fe = FunctionalExchange(fe_java)
                                        ce_id = ce.get_sid()
                                        fe_id = fe.get_sid()
                                        allocation_id = allocation.getSid()

                                        # Check if this allocation should be kept
                                        if (ce_id in imported_allocations and
                                            fe_id in imported_allocations[ce_id] and
                                            allocation_id in imported_allocations[ce_id][fe_id]):
                                            continue  # This allocation is in the import file

                                        # Get names for logging
                                        ce_name = ce.get_name()
                                        fe_name = fe.get_name()

                                        print(f"Deleting allocation: '{ce_name}' (ID: {ce_id}) -> '{fe_name}' (ID: {fe_id}) with allocation ID {allocation_id}")
                                        EObject.delete_e_object(ComponentExchangeFunctionalExchangeAllocation(allocation))
                                        deleted_count += 1

                                except Exception as e:
                                    print(f"Error processing allocation: {e}")
                        except Exception as e:
                            print(f"Error getting allocations for component exchange: {e}")

                # Recursively check children
                if hasattr(java_object, 'getOwnedLogicalComponents'):
                    children = java_object.getOwnedLogicalComponents()
                    for child in children:
                        search_and_delete_recursive(child)
            except Exception as e:
                print(f"Error in search_and_delete_recursive: {e}")

        # Start the recursive search from the logical component package
        lc_pkg_java = lc_pkg.get_java_object()
        search_and_delete_recursive(lc_pkg_java)

        print(f"Deleted {deleted_count} unused allocations")
        return deleted_count

    except Exception as e:
        print(f"Error in delete_unused_allocations: {e}")
        return 0
    
# Start the import
model.start_transaction()

try:
    # First collect all existing allocations
    # print("Collecting existing allocations...")
    existing_allocations = collect_all_allocations()
    total_existing = sum(len(v) for d in existing_allocations.values() for v in d.values())
    # print(f"Found {total_existing} existing allocations")

    # Dictionary to store imported allocations: {ce_id: {fe_id: [allocation_ids]}}
    imported_allocations = {}

    # Process each row in the Excel file
    for row in ws.iter_rows(min_row=2):
        # Extract data from the row
        ce_id = row[0].value
        ce_name = row[1].value
        fe_id = row[10].value
        fe_name = row[11].value
        allocation_id = row[16].value

        # print(f"\nProcessing allocation: {ce_name} -> {fe_name}")

        # Find component exchange
        ce = find_component_exchange_by_id(ce_id)
        if not ce:
            print(f"Skipping - Could not find component exchange with ID {ce_id}")
            continue

        # Find functional exchange
        fe = find_functional_exchange_by_id(fe_id)
        if not fe:
            print(f"Skipping - Could not find functional exchange with ID {fe_id}")
            continue

        # Check if allocation already exists
        if check_allocation_exists(ce, fe):
            # print(f"Allocation already exists between {ce_name} and {fe_name}")

            # Find the existing allocation to get its ID
            existing_allocation = None
            for alloc in ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations():
                if alloc.getTargetElement().getSid() == fe.get_sid():
                    existing_allocation = alloc
                    break

            if existing_allocation:
                allocation_id = existing_allocation.getSid()
                # print(f"Using existing allocation with ID: {allocation_id}")
            else:
                print(f"Could not find existing allocation ID")
                continue
        else:
            # Create the allocation between component exchange and functional exchange
            allocation = create_allocation(ce, fe, allocation_id)
            if not allocation:
                print(f"Skipping - Could not create allocation between {ce_name} and {fe_name}")
                continue
            # print(f"Created new allocation with ID: {allocation_id}")

        # Add to imported allocations
        if ce_id not in imported_allocations:
            imported_allocations[ce_id] = {}
        if fe_id not in imported_allocations[ce_id]:
            imported_allocations[ce_id][fe_id] = []
        imported_allocations[ce_id][fe_id].append(allocation_id)

        # Synchronize port allocations
        synchronize_port_allocations(ce, fe)

        # print(f"Successfully processed allocation {allocation_id}")

    # After ALL imports are complete, delete unused allocations
    # print("\nComparing existing allocations with imported ones...")

    # Count how many allocations should be deleted
    should_delete_count = 0
    for ce_id, fe_dict in existing_allocations.items():
        for fe_id, alloc_ids in fe_dict.items():
            if ce_id in imported_allocations and fe_id in imported_allocations[ce_id]:
                # Check which allocations are not in imported list
                imported_ids = set(imported_allocations[ce_id][fe_id])
                for alloc_id in alloc_ids:
                    if alloc_id not in imported_ids:
                        should_delete_count += 1
            else:
                # All allocations for this CE-FE pair should be deleted
                should_delete_count += len(alloc_ids)

    print(f"Should delete {should_delete_count} allocations")

    # Now actually delete unused allocations
    deleted_count = delete_unused_allocations(imported_allocations)
    print(f"Deleted {deleted_count} allocations that were not in the import file")

    if should_delete_count != deleted_count:
        print(f"Warning: Expected to delete {should_delete_count} allocations but deleted {deleted_count}")

except Exception as e:
    print("Error: " + str(e))
    model.rollback_transaction()
    raise


# else:
model.commit_transaction()

model.save()
print("Exchange allocations import completed successfully")