'''
Created on 25 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import Workbook
import os

# Path names
# aird_path = "/DVS/DVS/DVS.aird"
# aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
aird_path = "/test_ce/test_ce.aird"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()

# Create a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
folder_path = CapellaPlatform.getAbsolutePath(folder)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)

xlsx_file_name = folder_path + "/" + se.get_name() + "_Component_Exchanges.xlsx"

print("Writing " + xlsx_file_name)

# Create a workbook
wb = Workbook()

# Grab the active worksheet
ws = wb.active
ws.title = 'Component Exchanges'

# Create headers with port directions and kinds
ws.cell(row=1, column=1, value="Component From ID")
ws.cell(row=1, column=2, value="Component From Name")
ws.cell(row=1, column=3, value="Component From Port ID")
ws.cell(row=1, column=4, value="Component From Port Name")
ws.cell(row=1, column=5, value="Component From Port Direction")
ws.cell(row=1, column=6, value="Component From Port Kind")
ws.cell(row=1, column=7, value="Component Exchange ID")
ws.cell(row=1, column=8, value="Component Exchange Name")
ws.cell(row=1, column=9, value="Component Exchange Kind")
ws.cell(row=1, column=10, value="Component To ID")
ws.cell(row=1, column=11, value="Component To Name")
ws.cell(row=1, column=12, value="Component To Port ID")
ws.cell(row=1, column=13, value="Component To Port Name")
ws.cell(row=1, column=14, value="Component To Port Direction")
ws.cell(row=1, column=15, value="Component To Port Kind")  

# Track exported exchanges to avoid duplicates
exported_exchanges = set()
current_row = 2  # Start from row 2 (after headers)

def create_component_from_java(java_obj):
    """Create the appropriate component type from a Java object"""
    try:
        if java_obj.isActor():
            return LogicalActor(java_obj)
        else:
            return LogicalComponent(java_obj)
    except Exception as e:
        print(f"Error creating component: {e}")
        return None

def get_component_from_port(port):
    """Get the component that owns a port"""
    if not port:
        return None

    try:
        container = port.get_java_object().eContainer()
        if container:
            return create_component_from_java(container)
    except Exception as e:
        print(f"Error getting component from port: {e}")
        import traceback
        traceback.print_exc()
    return None

def get_port_direction(port):
    """Get the direction of a port as a string"""
    try:
        direction = port.get_orientation()
        if direction == "UNSET":
            return "UNSET"
        elif direction == "IN":
            return "IN"
        elif direction == "OUT":
            return "OUT"
        elif direction == "INOUT":
            return "INOUT"
        else:
            return "UNSET"
    except Exception as e:
        print(f"Error getting port direction: {e}")
        return "UNKNOWN"

def get_port_kind(port):
    """Get the kind of a port (STANDARD or FLOW)"""
    try:
        # Get the Java object
        port_java = port.get_java_object()

        # Check if it's a ComponentPort
        if hasattr(port_java, 'eClass') and port_java.eClass().getName() == 'ComponentPort':
            # Get the kind from the Java object
            kind = port_java.getKind()
            if kind:
                if kind.getName() == 'STANDARD':
                    return "STANDARD"
                elif kind.getName() == 'FLOW':
                    return "FLOW"

        # Default to STANDARD if we can't determine
        return "STANDARD"
    except Exception as e:
        print(f"Error getting port kind: {e}")
        return "STANDARD"

def process_component_exchanges(component, row):
    """Process component exchanges for a component and optionally its children"""
    global ws, exported_exchanges, current_row

    try:
        # Process exchanges owned by this component
        java_obj = component.get_java_object()
        exchanges = java_obj.getOwnedComponentExchanges()

        for ce in exchanges:
            ce_obj = ComponentExchange(ce)
            ce_id = ce_obj.get_id()

            # Skip if we've already exported this exchange
            if ce_id in exported_exchanges:
                continue

            # Get source and target ports
            source_port = ce_obj.get_source_port()
            target_port = ce_obj.get_target_port()

            # Get source and target components from ports
            source_component = get_component_from_port(source_port)
            target_component = get_component_from_port(target_port)

            if source_component and target_component and source_port and target_port:
                # Get port directions
                source_port_dir = get_port_direction(source_port)
                target_port_dir = get_port_direction(target_port)

                # Get port kinds
                source_port_kind = get_port_kind(source_port)
                target_port_kind = get_port_kind(target_port)

                # Write exchange information
                ws.cell(row=current_row, column=1, value=source_component.get_id())
                ws.cell(row=current_row, column=2, value=source_component.get_name())
                ws.cell(row=current_row, column=3, value=source_port.get_id())
                ws.cell(row=current_row, column=4, value=source_port.get_name())
                ws.cell(row=current_row, column=5, value=source_port_dir)
                ws.cell(row=current_row, column=6, value=source_port_kind)
                ws.cell(row=current_row, column=7, value=ce_obj.get_id())
                ws.cell(row=current_row, column=8, value=ce_obj.get_name())
                ws.cell(row=current_row, column=9, value=ce_obj.get_kind())
                ws.cell(row=current_row, column=10, value=target_component.get_id())
                ws.cell(row=current_row, column=11, value=target_component.get_name())
                ws.cell(row=current_row, column=12, value=target_port.get_id())
                ws.cell(row=current_row, column=13, value=target_port.get_name())
                ws.cell(row=current_row, column=14, value=target_port_dir)
                ws.cell(row=current_row, column=15, value=target_port_kind)

                current_row += 1
                exported_exchanges.add(ce_id)
                print(f"Exported exchange: {ce_obj.get_name()} between {source_component.get_name()} and {target_component.get_name()}")

        if java_obj:
            children = java_obj.getOwnedLogicalComponents()
            for child in children:
                child_component = create_component_from_java(child)
                if child_component:
                    process_component_exchanges(child_component, row)

    except Exception as e:
        print(f"Error processing component {component.get_name()}: {e}")
        import traceback
        traceback.print_exc()

    return current_row

# Export all component exchanges
if la:
    print("Exporting all component exchanges...")

    # Process the logical component package directly
    lc_pkg = la.get_logical_component_pkg()
    process_component_exchanges(lc_pkg, current_row)

try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

CapellaPlatform.refresh(folder)
print(f"Export completed. Exported data to {xlsx_file_name}")