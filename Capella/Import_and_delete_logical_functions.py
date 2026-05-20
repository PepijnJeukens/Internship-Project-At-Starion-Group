'''
Created on 13 Apr 2026

@author: p.jeukens
'''
# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
aird_path = "/delete_test/delete_test.aird"
xlsx_path = "/DVS/results/DVS_Logical_Functions2.xlsx"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()
lf_pkg = la.get_logical_function_pkg()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Read " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

def find_function_by_id(parent, function_id):
    """Find a function by its ID in the children of a parent function."""
    for function in parent.get_owned_functions():
        try:
            if function.get_sid() == function_id:
                return function
        except AttributeError:
            if function.get_java_object().getSid() == function_id:
                return function
    return None

def set_function_kind(function, kind):
    """Set the kind of function using the proper Function.set_kind() method"""
    try:
        # Convert to uppercase to match the expected values
        kind = kind.upper()

        # Use the set_kind method from the Function class
        function.set_kind(kind)
        return True
    except Exception as e:
        print(f"Error setting function kind for {function.get_name()}: {e}")
        return False

def collect_all_function_ids(function):
    """Recursively collect all function IDs from a function and its children"""
    function_ids = set()

    try:
        # Add the current function's ID
        function_ids.add(function.get_sid())

        # Recursively collect IDs from children
        for child in function.get_owned_functions():
            function_ids.update(collect_all_function_ids(child))

    except Exception as e:
        print(f"Error collecting function IDs: {e}")

    return function_ids

def delete_function_and_children(function):
    """Delete a function and all its children recursively, printing all deleted items"""
    try:
        # Get the function name and ID for printing
        function_name = function.get_name()
        function_id = function.get_sid()

        # Collect all children to print
        all_to_delete = [function]
        stack = [function]

        # Collect all children recursively
        while stack:
            current = stack.pop()
            try:
                for child in current.get_owned_functions():
                    all_to_delete.append(child)
                    stack.append(child)
            except:
                continue

        # Print all functions being deleted
        print(f"Deleting the following functions:")
        for func in all_to_delete:
            print(f"  - {func.get_name()} (ID: {func.get_sid()})")

        # Delete the main function (which will delete all its children)
        EObject.delete_e_object(function)

    except Exception as e:
        print(f"Error deleting function {function_id}: {e}")

def delete_unused_functions(parent_function, imported_ids):
    """Delete functions that exist in the model but were not imported"""
    try:
        # Get all functions from the parent
        children = list(parent_function.get_owned_functions())  # Convert to list to avoid modification during iteration

        # Create a list of functions to delete
        to_delete = []

        for child in children:
            child_id = child.get_sid()
            if child_id not in imported_ids:
                # This function was not imported, mark it and all its children for deletion
                to_delete.append(child)

        # Delete the functions (in reverse order to avoid index issues)
        for function in reversed(to_delete):
            try:
                delete_function_and_children(function)
            except Exception as e:
                print(f"Error deleting function {function.get_sid()}: {e}")

        # Recursively check children of imported functions
        for child in children:
            child_id = child.get_sid()
            if child_id in imported_ids:
                # This function was imported, check its children
                delete_unused_functions(child, imported_ids)

    except Exception as e:
        print(f"Error in delete_unused_functions: {e}")

# Start the import
model.start_transaction()

try:
    # Check if the root logical function exists
    rlf = None
    for lf in lf_pkg.get_owned_logical_functions():
        if lf.get_name() == "Root Logical Function":
            rlf = lf
            break
    # Set to track imported function IDs
    imported_function_ids = set()

    # If RLF doesn't exist, create it
    if rlf is None:
        # print("Creating Root Logical Function")
        rlf = LogicalFunction()
        lf_pkg.get_owned_logical_functions().add(rlf)
        rlf.set_name("Root Logical Function")
        # Set an ID for RLF if needed
        rlf.get_java_object().setSid("RLF_ID")
    else:
        # print("Root Logical Function already exists")
        # Add RLF to imported IDs
        imported_function_ids.add(rlf.get_sid())
    
    # First, collect all existing function IDs in the model
    # print("Collecting existing function IDs...")
    existing_function_ids = set()
    if rlf:
        existing_function_ids.update(collect_all_function_ids(rlf))
    # print(f"Found {len(existing_function_ids)} existing functions in the model")

    # Determine the maximum depth from the Excel file
    max_col = ws.max_column
    max_depth = max_col // 3  # 3 columns per level (ID, Name, Kind)
    # print(f"Maximum depth detected from Excel: {max_depth}")

    # Create a list to keep track of parent functions at each level
    parent_functions = [None] * max_depth
    
    for row in ws.iter_rows(min_row=2):
        # Create a list of function data for each level
        functions = []
        for level in range(max_depth):
            col = level * 3  # 3 columns per level (ID, Name, Kind)
            function_id = row[col].value if col < len(row) and row[col].value else None
            function_name = row[col + 1].value if col + 1 < len(row) and row[col + 1].value else None
            function_kind = row[col + 2].value if col + 2 < len(row) and row[col + 2].value else None
            functions.append((function_id, function_name, function_kind))

        # Import the hierarchy
        for level in range(max_depth):
            function_id, function_name, function_kind = functions[level]

            if not function_id or not function_name:
                continue  # Skip empty entries

            # For level 0, parent is RLF
            if level == 0:
                parent_function = rlf
            else:
                parent_function = parent_functions[level - 1]

            if parent_function is None:
                print(f"Parent function not found for {function_name} at level {level+1}")
                continue

            # Check if the function already exists
            existing_function = find_function_by_id(parent_function, function_id)

            if existing_function is None:
                # print(f"Creating new function: {function_name} with ID: {function_id} at level {level+1}")
                function = LogicalFunction()
                parent_function.get_owned_functions().add(function)
                org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(function.get_java_object())
                function.set_name(function_name)
                function.get_java_object().setSid(function_id)

                # Set the function kind if provided
                if function_kind:
                    success = set_function_kind(function, function_kind)
                    if not success:
                        print(f"Failed to set function kind to {function_kind} for {function_name}")
                else:
                    print(f"No function kind specified for {function_name}, defaulting to FUNCTION")
            else:
                function = existing_function
                # print(f"Function {function_name} with ID: {function_id} already exists at level {level+1}")

                # Update the function kind if provided
                if function_kind:
                    success = set_function_kind(function, function_kind)
                    if not success:

                        print(f"Failed to update function kind to {function_kind} for {function_name}")

            # Add to imported IDs
            imported_function_ids.add(function_id)

            # Only set parent for next level if this is not the last level
            if level < max_depth - 1:
                parent_functions[level] = function

    # After import, check for functions that need to be deleted
    print("\nChecking for functions to delete...")
    unused_ids = existing_function_ids - imported_function_ids
    print(f"Found {len(unused_ids)} functions that were not imported and will be deleted")

    # Delete unused functions from the root logical function
    if rlf:
        delete_unused_functions(rlf, imported_function_ids)

except Exception as e:
    print("Error: " + str(e))
    model.rollback_transaction()
    raise

# else:
model.commit_transaction()

model.save()
print("Logical functions import completed successfully")