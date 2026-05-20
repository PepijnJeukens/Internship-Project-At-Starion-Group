'''
Created on 26 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
# aird_path = "/Import_test/Import_test.aird"
# xlsx_path = "/DVS/results/DVS_Exchange_Allocations.xlsx"
# aird_path = "/Import_test_IFES/Import_test_IFES.aird"
# xlsx_path = "/In-Flight Entertainment System/results/In-Flight Entertainment System_Exchange_Allocations.xlsx"
aird_path = "/import/import.aird"
xlsx_path = "/test_ce/results/test_ce_Exchange_Allocations.xlsx"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()
lc_pkg = la.get_logical_component_pkg()
lf_pkg = la.get_logical_function_pkg()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Reading " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

def find_component_exchange_by_id(component_id):
    """Find a component exchange by its ID in the model using recursive search"""
    try:
        # Start from the logical component package
        lc_pkg_java = lc_pkg.get_java_object()

        # Use a stack for iterative depth-first search
        stack = [lc_pkg_java]

        while stack:
            current = stack.pop()

            try:
                # Check exchanges owned by this object
                if hasattr(current, 'getOwnedComponentExchanges'):
                    exchanges = current.getOwnedComponentExchanges()
                    for ce in exchanges:
                        if ce.getSid() == component_id:
                            return ComponentExchange(ce)

                # Add children to stack for processing
                if hasattr(current, 'getOwnedLogicalComponents'):
                    children = current.getOwnedLogicalComponents()
                    for child in reversed(children):
                        stack.append(child)
            except:
                continue

        print(f"Could not find component exchange with ID {component_id}")
        return None
    except Exception as e:
        print(f"Error finding component exchange by ID {component_id}: {e}")
        return None

def find_functional_exchange_by_id(exchange_id):
    """Find a functional exchange by its ID in the model"""
    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for fe in all_exchanges:
            if fe.get_sid() == exchange_id:
                return fe

        print(f"Could not find functional exchange with ID {exchange_id}")
        return None
    except Exception as e:
        print(f"Error finding functional exchange by ID {exchange_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def find_functional_exchange_port(fe, expected_port_id):
    """Find a functional exchange port by ID from the functional exchange"""
    try:
        # Get source and target ports from the functional exchange
        source_port = fe.get_source_port()
        target_port = fe.get_target_port()

        # Check if either port matches the expected ID
        if source_port and source_port.get_sid() == expected_port_id:
            return source_port
        if target_port and target_port.get_sid() == expected_port_id:
            return target_port

        print(f"Could not find functional exchange port with ID {expected_port_id} in exchange {fe.get_name()}")
        return None
    except Exception as e:
        print(f"Error finding functional exchange port: {e}")
        return None

def check_allocation_exists(ce, fe):
    """Check if an allocation between component exchange and functional exchange already exists"""
    try:
        # Check all allocations owned by the component exchange
        allocations = ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations()
        for allocation in allocations:
            target = allocation.getTargetElement()
            if target and target.getSid() == fe.get_sid():
                return True
        return False
    except Exception as e:
        print(f"Error checking if allocation exists: {e}")
        return False

def check_port_allocation_exists(fe_port, ce_port):
    """Check if a port allocation already exists between functional and component ports"""
    try:
        if fe_port and ce_port:
            # Check if the functional port already has this component port as allocator
            allocator = fe_port.get_allocator_component_port()
            if allocator and allocator.get_sid() == ce_port.get_sid():
                return True
        return False
    except Exception as e:
        print(f"Error checking if port allocation exists: {e}")
        return False

def create_allocation(ce, fe, allocation_id):
    """Create an allocation between component exchange and functional exchange"""
    try:
        # Check if allocation already exists
        if check_allocation_exists(ce, fe):
            print(f"Allocation already exists between {ce.get_name()} and {fe.get_name()}")
            return None

        # Create the allocation
        allocation = ComponentExchangeFunctionalExchangeAllocation()

        # Add the allocation to the component exchange's owned allocations
        ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations().add(allocation.get_java_object())
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(allocation.get_java_object())

        # Set the ID for the allocation
        allocation.get_java_object().setSid(allocation_id)

        # Set source as component exchange and target as functional exchange
        allocation.get_java_object().setSourceElement(ce.get_java_object())
        allocation.get_java_object().setTargetElement(fe.get_java_object())

        print(f"Successfully created allocation between {ce.get_name()} and {fe.get_name()}")
        return allocation
    except Exception as e:
        print(f"Error creating allocation: {e}")
        import traceback
        traceback.print_exc()
        return None

def synchronize_port_allocations(ce, fe):
    """Synchronize port allocations between component exchange and functional exchange"""
    try:
        # Use the Capella API method to synchronize port allocations
        org.polarsys.capella.core.model.helpers.ComponentExchangeExt.synchronizePortAllocations(
            ce.get_java_object(),
            fe.get_java_object()
        )
        print(f"Successfully synchronized port allocations between {ce.get_name()} and {fe.get_name()}")
        return True
    except Exception as e:
        print(f"Error synchronizing port allocations: {e}")
        import traceback
        traceback.print_exc()
        return False

# Start the import
model.start_transaction()

try:
    # Process each row in the Excel file
    for row in ws.iter_rows(min_row=2):
        # Extract data from the row
        ce_id = row[0].value
        ce_name = row[1].value
        ce_source_port_id = row[2].value
        ce_source_port_name = row[3].value
        source_comp_id = row[4].value
        source_comp_name = row[5].value
        ce_target_port_id = row[6].value
        ce_target_port_name = row[7].value
        target_comp_id = row[8].value
        target_comp_name = row[9].value
        fe_id = row[10].value
        fe_name = row[11].value
        fe_source_port_id = row[12].value
        fe_source_port_name = row[13].value
        fe_target_port_id = row[14].value
        fe_target_port_name = row[15].value
        allocation_id = row[16].value
        source_func_id = row[17].value
        target_func_id = row[18].value

        print(f"\nProcessing allocation: {ce_name} -> {fe_name}")

        # Find component exchange
        ce = find_component_exchange_by_id(ce_id)
        if not ce:
            print(f"Skipping - Could not find component exchange with ID {ce_id}")
            continue

        # Find functional exchange
        fe = find_functional_exchange_by_id(fe_id)
        if not fe:
            print(f"Skipping - Could not find functional exchange with ID {fe_id}")
            continue

        # Create the allocation between component exchange and functional exchange
        allocation = create_allocation(ce, fe, allocation_id)
        if not allocation:
            print(f"Skipping - Could not create allocation between {ce_name} and {fe_name}")
            continue

        # Synchronize port allocations
        synchronize_port_allocations(ce, fe)

        print(f"Successfully processed allocation {allocation_id}")

except Exception as e:
    print("Error: " + str(e))
    import traceback
    traceback.print_exc()
    model.rollback_transaction()
    raise

# else:
model.commit_transaction()

model.save()
print("Exchange allocations import completed successfully")