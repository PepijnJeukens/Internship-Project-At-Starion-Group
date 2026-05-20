'''
Created on 09 Apr 2026

@author: p.jeukens
'''

""" 
This script contains all functions necessary to import the Logical Architecture layer from Capella to Excel
It can import systems (components/actors), functions, the link between systems and functions, functional exchanges,
component exchanges, the link between functional and component exchanges, functional chains and capabilities
"""

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

from openpyxl import load_workbook

def get_model_and_workbook(aird_path, xlsx_path):
    """Open the model and load the workbook"""
    model = CapellaModel()
    model.open(aird_path)

    # Get system engineering
    se = model.get_system_engineering()
    la = se.get_logical_architecture()
    lc_pkg = la.get_logical_component_pkg()
    lf_pkg = la.get_logical_function_pkg()

    # Find the Root Logical Function
    rlf = None
    for lf in lf_pkg.get_owned_logical_functions():
        if lf.get_name() == "Root Logical Function":
            rlf = lf
            break

    # Load the workbook
    xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
    xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)
    print("Read " + xlsx_file_name)
    wb = load_workbook(xlsx_file_name)

    return model, se, la, lc_pkg, lf_pkg, rlf, wb, xlsx_file_name

def find_component_by_id(component_input, component_id):
    """
    Find a component by its ID.
    Works with either a Java object or a list of components.
    """
    try:
        # Case 1: component_input is a Java object (for recursive search)
        if hasattr(component_input, 'getSid') or hasattr(component_input, 'getOwnedLogicalComponents'):
            java_obj = component_input
            # Check if this object matches the ID
            if hasattr(java_obj, 'getSid') and java_obj.getSid() == component_id:
                return java_obj

            # Check children recursively
            if hasattr(java_obj, 'getOwnedLogicalComponents'):
                children = java_obj.getOwnedLogicalComponents()
                for child in children:
                    found = find_component_by_id(child, component_id)
                    if found:
                        return found
            return None

        # Case 2: component_input is a list of components
        elif hasattr(component_input, '__iter__'):
            for component in component_input:
                try:
                    if component.get_sid() == component_id:
                        return component
                except AttributeError:
                    if hasattr(component, 'get_java_object') and component.get_java_object().getSid() == component_id:
                        return component
            return None

        # Unknown input type
        return None
    except Exception as e:
        print(f"Error finding component by ID {component_id}: {e}")
        return None
def get_component_by_id(java_object, component_id):
    """Find a component by its ID in the children of a Java object."""
    for child in java_object.getOwnedLogicalComponents():
        if child.getSid() == component_id:
            return child
    return None

def find_component_recursive(java_object, component_id):
    """Recursively find a component by its ID in the hierarchy."""
    # First check direct children
    component = get_component_by_id(java_object, component_id)
    if component:
        return component

    # If not found, recursively check children
    children = java_object.getOwnedLogicalComponents()
    for child in children:
        found = find_component_recursive(child, component_id)
        if found:
            return found

    return None

def find_function_by_id(parent, function_id):
    """Find a function by ID in the hierarchy"""
    if parent.get_sid() == function_id:
        return parent

    for function in parent.get_owned_functions():
        if function.get_sid() == function_id:
            return function
        found = find_function_by_id(function, function_id)
        if found:
            return found
    return None

def set_function_kind(function, kind):
    """Set the kind of function using the proper Function.set_kind() method"""
    try:
        # Convert to uppercase to match the expected values
        kind = kind.upper()

        # Use the set_kind method from the Function class
        function.set_kind(kind)
        return True
    except Exception as e:
        print(f"Error setting function kind for {function.get_name()}: {e}")
        return False

def get_or_create_rlf(lf_pkg):
    """Get or create the Root Logical Function"""
    # Check if the root logical function exists
    rlf = None
    for lf in lf_pkg.get_owned_logical_functions():
        if lf.get_name() == "Root Logical Function":
            rlf = lf
            break

    # If RLF doesn't exist, create it
    if rlf is None:
        print("Creating Root Logical Function")
        rlf = LogicalFunction()
        lf_pkg.get_owned_logical_functions().add(rlf)
        rlf.set_name("Root Logical Function")
        # Set an ID for RLF if needed
        rlf.get_java_object().setSid("RLF_ID")
    return rlf

def import_systems(wb, model, lc_pkg):
    """Import systems from the Systems worksheet"""
    try:
        ws = wb["Systems"]

        # Determine the maximum depth from the Excel file
        max_col = ws.max_column
        max_depth = max_col // 3

        # Create a list to keep track of parent components at each level
        parent_components = [None] * max_depth

        for row in ws.iter_rows(min_row=2):
            # Create a list of component data for each level
            components = []
            for level in range(max_depth):
                col = level * 3
                component_id = row[col].value if col < len(row) and row[col].value else None
                component_name = row[col + 1].value if col + 1 < len(row) and row[col + 1].value else None
                component_type = row[col + 2].value if col + 2 < len(row) and row[col + 2].value else None
                components.append((component_id, component_name, component_type))

            # Import the hierarchy
            for level in range(max_depth):
                component_id, component_name, component_type = components[level]

                if not component_id:
                    continue  # Skip empty entries

                if level == 0:
                    # Top level component
                    existing_component = find_component_by_id(lc_pkg.get_owned_logical_components(), component_id)
                    if existing_component is None:
                        if component_type == "Actor":
                            component = LogicalActor()
                        else:
                            component = LogicalComponent()
                        lc_pkg.get_owned_logical_components().add(component)
                        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(component.get_java_object())
                        component.set_name(component_name)
                        component.set_sid(component_id)
                    else:
                        component = existing_component
                else:
                    # Child component
                    parent_component = parent_components[level - 1]
                    if parent_component is None:
                        print(f"Parent component not found for {component_name} at level {level}")
                        continue

                    java_parent = parent_component.get_java_object()
                    existing_component = get_component_by_id(java_parent, component_id)

                    if existing_component is None:
                        if component_type == "Actor":
                            component = LogicalActor()
                        else:
                            component = LogicalComponent()
                        java_parent.getOwnedLogicalComponents().add(component.get_java_object())
                        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(component.get_java_object())
                        component.set_name(component_name)
                        component.set_sid(component_id)
                    else:
                        if component_type == "Actor":
                            component = LogicalActor(existing_component)
                        else:
                            component = LogicalComponent(existing_component)

                parent_components[level] = component

        print(f"Imported Systems worksheet successfully")
        return True
    except Exception as e:
        print(f"Error importing systems: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def import_functions(wb, model, lf_pkg):
    """Import functions from the Functions worksheet"""
    try:
        ws = wb["Functions"]

        # Determine the maximum depth from the Excel file
        max_col = ws.max_column
        max_depth = max_col // 3  # 3 columns per level (ID, Name, Kind)

        # Create a list to keep track of parent functions at each level
        parent_functions = [None] * max_depth

        # Get or create RLF
        rlf = get_or_create_rlf(lf_pkg)

        for row in ws.iter_rows(min_row=2):
            # Create a list of function data for each level
            functions = []
            for level in range(max_depth):
                col = level * 3  # 3 columns per level (ID, Name, Kind)
                function_id = row[col].value if col < len(row) and row[col].value else None
                function_name = row[col + 1].value if col + 1 < len(row) and row[col + 1].value else None
                function_kind = row[col + 2].value if col + 2 < len(row) and row[col + 2].value else None
                functions.append((function_id, function_name, function_kind))

            # Import the hierarchy
            for level in range(max_depth):
                function_id, function_name, function_kind = functions[level]

                if not function_id or not function_name:
                    continue  # Skip empty entries

                # For level 0, parent is RLF
                if level == 0:
                    parent_function = rlf
                else:
                    parent_function = parent_functions[level - 1]

                if parent_function is None:
                    print(f"Parent function not found for {function_name} at level {level+1}")
                    continue

                # Check if the function already exists
                existing_function = find_function_by_id(parent_function, function_id)

                if existing_function is None:
                    function = LogicalFunction()
                    parent_function.get_owned_functions().add(function)
                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(function.get_java_object())
                    function.set_name(function_name)
                    function.get_java_object().setSid(function_id)

                    # Set the function kind if provided
                    if function_kind:
                        success = set_function_kind(function, function_kind)
                        if not success:
                            print(f"Failed to set function kind to {function_kind} for {function_name}")
                    else:
                        print(f"No function kind specified for {function_name}, defaulting to FUNCTION")
                else:
                    function = existing_function

                    # Update the function kind if provided
                    if function_kind:
                        success = set_function_kind(function, function_kind)
                        if not success:
                            print(f"Failed to update function kind to {function_kind} for {function_name}")

                # Only set parent for next level if this is not the last level
                if level < max_depth - 1:
                    parent_functions[level] = function

        print(f"Imported Functions worksheet successfully")
        return True
    except Exception as e:
        print(f"Error importing functions: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def import_link_functions_systems(wb, model, lc_pkg, lf_pkg, rlf):
    """Import links between systems and functions from the Link Systems and Functions worksheet"""
    try:
        ws = wb["Link Systems and Functions"]

        # Find the column where "Function ID" appears to determine system depth
        function_id_col = None
        for cell in ws[1]:
            if cell.value == "Function ID":
                function_id_col = cell.column
                break

        if function_id_col is None:
            print("Could not find 'Function ID' column in the Excel file.")
            return False

        # Calculate system depth
        system_max_depth = (function_id_col - 1) // 2

        # Store the current system path for each row
        current_system_java = None
        current_system_id = None
        current_system_level = 0

        for row in ws.iter_rows(min_row=2):
            # Check if this row contains a system (has value in any system column)
            system_found = False
            system_id = None
            system_col = -1

            # Find which system column has a value
            for level in range(system_max_depth):
                col = level * 2
                if col < len(row) and row[col].value:
                    system_id = row[col].value
                    system_col = col
                    system_found = True
                    break  # Use the first system found (highest level)

            # If we found a system, determine its level and find it
            if system_found:
                # Level is determined by column position: col 0-1 = level 0, col 2-3 = level 1, etc.
                current_system_level = system_col // 2
                current_system_id = system_id

                # Find the system using recursive search
                java_lc_pkg = lc_pkg.get_java_object()
                current_system_java = find_component_recursive(java_lc_pkg, current_system_id)

                if not current_system_java:
                    print(f"System with ID {current_system_id} not found at level {current_system_level + 1}")
                    continue

                # Create a Python wrapper for display purposes
                if current_system_java.isActor():
                    current_system = LogicalActor(current_system_java)
                else:
                    current_system = LogicalComponent(current_system_java)


            # If this row contains a function (has value in function ID column)
            if function_id_col <= len(row) and row[function_id_col-1].value:
                function_id = row[function_id_col-1].value
                function_name = row[function_id_col].value if function_id_col < len(row) else ""

                if current_system_java:
                    # Find the function
                    function = find_function_by_id(rlf, function_id)

                    if function:
                        try:
                            # Check if the allocation already exists
                            allocated_functions = current_system_java.getAllocatedLogicalFunctions()
                            already_allocated = False
                            for allocated_func in allocated_functions:
                                if allocated_func.getSid() == function_id:
                                    already_allocated = True
                                    break

                            if not already_allocated:
                                # Create the allocation using the exact same method as the working script
                                try:
                                    # Use the Capella Java API directly
                                    cfa = org.polarsys.capella.core.data.fa.FaFactory.eINSTANCE.createComponentFunctionalAllocation()
                                    cfa.setSourceElement(current_system_java)
                                    cfa.setTargetElement(function.get_java_object())

                                    # Add to the correct containment feature
                                    current_system_java.getOwnedFunctionalAllocation().add(cfa)

                                    # Create the service
                                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(cfa)

                                except Exception as e:
                                    print(f"Failed to create allocation: {str(e)}")
                                    print(f"Failed to allocate function '{function.get_name()}' to system '{current_system.get_name()}': {str(e)}")

                        except Exception as e:
                            print(f"Failed to check existing allocations for function '{function.get_name()}' to system '{current_system.get_name()}': {str(e)}")
                    else:
                        print(f"Function with ID {function_id} and name {function_name} not found.")
                else:
                    print("No current system defined for function allocation.")

        print(f"Imported Link Systems and Functions worksheet successfully")
        return True
    except Exception as e:
        print(f"Error importing link functions systems: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def find_functional_exchange_by_id_existance(exchange_id, se):
    """Find a functional exchange by its ID in the model"""
    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for fe in all_exchanges:
            if fe.get_sid() == exchange_id:
                return fe
        return None
    except Exception as e:
        print(f"Error finding functional exchange by ID {exchange_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def find_or_create_function_port(function, is_input, port_name, port_id):
    """Find an existing function port or create a new one"""
    try:
        if is_input:
            # Check existing input ports
            for port in function.get_inputs():
                if port.get_sid() == port_id:
                    return port
            # Create new input port if not found
            port = FunctionInputPort()
            function.get_inputs().add(port)
            port.set_name(port_name)
            port.get_java_object().setSid(port_id)
            return port
        else:
            # Check existing output ports
            for port in function.get_outputs():
                if port.get_sid() == port_id:
                    return port
            # Create new output port if not found
            port = FunctionOutputPort()
            function.get_outputs().add(port)
            port.set_name(port_name)
            port.get_java_object().setSid(port_id)
            return port
    except Exception as e:
        print(f"Error finding or creating function port {port_name}: {e}")
        return None

def check_functional_exchange_exists(exchange_id, se):
    """Check if a functional exchange with the given ID already exists"""
    try:
        # Find the functional exchange by ID
        existing_fe = find_functional_exchange_by_id_existance(exchange_id, se)
        if existing_fe:
            return True
        return False
    except Exception as e:
        print(f"Error checking if functional exchange exists: {e}")
        return False

def import_functional_exchanges(wb, model, se, lf_pkg):
    """Import functional exchanges from the Functional Exchanges worksheet"""
    try:
        ws = wb["Functional Exchanges"]

        # Find the Root Logical Function
        rlf = None
        for lf in lf_pkg.get_owned_logical_functions():
            if lf.get_name() == "Root Logical Function":
                rlf = lf
                break

        if rlf is None:
            print("Root Logical Function not found!")
            return False

        # Process each row in the Excel file
        for row in ws.iter_rows(min_row=2):
            # Extract data from the row
            source_func_id = row[0].value
            source_func_name = row[1].value
            source_port_id = row[2].value
            source_port_name = row[3].value
            fe_id = row[4].value
            fe_name = row[5].value
            target_func_id = row[6].value
            target_func_name = row[7].value
            target_port_id = row[8].value
            target_port_name = row[9].value

            # Check if the functional exchange already exists by ID
            if check_functional_exchange_exists(fe_id, se):
                continue

            # Find source and target functions
            source_function = find_function_by_id(rlf, source_func_id)
            target_function = find_function_by_id(rlf, target_func_id)

            if not source_function or not target_function:
                print(f"Could not find source or target function for exchange {fe_name}")
                continue

            # Find or create source port (output port for source function)
            source_port = find_or_create_function_port(source_function, False, source_port_name, source_port_id)
            if not source_port:
                print(f"Failed to create/get source port {source_port_name}")
                continue

            # Find or create target port (input port for target function)
            target_port = find_or_create_function_port(target_function, True, target_port_name, target_port_id)
            if not target_port:
                print(f"Failed to create/get target port {target_port_name}")
                continue

            # Create the functional exchange
            fe = FunctionalExchange()

            # Set source and target ports
            fe.set_source_port(source_port)
            fe.set_target_port(target_port)

            source_function.get_owned_functional_exchanges().add(fe)
            org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fe.get_java_object())

            fe.set_name(fe_name)
            fe.get_java_object().setSid(fe_id)

        print(f"Imported Functional Exchanges worksheet successfully")
        return True
    except Exception as e:
        print(f"Error importing functional exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def create_component_from_java(java_obj):
    """Create the appropriate component type from a Java object"""
    try:
        if hasattr(java_obj, 'isActor') and java_obj.isActor():
            return LogicalActor(java_obj)
        else:
            return LogicalComponent(java_obj)
    except Exception as e:
        print(f"Error creating component: {e}")
        return None

def find_or_create_component_port(component, port_id, port_name, port_direction, port_kind):
    """Find an existing component port or create a new one with the specified direction and kind"""
    try:
        java_obj = component.get_java_object()

        # Check existing ports
        for port in java_obj.getOwnedFeatures():
            if hasattr(port, 'getSid') and port.getSid() == port_id:
                # Found existing port, update its direction and kind
                set_port_direction(ComponentPort(port), port_direction)
                set_port_kind(ComponentPort(port), port_kind)
                return ComponentPort(port)

        # Create new port if not found
        port = ComponentPort()
        java_obj.getOwnedFeatures().add(port.get_java_object())
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(port.get_java_object())

        port.set_name(port_name)
        port.get_java_object().setSid(port_id)

        # Set the port direction
        set_port_direction(port, port_direction)

        # Set the port kind
        set_port_kind(port, port_kind)

        return port
    except Exception as e:
        print(f"Error finding or creating component port {port_name}: {e}")
        return None

def set_port_direction(port, direction):
    """Set the direction of a port using the correct method"""
    try:
        # Use the correct method to set orientation
        if direction == "IN":
            port.set_orientation("IN")
        elif direction == "OUT":
            port.set_orientation("OUT")
        elif direction == "INOUT":
            port.set_orientation("INOUT")
        elif direction == "UNSET":
            port.set_orientation("UNSET")
        else:
            print(f"Unknown port direction: {direction}")
    except Exception as e:
        print(f"Error setting port direction: {e}")

def set_port_kind(port, kind):
    """Set the kind of a port (STANDARD or FLOW)"""
    try:
        # Set the kind based on the input
        if kind == "STANDARD":
            port.set_kind("STANDARD")
        elif kind == "FLOW":
            port.set_kind("FLOW")
        else:
            # Default to STANDARD if unknown kind
            port.set_kind("STANDARD")
    except Exception as e:
        print(f"Error setting port kind: {e}")

def get_ancestor_path(component):
    """Get the path from a component to the root as a list of component IDs"""
    path = []
    current = component.get_java_object()
    while current:
        try:
            # Add the current component's ID to the path
            path.append(current.getSid())
            container = current.eContainer()
            if container and hasattr(container, 'getSid'):
                current = container
            else:
                current = None
        except:
            current = None
    return path

def find_common_ancestor(component1, component2, lc_pkg):
    """Find the common ancestor of two components by comparing IDs"""
    try:
        # Get the paths from each component to the root as lists of IDs
        path1 = get_ancestor_path(component1)
        path2 = get_ancestor_path(component2)

        # Find all common ancestor IDs
        common_ancestors = []
        for i in range(len(path1)):
            for j in range(len(path2)):
                if path1[i] == path2[j]:
                    common_ancestors.append(path1[i])

        # If no common ancestors found, use the logical component package
        if not common_ancestors:
            return lc_pkg

        # Use the first common ancestor as you specified
        common_ancestor_id = common_ancestors[0]

        # Find the component with the common ancestor ID
        def find_component_by_id_recursive(java_object, target_id):
            try:
                # Check if this object matches the ID
                if java_object.getSid() == target_id:
                    return java_object

                # Check children
                try:
                    children = java_object.getOwnedLogicalComponents()
                    for child in children:
                        found = find_component_by_id_recursive(child, target_id)
                        if found:
                            return found
                except:
                    pass
            except:
                pass
            return None

        # Search for the component with the common ancestor ID
        common_ancestor_java = find_component_by_id_recursive(lc_pkg.get_java_object(), common_ancestor_id)

        if common_ancestor_java:
            # Create a component object from the Java object
            try:
                # Try to create a LogicalComponent first
                return LogicalComponent(common_ancestor_java)
            except:
                try:
                    # If that fails, try to create a LogicalActor
                    return LogicalActor(common_ancestor_java)
                except:
                    # If both fail, return the logical component package as fallback
                    return lc_pkg
        else:
            # Fallback to logical component package if we can't find the component
            return lc_pkg

    except Exception as e:
        print(f"Error finding common ancestor: {e}")
        return lc_pkg

def import_component_exchanges(wb, model, lc_pkg):
    """Import component exchanges from the Component Exchanges worksheet"""
    try:
        ws = wb["Component Exchanges"]

        # Process each row in the Excel file
        for row in ws.iter_rows(min_row=2):
            # Extract data from the row
            source_comp_id = row[0].value
            source_comp_name = row[1].value
            source_port_id = row[2].value
            source_port_name = row[3].value
            source_port_dir = row[4].value
            source_port_kind = row[5].value
            ce_id = row[6].value
            ce_name = row[7].value
            ce_kind = row[8].value
            target_comp_id = row[9].value
            target_comp_name = row[10].value
            target_port_id = row[11].value
            target_port_name = row[12].value
            target_port_dir = row[13].value
            target_port_kind = row[14].value

            # Find source and target components
            source_java = find_component_by_id(lc_pkg.get_java_object(), source_comp_id)
            target_java = find_component_by_id(lc_pkg.get_java_object(), target_comp_id)

            if not source_java or not target_java:
                print(f"Could not find source or target component for exchange {ce_name}")
                continue

            # Create component objects
            source_component = create_component_from_java(source_java)
            target_component = create_component_from_java(target_java)

            if not source_component or not target_component:
                print(f"Could not create component objects for exchange {ce_name}")
                continue

            # Find or create source port with direction and kind
            source_port = find_or_create_component_port(source_component, source_port_id, source_port_name, source_port_dir, source_port_kind)
            if not source_port:
                print(f"Failed to create/get source port {source_port_name}")
                continue

            # Find or create target port with direction and kind
            target_port = find_or_create_component_port(target_component, target_port_id, target_port_name, target_port_dir, target_port_kind)
            if not target_port:
                print(f"Failed to create/get target port {target_port_name}")
                continue

            # Rest of the function remains the same...
            # Find the common ancestor to store the component exchange
            common_ancestor = find_common_ancestor(source_component, target_component, lc_pkg)

            # Check if the component exchange already exists in the common ancestor
            existing_ce = None
            common_ancestor_java = common_ancestor.get_java_object()
            for ce in common_ancestor_java.getOwnedComponentExchanges():
                if ce.getSid() == ce_id:
                    existing_ce = ComponentExchange(ce)
                    break

            if existing_ce:
                continue

            # Create the component exchange
            ce = ComponentExchange()

            # Set the exchange kind
            if ce_kind:
                ce.set_kind(ce_kind)

            # Set source and target ports
            ce.get_java_object().setSource(source_port.get_java_object())
            ce.get_java_object().setTarget(target_port.get_java_object())

            # Add the exchange to the common ancestor
            common_ancestor.get_java_object().getOwnedComponentExchanges().add(ce.get_java_object())
            org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(ce.get_java_object())

            # Set basic properties
            ce.set_name(ce_name)
            ce.get_java_object().setSid(ce_id)

        print(f"Imported Component Exchanges worksheet successfully")
        return True
    except Exception as e:
        print(f"Error importing component exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
   
def find_component_exchange_by_id(ce_id, lc_pkg):
    """Find a component exchange by its ID in the model using recursive search"""
    try:
        # Start from the logical component package
        lc_pkg_java = lc_pkg.get_java_object()

        # Use a stack for iterative depth-first search
        stack = [lc_pkg_java]

        while stack:
            current = stack.pop()

            try:
                # Check exchanges owned by this object
                if hasattr(current, 'getOwnedComponentExchanges'):
                    exchanges = current.getOwnedComponentExchanges()
                    for ce in exchanges:
                        if ce.getSid() == ce_id:
                            return ComponentExchange(ce)

                # Add children to stack for processing
                if hasattr(current, 'getOwnedLogicalComponents'):
                    children = current.getOwnedLogicalComponents()
                    for child in reversed(children):
                        stack.append(child)
            except:
                continue

        print(f"Could not find component exchange with ID {ce_id}")
        return None
    except Exception as e:
        print(f"Error finding component exchange by ID {ce_id}: {e}")
        return None

def find_functional_exchange_by_id(exchange_id, se):
    """Find a functional exchange by its ID in the model"""
    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for fe in all_exchanges:
            if fe.get_sid() == exchange_id:
                return fe

        print(f"Could not find functional exchange with ID {exchange_id}")
        return None
    except Exception as e:
        print(f"Error finding functional exchange by ID {exchange_id}: {e}")
        import traceback
        traceback.print_exc()
        return None

def check_allocation_exists(ce, fe):
    """Check if an allocation between component exchange and functional exchange already exists"""
    try:
        # Check all allocations owned by the component exchange
        allocations = ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations()
        for allocation in allocations:
            target = allocation.getTargetElement()
            if target and target.getSid() == fe.get_sid():
                return True
        return False
    except Exception as e:
        print(f"Error checking if allocation exists: {e}")
        return False

def create_allocation(ce, fe, allocation_id):
    """Create an allocation between component exchange and functional exchange"""
    try:
        # Check if allocation already exists
        if check_allocation_exists(ce, fe):
            return None

        # Create the allocation
        allocation = ComponentExchangeFunctionalExchangeAllocation()

        # Add the allocation to the component exchange's owned allocations
        ce.get_java_object().getOwnedComponentExchangeFunctionalExchangeAllocations().add(allocation.get_java_object())
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(allocation.get_java_object())

        # Set the ID for the allocation
        allocation.get_java_object().setSid(allocation_id)

        # Set source as component exchange and target as functional exchange
        allocation.get_java_object().setSourceElement(ce.get_java_object())
        allocation.get_java_object().setTargetElement(fe.get_java_object())

        return allocation
    except Exception as e:
        print(f"Error creating allocation: {e}")
        import traceback
        traceback.print_exc()
        return None

def synchronize_port_allocations(ce, fe):
    """Synchronize port allocations between component exchange and functional exchange"""
    try:
        # Use the Capella API method to synchronize port allocations
        org.polarsys.capella.core.model.helpers.ComponentExchangeExt.synchronizePortAllocations(
            ce.get_java_object(),
            fe.get_java_object()
        )
        return True
    except Exception as e:
        print(f"Error synchronizing port allocations: {e}")
        import traceback
        traceback.print_exc()
        return False   
    
def import_link_exchanges(wb, model, se, lc_pkg, lf_pkg):
    """Import link exchanges from the Link Exchanges worksheet"""
    try:
        ws = wb["Link Exchanges"]

        # Process each row in the Excel file
        for row in ws.iter_rows(min_row=2):
            # Extract data from the row
            ce_id = row[0].value
            ce_name = row[1].value
            ce_source_port_id = row[2].value
            ce_source_port_name = row[3].value
            source_comp_id = row[4].value
            source_comp_name = row[5].value
            ce_target_port_id = row[6].value
            ce_target_port_name = row[7].value
            target_comp_id = row[8].value
            target_comp_name = row[9].value
            fe_id = row[10].value
            fe_name = row[11].value
            fe_source_port_id = row[12].value
            fe_source_port_name = row[13].value
            fe_target_port_id = row[14].value
            fe_target_port_name = row[15].value
            allocation_id = row[16].value
            source_func_id = row[17].value
            target_func_id = row[18].value

            # Find component exchange
            ce = find_component_exchange_by_id(ce_id, lc_pkg)
            if not ce:
                print(f"Skipping - Could not find component exchange with ID {ce_id}")
                continue

            # Find functional exchange
            fe = find_functional_exchange_by_id(fe_id, se)
            if not fe:
                print(f"Skipping - Could not find functional exchange with ID {fe_id}")
                continue

            # Check if allocation already exists
            if check_allocation_exists(ce, fe):
                continue

            # Create the allocation between component exchange and functional exchange
            allocation = create_allocation(ce, fe, allocation_id)
            if not allocation:
                print(f"Skipping - Could not create allocation between {ce_name} and {fe_name}")
                continue

            # Synchronize port allocations
            if not synchronize_port_allocations(ce, fe):
                print(f"Warning: Could not synchronize port allocations between {ce_name} and {fe_name}")


        print(f"Imported Link Exchanges worksheet successfully")
        return True
    except Exception as e:
        print(f"Error importing link exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def find_function_by_id_global(function_id, se):
    """Find a function by its ID by searching through all functions in the model"""
    try:
        # First try the hierarchical search starting from RLF
        rlf = None
        for lf in se.get_logical_architecture().get_logical_function_pkg().get_owned_logical_functions():
            if lf.get_name() == "Root Logical Function":
                rlf = lf
                break

        if rlf:
            found = find_function_by_id(rlf, function_id)
            if found:
                return found

        # If not found in hierarchy, search through all functions
        all_functions = se.get_all_contents_by_type(Function)
        for func in all_functions:
            if func.get_sid() == function_id:
                return func
        return None
    except Exception as e:
        print(f"Error finding function by ID {function_id}: {e}")
        return None
    
def functional_chain_exists(fc_id, se):
    """Check if a functional chain already exists"""
    try:
        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)
        for chain in all_chains:
            if chain.get_sid() == fc_id:
                return True
        return False
    except Exception as e:
        print(f"Error checking if functional chain exists: {e}")
        return False

def find_function_involvement(function_involvements, function_id):
    """Find a function involvement by function ID"""
    for fcif in function_involvements:
        try:
            involved = fcif.getInvolvedElement()
            if involved and involved.getSid() == function_id:
                return fcif
        except Exception as e:
            print(f"Error checking function involvement: {e}")
    return None

def create_function_involvement(fc, func, involvement_id=None):
    """Create a functional chain involvement function"""
    try:
        # Use the Capella factory to create the involvement
        factory = org.polarsys.capella.core.data.fa.FaFactory.eINSTANCE
        fcif = factory.createFunctionalChainInvolvementFunction()
        fcif.setInvolved(func.get_java_object())
        fc.get_java_object().getOwnedFunctionalChainInvolvements().add(fcif)
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fcif)
        if involvement_id:
            fcif.setSid(involvement_id)
        return fcif
    except Exception as e:
        print(f"Error creating function involvement: {e}")
        return None

def create_link_involvement(fc, exchange, source_fcif, target_fcif, involvement_id=None):
    """Create a functional chain involvement link"""
    try:
        # Use the Capella factory to create the involvement
        factory = org.polarsys.capella.core.data.fa.FaFactory.eINSTANCE
        fcil = factory.createFunctionalChainInvolvementLink()
        fcil.setInvolved(exchange.get_java_object())
        fcil.setSource(source_fcif)
        fcil.setTarget(target_fcif)
        fc.get_java_object().getOwnedFunctionalChainInvolvements().add(fcil)
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fcil)
        if involvement_id:
            fcil.setSid(involvement_id)
        return fcil
    except Exception as e:
        print(f"Error creating link involvement: {e}")
        return None

def import_functional_chains(wb, model, se, lf_pkg):
    """Import functional chains from the Functional Chains worksheet"""
    try:
        ws = wb["Functional Chains"]

        # Find the column indices for the different sections
        header_row = ws[1]
        function_start_col = 7  # Functions start at column 7 (after FC ID, Name, Start/End Function info)
        exchange_start_col = None

        # Find where the Exchange section starts
        for cell in header_row:
            if cell.value and "Exchange 1 ID" in str(cell.value):
                exchange_start_col = cell.column
                break

        # Find the Root Logical Function
        rlf = None
        for lf in lf_pkg.get_owned_logical_functions():
            if lf.get_name() == "Root Logical Function":
                rlf = lf
                break

        if not rlf:
            print("Root Logical Function not found!")
            return False

        # Process each row in the Excel file (skip header row)
        for row in ws.iter_rows(min_row=2):
            # Extract basic functional chain info
            fc_id = row[0].value
            fc_name = row[1].value

            if not fc_id:
                continue

            # Check if the functional chain already exists
            if functional_chain_exists(fc_id, se):
                continue

            # Create a new functional chain
            fc = FunctionalChain()
            fc.get_java_object().setSid(fc_id)
            fc.set_name(fc_name)

            # Get start and end function info from Excel
            start_func_id = row[2].value
            start_func_name = row[3].value
            end_func_id = row[4].value
            end_func_name = row[5].value

            # Find start and end functions using the global search
            start_func = find_function_by_id_global(start_func_id, se) if start_func_id else None
            end_func = find_function_by_id_global(end_func_id, se) if end_func_id else None


            if not start_func:
                print(f"Could not find start function with ID {start_func_id} and name {start_func_name}")
                continue
            if not end_func:
                print(f"Could not find end function with ID {end_func_id} and name {end_func_name}")
                continue

            # Process functions and their involvements
            function_col = function_start_col - 1  # Convert to 0-based index
            function_involvements = []
            involved_functions = []
            function_order = []  # To keep track of the order of functions

            # Process functions until we reach the exchange section
            while function_col < exchange_start_col - 1:  # Stop before the exchange section
                try:
                    # Get function info
                    func_id = row[function_col].value
                    func_name = row[function_col + 1].value
                    func_involvement_id = row[function_col + 2].value

                    if not func_id:
                        break

                    # Find the function
                    func = find_function_by_id_global(func_id, se)
                    if not func:
                        print(f"Could not find function with ID {func_id} and name {func_name}")
                        function_col += 3
                        continue

                    # Create function involvement
                    fcif = create_function_involvement(fc, func, func_involvement_id)
                    if fcif:
                        function_involvements.append(fcif)
                        involved_functions.append(func)
                        function_order.append(func_id)

                    function_col += 3
                except Exception as e:
                    print(f"Error processing function: {e}")
                    function_col += 3
                    continue

            # Find the start and end function involvements
            start_fcif = find_function_involvement(function_involvements, start_func_id)
            end_fcif = find_function_involvement(function_involvements, end_func_id)

            if not start_fcif:
                print(f"Could not find start function involvement for function ID {start_func_id}")
                continue
            if not end_fcif:
                print(f"Could not find end function involvement for function ID {end_func_id}")
                continue

            # Process exchanges and their involvements
            exchange_col = exchange_start_col - 1  # Convert to 0-based index

            # Create a path from start to end function
            # First, find the path from start to end function
            path = [start_fcif]
            current_fcif = start_fcif

            # Try to find a path from start to end
            while current_fcif != end_fcif:
                found_next = False
                for fcif in function_involvements:
                    if fcif == current_fcif:
                        continue

                    # Check if this function is connected to the current one
                    next_func_id = fcif.getInvolvedElement().getSid()

                    # Check if this function is in the Excel function list
                    found_in_excel = False
                    for i in range(0, len(row), 3):
                        if i >= function_start_col - 1 and i < exchange_start_col - 1:
                            if row[i].value == next_func_id:
                                found_in_excel = True
                                break

                    if found_in_excel and next_func_id not in [f.getInvolvedElement().getSid() for f in path]:
                        path.append(fcif)
                        current_fcif = fcif
                        found_next = True
                        break

                if not found_next:
                    # If we can't find a path, just add the end function
                    if end_fcif not in path:
                        path.append(end_fcif)
                    break

            # Now create links between consecutive functions in the path
            exchange_index = 0
            while exchange_col < len(row):
                try:
                    # Get exchange info
                    exchange_id = row[exchange_col].value
                    exchange_name = row[exchange_col + 1].value
                    exchange_involvement_id = row[exchange_col + 2].value

                    if not exchange_id:
                        break

                    # Find the functional exchange
                    exchange = find_functional_exchange_by_id(exchange_id, se)
                    if not exchange:
                        print(f"Could not find functional exchange with ID {exchange_id} and name {exchange_name}")
                        exchange_col += 3
                        continue

                    # Create link between consecutive functions in the path
                    if exchange_index < len(path) - 1:
                        source_fcif = path[exchange_index]
                        target_fcif = path[exchange_index + 1]
                        create_link_involvement(fc, exchange, source_fcif, target_fcif, exchange_involvement_id)
                        exchange_index += 1
                    else:
                        # If we have more exchanges than path segments, use start and end
                        create_link_involvement(fc, exchange, start_fcif, end_fcif, exchange_involvement_id)

                    exchange_col += 3
                except Exception as e:
                    print(f"Error processing exchange: {e}")
                    exchange_col += 3
                    continue

            # Add the functional chain to the root logical function
            try:
                rlf.get_java_object().getOwnedFunctionalChains().add(fc.get_java_object())
                org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fc.get_java_object())
                fc.get_java_object().setSid(fc_id)
                fc.set_name(fc_name)
            except Exception as e:
                print(f"Error adding functional chain to root logical function: {e}")

            # Add the functional chain to the involved functions
            for func in involved_functions:
                try:
                    # Add the functional chain to the function's owned functional chains
                    func.get_java_object().getOwnedFunctionalChains().add(fc.get_java_object())
                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fc.get_java_object())
                    fc.get_java_object().setSid(fc_id)
                    fc.set_name(fc_name)
                except Exception as e:
                    print(f"Error adding functional chain to function {func.get_name()}: {e}")

        print(f"Imported Functional Chains worksheet successfully")
        return True
    except Exception as e:
        print(f"Error importing functional chains: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def get_java_object(obj):
    """Get the Java object from a Python wrapper or return the object itself"""
    try:
        if hasattr(obj, 'get_java_object'):
            return obj.get_java_object()
        return obj
    except Exception as e:
        print(f"Error getting Java object: {str(e)}")
        return None    

def get_component_id(component):
    """Get the ID of a component, handling both LogicalComponent and LogicalActor."""
    try:
        if hasattr(component, 'get_sid'):
            return component.get_sid()
        elif hasattr(component, 'getSid'):
            return component.getSid()
        elif hasattr(component, 'get_java_object'):
            java_obj = component.get_java_object()
            if hasattr(java_obj, 'getSid'):
                return java_obj.getSid()
    except Exception as e:
        print(f"Error getting component ID: {e}")
    return None

def find_capability_realization_by_id(capability_realization_pkg, cr_id):
    """Find a capability realization by its ID"""
    try:
        java_pkg = get_java_object(capability_realization_pkg)
        if not java_pkg:
            return None

        for cr in java_pkg.getOwnedCapabilityRealizations():
            if cr.getSid() == cr_id:
                return CapabilityRealization(cr)
        return None
    except Exception as e:
        print(f"Error finding capability realization by ID: {e}")
        return None

def find_functional_chain_by_id(fc_id, se):
    """Find a functional chain by its ID by searching through all functions"""
    try:
        # Get all functions in the model
        all_functions = se.get_all_contents_by_type(Function)

        # Search through all functions for owned functional chains
        for func in all_functions:
            java_func = get_java_object(func)
            for chain in java_func.getOwnedFunctionalChains():
                if chain.getSid() == fc_id:
                    return FunctionalChain(chain)

        return None
    except Exception as e:
        print(f"Error finding functional chain by ID: {e}")
        return None

def find_logical_function_by_id(func_id, se):
    """Find a logical function by its ID"""
    try:
        # Get all logical functions from the model
        all_functions = se.get_all_contents_by_type(LogicalFunction)
        for func in all_functions:
            if get_component_id(func) == func_id:
                return func
        return None
    except Exception as e:
        print(f"Error finding logical function by ID: {e}")
        return None

def find_logical_component_by_id(comp_id, se):
    """Find a logical component by its ID"""
    try:
        # Get all logical components from the model
        all_components = se.get_all_contents_by_type(LogicalComponent)
        for comp in all_components:
            if get_component_id(comp) == comp_id:
                return comp
        return None
    except Exception as e:
        print(f"Error finding logical component by ID: {e}")
        return None

def find_logical_actor_by_id(actor_id, se):
    """Find a logical actor by its ID"""
    try:
        # Get all logical actors from the model
        all_actors = se.get_all_contents_by_type(LogicalActor)
        for actor in all_actors:
            if get_component_id(actor) == actor_id:
                return actor
        return None
    except Exception as e:
        print(f"Error finding logical actor by ID: {e}")
        return None

def create_function_capability_involvement(cr, element, involvement_id=None):
    """Create an involvement between a function and a capability realization"""
    try:
        # Create the involvement object using the factory
        factory = org.polarsys.capella.core.data.interaction.InteractionFactory.eINSTANCE
        involvement = factory.createAbstractFunctionAbstractCapabilityInvolvement()

        # Get the Java objects
        java_cr = get_java_object(cr)
        java_element = get_java_object(element)

        # Add the involvement to the capability realization
        java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements().add(involvement)

        # Set the involved element
        involvement.setInvolved(java_element)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)

        if involvement_id:
            involvement.setSid(involvement_id)

        return involvement
    except Exception as e:
        print(f"Error creating function/capability involvement: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_functional_chain_capability_involvement(cr, fc, involvement_id=None):
    """Create an involvement between a functional chain and a capability realization"""
    try:
        # Create the involvement object using the factory
        factory = org.polarsys.capella.core.data.interaction.InteractionFactory.eINSTANCE
        involvement = factory.createFunctionalChainAbstractCapabilityInvolvement()

        # Get the Java objects
        java_cr = get_java_object(cr)
        java_element = get_java_object(fc)

        # Add the involvement to the capability realization
        java_cr.getOwnedFunctionalChainAbstractCapabilityInvolvements().add(involvement)

        # Set the involved element
        involvement.setInvolved(java_element)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)

        if involvement_id:
            involvement.setSid(involvement_id)

        return involvement
    except Exception as e:
        print(f"Error creating function/capability involvement: {e}")
        import traceback
        traceback.print_exc()
        return None

def add_component_to_capability_realization(cr, component, involvement_id=None):
    """Add a component or actor to a capability realization"""
    try:
        # Get the Java objects
        java_cr = get_java_object(cr)
        java_component = get_java_object(component)

        # Check if the component is already involved
        already_involved = False
        for involved in java_cr.getInvolvedComponents():
            if involved.getSid() == java_component.getSid():
                already_involved = True
                break

        if already_involved:
            return True

        # Create the involvement using the factory
        factory = org.polarsys.capella.core.data.capellacommon.CapellacommonFactory.eINSTANCE
        involvement = factory.createCapabilityRealizationInvolvement()

        # Set the involved element
        involvement.setInvolved(java_component)

        # Add the involvement to the capability realization
        java_cr.getOwnedCapabilityRealizationInvolvements().add(involvement)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)

        # Set the involvement ID if provided
        if involvement_id:
            involvement.setSid(involvement_id)

        return True
    except Exception as e:
        print(f"Error adding component to capability realization: {e}")
        import traceback
        traceback.print_exc()
        return False

def import_capabilities(wb, model, se, la):
    """Import capabilities from the Capabilities worksheet"""
    try:
        ws = wb["Capabilities"]

        # Find max functional chains, logical functions, and involved components from headers
        max_functional_chains = 0
        max_logical_functions = 0
        max_involved_components = 0

        function_start_col = 3
        for col in range(function_start_col, ws.max_column + 1, 3):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "Functional Chain" in cell_value:
                max_functional_chains += 1

        function_end_col = function_start_col + max_functional_chains * 3
        for col in range(function_end_col, ws.max_column + 1, 3):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "Logical Function" in cell_value:
                max_logical_functions += 1

        component_start_col = function_end_col + max_logical_functions * 3
        for col in range(component_start_col, ws.max_column + 1, 4):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and "Involved Component" in cell_value:
                max_involved_components += 1

        # Get the capability realization package
        capability_realization_pkg = la.get_capability_realization_pkg()
        if not capability_realization_pkg:
            print("No Capability Realization Package found in Logical Architecture")
            return False

        # Process each row in the Excel file (skip header row)
        for row in ws.iter_rows(min_row=2):
            try:
                # Extract basic capability realization info
                cr_id = row[0].value
                cr_name = row[1].value

                if not cr_id:
                    continue

                # Check if the capability realization already exists
                existing_cr = find_capability_realization_by_id(capability_realization_pkg, cr_id)
                if existing_cr:
                    continue

                # Create a new capability realization
                cr = CapabilityRealization()
                java_cr = get_java_object(cr)

                # Add to the capability realization package first
                java_pkg = get_java_object(capability_realization_pkg)
                java_pkg.getOwnedCapabilityRealizations().add(java_cr)
                org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(java_cr)

                # Now set the ID and name
                java_cr.setSid(cr_id)
                cr.set_name(cr_name)

                # Process functional chains
                for i in range(max_functional_chains):
                    fc_id_col = function_start_col + i*3 - 1  # Convert to 0-based index
                    fc_name_col = function_start_col + i*3
                    fc_involvement_col = function_start_col + i*3 + 1

                    if fc_id_col >= len(row) or not row[fc_id_col].value:
                        continue

                    fc_id = row[fc_id_col].value
                    fc_name = row[fc_name_col].value if fc_name_col < len(row) else ""
                    fc_involvement_id = row[fc_involvement_col].value if fc_involvement_col < len(row) else ""

                    if not fc_id:
                        continue

                    # Find the functional chain
                    fc = find_functional_chain_by_id(fc_id, se)
                    if not fc:
                        print(f"WARNING: Could not find functional chain with ID {fc_id} and name '{fc_name}'. Skipping this chain.")
                        continue

                    # Create involvement between functional chain and capability realization
                    involvement = create_functional_chain_capability_involvement(cr, fc, fc_involvement_id)
                    if not involvement:
                        print(f"Failed to create involvement for functional chain {fc_name}")

                # Process logical functions
                for i in range(max_logical_functions):
                    func_id_col = function_end_col + i*3 - 1  # Convert to 0-based index
                    func_name_col = function_end_col + i*3
                    func_involvement_col = function_end_col + i*3 + 1

                    if func_id_col >= len(row) or not row[func_id_col].value:
                        continue

                    func_id = row[func_id_col].value
                    func_name = row[func_name_col].value if func_name_col < len(row) else ""
                    func_involvement_id = row[func_involvement_col].value if func_involvement_col < len(row) else ""

                    if not func_id:
                        continue

                    # Find the logical function
                    func = find_logical_function_by_id(func_id, se)
                    if not func:
                        print(f"WARNING: Could not find logical function with ID {func_id} and name '{func_name}'. Skipping this function.")
                        continue

                    # Create involvement between logical function and capability realization
                    involvement = create_function_capability_involvement(cr, func, func_involvement_id)
                    if not involvement:
                        print(f"Failed to create involvement for logical function {func_name}")

                # Process involved components
                for i in range(max_involved_components):
                    comp_id_col = component_start_col + i*4 - 1  # Convert to 0-based index
                    comp_name_col = component_start_col + i*4
                    comp_type_col = component_start_col + i*4 + 1
                    comp_involvement_col = component_start_col + i*4 + 2

                    if comp_id_col >= len(row) or not row[comp_id_col].value:
                        continue

                    comp_id = row[comp_id_col].value
                    comp_name = row[comp_name_col].value if comp_name_col < len(row) else ""
                    comp_type = row[comp_type_col].value if comp_type_col < len(row) else ""
                    comp_involvement_id = row[comp_involvement_col].value if comp_involvement_col < len(row) else ""

                    if not comp_id:
                        continue

                    # Find the component or actor based on type
                    component = None
                    if comp_type == "actor":
                        component = find_logical_actor_by_id(comp_id, se)
                    else:
                        component = find_logical_component_by_id(comp_id, se)

                    if not component:
                        print(f"WARNING: Could not find {comp_type} with ID {comp_id} and name '{comp_name}'. Skipping this component.")
                        continue

                    # Add component/actor to capability realization
                    success = add_component_to_capability_realization(cr, component, comp_involvement_id)
                    if not success:
                        print(f"Failed to add {comp_type} {comp_name} to capability realization {cr_name}")

            except Exception as e:
                print(f"Error processing row: {str(e)}")
                import traceback
                traceback.print_exc()
                continue

        print(f"Imported Capabilities worksheet successfully")
        return True
    except Exception as e:
        print(f"Error importing capabilities: {str(e)}")
        return False

def full_import(aird_path, xlsx_path):
    """
    Perform a full import of all model elements from a single Excel workbook
    with multiple worksheets.
    """
    try:
        # Get model and workbook
        model, se, la, lc_pkg, lf_pkg, rlf, wb, xlsx_file_name = get_model_and_workbook(aird_path, xlsx_path)

        # Start the import
        model.start_transaction()

        # Import all elements
        success = True
        success &= import_systems(wb, model, lc_pkg)
        success &= import_functions(wb, model, lf_pkg)
        success &= import_link_functions_systems(wb, model, lc_pkg, lf_pkg, rlf)
        success &= import_functional_exchanges(wb, model, se, lf_pkg)
        success &= import_component_exchanges(wb, model, lc_pkg)
        success &= import_link_exchanges(wb, model, se, lc_pkg, lf_pkg)
        success &= import_functional_chains(wb, model, se, lf_pkg)
        success &= import_capabilities(wb, model, se, la)

        if success:
            model.commit_transaction()
            model.save()
            print("Full import completed successfully!")
            return True
        else:
            model.rollback_transaction()
            print("Some imports failed. Check error messages above.")
            return False

    except Exception as e:
        print(f"Error in full import: {str(e)}")
        import traceback
        traceback.print_exc()
        model.rollback_transaction()
        return False