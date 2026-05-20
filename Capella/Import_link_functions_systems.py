'''
Created on 24 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
aird_path = "/Import_test/Import_test.aird"
xlsx_path = "/DVS/results/DVS_Function_System_Links.xlsx"
# aird_path = "/Import_test_IFES/Import_test_IFES.aird"
# xlsx_path = "/In-Flight Entertainment System/results/In-Flight Entertainment System_Function_System_Links.xlsx"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()

# Get the logical component package and root logical function
lc_pkg = la.get_logical_component_pkg()
lf_pkg = la.get_logical_function_pkg()

# Find the Root Logical Function
rlf = None
for lf in lf_pkg.get_owned_logical_functions():
    if lf.get_name() == "Root Logical Function":
        rlf = lf
        break

if rlf is None:
    print("Root Logical Function not found!")
    exit()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Read " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

def get_component_by_id(java_object, component_id):
    """Find a component by its ID in the children of a Java object."""
    for child in java_object.getOwnedLogicalComponents():
        if child.getSid() == component_id:
            return child
    return None

def find_component_recursive(java_object, component_id):
    """Recursively find a component by its ID in the hierarchy."""
    # First check direct children
    component = get_component_by_id(java_object, component_id)
    if component:
        return component

    # If not found, recursively check children
    children = java_object.getOwnedLogicalComponents()
    for child in children:
        found = find_component_recursive(child, component_id)
        if found:
            return found

    return None

def find_function_by_id(parent, function_id):
    """Find a function by ID in the hierarchy"""
    if parent.get_sid() == function_id:
        return parent

    for function in parent.get_owned_functions():
        if function.get_sid() == function_id:
            return function
        found = find_function_by_id(function, function_id)
        if found:
            return found
    return None

# Find the column where "Function ID" appears to determine system depth
function_id_col = None
for cell in ws[1]:
    if cell.value == "Function ID":
        function_id_col = cell.column
        break

if function_id_col is None:
    print("Could not find 'Function ID' column in the Excel file.")
    exit()

# Calculate system depth
system_max_depth = (function_id_col - 1) // 2

print(f"Maximum system depth: {system_max_depth}")

model.start_transaction()

try:
    # Store the current system path for each row
    current_system_java = None
    current_system_id = None
    current_system_level = 0

    for row in ws.iter_rows(min_row=2):
        # Check if this row contains a system (has value in any system column)
        system_found = False
        system_id = None
        system_col = -1

        # Find which system column has a value
        for level in range(system_max_depth):
            col = level * 2
            if col < len(row) and row[col].value:
                system_id = row[col].value
                system_col = col
                system_found = True
                break  # Use the first system found (highest level)

        # If we found a system, determine its level and find it
        if system_found:
            # Level is determined by column position: col 0-1 = level 0, col 2-3 = level 1, etc.
            current_system_level = system_col // 2
            current_system_id = system_id

            # Find the system using recursive search
            java_lc_pkg = lc_pkg.get_java_object()
            current_system_java = find_component_recursive(java_lc_pkg, current_system_id)

            if not current_system_java:
                print(f"System with ID {current_system_id} not found at level {current_system_level + 1}")
                continue

            # Create a Python wrapper for display purposes
            if current_system_java.isActor():
                current_system = LogicalActor(current_system_java)
            else:
                current_system = LogicalComponent(current_system_java)

            print(f"Found system: {current_system.get_name()} at level {current_system_level + 1}")
    
    
        # If this row contains a function (has value in function ID column)
        if function_id_col <= len(row) and row[function_id_col-1].value:
            function_id = row[function_id_col-1].value
            function_name = row[function_id_col].value if function_id_col < len(row) else ""

            if current_system_java:
                # Find the function
                function = find_function_by_id(rlf, function_id)

                if function:
                    try:
                        # Check if the allocation already exists
                        allocated_functions = current_system_java.getAllocatedLogicalFunctions()
                        already_allocated = False
                        for allocated_func in allocated_functions:
                            if allocated_func.getSid() == function_id:
                                already_allocated = True
                                break

                        if not already_allocated:
                            # Create the allocation
                            cfa = create_e_object("http://www.polarsys.org/capella/core/fa/" + capella_version(), "ComponentFunctionalAllocation")
                            cfa.setSourceElement(current_system_java)
                            cfa.setTargetElement(function.get_java_object())
                            current_system_java.getOwnedFunctionalAllocation().add(cfa)
                            org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(cfa)
                            print(f"Allocated function '{function.get_name()}' to system '{current_system.get_name()}'.")
                        else:
                            print(f"Function '{function.get_name()}' is already allocated to system '{current_system.get_name()}'.")
                    except Exception as e:
                        print(f"Failed to allocate function '{function.get_name()}' to system '{current_system.get_name()}': {str(e)}")
                else:
                    print(f"Function with ID {function_id} and name {function_name} not found.")
            else:
                print("No current system defined for function allocation.")

except Exception as e:
    print("Error: " + str(e))
    import traceback
    traceback.print_exc()
    model.rollback_transaction()
    raise

# else:
model.commit_transaction()

model.save()
print("Functions and systems linked successfully")