'''
Created on 27 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import Workbook
import os

# Path names
# aird_path = "/DVS/DVS/DVS.aird"
aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering and Logical Architecture
se = model.get_system_engineering()
la = se.get_logical_architecture()

# Get the root logical function
lf_pkg = la.get_logical_function_pkg()
rlf = None
for lf in lf_pkg.get_owned_logical_functions():
    if lf.get_name() == "Root Logical Function":
        rlf = lf
        break

if not rlf:
    print("Root Logical Function not found!")
    exit()

# Collect all logical function IDs
logical_function_ids = set()

def collect_logical_function_ids(function):
    """Recursively collect all logical function IDs"""
    try:
        # Add current function ID
        logical_function_ids.add(function.get_id())

        # Process children
        children = function.get_java_object().getOwnedFunctions()
        for child in children:
            child_function = LogicalFunction(child)
            collect_logical_function_ids(child_function)
    except Exception as e:
        print(f"Error collecting function IDs: {e}")

# Start collecting from the root logical function
collect_logical_function_ids(rlf)
print(f"Collected {len(logical_function_ids)} logical function IDs")

# Create a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
folder_path = CapellaPlatform.getAbsolutePath(folder)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)

# Set the correct file name format
xlsx_file_name = folder_path + "/" + se.get_name() + "_Functional_Chains.xlsx"

print("Writing " + xlsx_file_name)

# Create a workbook
wb = Workbook()
ws = wb.active
ws.title = 'Functional Chains'

# Helper function to find start and end functions of a functional chain
def get_start_end_functions(fc):
    """Find the start and end functions of a functional chain"""
    start_func = None
    end_func = None

    try:
        # Get all functional chain involvements
        involvements = fc.get_java_object().getOwnedFunctionalChainInvolvements()

        # Find all function involvements
        function_involvements = []
        for involvement in involvements:
            if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                function_involvements.append(involvement)

        # Find all link involvements
        link_involvements = []
        for involvement in involvements:
            if involvement.eClass().getName() == "FunctionalChainInvolvementLink":
                link_involvements.append(involvement)

        # If we have function involvements but no links, all functions are start and end
        if len(function_involvements) > 0 and len(link_involvements) == 0:
            start_func = function_involvements[0].getInvolvedElement()
            end_func = function_involvements[-1].getInvolvedElement()
        else:
            # Find functions that are sources but not targets of any link
            sources = set()
            targets = set()

            for link in link_involvements:
                if link.getSource() and hasattr(link.getSource(), 'getInvolvedElement'):
                    sources.add(link.getSource().getInvolvedElement())
                if link.getTarget() and hasattr(link.getTarget(), 'getInvolvedElement'):
                    targets.add(link.getTarget().getInvolvedElement())

            # Start functions are those that are sources but not targets
            for func_inv in function_involvements:
                func = func_inv.getInvolvedElement()
                if func in sources and func not in targets:
                    start_func = func
                    break

            # End functions are those that are targets but not sources
            for func_inv in function_involvements:
                func = func_inv.getInvolvedElement()
                if func in targets and func not in sources:
                    end_func = func
                    break

            # If we couldn't find start/end functions, use first/last
            if not start_func and len(function_involvements) > 0:
                start_func = function_involvements[0].getInvolvedElement()
            if not end_func and len(function_involvements) > 0:
                end_func = function_involvements[-1].getInvolvedElement()

    except Exception as e:
        print(f"Error finding start/end functions: {e}")
        # Fallback to first and last functions
        involved_functions = fc.get_involved_functions()
        if len(involved_functions) > 0:
            start_func = involved_functions[0].get_java_object()
            end_func = involved_functions[-1].get_java_object()

    return start_func, end_func

# Helper function to get ordered functions in a functional chain
def get_ordered_functions(fc):
    """Get functions in a functional chain in order from start to end"""
    try:
        # Get start and end functions
        start_func, end_func = get_start_end_functions(fc)

        # Get all function involvements
        involvements = fc.get_java_object().getOwnedFunctionalChainInvolvements()
        function_involvements = []
        for involvement in involvements:
            if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                function_involvements.append(involvement)

        # Get all link involvements
        link_involvements = []
        for involvement in involvements:
            if involvement.eClass().getName() == "FunctionalChainInvolvementLink":
                link_involvements.append(involvement)

        # If no links, return all functions in order
        if len(link_involvements) == 0:
            return [fi.getInvolvedElement() for fi in function_involvements]

        # Build a graph of function connections
        graph = {}
        for link in link_involvements:
            source = link.getSource()
            target = link.getTarget()

            if source and target:
                source_func = source.getInvolvedElement()
                target_func = target.getInvolvedElement()

                if source_func.getId() not in graph:
                    graph[source_func.getId()] = []
                graph[source_func.getId()].append(target_func.getId())

                if target_func.getId() not in graph:
                    graph[target_func.getId()] = []

        # Find a path from start to end
        start_id = start_func.getId()
        end_id = end_func.getId()

        # Use BFS to find the path
        queue = [[start_id]]
        visited = set()

        while queue:
            path = queue.pop(0)
            node = path[-1]

            if node == end_id:
                # Found the path
                ordered_func_ids = path
                # Convert IDs to function objects
                ordered_functions = []
                for func_id in ordered_func_ids:
                    for fi in function_involvements:
                        if fi.getInvolvedElement().getId() == func_id:
                            ordered_functions.append(fi.getInvolvedElement())
                            break
                return ordered_functions

            if node not in visited:
                visited.add(node)
                for neighbor in graph.get(node, []):
                    if neighbor not in visited:
                        new_path = list(path)
                        new_path.append(neighbor)
                        queue.append(new_path)

        # If no path found, return all functions in arbitrary order
        return [fi.getInvolvedElement() for fi in function_involvements]

    except Exception as e:
        print(f"Error getting ordered functions: {e}")
        # Fallback to all functions in arbitrary order
        return fc.get_involved_functions()

# Find all functional chains in the model
all_functional_chains = se.get_all_contents_by_type(FunctionalChain)

# Filter to only include functional chains where ALL involved functions are in the logical architecture
logical_functional_chains = []
for fc in all_functional_chains:
    try:
        # Get all involved functions
        involved_functions = fc.get_involved_functions()

        # If no functions are involved, skip this chain
        if not involved_functions or len(involved_functions) == 0:
            continue

        # Check if ALL involved functions are in the logical architecture
        all_in_logical = True
        for func in involved_functions:
            func_id = func.get_id()
            if func_id not in logical_function_ids:
                all_in_logical = False
                break

        # Only add to our list if all functions are in the logical architecture
        if all_in_logical:
            logical_functional_chains.append(fc)
    except Exception as e:
        print(f"Error checking functional chain {fc.get_name()}: {e}")
        continue

print(f"Found {len(logical_functional_chains)} logical functional chains")

# Determine the maximum number of functions and exchanges in any chain
max_functions = 0
max_exchanges = 0

for fc in logical_functional_chains:
    try:
        # Get ordered functions
        ordered_functions = get_ordered_functions(fc)
        involved_exchanges = fc.get_involved_functional_exchanges()

        if len(ordered_functions) > max_functions:
            max_functions = len(ordered_functions)
        if len(involved_exchanges) > max_exchanges:
            max_exchanges = len(involved_exchanges)
    except Exception as e:
        print(f"Error determining max functions/exchanges for {fc.get_name()}: {e}")

# Create headers
ws.cell(row=1, column=1, value="Functional Chain ID")
ws.cell(row=1, column=2, value="Functional Chain Name")
ws.cell(row=1, column=3, value="Start Function ID")
ws.cell(row=1, column=4, value="Start Function Name")
ws.cell(row=1, column=5, value="End Function ID")
ws.cell(row=1, column=6, value="End Function Name")

# Add headers for functions and their involvements
function_start_col = 7
for i in range(max_functions):
    ws.cell(row=1, column=function_start_col + i*3, value=f"Function {i+1} ID")
    ws.cell(row=1, column=function_start_col + i*3 + 1, value=f"Function {i+1} Name")
    ws.cell(row=1, column=function_start_col + i*3 + 2, value=f"Function {i+1} Involvement ID")

# Add headers for exchanges and their involvements after the last function
exchange_start_col = function_start_col + max_functions*3
for i in range(max_exchanges):
    ws.cell(row=1, column=exchange_start_col + i*3, value=f"Exchange {i+1} ID")
    ws.cell(row=1, column=exchange_start_col + i*3 + 1, value=f"Exchange {i+1} Name")
    ws.cell(row=1, column=exchange_start_col + i*3 + 2, value=f"Exchange {i+1} Involvement ID")

# Track the next available row
next_row = 2

# Export functional chains with organized information
for fc in logical_functional_chains:
    try:
        # Get basic info
        fc_id = fc.get_id()
        fc_name = fc.get_name()

        # Write basic chain info
        ws.cell(row=next_row, column=1, value=fc_id)
        ws.cell(row=next_row, column=2, value=fc_name)

        # Get start and end functions
        start_func, end_func = get_start_end_functions(fc)

        # Write start and end function info
        if start_func:
            start_func_id = start_func.getId()
            start_func_name = start_func.getName()
            ws.cell(row=next_row, column=3, value=start_func_id)
            ws.cell(row=next_row, column=4, value=start_func_name)
        else:
            ws.cell(row=next_row, column=3, value="")
            ws.cell(row=next_row, column=4, value="")

        if end_func:
            end_func_id = end_func.getId()
            end_func_name = end_func.getName()
            ws.cell(row=next_row, column=5, value=end_func_id)
            ws.cell(row=next_row, column=6, value=end_func_name)
        else:
            ws.cell(row=next_row, column=5, value="")
            ws.cell(row=next_row, column=6, value="")

        # Get all involvements
        all_involvements = {}
        try:
            involvements = fc.get_java_object().getOwnedFunctionalChainInvolvements()
            for involvement in involvements:
                if hasattr(involvement, 'getId'):
                    involvement_id = involvement.getId()
                    if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                        involved_element = involvement.getInvolvedElement()
                        if involved_element and hasattr(involved_element, 'getId'):
                            all_involvements[involved_element.getId()] = involvement_id
                    elif involvement.eClass().getName() == "FunctionalChainInvolvementLink":
                        involved_element = involvement.getInvolvedElement()
                        if involved_element and hasattr(involved_element, 'getId'):
                            all_involvements[involved_element.getId()] = involvement_id
        except Exception as e:
            print(f"Error getting involvements for {fc.get_name()}: {e}")

        # Get ordered functions
        ordered_functions = get_ordered_functions(fc)

        # Write ordered functions with their involvements
        for i, func in enumerate(ordered_functions):
            try:
                func_id = func.getId()
                func_name = func.getName()

                # Write function info
                ws.cell(row=next_row, column=function_start_col + i*3, value=func_id)
                ws.cell(row=next_row, column=function_start_col + i*3 + 1, value=func_name)

                # Write function involvement ID if available
                if func_id in all_involvements:
                    ws.cell(row=next_row, column=function_start_col + i*3 + 2, value=all_involvements[func_id])
                else:
                    ws.cell(row=next_row, column=function_start_col + i*3 + 2, value="")
            except Exception as e:
                print(f"Error processing function {i+1} in {fc.get_name()}: {e}")

        # Get involved exchanges and write with their involvements
        involved_exchanges = fc.get_involved_functional_exchanges()
        for i, exchange in enumerate(involved_exchanges):
            try:
                exchange_id = exchange.get_id()
                exchange_name = exchange.get_name()

                # Write exchange info
                ws.cell(row=next_row, column=exchange_start_col + i*3, value=exchange_id)
                ws.cell(row=next_row, column=exchange_start_col + i*3 + 1, value=exchange_name)

                # Write exchange involvement ID if available
                if exchange_id in all_involvements:
                    ws.cell(row=next_row, column=exchange_start_col + i*3 + 2, value=all_involvements[exchange_id])
                else:
                    ws.cell(row=next_row, column=exchange_start_col + i*3 + 2, value="")
            except Exception as e:
                print(f"Error processing exchange {i+1} in {fc.get_name()}: {e}")

        next_row += 1

    except Exception as e:
        print(f"Error processing functional chain {fc.get_name()}: {e}")
        import traceback
        traceback.print_exc()
        continue

# Save and refresh
try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

CapellaPlatform.refresh(folder)
print(f"Export completed. Exported data to {xlsx_file_name}")