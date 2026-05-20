'''
Created on 26 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
# aird_path = "/Import_test/Import_test.aird"
# xlsx_path = "/DVS/results/DVS_Component_Exchanges.xlsx"
# aird_path = "/Import_test_IFES/Import_test_IFES.aird"
# xlsx_path = "/In-Flight Entertainment System/results/In-Flight Entertainment System_Component_Exchanges.xlsx"
aird_path = "/import/import.aird"
xlsx_path = "/test_ce/results/test_ce_Component_Exchanges.xlsx"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()
lc_pkg = la.get_logical_component_pkg()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Reading " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

def create_component_from_java(java_obj):
    """Create the appropriate component type from a Java object"""
    try:
        if java_obj.isActor():
            return LogicalActor(java_obj)
        else:
            return LogicalComponent(java_obj)
    except Exception as e:
        print(f"Error creating component: {e}")
        return None

def find_component_by_id(component_id):
    """Find a component by its ID in the model using recursive search"""
    try:
        # Start from the logical component package
        def search_recursive(java_object):
            try:
                # Check if this object matches the ID
                if java_object.getSid() == component_id:
                    return java_object

                # Check children
                children = java_object.getOwnedLogicalComponents()
                for child in children:
                    found = search_recursive(child)
                    if found:
                        return found
            except:
                pass
            return None

        lc_pkg_java = lc_pkg.get_java_object()
        return search_recursive(lc_pkg_java)
    except Exception as e:
        print(f"Error finding component by ID {component_id}: {e}")
        return None

def find_or_create_port(component, port_id, port_name, port_direction, port_kind):
    """Find an existing port or create a new one with the specified direction and kind"""
    try:
        java_obj = component.get_java_object()

        # Check existing ports
        for port in java_obj.getOwnedFeatures():
            if hasattr(port, 'getSid') and port.getSid() == port_id:
                # Found existing port, set its direction and kind
                set_port_direction(ComponentPort(port), port_direction)
                set_port_kind(ComponentPort(port), port_kind)
                return ComponentPort(port)

        # Create new port if not found
        print(f"Creating new port {port_name} with direction {port_direction} and kind {port_kind}")
        port = ComponentPort()
        java_obj.getOwnedFeatures().add(port.get_java_object())
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(port.get_java_object())
        
        port.set_name(port_name)
        port.get_java_object().setSid(port_id)

        # Set the port direction
        set_port_direction(port, port_direction)

        # Set the port kind
        set_port_kind(port, port_kind)

        return port
    except Exception as e:
        print(f"Error finding or creating port {port_name}: {e}")
        return None

def set_port_direction(port, direction):
    """Set the direction of a port using the correct method"""
    try:
        # Use the correct method to set orientation
        if direction == "IN":
            port.set_orientation("IN")
        elif direction == "OUT":
            port.set_orientation("OUT")
        elif direction == "INOUT":
            port.set_orientation("INOUT")
        elif direction == "UNSET":
            port.set_orientation("UNSET")
        else:
            print(f"Unknown port direction: {direction}")
    except Exception as e:
        print(f"Error setting port direction: {e}")
        import traceback
        traceback.print_exc()

def set_port_kind(port, kind):
    """Set the kind of a port (STANDARD or FLOW)"""
    try:
        # Set the kind based on the input
        if kind == "STANDARD":
            port.set_kind("STANDARD")
        elif kind == "FLOW":
            port.set_kind("FLOW")
        else:
            # Default to STANDARD if unknown kind
            port.set_kind("STANDARD")
    except Exception as e:
        print(f"Error setting port kind: {e}")
        import traceback
        traceback.print_exc()

def get_ancestor_path(component):
    """Get the path from a component to the root as a list of component IDs"""
    path = []
    current = component
    while current:
        try:
            # Add the current component's ID to the path
            path.append(current.getSid())
            container = current.eContainer()
            if container and hasattr(container, 'getSid'):
                current = container
            else:
                current = None
        except:
            current = None
    return path

def find_common_ancestor(component1, component2):
    """Find the common ancestor of two components by comparing IDs"""
    try:
        # Get the paths from each component to the root as lists of IDs
        path1 = get_ancestor_path(component1.get_java_object())
        path2 = get_ancestor_path(component2.get_java_object())

        # Find all common ancestor IDs
        common_ancestors = []
        for i in range(len(path1)):
            for j in range(len(path2)):
                if path1[i] == path2[j]:
                    common_ancestors.append(path1[i])

        # If no common ancestors found, use the logical component package
        if not common_ancestors:
            return lc_pkg

        # Use the first common ancestor as you specified
        common_ancestor_id = common_ancestors[0]

        # Find the component with the common ancestor ID
        def find_component_by_id_recursive(java_object, target_id):
            try:
                # Check if this object matches the ID
                if java_object.getSid() == target_id:
                    return java_object

                # Check children
                try:
                    children = java_object.getOwnedLogicalComponents()
                    for child in children:
                        found = find_component_by_id_recursive(child, target_id)
                        if found:
                            return found
                except:
                    pass
            except:
                pass
            return None

        # Search for the component with the common ancestor ID
        common_ancestor_java = find_component_by_id_recursive(lc_pkg.get_java_object(), common_ancestor_id)

        if common_ancestor_java:
            # Create a component object from the Java object
            try:
                # Try to create a LogicalComponent first
                return LogicalComponent(common_ancestor_java)
            except:
                try:
                    # If that fails, try to create a LogicalActor
                    return LogicalActor(common_ancestor_java)
                except:
                    # If both fail, return the logical component package as fallback
                    return lc_pkg
        else:
            # Fallback to logical component package if we can't find the component
            return lc_pkg

    except Exception as e:
        print(f"Error finding common ancestor: {e}")
        import traceback
        traceback.print_exc()
        return lc_pkg

# Start the import
model.start_transaction()

try:
    # Process each row in the Excel file
    for row in ws.iter_rows(min_row=2):
        # Extract data from the row
        source_comp_id = row[0].value
        source_comp_name = row[1].value
        source_port_id = row[2].value
        source_port_name = row[3].value
        source_port_dir = row[4].value
        source_port_kind = row[5].value
        ce_id = row[6].value
        ce_name = row[7].value
        ce_kind = row[8].value
        target_comp_id = row[9].value
        target_comp_name = row[10].value
        target_port_id = row[11].value
        target_port_name = row[12].value
        target_port_dir = row[13].value
        target_port_kind = row[14].value  

        print(f"Processing exchange: {ce_name} between {source_comp_name} and {target_comp_name}")

        # Find source and target components
        source_java = find_component_by_id(source_comp_id)
        target_java = find_component_by_id(target_comp_id)

        if not source_java or not target_java:
            print(f"Could not find source or target component for exchange {ce_name}")
            continue

        # Create component objects
        source_component = create_component_from_java(source_java)
        target_component = create_component_from_java(target_java)

        if not source_component or not target_component:
            print(f"Could not create component objects for exchange {ce_name}")
            continue

        # Find or create source port with direction and kind
        source_port = find_or_create_port(source_component, source_port_id, source_port_name, source_port_dir, source_port_kind)
        if not source_port:
            print(f"Failed to create/get source port {source_port_name}")
            continue

        # Find or create target port with direction and kind
        target_port = find_or_create_port(target_component, target_port_id, target_port_name, target_port_dir, target_port_kind)
        if not target_port:
            print(f"Failed to create/get target port {target_port_name}")
            continue

        # Find the common ancestor to store the component exchange
        common_ancestor = find_common_ancestor(source_component, target_component)
        print(f"Storing component exchange under {common_ancestor.get_name()}")

        # Check if the component exchange already exists in the common ancestor
        existing_ce = None
        common_ancestor_java = common_ancestor.get_java_object()
        for ce in common_ancestor_java.getOwnedComponentExchanges():
            if ce.getSid() == ce_id:
                existing_ce = ComponentExchange(ce)
                print(f"Component exchange {ce_name} already exists in {common_ancestor.get_name()}")
                break

        if existing_ce:
            continue

        # Create the component exchange
        print(f"Creating component exchange: {ce_name}")
        ce = ComponentExchange()

        # Set the exchange kind
        if ce_kind:
            ce.set_kind(ce_kind)

        # Set source and target ports
        ce.get_java_object().setSource(source_port.get_java_object())
        ce.get_java_object().setTarget(target_port.get_java_object())

        # Add the exchange to the common ancestor
        common_ancestor.get_java_object().getOwnedComponentExchanges().add(ce.get_java_object())
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(ce.get_java_object())
        
        # Set basic properties
        ce.set_name(ce_name)
        ce.get_java_object().setSid(ce_id)

        print(f"Successfully created exchange {ce_name}")

except Exception as e:
    print("Error: " + str(e))
    import traceback
    traceback.print_exc()
    model.rollback_transaction()
    raise

# else:
model.commit_transaction()

model.save()
print("Component exchanges import completed successfully")