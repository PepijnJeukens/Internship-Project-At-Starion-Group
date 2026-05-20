'''
Created on 10 Apr 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import *

#Path names
aird_path = "/delete_test/delete_test.aird"
xlsx_path = "/DVS/results/DVS_Logical_System2.xlsx"

model = CapellaModel()
model.open(aird_path)

# gets the SystemEngineering
se = model.get_system_engineering()
lc_pkg = se.get_logical_architecture().get_logical_component_pkg()

# create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Read " + xlsx_file_name)

# load the workbook
wb = load_workbook(xlsx_file_name)

# grab the active worksheet
ws = wb.active

def find_component_by_id(components, component_id):
    """Find a component by its ID in a list of components."""
    for component in components:
        try:
            if component.get_sid() == component_id:
                return component
        except AttributeError:
            if component.get_java_object().getSid() == component_id:
                return component
    return None

def get_component_by_id(java_object, component_id):
    """Find a component by its ID in the children of a Java object."""
    for child in java_object.getOwnedLogicalComponents():
        if child.getSid() == component_id:
            return child
    return None

def collect_all_component_ids(component):
    """Recursively collect all component IDs from a component and its children"""
    component_ids = set()

    try:
        # Add the current component's ID
        component_ids.add(component.get_sid())

        # Recursively collect IDs from children
        java_obj = component.get_java_object()
        for child in java_obj.getOwnedLogicalComponents():
            child_component = LogicalComponent(child) if not child.isActor() else LogicalActor(child)
            component_ids.update(collect_all_component_ids(child_component))

    except Exception as e:
        print(f"Error collecting component IDs: {e}")

    return component_ids

def delete_component_and_children(component):
    """Delete a component and all its children recursively, printing all deleted items"""
    try:
        # Get the Python wrapper for the component
        if component.isActor():
            comp_to_delete = LogicalActor(component)
        else:
            comp_to_delete = LogicalComponent(component)

        # Collect all children to print
        all_to_delete = [comp_to_delete]
        stack = [component]

        # Collect all children recursively
        while stack:
            current = stack.pop()
            try:
                for child in current.getOwnedLogicalComponents():
                    child_comp = LogicalComponent(child) if not child.isActor() else LogicalActor(child)
                    all_to_delete.append(child_comp)
                    stack.append(child)
            except:
                continue

        # Print all components being deleted
        print(f"Deleting the following components:")
        for comp in all_to_delete:
            print(f"  - {comp.get_name()} (ID: {comp.get_sid()})")

        # Delete the main component (which will delete all its children)
        EObject.delete_e_object(comp_to_delete)

    except Exception as e:
        print(f"Error deleting component {component.getSid()}: {e}")

def delete_unused_components(parent_component, imported_ids):
    """Delete components that exist in the model but were not imported"""
    try:
        # Get all components from the parent
        java_parent = parent_component.get_java_object()
        children = list(java_parent.getOwnedLogicalComponents())  # Convert to list to avoid modification during iteration

        # Create a list of components to delete
        to_delete = []

        for child in children:
            child_id = child.getSid()
            if child_id not in imported_ids:
                # This component was not imported, mark it and all its children for deletion
                to_delete.append(child)

        # Delete the components (in reverse order to avoid index issues)
        for component in reversed(to_delete):
            try:
                delete_component_and_children(component)
            except Exception as e:
                print(f"Error deleting component {component.getSid()}: {e}")

        # Recursively check children of imported components
        for child in children:
            child_id = child.getSid()
            if child_id in imported_ids:
                # This component was imported, check its children
                child_component = LogicalComponent(child) if not child.isActor() else LogicalActor(child)
                delete_unused_components(child_component, imported_ids)

    except Exception as e:
        print(f"Error in delete_unused_components: {e}")

# Start the import
model.start_transaction()

try:
    # First, collect all existing component IDs in the model
    # print("Collecting existing component IDs...")
    existing_component_ids = set()
    for component in lc_pkg.get_owned_logical_components():
        existing_component_ids.update(collect_all_component_ids(component))
    # print(f"Found {len(existing_component_ids)} existing components in the model")

    # Set to track imported component IDs (only those explicitly in the import file)
    imported_component_ids = set()

    # Determine the maximum depth from the Excel file
    max_col = ws.max_column
    max_depth = max_col // 3
    # print(f"Maximum depth detected from Excel: {max_depth}")

    # Create a list to keep track of parent components at each level
    parent_components = [None] * max_depth

    for row in ws.iter_rows(min_row=2):
        # Create a list of component data for each level
        components = []
        for level in range(max_depth):
            col = level * 3
            component_id = row[col].value if row[col].value else None
            component_name = row[col + 1].value if col + 1 < len(row) and row[col + 1].value else None
            component_type = row[col + 2].value if col + 2 < len(row) and row[col + 2].value else None
            components.append((component_id, component_name, component_type))

        # Import the hierarchy
        for level in range(max_depth):
            component_id, component_name, component_type = components[level]

            if not component_id:
                continue  # Skip empty entries

            if level == 0:
                # Top level component
                existing_component = find_component_by_id(lc_pkg.get_owned_logical_components(), component_id)
                if existing_component is None:
                    # print(f"Creating new top-level component: {component_name} with ID: {component_id} as {component_type}")
                    if component_type == "Actor":
                        component = LogicalActor()
                    else:
                        component = LogicalComponent()
                    lc_pkg.get_owned_logical_components().add(component)
                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(component.get_java_object())
                    component.set_name(component_name)
                    component.set_sid(component_id)
                else:
                    component = existing_component
                    # print(f"Top-level component: {component_name} with ID: {component_id} already exists in the model.")

                # Add ONLY this component's ID to imported IDs (not its children)
                imported_component_ids.add(component_id)

            else:
                # Child component
                parent_component = parent_components[level - 1]
                if parent_component is None:
                    # print(f"Parent component not found for {component_name} at level {level}")
                    continue

                java_parent = parent_component.get_java_object()
                existing_component = get_component_by_id(java_parent, component_id)

                if existing_component is None:
                    # print(f"Creating new component: {component_name} with ID: {component_id} under {parent_component.get_name()} as {component_type}")
                    if component_type == "Actor":
                        component = LogicalActor()
                    else:
                        component = LogicalComponent()
                    java_parent.getOwnedLogicalComponents().add(component.get_java_object())
                    org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(component.get_java_object())
                    component.set_name(component_name)
                    component.set_sid(component_id)
                else:
                    if component_type == "Actor":
                        component = LogicalActor(existing_component)
                    else:
                        component = LogicalComponent(existing_component)
                    # print(f"Component {component_name} with ID: {component_id} already exists under {parent_component.get_name()}")

                # Add ONLY this component's ID to imported IDs (not its children)
                imported_component_ids.add(component_id)

            parent_components[level] = component

    # After import, check for components that need to be deleted
    print("\nChecking for components to delete...")
    unused_ids = existing_component_ids - imported_component_ids
    print(f"Found {len(unused_ids)} components that were not imported and will be deleted")

    # Delete unused components from the logical component package
    delete_unused_components(lc_pkg, imported_component_ids)

except Exception as e:
    print("Error: " + str(e))
    import traceback
    traceback.print_exc()
    model.rollback_transaction()
    raise

else:
    model.commit_transaction()

# At the end of your script, after the transaction is committed
model.save()
print("Model saved and cleaned up")