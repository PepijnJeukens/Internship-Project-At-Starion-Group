'''
Created on 25 Mar 2026

@author: p.jeukens
'''
# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import Workbook
import os

# Path names
aird_path = "/DVS/DVS/DVS.aird"
# aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()

# Get the logical function package
lf_pkg = la.get_logical_function_pkg()

# Find the Root Logical Function
rlf = None
for lf in lf_pkg.get_owned_logical_functions():
    if lf.get_name() == "Root Logical Function":
        rlf = lf
        break

if rlf is None:
    print("Root Logical Function not found!")
    exit()

# Create a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
folder_path = CapellaPlatform.getAbsolutePath(folder)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)

xlsx_file_name = folder_path + "/" + se.get_name() + "_Logical_Functional_Exchanges.xlsx"

print("Writing " + xlsx_file_name)

# Create a workbook
wb = Workbook()

# Grab the active worksheet
ws = wb.active
ws.title = 'Logical Functional Exchanges'

# Create headers
ws.cell(row=1, column=1, value="Function From ID")
ws.cell(row=1, column=2, value="Function From Name")
ws.cell(row=1, column=3, value="Function From Port ID")
ws.cell(row=1, column=4, value="Function From Port Name")
ws.cell(row=1, column=5, value="Functional Exchange ID")
ws.cell(row=1, column=6, value="Functional Exchange Name")
ws.cell(row=1, column=7, value="Function To ID")
ws.cell(row=1, column=8, value="Function To Name")
ws.cell(row=1, column=9, value="Function To Port ID")
ws.cell(row=1, column=10, value="Function To Port Name")

def export_functional_exchanges(function, row):
    """Recursively export functional exchanges for a function and its children."""
    global ws

    try:
        # Get all outgoing functional exchanges
        outgoing_exchanges = function.get_outgoing()

        for fe in outgoing_exchanges:
            # Get source and target information
            source_function = function
            target_function = fe.get_target_function()

            # Get source and target ports
            source_port = fe.get_source_port()
            target_port = fe.get_target_port()

            if target_function and source_port and target_port:
                # Write exchange information
                ws.cell(row=row, column=1, value=source_function.get_id())
                ws.cell(row=row, column=2, value=source_function.get_name())
                ws.cell(row=row, column=3, value=source_port.get_id())
                ws.cell(row=row, column=4, value=source_port.get_name())
                ws.cell(row=row, column=5, value=fe.get_id())
                ws.cell(row=row, column=6, value=fe.get_name())
                ws.cell(row=row, column=7, value=target_function.get_id())
                ws.cell(row=row, column=8, value=target_function.get_name())
                ws.cell(row=row, column=9, value=target_port.get_id())
                ws.cell(row=row, column=10, value=target_port.get_name())
                row += 1

        # Process children
        for child in function.get_owned_functions():
            row = export_functional_exchanges(child, row)

    except Exception as e:
        print(f"Error processing function {function.get_name()}: {e}")
        import traceback
        traceback.print_exc()

    return row

# Export functional exchanges starting from the root logical function
row = 2
if rlf:
    print("Exporting functional exchanges...")
    row = export_functional_exchanges(rlf, row)

try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

CapellaPlatform.refresh(folder)