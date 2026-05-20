'''
Created on 26 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import Workbook
import os

# Path names
# aird_path = "/DVS/DVS/DVS.aird"
# aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"
aird_path = "/test_ce/test_ce.aird"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()
lf_pkg = la.get_logical_function_pkg()

# Create a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
folder_path = CapellaPlatform.getAbsolutePath(folder)

if not os.path.exists(folder_path):
    os.makedirs(folder_path)

xlsx_file_name = folder_path + "/" + se.get_name() + "_Exchange_Allocations.xlsx"

print("Writing " + xlsx_file_name)

# Create a workbook
wb = Workbook()

# Grab the active worksheet
ws = wb.active
ws.title = 'Exchange Allocations'

# Create headers with all necessary columns including component IDs and names
ws.cell(row=1, column=1, value="Component Exchange ID")
ws.cell(row=1, column=2, value="Component Exchange Name")
ws.cell(row=1, column=3, value="Component Source Port ID")
ws.cell(row=1, column=4, value="Component Source Port Name")
ws.cell(row=1, column=5, value="Source Component ID")
ws.cell(row=1, column=6, value="Source Component Name")
ws.cell(row=1, column=7, value="Component Target Port ID")
ws.cell(row=1, column=8, value="Component Target Port Name")
ws.cell(row=1, column=9, value="Target Component ID")
ws.cell(row=1, column=10, value="Target Component Name")
ws.cell(row=1, column=11, value="Functional Exchange ID")
ws.cell(row=1, column=12, value="Functional Exchange Name")
ws.cell(row=1, column=13, value="Functional Source Port ID")
ws.cell(row=1, column=14, value="Functional Source Port Name")
ws.cell(row=1, column=15, value="Functional Target Port ID")
ws.cell(row=1, column=16, value="Functional Target Port Name")
ws.cell(row=1, column=17, value="Allocation ID")
ws.cell(row=1, column=18, value="Source Function ID")
ws.cell(row=1, column=19, value="Target Function ID")

# Track exported allocations to avoid duplicates
exported_allocations = set()
current_row = 2  # Start from row 2 (after headers)

def create_component_from_java(java_obj):
    """Create the appropriate component type from a Java object"""
    try:
        if hasattr(java_obj, 'isActor') and java_obj.isActor():
            return LogicalActor(java_obj)
        else:
            return LogicalComponent(java_obj)
    except Exception as e:
        print(f"Error creating component: {e}")
        return None

def get_component_exchange_ports(ce):
    """Get the source and target ports of a component exchange"""
    try:
        source_port = ce.get_source_port()
        target_port = ce.get_target_port()
        return source_port, target_port
    except Exception as e:
        print(f"Error getting component exchange ports: {e}")
        return None, None

def get_functional_exchange_ports(fe):
    """Get the source and target ports of a functional exchange"""
    try:
        source_port = fe.get_source_port()
        target_port = fe.get_target_port()
        return source_port, target_port
    except Exception as e:
        print(f"Error getting functional exchange ports: {e}")
        return None, None

def get_component_for_port(port):
    """Get the component that owns a port"""
    try:
        if not port:
            return None

        # Get the container of the port
        container = port.get_java_object().eContainer()
        if container:
            return create_component_from_java(container)
        return None
    except Exception as e:
        print(f"Error getting component for port: {e}")
        return None

def get_function_for_port(port):
    """Get the function that owns a port"""
    try:
        if not port:
            return None

        # Get the Java object of the port
        port_java = port.get_java_object()

        # Get the container (function) of the port
        container = port_java.eContainer()
        if container and hasattr(container, 'eClass') and 'Function' in container.eClass().getName():
            return Function(container)

        # If we can't get the function directly, try to find it
        all_functions = se.get_all_contents_by_type(Function)
        for func in all_functions:
            try:
                # Check input ports
                for p in func.get_inputs():
                    if p.get_id() == port.get_id():
                        return func
                # Check output ports
                for p in func.get_outputs():
                    if p.get_id() == port.get_id():
                        return func
            except:
                continue

        return None
    except Exception as e:
        print(f"Error getting function for port: {e}")
        return None

def process_container(container, recursive=True):
    """Process a container (package or component) and optionally its children"""
    global ws, exported_allocations, current_row

    try:
        # Process exchanges owned by this container
        java_obj = container.get_java_object()
        exchanges = java_obj.getOwnedComponentExchanges()

        for ce in exchanges:
            ce_obj = ComponentExchange(ce)
            ce_id = ce_obj.get_id()

            # Get allocations for this component exchange
            allocations = ce_obj.get_owned_component_exchange_functional_exchange_allocations()

            for allocation in allocations:
                allocation_id = allocation.get_id()

                # Skip if we've already exported this allocation
                if allocation_id in exported_allocations:
                    continue

                # Get the functional exchange
                fe = allocation.get_allocated_functional_exchange()

                if fe:
                    # Get component exchange ports
                    ce_source_port, ce_target_port = get_component_exchange_ports(ce_obj)

                    # Get functional exchange ports using the direct API methods
                    fe_source_port = fe.get_source_port()
                    fe_target_port = fe.get_target_port()

                    # Get components for the ports
                    source_component = get_component_for_port(ce_source_port) if ce_source_port else None
                    target_component = get_component_for_port(ce_target_port) if ce_target_port else None

                    # Get functions for the ports
                    source_function = get_function_for_port(fe_source_port) if fe_source_port else None
                    target_function = get_function_for_port(fe_target_port) if fe_target_port else None

                    # Write allocation information
                    ws.cell(row=current_row, column=1, value=ce_id)
                    ws.cell(row=current_row, column=2, value=ce_obj.get_name())
                    ws.cell(row=current_row, column=3, value=ce_source_port.get_id() if ce_source_port else "")
                    ws.cell(row=current_row, column=4, value=ce_source_port.get_name() if ce_source_port else "")
                    ws.cell(row=current_row, column=5, value=source_component.get_id() if source_component else "")
                    ws.cell(row=current_row, column=6, value=source_component.get_name() if source_component else "")
                    ws.cell(row=current_row, column=7, value=ce_target_port.get_id() if ce_target_port else "")
                    ws.cell(row=current_row, column=8, value=ce_target_port.get_name() if ce_target_port else "")
                    ws.cell(row=current_row, column=9, value=target_component.get_id() if target_component else "")
                    ws.cell(row=current_row, column=10, value=target_component.get_name() if target_component else "")
                    ws.cell(row=current_row, column=11, value=fe.get_id())
                    ws.cell(row=current_row, column=12, value=fe.get_name())

                    # Export functional source port ID and name
                    ws.cell(row=current_row, column=13, value=fe_source_port.get_id() if fe_source_port else "")
                    ws.cell(row=current_row, column=14, value=fe_source_port.get_name() if fe_source_port else "")

                    # Export functional target port ID and name
                    ws.cell(row=current_row, column=15, value=fe_target_port.get_id() if fe_target_port else "")
                    ws.cell(row=current_row, column=16, value=fe_target_port.get_name() if fe_target_port else "")

                    ws.cell(row=current_row, column=17, value=allocation_id)
                    ws.cell(row=current_row, column=18, value=source_function.get_id() if source_function else "")
                    ws.cell(row=current_row, column=19, value=target_function.get_id() if target_function else "")

                    exported_allocations.add(allocation_id)
                    current_row += 1
                    print(f"Exported allocation: {ce_obj.get_name()} -> {fe.get_name()}")

        # Process children if recursive flag is True
        if recursive:
            children = java_obj.getOwnedLogicalComponents()
            for child in children:
                child_component = create_component_from_java(child)
                if child_component:
                    process_container(child_component, recursive=False)  # Process children non-recursively

    except Exception as e:
        print(f"Error processing container: {e}")
        import traceback
        traceback.print_exc()

# Export all exchange allocations
if la:
    print("Exporting all exchange allocations...")

    # Process the logical component package directly
    lc_pkg = la.get_logical_component_pkg()
    process_container(lc_pkg)

try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

CapellaPlatform.refresh(folder)
print(f"Export completed. Exported data to {xlsx_file_name}")