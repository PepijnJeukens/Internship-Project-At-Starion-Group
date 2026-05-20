'''
Created on 24 Mar 2026

@author: p.jeukens
'''
# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
aird_path = "/Import_test/Import_test.aird"
xlsx_path = "/DVS/results/DVS_Logical_Functions.xlsx"
# aird_path = "/Import_test_IFES/Import_test_IFES.aird"
# xlsx_path = "/In-Flight Entertainment System/results/In-Flight Entertainment System_Logical_Functions.xlsx"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()
lf_pkg = la.get_logical_function_pkg()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Read " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

def find_function_by_id(parent, function_id):
    """Find a function by its ID in the children of a parent function."""
    for function in parent.get_owned_functions():
        try:
            if function.get_sid() == function_id:
                return function
        except AttributeError:
            if function.get_java_object().getSid() == function_id:
                return function
    return None

def set_function_kind(function, kind):
    """Set the kind of function using the proper Function.set_kind() method"""
    try:
        # Convert to uppercase to match the expected values
        kind = kind.upper()

        # Use the set_kind method from the Function class
        function.set_kind(kind)
        return True
    except Exception as e:
        print(f"Error setting function kind for {function.get_name()}: {e}")
        return False

# Start the import
model.start_transaction()

try:
    # Determine the maximum depth from the Excel file
    max_col = ws.max_column
    max_depth = max_col // 3  # 3 columns per level (ID, Name, Kind)
    print(f"Maximum depth detected from Excel: {max_depth}")

    # Create a list to keep track of parent functions at each level
    parent_functions = [None] * max_depth

    # Check if the root logical function exists
    rlf = None
    for lf in lf_pkg.get_owned_logical_functions():
        if lf.get_name() == "Root Logical Function":
            rlf = lf
            break

    # If RLF doesn't exist, create it
    if rlf is None:
        print("Creating Root Logical Function")
        rlf = LogicalFunction()
        lf_pkg.get_owned_logical_functions().add(rlf)
        rlf.set_name("Root Logical Function")
        # Set an ID for RLF if needed
        rlf.get_java_object().setSid("RLF_ID")
    else:
        print("Root Logical Function already exists")

    for row in ws.iter_rows(min_row=2):
        # Create a list of function data for each level
        functions = []
        for level in range(max_depth):
            col = level * 3  # 3 columns per level (ID, Name, Kind)
            function_id = row[col].value if col < len(row) and row[col].value else None
            function_name = row[col + 1].value if col + 1 < len(row) and row[col + 1].value else None
            function_kind = row[col + 2].value if col + 2 < len(row) and row[col + 2].value else None
            functions.append((function_id, function_name, function_kind))

        # Import the hierarchy
        for level in range(max_depth):
            function_id, function_name, function_kind = functions[level]

            if not function_id or not function_name:
                continue  # Skip empty entries

            # For level 0, parent is RLF
            if level == 0:
                parent_function = rlf
            else:
                parent_function = parent_functions[level - 1]

            if parent_function is None:
                print(f"Parent function not found for {function_name} at level {level+1}")
                continue

            # Check if the function already exists
            existing_function = find_function_by_id(parent_function, function_id)

            if existing_function is None:
                print(f"Creating new function: {function_name} with ID: {function_id} at level {level+1}")
                function = LogicalFunction()
                parent_function.get_owned_functions().add(function)
                org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(function.get_java_object())
                function.set_name(function_name)
                function.get_java_object().setSid(function_id)

                # Set the function kind if provided
                if function_kind:
                    success = set_function_kind(function, function_kind)
                    if success:
                        print(f"Set function kind to {function_kind} for {function_name}")
                    else:
                        print(f"Failed to set function kind to {function_kind} for {function_name}")
                else:
                    print(f"No function kind specified for {function_name}, defaulting to FUNCTION")
            else:
                function = existing_function
                print(f"Function {function_name} with ID: {function_id} already exists at level {level+1}")

                # Update the function kind if provided
                if function_kind:
                    success = set_function_kind(function, function_kind)
                    if success:
                        print(f"Updated function kind to {function_kind} for {function_name}")
                    else:
                        print(f"Failed to update function kind to {function_kind} for {function_name}")

            # Only set parent for next level if this is not the last level
            if level < max_depth - 1:
                parent_functions[level] = function


except Exception as e:
    print("Error: " + str(e))
    import traceback
    traceback.print_exc()
    model.rollback_transaction()
    raise

# else:
model.commit_transaction()

model.save()
print("Logical functions import completed successfully")