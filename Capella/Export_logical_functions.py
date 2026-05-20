'''
Created on 24 Mar 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import Workbook
import os

# Path names
aird_path = "/DVS/DVS/DVS.aird"
# aird_path = "/In-Flight Entertainment System/In-Flight Entertainment System.aird"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()

# Gets the root logical function
lf_pkg = se.get_logical_architecture().get_logical_function_pkg()
rlf = None
for lf in lf_pkg.get_owned_logical_functions():
    if lf.get_name() == "Root Logical Function":
        rlf = lf
        break

if rlf is None:
    print("Root Logical Function not found!")
    exit()

# Create a folder in the project
model_path = CapellaPlatform.getModelPath(se)
project_name = model_path[0:(model_path.index("/", 1) + 1)]
project = CapellaPlatform.getProject(project_name)
folder = CapellaPlatform.getFolder(project, "results")
xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + "/" + se.get_name() + "_Logical_Functions.xlsx"

print("Writing " + xlsx_file_name)

def get_max_depth(function, current_depth=1):
    """Recursively find the maximum depth of the function hierarchy."""
    max_depth = current_depth
    try:
        children = function.get_java_object().getOwnedFunctions()
        for child in children:
            child_function = LogicalFunction(child)
            child_depth = get_max_depth(child_function, current_depth + 1)
            if child_depth > max_depth:
                max_depth = child_depth
    except Exception as e:
        print(f"Error getting children for {function.get_name()}: {e}")
    return max_depth

def get_function_kind(function):
    """Determine the kind of function (duplicate, function, gather, route, select, split)"""
    try:
        java_obj = function.get_java_object()
        kind = java_obj.getKind()
        kind = str(kind)
        
        if kind == "DUPLICATE":
            return "DUPLICATE"
        elif kind == "GATHER":
            return "GATHER"
        elif kind == "ROUTE":
            return "ROUTE"
        elif kind == "SELECT":
            return "SELECT" 
        elif kind == "SPLIT":
            return "SPLIT"
        else:
            return "FUNCTION"
    except Exception as e:
        print(f"Error determining function kind for {function.get_name()}: {e}")
        return "FUNCTION"

# Determine the maximum depth of the hierarchy under RLF
max_depth = 0
children = rlf.get_java_object().getOwnedFunctions()
for child in children:
    child_function = LogicalFunction(child)
    depth = get_max_depth(child_function, 1)  # Start from 1 since we're skipping RLF
    if depth > max_depth:
        max_depth = depth

print(f"Maximum depth of the function hierarchy: {max_depth}")

# Create a workbook
wb = Workbook()

# Grab the active worksheet
ws = wb.active
ws.title = 'Logical Functions'

# Create headers based on the maximum depth
for i in range(max_depth):
    col = i * 3 + 1  # 3 columns per level (ID, Name, Kind)
    ws.cell(row=1, column=col, value=f"{'Sub' * i}Function ID")
    ws.cell(row=1, column=col + 1, value=f"{'Sub' * i}Function Name")
    ws.cell(row=1, column=col + 2, value=f"{'Sub' * i}Function Kind")

def export_function(function, row, level=0):
    """Recursively export a function and its children."""
    global ws
    col = level * 3 + 1
    ws.cell(row=row, column=col, value=function.get_id())
    ws.cell(row=row, column=col + 1, value=function.get_name())

    # Get and write function kind
    function_kind = get_function_kind(function)
    ws.cell(row=row, column=col + 2, value=function_kind)

    # Move to the next row for children
    next_row = row + 1

    try:
        children = function.get_java_object().getOwnedFunctions()
        if children.size() > 0:
            for child in children:
                child_function = LogicalFunction(child)
                next_row = export_function(child_function, next_row, level + 1)
    except Exception as e:
        print(f"Error exporting children for {function.get_name()}: {e}")

    return next_row

# Export functions starting from the root logical function
row = 2
children = rlf.get_java_object().getOwnedFunctions()
for child in children:
    child_function = LogicalFunction(child)
    row = export_function(child_function, row)

try:
    wb.save(filename=xlsx_file_name)
    print("File saved successfully.")
except PermissionError:
    print(f"Permission denied: Could not save file to {xlsx_file_name}. Make sure the file is not open in another application.")
except Exception as e:
    print(f"An error occurred: {e}")

CapellaPlatform.refresh(folder)