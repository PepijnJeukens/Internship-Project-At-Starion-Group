'''
Created on 13 Apr 2026

@author: p.jeukens
'''
# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
aird_path = "/delete_test/delete_test.aird"
xlsx_path = "/DVS/results/DVS_Functional_Chains.xlsx"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering and Logical Architecture
se = model.get_system_engineering()
la = se.get_logical_architecture()

# Get the root logical function
lf_pkg = la.get_logical_function_pkg()
rlf = None
for lf in lf_pkg.get_owned_logical_functions():
    if lf.get_name() == "Root Logical Function":
        rlf = lf
        break

if not rlf:
    print("Root Logical Function not found!")
    exit()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Reading " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

# Find the column indices for the different sections
header_row = ws[1]
function_start_col = 7  # Functions start at column 7 (after FC ID, Name, Start/End Function info)
exchange_start_col = None

# Find where the Exchange section starts
for cell in header_row:
    if cell.value and "Exchange 1 ID" in str(cell.value):
        exchange_start_col = cell.column
        break

# Helper function to find a function by ID
def find_function_by_id(function_id):
    try:
        # Get all functions in the model
        all_functions = se.get_all_contents_by_type(Function)
        for func in all_functions:
            if func.get_sid() == function_id:
                return func
        return None
    except Exception as e:
        print(f"Error finding function by ID {function_id}: {e}")
        return None

# Helper function to find a functional exchange by ID
def find_functional_exchange_by_id(exchange_id):
    try:
        # Get all functional exchanges in the model
        all_exchanges = se.get_all_contents_by_type(FunctionalExchange)
        for exchange in all_exchanges:
            if exchange.get_sid() == exchange_id:
                return exchange
        return None
    except Exception as e:
        print(f"Error finding functional exchange by ID {exchange_id}: {e}")
        return None

# Helper function to check if a functional chain already exists
def functional_chain_exists(fc_id):
    try:
        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)
        for chain in all_chains:
            if chain.get_sid() == fc_id:
                return chain
        return None
    except Exception as e:
        print(f"Error checking if functional chain exists: {e}")
        return None



# Helper function to find a function involvement by function ID
def find_function_involvement(function_involvements, function_id):
    """Find a function involvement by function ID"""
    for fcif in function_involvements:
        try:
            involved = fcif.getInvolvedElement()
            if involved and involved.getSid() == function_id:
                return fcif
        except Exception as e:
            print(f"Error checking function involvement: {e}")
    return None

# Helper function to create a functional chain involvement function
def create_function_involvement(fc, func, involvement_id=None):
    """Create a functional chain involvement function"""
    try:
        fcif = create_e_object("http://www.polarsys.org/capella/core/fa/" + capella_version(), "FunctionalChainInvolvementFunction")
        fcif.setInvolved(func.get_java_object())
        fc.get_java_object().getOwnedFunctionalChainInvolvements().add(fcif)
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fcif)
        if involvement_id:
            fcif.setSid(involvement_id)
        return fcif
    except Exception as e:
        print(f"Error creating function involvement: {e}")
        return None

# Helper function to create a functional chain involvement link
def create_link_involvement(fc, exchange, source_fcif, target_fcif, involvement_id=None):
    """Create a functional chain involvement link"""
    try:
        fcil = create_e_object("http://www.polarsys.org/capella/core/fa/" + capella_version(), "FunctionalChainInvolvementLink")
        fcil.setInvolved(exchange.get_java_object())
        fcil.setSource(source_fcif)
        fcil.setTarget(target_fcif)
        fc.get_java_object().getOwnedFunctionalChainInvolvements().add(fcil)
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fcil)
        if involvement_id:
            fcil.setSid(involvement_id)
        return fcil
    except Exception as e:
        print(f"Error creating link involvement: {e}")
        return None
    
# Helper function to collect all existing functional chains
def collect_all_functional_chains():
    """Collect all existing functional chains in the model"""
    chains = {}

    try:
        print("Starting to collect all existing functional chains...")

        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)
        for chain in all_chains:
            chain_id = chain.get_sid()
            chains[chain_id] = chain
            print(f"Found functional chain: {chain.get_name()} (ID: {chain_id})")

        total_count = len(chains)
        print(f"Found {total_count} existing functional chains")
        return chains
    except Exception as e:
        print(f"Error in collect_all_functional_chains: {e}")
        return {}    

# Helper function to delete unused functional chains
def delete_unused_functional_chains(imported_chains):
    """Delete functional chains that exist in the model but were not imported"""
    try:
        deleted_count = 0

        print("Starting to delete unused functional chains...")

        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)
        for chain in all_chains:
            chain_id = chain.get_sid()

            # Check if this chain should be kept
            if chain_id in imported_chains:
                continue  # This chain is in the import file

            # Get name for logging
            chain_name = chain.get_name()

            print(f"Deleting functional chain: '{chain_name}' (ID: {chain_id})")
            EObject.delete_e_object(chain)
            deleted_count += 1

        print(f"Deleted {deleted_count} unused functional chains")
        return deleted_count
    except Exception as e:
        print(f"Error in delete_unused_functional_chains: {e}")
        return 0

# Start the import
model.start_transaction()

try:
    # First collect all existing functional chains
    # print("Collecting existing functional chains...")
    existing_chains = collect_all_functional_chains()

    # Dictionary to store imported functional chain IDs
    imported_chains = {}

    # Process each row in the Excel file (skip header row)
    for row in ws.iter_rows(min_row=2):
        # Extract basic functional chain info
        fc_id = row[0].value
        fc_name = row[1].value

        if not fc_id:
            continue

        # print(f"Processing functional chain: {fc_name}")

        # Check if the functional chain already exists
        existing_fc = functional_chain_exists(fc_id)
        if existing_fc:
            # print(f"Functional chain with ID {fc_id} and name '{fc_name}' already exists.")
            imported_chains[fc_id] = existing_fc
            continue

        # Create a new functional chain
        fc = FunctionalChain()
        fc.get_java_object().setSid(fc_id)
        fc.set_name(fc_name)

        # Add to imported chains
        imported_chains[fc_id] = fc

        # Get start and end function info from Excel
        start_func_id = row[2].value
        start_func_name = row[3].value
        end_func_id = row[4].value
        end_func_name = row[5].value

        # Find start and end functions
        start_func = find_function_by_id(start_func_id) if start_func_id else None
        end_func = find_function_by_id(end_func_id) if end_func_id else None

        if not start_func:
            print(f"Could not find start function with ID {start_func_id} and name {start_func_name}")
            continue
        if not end_func:
            print(f"Could not find end function with ID {end_func_id} and name {end_func_name}")
            continue

        # Process functions and their involvements
        function_col = function_start_col - 1  # Convert to 0-based index
        function_involvements = []
        involved_functions = []
        function_order = []  # To keep track of the order of functions

        # Process functions until we reach the exchange section
        while function_col < exchange_start_col - 1:  # Stop before the exchange section
            try:
                # Get function info
                func_id = row[function_col].value
                func_name = row[function_col + 1].value
                func_involvement_id = row[function_col + 2].value

                if not func_id:
                    break

                # Find the function
                func = find_function_by_id(func_id)
                if not func:
                    print(f"Could not find function with ID {func_id} and name {func_name}")
                    function_col += 3
                    continue

                # Create function involvement
                fcif = create_function_involvement(fc, func, func_involvement_id)
                if fcif:
                    function_involvements.append(fcif)
                    involved_functions.append(func)
                    function_order.append(func_id)

                function_col += 3
            except Exception as e:
                print(f"Error processing function: {e}")
                function_col += 3
                continue

        # Find the start and end function involvements
        start_fcif = find_function_involvement(function_involvements, start_func_id)
        end_fcif = find_function_involvement(function_involvements, end_func_id)

        if not start_fcif:
            print(f"Could not find start function involvement for function ID {start_func_id}")
            continue
        if not end_fcif:
            print(f"Could not find end function involvement for function ID {end_func_id}")
            continue

        # Process exchanges and their involvements
        exchange_col = exchange_start_col - 1  # Convert to 0-based index

        # Create a path from start to end function
        # First, find the path from start to end function
        path = [start_fcif]
        current_fcif = start_fcif

        # Try to find a path from start to end
        while current_fcif != end_fcif:
            found_next = False
            for fcif in function_involvements:
                if fcif == current_fcif:
                    continue

                # Check if this function is connected to the current one
                current_func_id = current_fcif.getInvolvedElement().getSid()
                next_func_id = fcif.getInvolvedElement().getSid()

                # Check if this function is in the Excel function list
                found_in_excel = False
                for i in range(0, len(row), 3):
                    if i >= function_start_col - 1 and i < exchange_start_col - 1:
                        if row[i].value == next_func_id:
                            found_in_excel = True
                            break

                if found_in_excel and next_func_id not in [f.getInvolvedElement().getSid() for f in path]:
                    path.append(fcif)
                    current_fcif = fcif
                    found_next = True
                    break

            if not found_next:
                # If we can't find a path, just add the end function
                if end_fcif not in path:
                    path.append(end_fcif)
                break

        # Now create links between consecutive functions in the path
        exchange_index = 0
        while exchange_col < len(row):
            try:
                # Get exchange info
                exchange_id = row[exchange_col].value
                exchange_name = row[exchange_col + 1].value
                exchange_involvement_id = row[exchange_col + 2].value

                if not exchange_id:
                    break

                # Find the functional exchange
                exchange = find_functional_exchange_by_id(exchange_id)
                if not exchange:
                    print(f"Could not find functional exchange with ID {exchange_id} and name {exchange_name}")
                    exchange_col += 3
                    continue

                # Create link between consecutive functions in the path
                if exchange_index < len(path) - 1:
                    source_fcif = path[exchange_index]
                    target_fcif = path[exchange_index + 1]
                    create_link_involvement(fc, exchange, source_fcif, target_fcif, exchange_involvement_id)
                    exchange_index += 1
                else:
                    # If we have more exchanges than path segments, use start and end
                    create_link_involvement(fc, exchange, start_fcif, end_fcif, exchange_involvement_id)

                exchange_col += 3
            except Exception as e:
                print(f"Error processing exchange: {e}")
                exchange_col += 3
                continue

        # Add the functional chain to the root logical function
        try:
            rlf.get_java_object().getOwnedFunctionalChains().add(fc.get_java_object())
            org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fc.get_java_object())
            fc.get_java_object().setSid(fc_id)
            fc.set_name(fc_name)
            # print(f"Added functional chain to root logical function")
        except Exception as e:
            print(f"Error adding functional chain to root logical function: {e}")

        # Add the functional chain to the involved functions
        for func in involved_functions:
            try:
                # Add the functional chain to the function's owned functional chains
                func.get_java_object().getOwnedFunctionalChains().add(fc.get_java_object())
                org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(fc.get_java_object())
                fc.get_java_object().setSid(fc_id)
                fc.set_name(fc_name)
            except Exception as e:
                print(f"Error adding functional chain to function {func.get_name()}: {e}")

        print(f"Successfully imported functional chain: {fc_name}")

    # After import, delete unused functional chains
    print("\nDeleting unused functional chains...")
    deleted_count = delete_unused_functional_chains(imported_chains)
    print(f"Deleted {deleted_count} functional chains that were not in the import file")

except Exception as e:
    print("Error: " + str(e))
    model.rollback_transaction()
    raise

# else:
model.commit_transaction()

model.save()
print("Functional chains import completed successfully")