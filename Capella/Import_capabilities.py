'''
Created on 02 Apr 2026

@author: p.jeukens
'''

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook
import os
import time

# Path names
aird_path = "/Import_test/Import_test.aird"
xlsx_path = "/DVS/results/DVS_LA_Capabilities.xlsx"
# aird_path = "/Import_test_IFES/Import_test_IFES.aird"
# xlsx_path = "/In-Flight Entertainment System/results/In-Flight Entertainment System_LA_Capabilities.xlsx"

def get_java_object(obj):
    """Get the Java object from a Python wrapper or return the object itself"""
    try:
        if hasattr(obj, 'get_java_object'):
            return obj.get_java_object()
        return obj
    except Exception as e:
        print(f"Error getting Java object: {str(e)}")
        return None

def get_component_id(component):
    """Get the ID of a component, handling both LogicalComponent and LogicalActor."""
    try:
        if hasattr(component, 'get_sid'):
            return component.get_sid()
        elif hasattr(component, 'getSid'):
            return component.getSid()
        elif hasattr(component, 'get_java_object'):
            java_obj = component.get_java_object()
            if hasattr(java_obj, 'getSid'):
                return java_obj.getSid()
    except Exception as e:
        print(f"Error getting component ID: {e}")
    return None

def get_component_name(component):
    """Get the name of a component, handling both LogicalComponent and LogicalActor."""
    try:
        if hasattr(component, 'get_name'):
            return component.get_name()
        elif hasattr(component, 'getName'):
            return component.getName()
        elif hasattr(component, 'get_java_object'):
            java_obj = component.get_java_object()
            if hasattr(java_obj, 'getName'):
                return java_obj.getName()
    except Exception as e:
        print(f"Error getting component name: {e}")
    return None

def find_capability_realization_by_id(capability_realization_pkg, cr_id):
    """Find a capability realization by its ID"""
    try:
        java_pkg = get_java_object(capability_realization_pkg)
        if not java_pkg:
            return None

        for cr in java_pkg.getOwnedCapabilityRealizations():
            if cr.getSid() == cr_id:
                return CapabilityRealization(cr)
        return None
    except Exception as e:
        print(f"Error finding capability realization by ID: {e}")
        return None

def find_functional_chain_by_id(fc_id):
    """Find a functional chain by its ID by searching through all functions"""
    try:
        # Get all functions in the model
        all_functions = se.get_all_contents_by_type(Function)

        # Search through all functions for owned functional chains
        for func in all_functions:
            java_func = get_java_object(func)
            for chain in java_func.getOwnedFunctionalChains():
                if chain.getSid() == fc_id:
                    return FunctionalChain(chain)

        return None
    except Exception as e:
        print(f"Error finding functional chain by ID: {e}")
        return None

def find_logical_function_by_id(func_id):
    """Find a logical function by its ID"""
    try:
        # Get all logical functions from the model
        all_functions = se.get_all_contents_by_type(LogicalFunction)
        for func in all_functions:
            if get_component_id(func) == func_id:
                return func
        return None
    except Exception as e:
        print(f"Error finding logical function by ID: {e}")
        return None

def find_logical_component_by_id(comp_id):
    """Find a logical component by its ID"""
    try:
        # Get all logical components from the model
        all_components = se.get_all_contents_by_type(LogicalComponent)
        for comp in all_components:
            if get_component_id(comp) == comp_id:
                return comp
        return None
    except Exception as e:
        print(f"Error finding logical component by ID: {e}")
        return None

def find_logical_actor_by_id(actor_id):
    """Find a logical actor by its ID"""
    try:
        # Get all logical actors from the model
        all_actors = se.get_all_contents_by_type(LogicalActor)
        for actor in all_actors:
            if get_component_id(actor) == actor_id:
                return actor
        return None
    except Exception as e:
        print(f"Error finding logical actor by ID: {e}")
        return None

def create_function_capability_involvement(cr, element, involvement_id=None):
    """Create an involvement between a function and a capability realization"""
    try:
        # Create the involvement object using the factory
        factory = org.polarsys.capella.core.data.interaction.InteractionFactory.eINSTANCE
        involvement = factory.createAbstractFunctionAbstractCapabilityInvolvement()

        # Get the Java objects
        java_cr = get_java_object(cr)
        java_element = get_java_object(element)

        # Add the involvement to the capability realization
        java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements().add(involvement)

        # Set the involved element
        involvement.setInvolved(java_element)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)
        
        if involvement_id:
            involvement.setSid(involvement_id)

        return involvement
    except Exception as e:
        print(f"Error creating function/capability involvement: {e}")
        import traceback
        traceback.print_exc()
        return None

def create_functional_chain_capability_involvement(cr, fc, involvement_id=None):
    """Create an involvement between a functional chain and a capability realization"""
    try:
        # Create the involvement object using the factory
        factory = org.polarsys.capella.core.data.interaction.InteractionFactory.eINSTANCE
        involvement = factory.createFunctionalChainAbstractCapabilityInvolvement()

        # Get the Java objects
        java_cr = get_java_object(cr)
        java_element = get_java_object(fc)

        # Add the involvement to the capability realization
        java_cr.getOwnedFunctionalChainAbstractCapabilityInvolvements().add(involvement)

        # Set the involved element
        involvement.setInvolved(java_element)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)
        
        if involvement_id:
            involvement.setSid(involvement_id)

        return involvement
    except Exception as e:
        print(f"Error creating function/capability involvement: {e}")
        import traceback
        traceback.print_exc()
        return None

def add_component_to_capability_realization(cr, component, involvement_id=None):
    """Add a component or actor to a capability realization"""
    try:
        # Get the Java objects
        java_cr = get_java_object(cr)
        java_component = get_java_object(component)

        # Check if the component is already involved
        already_involved = False
        for involved in java_cr.getInvolvedComponents():
            if involved.getSid() == java_component.getSid():
                already_involved = True
                break

        if already_involved:
            print(f"Component/actor {get_component_name(component)} is already involved with capability realization")
            return True

        # Create the involvement using the factory
        factory = org.polarsys.capella.core.data.capellacommon.CapellacommonFactory.eINSTANCE
        involvement = factory.createCapabilityRealizationInvolvement()

        # Set the involved element
        involvement.setInvolved(java_component)

        # Add the involvement to the capability realization
        java_cr.getOwnedCapabilityRealizationInvolvements().add(involvement)

        # Use creation service
        org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(involvement)

        # Set the involvement ID if provided
        if involvement_id:
            involvement.setSid(involvement_id)
        
        return True
    except Exception as e:
        print(f"Error adding component to capability realization: {e}")
        import traceback
        traceback.print_exc()
        return False

# Open the model with proper error handling
try:
    model = CapellaModel()
    model.open(aird_path)
    print("Model opened successfully")
except Exception as e:
    print(f"Failed to open model: {str(e)}")
    raise

# Gets the System Engineering and Logical Architecture
se = model.get_system_engineering()
la = se.get_logical_architecture()

# Get the capability realization package
capability_realization_pkg = la.get_capability_realization_pkg()
if not capability_realization_pkg:
    print("No Capability Realization Package found in Logical Architecture")
    exit()

# Load the workbook with error handling
try:
    xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
    xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)
    print("Reading " + xlsx_file_name)

    wb = load_workbook(xlsx_file_name)
    ws = wb.active
except Exception as e:
    print(f"Error loading workbook: {str(e)}")
    exit()

# Find the column indices for the different sections
header_row = ws[1]

# Find max functional chains, logical functions, and involved components from headers
max_functional_chains = 0
max_logical_functions = 0
max_involved_components = 0

function_start_col = 3
for col in range(function_start_col, ws.max_column + 1, 3):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value and "Functional Chain" in cell_value:
        max_functional_chains += 1

function_end_col = function_start_col + max_functional_chains * 3
for col in range(function_end_col, ws.max_column + 1, 3):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value and "Logical Function" in cell_value:
        max_logical_functions += 1

component_start_col = function_end_col + max_logical_functions * 3
for col in range(component_start_col, ws.max_column + 1, 4):
    cell_value = ws.cell(row=1, column=col).value
    if cell_value and "Involved Component" in cell_value:
        max_involved_components += 1

print(f"Found {max_functional_chains} functional chains, {max_logical_functions} logical functions, and {max_involved_components} involved components in the Excel file")

# Start the import with proper transaction handling
model.start_transaction()

try:
    # Process each row in the Excel file (skip header row)
    for row in ws.iter_rows(min_row=2):
        try:
            # Extract basic capability realization info
            cr_id = row[0].value
            cr_name = row[1].value

            if not cr_id:
                continue

            print(f"Processing capability realization: {cr_name}")

            # Check if the capability realization already exists
            existing_cr = find_capability_realization_by_id(capability_realization_pkg, cr_id)
            if existing_cr:
                print(f"Capability realization with ID {cr_id} and name '{cr_name}' already exists. Skipping import.")
                continue

            # Create a new capability realization
            cr = CapabilityRealization()
            java_cr = get_java_object(cr)

            # Add to the capability realization package first
            java_pkg = get_java_object(capability_realization_pkg)
            java_pkg.getOwnedCapabilityRealizations().add(java_cr)
            org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(java_cr)

            # Now set the ID and name
            java_cr.setSid(cr_id)
            cr.set_name(cr_name)

            # Process functional chains
            found_chains = 0
            for i in range(max_functional_chains):
                fc_id_col = function_start_col + i*3 - 1  # Convert to 0-based index
                fc_name_col = function_start_col + i*3
                fc_involvement_col = function_start_col + i*3 + 1

                if fc_id_col >= len(row) or not row[fc_id_col].value:
                    continue

                fc_id = row[fc_id_col].value
                fc_name = row[fc_name_col].value if fc_name_col < len(row) else ""
                fc_involvement_id = row[fc_involvement_col].value if fc_involvement_col < len(row) else ""

                if not fc_id:
                    continue

                # Find the functional chain
                fc = find_functional_chain_by_id(fc_id)
                if not fc:
                    print(f"WARNING: Could not find functional chain with ID {fc_id} and name '{fc_name}'. Skipping this chain.")
                    continue

                # Create involvement between functional chain and capability realization
                involvement = create_functional_chain_capability_involvement(cr, fc, fc_involvement_id)
                if involvement:
                    found_chains += 1
                    print(f"Added functional chain {fc_name} to capability realization {cr_name}")
                else:
                    print(f"Failed to create involvement for functional chain {fc_name}")

            # Process logical functions
            found_functions = 0
            for i in range(max_logical_functions):
                func_id_col = function_end_col + i*3 - 1  # Convert to 0-based index
                func_name_col = function_end_col + i*3
                func_involvement_col = function_end_col + i*3 + 1

                if func_id_col >= len(row) or not row[func_id_col].value:
                    continue

                func_id = row[func_id_col].value
                func_name = row[func_name_col].value if func_name_col < len(row) else ""
                func_involvement_id = row[func_involvement_col].value if func_involvement_col < len(row) else ""

                if not func_id:
                    continue

                # Find the logical function
                func = find_logical_function_by_id(func_id)
                if not func:
                    print(f"WARNING: Could not find logical function with ID {func_id} and name '{func_name}'. Skipping this function.")
                    continue

                # Create involvement between logical function and capability realization
                involvement = create_function_capability_involvement(cr, func, func_involvement_id)
                if involvement:
                    found_functions += 1
                    print(f"Added logical function {func_name} to capability realization {cr_name}")
                else:
                    print(f"Failed to create involvement for logical function {func_name}")

            # Process involved components
            found_components = 0
            for i in range(max_involved_components):
                comp_id_col = component_start_col + i*4 - 1  # Convert to 0-based index
                comp_name_col = component_start_col + i*4
                comp_type_col = component_start_col + i*4 + 1
                comp_involvement_col = component_start_col + i*4 + 2

                if comp_id_col >= len(row) or not row[comp_id_col].value:
                    continue

                comp_id = row[comp_id_col].value
                comp_name = row[comp_name_col].value if comp_name_col < len(row) else ""
                comp_type = row[comp_type_col].value if comp_type_col < len(row) else ""
                comp_involvement_id = row[comp_involvement_col].value if comp_involvement_col < len(row) else ""

                if not comp_id:
                    continue

                # Find the component or actor based on type
                component = None
                if comp_type == "actor":
                    component = find_logical_actor_by_id(comp_id)
                else:
                    component = find_logical_component_by_id(comp_id)

                if not component:
                    print(f"WARNING: Could not find {comp_type} with ID {comp_id} and name '{comp_name}'. Skipping this component.")
                    continue

                # Add component/actor to capability realization
                success = add_component_to_capability_realization(cr, component, comp_involvement_id)
                if success:
                    found_components += 1
                    print(f"Added {comp_type} {comp_name} to capability realization {cr_name}")
                else:
                    print(f"Failed to add {comp_type} {comp_name} to capability realization {cr_name}")

            print(f"Successfully imported capability realization: {cr_name} with {found_chains} functional chains, {found_functions} logical functions, and {found_components} involved components")

        except Exception as e:
            print(f"Error processing row: {str(e)}")
            import traceback
            traceback.print_exc()
            continue

    # Commit the transaction
    model.commit_transaction()
    print("Transaction committed successfully")

    # Save the model with delay to ensure all operations complete
    model.save()
    print("Model saved successfully")

    # Refresh the project
    project_name = aird_path.split('/')[-1].replace('.aird', '')
    project = CapellaPlatform.getProject(project_name)
    if project:
        time.sleep(1)  # Give Capella time to process the save
        CapellaPlatform.refresh(project)
        print("Project refreshed successfully")
    else:
        print(f"Could not find project {project_name} for refresh")

except Exception as e:
    print("Error: " + str(e))
    import traceback
    traceback.print_exc()
    model.rollback_transaction()
    raise

print("Import completed successfully")