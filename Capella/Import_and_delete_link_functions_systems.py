'''
Created on 13 Apr 2026

@author: p.jeukens
'''
# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

# include needed to read/write xlsx files
from openpyxl import load_workbook

# Path names
aird_path = "/delete_test/delete_test.aird"
xlsx_path = "/DVS/results/DVS_Function_System_Links2.xlsx"

model = CapellaModel()
model.open(aird_path)

# Gets the System Engineering
se = model.get_system_engineering()
la = se.get_logical_architecture()

# Get the logical component package and root logical function
lc_pkg = la.get_logical_component_pkg()
lf_pkg = la.get_logical_function_pkg()

# Find the Root Logical Function
rlf = None
for lf in lf_pkg.get_owned_logical_functions():
    if lf.get_name() == "Root Logical Function":
        rlf = lf
        break

if rlf is None:
    print("Root Logical Function not found!")
    exit()

# Create a folder in the project
xlsx_file = CapellaPlatform.getWorkspaceFile(xlsx_path)
xlsx_file_name = CapellaPlatform.getAbsolutePath(xlsx_file)

print("Read " + xlsx_file_name)

# Load the workbook
wb = load_workbook(xlsx_file_name)

# Grab the active worksheet
ws = wb.active

def get_component_by_id(java_object, component_id):
    """Find a component by its ID in the children of a Java object."""
    for child in java_object.getOwnedLogicalComponents():
        if child.getSid() == component_id:
            return child
    return None

def find_component_recursive(java_object, component_id):
    """Recursively find a component by its ID in the hierarchy."""
    # First check direct children
    component = get_component_by_id(java_object, component_id)
    if component:
        return component

    # If not found, recursively check children
    children = java_object.getOwnedLogicalComponents()
    for child in children:
        found = find_component_recursive(child, component_id)
        if found:
            return found

    return None

def find_function_by_id(parent, function_id):
    """Find a function by ID in the hierarchy"""
    if parent.get_sid() == function_id:
        return parent

    for function in parent.get_owned_functions():
        if function.get_sid() == function_id:
            return function
        found = find_function_by_id(function, function_id)
        if found:
            return found
    return None

def collect_all_allocations():
    """Collect all existing system-function allocations in the model"""
    allocations = {}

    # Get all components
    java_lc_pkg = lc_pkg.get_java_object()
    components = [java_lc_pkg.getOwnedLogicalComponents()]

    while components:
        current_level = components.pop()
        for component in current_level:
            # Check for allocations
            allocated_functions = component.getAllocatedLogicalFunctions()
            for func in allocated_functions:
                system_id = component.getSid()
                function_id = func.getSid()
                if system_id not in allocations:
                    allocations[system_id] = []
                allocations[system_id].append(function_id)

            # Add children to the processing queue
            children = component.getOwnedLogicalComponents()
            if children:
                components.append(children)

    return allocations

def delete_unused_allocations(imported_allocations):
    """Delete allocations that exist in the model but were not imported"""
    try:
        # Get all components
        java_lc_pkg = lc_pkg.get_java_object()
        components = [java_lc_pkg.getOwnedLogicalComponents()]

        deleted_count = 0

        while components:
            current_level = components.pop()
            for component in current_level:
                # Get the system name for logging
                system_name = "Unknown"
                try:
                    if component.isActor():
                        system = LogicalActor(component)
                    else:
                        system = LogicalComponent(component)
                    system_name = system.get_name()
                except:
                    pass

                # Check for allocations to delete
                allocated_functions = list(component.getAllocatedLogicalFunctions())
                for func in allocated_functions:
                    system_id = component.getSid()
                    function_id = func.getSid()

                    # Check if this allocation should be kept
                    if system_id in imported_allocations and function_id in imported_allocations[system_id]:
                        continue  # This allocation is in the import file

                    # This allocation is not in the import file, delete it
                    try:
                        # Find the allocation object
                        allocations = component.getOwnedFunctionalAllocation()
                        for allocation in allocations:
                            if allocation.getTargetElement().getSid() == function_id:
                                # Get function name for logging
                                function_name = "Unknown"
                                try:
                                    function = LogicalFunction(func)
                                    function_name = function.get_name()
                                except:
                                    pass

                                print(f"Deleting allocation: System '{system_name}' (ID: {system_id}) -> Function '{function_name}' (ID: {function_id})")
                                component.getOwnedFunctionalAllocation().remove(allocation)
                                deleted_count += 1
                                break
                    except Exception as e:
                        print(f"Error deleting allocation between {system_name} and function {function_id}: {str(e)}")

                # Add children to the processing queue
                children = component.getOwnedLogicalComponents()
                if children:
                    components.append(children)

        print(f"Deleted {deleted_count} unused allocations")
        return deleted_count

    except Exception as e:
        print(f"Error in delete_unused_allocations: {str(e)}")
        return 0

# Find the column where "Function ID" appears to determine system depth
function_id_col = None
for cell in ws[1]:
    if cell.value == "Function ID":
        function_id_col = cell.column
        break

if function_id_col is None:
    print("Could not find 'Function ID' column in the Excel file.")
    exit()

# Calculate system depth
system_max_depth = (function_id_col - 1) // 2

# print(f"Maximum system depth: {system_max_depth}")

# Start transaction
model.start_transaction()

try:
    # First collect all existing allocations
    # print("Collecting existing allocations...")
    existing_allocations = collect_all_allocations()
    # print(f"Found {sum(len(v) for v in existing_allocations.values())} existing allocations")

    # Dictionary to store imported allocations: {system_id: [function_ids]}
    imported_allocations = {}

    # Store the current system path for each row
    current_system_java = None
    current_system_id = None
    current_system_level = 0

    for row in ws.iter_rows(min_row=2):
        # Check if this row contains a system (has value in any system column)
        system_found = False
        system_id = None
        system_col = -1

        # Find which system column has a value
        for level in range(system_max_depth):
            col = level * 2
            if col < len(row) and row[col].value:
                system_id = row[col].value
                system_col = col
                system_found = True
                break  # Use the first system found (highest level)

        # If we found a system, determine its level and find it
        if system_found:
            # Level is determined by column position: col 0-1 = level 0, col 2-3 = level 1, etc.
            current_system_level = system_col // 2
            current_system_id = system_id

            # Find the system using recursive search
            java_lc_pkg = lc_pkg.get_java_object()
            current_system_java = find_component_recursive(java_lc_pkg, current_system_id)

            if not current_system_java:
                print(f"System with ID {current_system_id} not found at level {current_system_level + 1}")
                continue

            # Create a Python wrapper for display purposes
            if current_system_java.isActor():
                current_system = LogicalActor(current_system_java)
            else:
                current_system = LogicalComponent(current_system_java)

            # print(f"Found system: {current_system.get_name()} at level {current_system_level + 1}")

        # If this row contains a function (has value in function ID column)
        if function_id_col <= len(row) and row[function_id_col-1].value:
            function_id = row[function_id_col-1].value
            function_name = row[function_id_col].value if function_id_col < len(row) else ""

            if current_system_java:
                # Find the function
                function = find_function_by_id(rlf, function_id)

                if function:
                    try:
                        # Check if the allocation already exists
                        allocated_functions = current_system_java.getAllocatedLogicalFunctions()
                        already_allocated = False
                        for allocated_func in allocated_functions:
                            if allocated_func.getSid() == function_id:
                                already_allocated = True
                                break

                        if not already_allocated:
                            # Create the allocation
                            cfa = create_e_object("http://www.polarsys.org/capella/core/fa/" + capella_version(), "ComponentFunctionalAllocation")
                            cfa.setSourceElement(current_system_java)
                            cfa.setTargetElement(function.get_java_object())
                            current_system_java.getOwnedFunctionalAllocation().add(cfa)
                            org.polarsys.capella.core.model.helpers.CapellaElementExt.creationService(cfa)
                            # print(f"Allocated function '{function.get_name()}' to system '{current_system.get_name()}'.")

                            # Add to imported allocations
                            if current_system_id not in imported_allocations:
                                imported_allocations[current_system_id] = []
                            imported_allocations[current_system_id].append(function_id)
                        else:
                            # print(f"Function '{function.get_name()}' is already allocated to system '{current_system.get_name()}'.")

                            # Add to imported allocations
                            if current_system_id not in imported_allocations:
                                imported_allocations[current_system_id] = []
                            imported_allocations[current_system_id].append(function_id)
                    except Exception as e:
                        print(f"Failed to allocate function '{function.get_name()}' to system '{current_system.get_name()}': {str(e)}")
                else:
                    print(f"Function with ID {function_id} and name {function_name} not found.")
            else:
                print("No current system defined for function allocation.")

    # After import, delete unused allocations
    print("\nDeleting unused allocations...")
    deleted_count = delete_unused_allocations(imported_allocations)
    print(f"Deleted {deleted_count} allocations that were not in the import file")

except Exception as e:
    print("Error: " + str(e))
    model.rollback_transaction()
    raise

# else:
model.commit_transaction()

model.save()
print("Functions and systems linked successfully")