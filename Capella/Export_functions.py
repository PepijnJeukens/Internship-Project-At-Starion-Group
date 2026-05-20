'''
Created on 08 Apr 2026

@author: p.jeukens
'''

""" 
This script contains all functions necessary to export the Logical Architecture layer from Capella to Excel
It can export systems (components/actors), functions, the link between systems and functions, functional exchanges,
component exchanges, the link between functional and component exchanges, functional chains and capabilities
"""

# include needed for the Capella modeller API
include('workspace://Python4Capella/simplified_api/capella.py')
if False:
    from simplified_api.capella import *

# include needed for utilities
include('workspace://Python4Capella/utilities/CapellaPlatform.py')
if False:
    from utilities.CapellaPlatform import *

include('workspace://Python4Capella/simplified_api/requirement.py')
if False:
    from simplified_api.requirement import *    

from openpyxl import Workbook
import os
import xml.etree.ElementTree as ET
from xml.dom import minidom
import html
from html.parser import HTMLParser
import random
import string
from datetime import datetime
import pytz
import re

def get_model_and_setup(aird_path):
    """Open the model and set up project paths"""
    # Open the model
    model = CapellaModel()
    model.open(aird_path)

    # Get system engineering
    se = model.get_system_engineering()

    # Set up project paths
    model_path = CapellaPlatform.getModelPath(se)
    project_name = model_path[0:(model_path.index("/", 1) + 1)]
    project_name = project_name.replace("/", "")  # Clean up project name
    project = CapellaPlatform.getProject(project_name)
    folder = CapellaPlatform.getFolder(project, "results")

    return model, se, project_name, folder

def get_logical_layer(model):
    """Get logical architecture and its packages"""
    la = model.get_system_engineering().get_logical_architecture()

    # Get packages
    lc_pkg = la.get_logical_component_pkg()
    lf_pkg = la.get_logical_function_pkg()

    # Find root logical function
    rlf = None
    for lf in lf_pkg.get_owned_logical_functions():
        if lf.get_name() == "Root Logical Function":
            rlf = lf
            break

    return la, lc_pkg, lf_pkg, rlf

def create_master_workbook():
    """Create a new workbook with all required sheets"""
    wb = Workbook()

    # Remove default sheet if it exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create worksheets
    wb.create_sheet("Systems")
    wb.create_sheet("Functions")
    wb.create_sheet("Link Systems and Functions")
    wb.create_sheet("Functional Exchanges")
    wb.create_sheet("Component Exchanges")
    wb.create_sheet("Link Exchanges")
    wb.create_sheet("Functional Chains")
    wb.create_sheet("Capabilities")

    # Create worksheets for ReqIF exports
    wb.create_sheet("HEADER")
    wb.create_sheet("DATATYPES")
    wb.create_sheet("ENUM-VALUES")
    wb.create_sheet("SPEC-TYPES")
    wb.create_sheet("SPEC-ATTRIBUTES")
    wb.create_sheet("SPEC-OBJECTS")
    wb.create_sheet("ATTRIBUTE-VALUES")
    wb.create_sheet("SPEC-RELATIONS")
    wb.create_sheet("SPECIFICATIONS")
    wb.create_sheet("SPEC-HIERARCHY")

    return wb

def get_id(obj):
    """Get the ID of an object, handling both Python and Java objects."""
    try:
        return obj.get_id()
    except Exception: #Changed to except Exception from except AttributeError, so in case some things dont work change this back and deal with the get_id(involvement.getInvolvedElement()) differently
        try:
            return obj.get_java_object().getId()
        except Exception:
            return obj.getId()

def get_name(obj):
    """Get the name of an object, handling both Python and Java objects."""
    try:
        return obj.get_name()
    except Exception: #Changed to except Exception from except AttributeError, so in case some things dont work change this back and deal with the get name in the latest export functional chains differently
        try:
            return obj.get_java_object().getName()
        except Exception:
            return obj.getName()

def get_max_depth(java_object, current_depth=1, is_function=False):
    """Recursively find the maximum depth of a hierarchy"""
    max_depth = current_depth
    try:
        if is_function:
            children = java_object.getOwnedFunctions()
        else:
            children = java_object.getOwnedLogicalComponents()

        for child in children:
            if is_function:
                child_obj = LogicalFunction(child)
                child_depth = get_max_depth(child_obj.get_java_object(), current_depth + 1, True)
            else:
                child_depth = get_max_depth(child, current_depth + 1)

            if child_depth > max_depth:
                max_depth = child_depth
    except Exception as e:
        print(f"Error getting children: {e}")
    return max_depth

def get_function_kind(function):
    """Determine the kind of function"""
    try:
        java_obj = function.get_java_object()
        kind = str(java_obj.getKind())

        kind_mapping = {
            "DUPLICATE": "DUPLICATE",
            "GATHER": "GATHER",
            "ROUTE": "ROUTE",
            "SELECT": "SELECT",
            "SPLIT": "SPLIT"
        }

        return kind_mapping.get(kind, "FUNCTION")
    except Exception as e:
        print(f"Error determining function kind: {e}")
        return "FUNCTION"

def create_system_headers(ws, max_depth):
    """Create headers for the Systems worksheet"""
    for i in range(max_depth):
        col = i * 3 + 1
        ws.cell(row=1, column=col, value=f"{'Sub' * i}System ID")
        ws.cell(row=1, column=col + 1, value=f"{'Sub' * i}System Name")
        ws.cell(row=1, column=col + 2, value=f"{'Sub' * i}System Type")

def create_function_headers(ws, max_depth):
    """Create headers for the Functions worksheet"""
    for i in range(max_depth):
        col = i * 3 + 1
        ws.cell(row=1, column=col, value=f"{'Sub' * i}Function ID")
        ws.cell(row=1, column=col + 1, value=f"{'Sub' * i}Function Name")
        ws.cell(row=1, column=col + 2, value=f"{'Sub' * i}Function Kind")

def create_link_headers(ws, system_max_depth, function_max_depth):
    """Create headers for the Link Systems and Functions worksheet"""
    # System headers
    for i in range(system_max_depth):
        col = i * 2 + 1
        ws.cell(row=1, column=col, value=f"{'Sub' * i}System ID")
        ws.cell(row=1, column=col + 1, value=f"{'Sub' * i}System Name")

    # Function headers
    function_start_col = system_max_depth * 2 + 1
    for i in range(function_max_depth):
        col = function_start_col + i * 2
        ws.cell(row=1, column=col, value=f"{'Sub' * i}Function ID")
        ws.cell(row=1, column=col + 1, value=f"{'Sub' * i}Function Name")

def export_component(component, ws, row, level=0):
    """Recursively export a component and its children"""
    col = level * 3 + 1
    ws.cell(row=row, column=col, value=get_id(component))
    ws.cell(row=row, column=col + 1, value=get_name(component))
    ws.cell(row=row, column=col + 2, value="Actor" if component.get_java_object().isActor() else "Component")

    next_row = row + 1

    try:
        java_object = component.get_java_object()
        children = java_object.getOwnedLogicalComponents()
        if children.size() > 0:
            for child in children:
                child_component = LogicalActor(child) if child.isActor() else LogicalComponent(child)
                next_row = export_component(child_component, ws, next_row, level + 1)
    except Exception as e:
        print(f"Error exporting children for {get_name(component)}: {e}")

    return next_row

def export_function(function, ws, row, level=0):
    """Recursively export a function and its children"""
    col = level * 3 + 1
    ws.cell(row=row, column=col, value=get_id(function))
    ws.cell(row=row, column=col + 1, value=get_name(function))
    ws.cell(row=row, column=col + 2, value=get_function_kind(function))

    next_row = row + 1

    try:
        children = function.get_java_object().getOwnedFunctions()
        if children.size() > 0:
            for child in children:
                next_row = export_function(LogicalFunction(child), ws, next_row, level + 1)
    except Exception as e:
        print(f"Error exporting children for {get_name(function)}: {e}")

    return next_row

def export_system_functions(system, ws, row, system_level=0, function_start_col=None):
    """Export all functions allocated to a system and its children"""
    try:
        # Write system info
        col = system_level * 2 + 1
        ws.cell(row=row, column=col, value=get_id(system))
        ws.cell(row=row, column=col + 1, value=get_name(system))
        row += 1

        # Get allocated functions
        java_object = system.get_java_object()
        allocated_functions = java_object.getAllocatedLogicalFunctions()

        # Write each allocated function
        for func in allocated_functions:
            function = LogicalFunction(func)
            ws.cell(row=row, column=function_start_col, value=get_id(function))
            ws.cell(row=row, column=function_start_col + 1, value=get_name(function))
            row += 1

        # Process children
        children = java_object.getOwnedLogicalComponents()
        for child in children:
            child_system = LogicalActor(child) if child.isActor() else LogicalComponent(child)
            row = export_system_functions(child_system, ws, row, system_level + 1, function_start_col)

    except Exception as e:
        print(f"Error processing system {get_name(system)}: {e}")

    return row

def export_systems(wb):
    """Export all systems to the Systems worksheet"""
    try:
        # Get model and setup from workbook properties
        model = wb._custom_doc_props["model"]
        la, lc_pkg, _, _ = get_logical_layer(model)

        # Get worksheet and determine max depth
        ws_systems = wb["Systems"]
        max_depth = 0
        java_lc_pkg = lc_pkg.get_java_object()
        children = java_lc_pkg.getOwnedLogicalComponents()

        for child in children:
            depth = get_max_depth(child)
            if depth > max_depth:
                max_depth = depth

        create_system_headers(ws_systems, max_depth)

        # Export components
        row = 2
        components = lc_pkg.get_owned_logical_components()
        print(f"Exported the Systems")
        for lc in components:
            component = LogicalActor(lc.get_java_object()) if lc.get_java_object().isActor() else lc
            row = export_component(component, ws_systems, row)

        return True
    except Exception as e:
        print(f"Error exporting systems: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def export_functions(wb):
    """Export all functions to the Functions worksheet"""
    try:
        # Get model and setup from workbook properties
        model = wb._custom_doc_props["model"]
        _, _, lf_pkg, rlf = get_logical_layer(model)

        if rlf is None:
            print("Root Logical Function not found!")
            return False

        # Get worksheet and determine max depth
        ws_functions = wb["Functions"]
        max_depth = 0
        children = rlf.get_java_object().getOwnedFunctions()

        for child in children:
            depth = get_max_depth(LogicalFunction(child).get_java_object(), 1, True)
            if depth > max_depth:
                max_depth = depth

        create_function_headers(ws_functions, max_depth)

        # Export functions
        row = 2
        functions = []
        for child in children:
            functions.append(LogicalFunction(child))
            row = export_function(LogicalFunction(child), ws_functions, row)

        print(f"Exported the Functions")
        return True
    except Exception as e:
        print(f"Error exporting functions: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def export_links(wb):
    """Export system-function links to the Link Systems and Functions worksheet"""
    try:
        # Get model and setup from workbook properties
        model = wb._custom_doc_props["model"]
        la, lc_pkg, lf_pkg, rlf = get_logical_layer(model)

        # Get worksheet
        ws_links = wb["Link Systems and Functions"]

        # Determine max depths
        system_max_depth = 0
        java_lc_pkg = lc_pkg.get_java_object()
        children = java_lc_pkg.getOwnedLogicalComponents()
        for child in children:
            depth = get_max_depth(child)
            if depth > system_max_depth:
                system_max_depth = depth

        function_max_depth = 0
        if rlf:
            rlf_children = rlf.get_java_object().getOwnedFunctions()
            for child in rlf_children:
                depth = get_max_depth(LogicalFunction(child).get_java_object(), 1, True)
                if depth > function_max_depth:
                    function_max_depth = depth

        create_link_headers(ws_links, system_max_depth, function_max_depth)

        # Export links
        function_start_col = system_max_depth * 2 + 1
        row = 2
        components = lc_pkg.get_owned_logical_components()
        print(f"Exported the link between Systems and Functions")
        for lc in components:
            row = export_system_functions(lc, ws_links, row, 0, function_start_col)

        return True
    except Exception as e:
        print(f"Error exporting links: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def save_workbook(wb):
    """Save the workbook using properties stored in the workbook"""
    try:
        se = wb._custom_doc_props["se"]
        folder = wb._custom_doc_props["folder"]
        project_name = wb._custom_doc_props["project_name"]

        xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + "/" + project_name + "_complete_export.xlsx"
        print("Writing " + xlsx_file_name)
        wb.save(filename=xlsx_file_name)
        print("File saved successfully.")
        return True
    except Exception as e:
        print(f"Error saving workbook: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def setup_workbook(aird_path):
    """Set up the workbook with model information and return it"""
    try:
        # Get model and setup information
        model, se, project_name, folder = get_model_and_setup(aird_path)
        la, lc_pkg, lf_pkg, rlf = get_logical_layer(model)

        # Create the workbook
        wb = create_master_workbook()

        # Store model information in custom properties
        wb._custom_doc_props = {
            "model": model,
            "se": se,
            "project_name": project_name,
            "folder": folder,
            "la": la,
            "lc_pkg": lc_pkg,
            "lf_pkg": lf_pkg,
            "rlf": rlf
        }

        return wb
    except Exception as e:
        print(f"Error setting up workbook: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def create_functional_exchange_headers(ws):
    """Create headers for the Functional Exchanges worksheet"""
    ws.cell(row=1, column=1, value="Function From ID")
    ws.cell(row=1, column=2, value="Function From Name")
    ws.cell(row=1, column=3, value="Function From Port ID")
    ws.cell(row=1, column=4, value="Function From Port Name")
    ws.cell(row=1, column=5, value="Functional Exchange ID")
    ws.cell(row=1, column=6, value="Functional Exchange Name")
    ws.cell(row=1, column=7, value="Function To ID")
    ws.cell(row=1, column=8, value="Function To Name")
    ws.cell(row=1, column=9, value="Function To Port ID")
    ws.cell(row=1, column=10, value="Function To Port Name")

def export_functional_exchange(exchange, ws, row):
    """Export a single functional exchange"""
    try:
        # Get source and target information
        source_function = exchange.get_source_function()
        target_function = exchange.get_target_function()

        # Get source and target ports
        source_port = exchange.get_source_port()
        target_port = exchange.get_target_port()

        if target_function and source_port and target_port:
            # Write exchange information
            ws.cell(row=row, column=1, value=get_id(source_function))
            ws.cell(row=row, column=2, value=get_name(source_function))
            ws.cell(row=row, column=3, value=get_id(source_port))
            ws.cell(row=row, column=4, value=get_name(source_port))
            ws.cell(row=row, column=5, value=get_id(exchange))
            ws.cell(row=row, column=6, value=get_name(exchange))
            ws.cell(row=row, column=7, value=get_id(target_function))
            ws.cell(row=row, column=8, value=get_name(target_function))
            ws.cell(row=row, column=9, value=get_id(target_port))
            ws.cell(row=row, column=10, value=get_name(target_port))
            return row + 1
    except Exception as e:
        print(f"Error exporting functional exchange: {e}")
        return row

def export_functional_exchanges(function, ws, row):
    """Recursively export functional exchanges for a function and its children"""
    try:
        # Get all outgoing functional exchanges
        outgoing_exchanges = function.get_outgoing()

        for fe in outgoing_exchanges:
            row = export_functional_exchange(fe, ws, row)

        # Process children
        for child in function.get_owned_functions():
            row = export_functional_exchanges(LogicalFunction(child), ws, row)

    except Exception as e:
        print(f"Error processing function {get_name(function)}: {e}")

    return row

def export_functional_exchange_sheet(wb):
    """Export all functional exchanges to the Functional Exchanges worksheet"""
    try:
        # Get model and setup from workbook properties
        model = wb._custom_doc_props["model"]
        _, _, lf_pkg, rlf = get_logical_layer(model)

        if rlf is None:
            print("Root Logical Function not found!")
            return False

        # Get worksheet
        ws_exchanges = wb["Functional Exchanges"]

        # Create headers
        create_functional_exchange_headers(ws_exchanges)

        # Track exported exchanges to avoid duplicates
        exported_exchanges = set()
        row = 2

        # Export functional exchanges
        for child in rlf.get_java_object().getOwnedFunctions():
            row = export_functional_exchanges(LogicalFunction(child), ws_exchanges, row)

        print(f"Exported the Functional Exchanges")
        return True
    except Exception as e:
        print(f"Error exporting functional exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def create_component_exchange_headers(ws):
    """Create headers for the Component Exchanges worksheet"""
    ws.cell(row=1, column=1, value="Component From ID")
    ws.cell(row=1, column=2, value="Component From Name")
    ws.cell(row=1, column=3, value="Component From Port ID")
    ws.cell(row=1, column=4, value="Component From Port Name")
    ws.cell(row=1, column=5, value="Component From Port Direction")
    ws.cell(row=1, column=6, value="Component From Port Kind")
    ws.cell(row=1, column=7, value="Component Exchange ID")
    ws.cell(row=1, column=8, value="Component Exchange Name")
    ws.cell(row=1, column=9, value="Component Exchange Kind")
    ws.cell(row=1, column=10, value="Component To ID")
    ws.cell(row=1, column=11, value="Component To Name")
    ws.cell(row=1, column=12, value="Component To Port ID")
    ws.cell(row=1, column=13, value="Component To Port Name")
    ws.cell(row=1, column=14, value="Component To Port Direction")
    ws.cell(row=1, column=15, value="Component To Port Kind")

def get_component_from_port(port):
    """Get the component that owns a port"""
    if not port:
        return None

    try:
        container = port.get_java_object().eContainer()
        if container:
            return create_component_from_java(container)
    except Exception as e:
        print(f"Error getting component from port: {e}")
    return None

def get_port_direction(port):
    """Get the direction of a port as a string"""
    try:
        direction = port.get_orientation()
        if direction == "UNSET":
            return "UNSET"
        elif direction == "IN":
            return "IN"
        elif direction == "OUT":
            return "OUT"
        elif direction == "INOUT":
            return "INOUT"
        else:
            return "UNSET"
    except Exception as e:
        print(f"Error getting port direction: {e}")
        return "UNKNOWN"

def get_port_kind(port):
    """Get the kind of a port (STANDARD or FLOW)"""
    try:
        port_java = port.get_java_object()
        if hasattr(port_java, 'eClass') and port_java.eClass().getName() == 'ComponentPort':
            kind = port_java.getKind()
            if kind:
                if kind.getName() == 'STANDARD':
                    return "STANDARD"
                elif kind.getName() == 'FLOW':
                    return "FLOW"
        return "STANDARD"
    except Exception as e:
        print(f"Error getting port kind: {e}")
        return "STANDARD"

def export_component_exchange(exchange, ws, row, exported_exchanges):
    """Export a single component exchange"""
    try:
        ce_id = exchange.get_id()

        # Skip if we've already exported this exchange
        if ce_id in exported_exchanges:
            return row

        # Get source and target ports
        source_port = exchange.get_source_port()
        target_port = exchange.get_target_port()

        # Get source and target components from ports
        source_component = get_component_from_port(source_port)
        target_component = get_component_from_port(target_port)

        if source_component and target_component and source_port and target_port:
            # Get port directions
            source_port_dir = get_port_direction(source_port)
            target_port_dir = get_port_direction(target_port)

            # Get port kinds
            source_port_kind = get_port_kind(source_port)
            target_port_kind = get_port_kind(target_port)

            # Write exchange information
            ws.cell(row=row, column=1, value=get_id(source_component))
            ws.cell(row=row, column=2, value=get_name(source_component))
            ws.cell(row=row, column=3, value=get_id(source_port))
            ws.cell(row=row, column=4, value=get_name(source_port))
            ws.cell(row=row, column=5, value=source_port_dir)
            ws.cell(row=row, column=6, value=source_port_kind)
            ws.cell(row=row, column=7, value=exchange.get_id())
            ws.cell(row=row, column=8, value=exchange.get_name())
            ws.cell(row=row, column=9, value=exchange.get_kind())
            ws.cell(row=row, column=10, value=get_id(target_component))
            ws.cell(row=row, column=11, value=get_name(target_component))
            ws.cell(row=row, column=12, value=get_id(target_port))
            ws.cell(row=row, column=13, value=get_name(target_port))
            ws.cell(row=row, column=14, value=target_port_dir)
            ws.cell(row=row, column=15, value=target_port_kind)

            exported_exchanges.add(ce_id)
            return row + 1

    except Exception as e:
        print(f"Error exporting component exchange: {e}")
        return row

def process_component_exchanges(component, ws, row, exported_exchanges):
    """Process component exchanges for a component and optionally its children"""
    try:
        # Process exchanges owned by this component
        java_obj = component.get_java_object()
        exchanges = java_obj.getOwnedComponentExchanges()
        
        for ce in exchanges:
            ce_obj = ComponentExchange(ce)
            
            row = export_component_exchange(ce_obj, ws, row, exported_exchanges)

        if java_obj:
            children = java_obj.getOwnedLogicalComponents()
            for child in children:
                child_component = create_component_from_java(child)
                if child_component:
                    row = process_component_exchanges(child_component, ws, row, exported_exchanges)

    except Exception as e:
        print(f"Error processing component {get_name(component)}: {e}")

    return row

def export_component_exchanges(wb):
    """Export all component exchanges to the Component Exchanges worksheet"""
    try:
        # Get model and setup from workbook properties
        model = wb._custom_doc_props["model"]
        la = wb._custom_doc_props["la"]

        if not la:
            print("Logical Architecture not found!")
            return False

        # Get worksheet
        ws_exchanges = wb["Component Exchanges"]

        # Create headers
        create_component_exchange_headers(ws_exchanges)

        # Track exported exchanges to avoid duplicates
        exported_exchanges = set()

        # Export all component exchanges
        lc_pkg = la.get_logical_component_pkg()
        row = process_component_exchanges(lc_pkg, ws_exchanges, 2, exported_exchanges)

        print(f"Exported the Component Exchanges")
        return True
    except Exception as e:
        print(f"Error exporting component exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def create_link_exchanges_headers(ws):
    """Create headers for the Link Exchanges worksheet"""
    ws.cell(row=1, column=1, value="Component Exchange ID")
    ws.cell(row=1, column=2, value="Component Exchange Name")
    ws.cell(row=1, column=3, value="Component Source Port ID")
    ws.cell(row=1, column=4, value="Component Source Port Name")
    ws.cell(row=1, column=5, value="Source Component ID")
    ws.cell(row=1, column=6, value="Source Component Name")
    ws.cell(row=1, column=7, value="Component Target Port ID")
    ws.cell(row=1, column=8, value="Component Target Port Name")
    ws.cell(row=1, column=9, value="Target Component ID")
    ws.cell(row=1, column=10, value="Target Component Name")
    ws.cell(row=1, column=11, value="Functional Exchange ID")
    ws.cell(row=1, column=12, value="Functional Exchange Name")
    ws.cell(row=1, column=13, value="Functional Source Port ID")
    ws.cell(row=1, column=14, value="Functional Source Port Name")
    ws.cell(row=1, column=15, value="Functional Target Port ID")
    ws.cell(row=1, column=16, value="Functional Target Port Name")
    ws.cell(row=1, column=17, value="Allocation ID")
    ws.cell(row=1, column=18, value="Source Function ID")
    ws.cell(row=1, column=19, value="Target Function ID")

def create_component_from_java(java_obj):
    """Create the appropriate component type from a Java object"""
    try:
        if hasattr(java_obj, 'isActor') and java_obj.isActor():
            return LogicalActor(java_obj)
        else:
            return LogicalComponent(java_obj)
    except Exception as e:
        print(f"Error creating component: {e}")
        return None

def get_component_exchange_ports(ce):
    """Get the source and target ports of a component exchange"""
    try:
        source_port = ce.get_source_port()
        target_port = ce.get_target_port()
        return source_port, target_port
    except Exception as e:
        print(f"Error getting component exchange ports: {e}")
        return None, None

def get_functional_exchange_ports(fe):
    """Get the source and target ports of a functional exchange"""
    try:
        source_port = fe.get_source_port()
        target_port = fe.get_target_port()
        return source_port, target_port
    except Exception as e:
        print(f"Error getting functional exchange ports: {e}")
        return None, None

def get_component_for_port(port):
    """Get the component that owns a port"""
    try:
        if not port:
            return None

        container = port.get_java_object().eContainer()
        if container:
            return create_component_from_java(container)
        return None
    except Exception as e:
        print(f"Error getting component for port: {e}")
        return None

def get_function_for_port(port):
    """Get the function that owns a port"""
    try:
        if not port:
            return None

        port_java = port.get_java_object()
        container = port_java.eContainer()

        if container and hasattr(container, 'eClass') and 'Function' in container.eClass().getName():
            return Function(container)

        # Fallback: search through all functions
        se = wb._custom_doc_props["se"]
        all_functions = se.get_all_contents_by_type(Function)
        for func in all_functions:
            try:
                for p in func.get_inputs():
                    if p.get_id() == port.get_id():
                        return func
                for p in func.get_outputs():
                    if p.get_id() == port.get_id():
                        return func
            except:
                continue

        return None
    except Exception as e:
        print(f"Error getting function for port: {e}")
        return None

def export_link_exchange_allocation(allocation, ce_obj, ws, row, exported_allocations):
    """Export a single link exchange allocation"""
    try:
        allocation_id = allocation.get_id()

        # Skip if we've already exported this allocation
        if allocation_id in exported_allocations:
            return row

        # Get the functional exchange
        fe = allocation.get_allocated_functional_exchange()

        if fe:
            # Get component exchange ports
            ce_source_port, ce_target_port = get_component_exchange_ports(ce_obj)

            # Get functional exchange ports
            fe_source_port, fe_target_port = get_functional_exchange_ports(fe)

            # Get components for the ports
            source_component = get_component_for_port(ce_source_port) if ce_source_port else None
            target_component = get_component_for_port(ce_target_port) if ce_target_port else None

            # Get functions for the ports
            source_function = get_function_for_port(fe_source_port) if fe_source_port else None
            target_function = get_function_for_port(fe_target_port) if fe_target_port else None

            # Write allocation information
            ws.cell(row=row, column=1, value=get_id(ce_obj))
            ws.cell(row=row, column=2, value=ce_obj.get_name())
            ws.cell(row=row, column=3, value=get_id(ce_source_port) if ce_source_port else "")
            ws.cell(row=row, column=4, value=ce_source_port.get_name() if ce_source_port else "")
            ws.cell(row=row, column=5, value=get_id(source_component) if source_component else "")
            ws.cell(row=row, column=6, value=source_component.get_name() if source_component else "")
            ws.cell(row=row, column=7, value=get_id(ce_target_port) if ce_target_port else "")
            ws.cell(row=row, column=8, value=ce_target_port.get_name() if ce_target_port else "")
            ws.cell(row=row, column=9, value=get_id(target_component) if target_component else "")
            ws.cell(row=row, column=10, value=target_component.get_name() if target_component else "")
            ws.cell(row=row, column=11, value=get_id(fe))
            ws.cell(row=row, column=12, value=fe.get_name())

            # Export functional source port ID and name
            ws.cell(row=row, column=13, value=get_id(fe_source_port) if fe_source_port else "")
            ws.cell(row=row, column=14, value=fe_source_port.get_name() if fe_source_port else "")

            # Export functional target port ID and name
            ws.cell(row=row, column=15, value=get_id(fe_target_port) if fe_target_port else "")
            ws.cell(row=row, column=16, value=fe_target_port.get_name() if fe_target_port else "")

            ws.cell(row=row, column=17, value=allocation_id)
            ws.cell(row=row, column=18, value=get_id(source_function) if source_function else "")
            ws.cell(row=row, column=19, value=get_id(target_function) if target_function else "")

            exported_allocations.add(allocation_id)
            return row + 1

    except Exception as e:
        print(f"Error exporting link exchange allocation: {e}")
        return row

def process_container(container, ws, row, exported_allocations):
    """Process a container (package or component) and optionally its children"""
    try:
        # Process exchanges owned by this container
        java_obj = container.get_java_object()
        exchanges = java_obj.getOwnedComponentExchanges()

        for ce in exchanges:
            ce_obj = ComponentExchange(ce)

            # Get allocations for this component exchange
            allocations = ce_obj.get_owned_component_exchange_functional_exchange_allocations()

            for allocation in allocations:
                row = export_link_exchange_allocation(allocation, ce_obj, ws, row, exported_allocations)

        # Process children if they exist
        children = java_obj.getOwnedLogicalComponents()
        if children:
            for child in children:
                child_component = create_component_from_java(child)
                if child_component:
                    row = process_container(child_component, ws, row, exported_allocations)

    except Exception as e:
        print(f"Error processing container: {e}")

    return row

def export_link_exchanges(wb):
    """Export all link exchanges to the Link Exchanges worksheet"""
    try:
        # Get model and setup from workbook properties
        model = wb._custom_doc_props["model"]
        la = wb._custom_doc_props["la"]

        if not la:
            print("Logical Architecture not found!")
            return False

        # Get worksheet
        ws_links = wb["Link Exchanges"]

        # Create headers
        create_link_exchanges_headers(ws_links)

        # Track exported allocations to avoid duplicates
        exported_allocations = set()

        # Export all link exchanges
        lc_pkg = la.get_logical_component_pkg()
        row = process_container(lc_pkg, ws_links, 2, exported_allocations)

        print(f"Exported the link between Functional and Component Exchanges")
        return True
    except Exception as e:
        print(f"Error exporting link exchanges: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def create_functional_chain_headers(ws, max_functions, max_exchanges):
    """Create headers for the Functional Chains worksheet"""
    # Basic chain info
    ws.cell(row=1, column=1, value="Functional Chain ID")
    ws.cell(row=1, column=2, value="Functional Chain Name")

    # Function headers
    function_start_col = 3
    for i in range(max_functions):
        ws.cell(row=1, column=function_start_col + i*3, value=f"Function {i+1} ID")
        ws.cell(row=1, column=function_start_col + i*3 + 1, value=f"Function {i+1} Name")
        ws.cell(row=1, column=function_start_col + i*3 + 2, value=f"Function {i+1} Involvement ID")

    # Exchange headers
    exchange_start_col = function_start_col + max_functions*3
    for i in range(max_exchanges):
        ws.cell(row=1, column=exchange_start_col + i*7, value=f"Exchange {i+1} ID")
        ws.cell(row=1, column=exchange_start_col + i*7 + 1, value=f"Exchange {i+1} Name")
        ws.cell(row=1, column=exchange_start_col + i*7 + 2, value=f"Exchange {i+1} Involvement ID")
        ws.cell(row=1, column=exchange_start_col + i*7 + 3, value=f"Exchange {i+1} Source Function ID")
        ws.cell(row=1, column=exchange_start_col + i*7 + 4, value=f"Exchange {i+1} Source Function Name")
        ws.cell(row=1, column=exchange_start_col + i*7 + 5, value=f"Exchange {i+1} Target Function ID")
        ws.cell(row=1, column=exchange_start_col + i*7 + 6, value=f"Exchange {i+1} Target Function Name")

def collect_logical_function_ids(model):
    """Collect all logical function IDs"""
    try:
        # Get logical architecture
        la = model.get_system_engineering().get_logical_architecture()
        lf_pkg = la.get_logical_function_pkg()

        # Find root logical function
        rlf = None
        for lf in lf_pkg.get_owned_logical_functions():
            if lf.get_name() == "Root Logical Function":
                rlf = lf
                break

        if not rlf:
            print("Root Logical Function not found!")
            return set()

        # Collect all function IDs
        logical_function_ids = set()

        def collect_ids(function):
            try:
                logical_function_ids.add(function.get_id())
                for child in function.get_owned_functions():
                    collect_ids(child)
            except Exception as e:
                print(f"Error collecting function IDs: {e}")

        collect_ids(rlf)
        return logical_function_ids

    except Exception as e:
        print(f"Error collecting logical function IDs: {e}")
        return set()

def export_functional_chains(wb):
    """Export all functional chains to the Functional Chains worksheet with source/target info for exchanges"""
    try:
        # Get model and setup from workbook properties
        model = wb._custom_doc_props["model"]
        se = wb._custom_doc_props["se"]

        # Get worksheet
        ws_chains = wb["Functional Chains"]

        # Collect all logical function IDs
        logical_function_ids = collect_logical_function_ids(model)

        # Find all functional chains in the model
        all_functional_chains = se.get_all_contents_by_type(FunctionalChain)

        # Filter to only include functional chains where ALL involved functions are in the logical architecture
        logical_functional_chains = []
        for fc in all_functional_chains:
            try:
                involved_functions = fc.get_involved_functions()
                if not involved_functions:
                    continue

                # Check if ALL involved functions are in the logical architecture
                all_in_logical = all(get_id(func) in logical_function_ids for func in involved_functions)
                if all_in_logical:
                    logical_functional_chains.append(fc)
            except Exception as e:
                print(f"Error checking functional chain {get_name(fc)}: {e}")
                continue

        # Determine the maximum number of functions and exchanges in any chain
        max_functions = 0
        max_exchanges = 0

        for fc in logical_functional_chains:
            try:
                involved_functions = fc.get_involved_functions()
                involved_exchanges = fc.get_involved_functional_exchanges()
                max_functions = max(max_functions, len(involved_functions))
                max_exchanges = max(max_exchanges, len(involved_exchanges))
            except Exception as e:
                print(f"Error determining max functions/exchanges for {get_name(fc)}: {e}")

        # Create headers (with source/target columns for exchanges)
        create_functional_chain_headers(ws_chains, max_functions, max_exchanges)

        # Track the next available row
        next_row = 2

        # Export functional chains with ALL involved functions and exchanges
        for fc in logical_functional_chains:
            try:
                # Basic chain info
                fc_id = get_id(fc)
                fc_name = get_name(fc)

                # Write basic chain info
                ws_chains.cell(row=next_row, column=1, value=fc_id)
                ws_chains.cell(row=next_row, column=2, value=fc_name)

                # Get ALL involvements and build exchange-to-source/target mapping
                all_involvements = {}
                exchange_to_source_target = {}  # {exchange_id: (src_id, src_name, tgt_id, tgt_name)}
                try:
                    involvements = fc.get_java_object().getOwnedFunctionalChainInvolvements()
                    for involvement in involvements:
                        # Map involvement IDs for functions/exchanges
                        if hasattr(involvement, 'getId'):
                            involvement_id = involvement.getId()
                            involved_element = involvement.getInvolvedElement()
                            if involved_element:
                                element_id = get_id(involved_element)
                                all_involvements[element_id] = involvement_id

                        # For FunctionalChainInvolvementLink, extract source/target
                        if involvement.eClass().getName() == "FunctionalChainInvolvementLink":
                            exchange = involvement.getInvolved()
                            if exchange:
                                exchange_id = get_id(exchange)
                                source_fcif = involvement.getSource()
                                target_fcif = involvement.getTarget()

                                # Get the actual functions from the involvements
                                source_func = source_fcif.getInvolvedElement() if source_fcif else None
                                target_func = target_fcif.getInvolvedElement() if target_fcif else None

                                if source_func and target_func:
                                    exchange_to_source_target[exchange_id] = (
                                        get_id(source_func),
                                        get_name(source_func),
                                        get_id(target_func),
                                        get_name(target_func)
                                    )
                except Exception as e:
                    print(f"Error building exchange-to-source/target mapping for {fc_name}: {e}")

                # Export ALL involved functions (no ordering)
                involved_functions = fc.get_involved_functions()
                for i, func in enumerate(involved_functions):
                    try:
                        func_id = get_id(func)
                        func_name = get_name(func)

                        ws_chains.cell(row=next_row, column=3 + i*3, value=func_id)
                        ws_chains.cell(row=next_row, column=3 + i*3 + 1, value=func_name)
                        ws_chains.cell(row=next_row, column=3 + i*3 + 2, value=all_involvements.get(func_id, ""))
                    except Exception as e:
                        print(f"Error processing function in {fc_name}: {e}")

                # Export ALL involved exchanges WITH SOURCE/TARGET INFO
                involved_exchanges = fc.get_involved_functional_exchanges()
                exchange_start_col = 3 + max_functions * 3
                for i, exchange in enumerate(involved_exchanges):
                    try:
                        exchange_id = get_id(exchange)
                        exchange_name = get_name(exchange)

                        # Get source/target info from the mapping (default to empty strings)
                        source_func_id, source_func_name, target_func_id, target_func_name = exchange_to_source_target.get(
                            exchange_id, ("", "", "", "")
                        )

                        # Write exchange info (7 columns per exchange)
                        ws_chains.cell(row=next_row, column=exchange_start_col + i*7, value=exchange_id)
                        ws_chains.cell(row=next_row, column=exchange_start_col + i*7 + 1, value=exchange_name)
                        ws_chains.cell(row=next_row, column=exchange_start_col + i*7 + 2, value=all_involvements.get(exchange_id, ""))

                        # Write source/target info
                        ws_chains.cell(row=next_row, column=exchange_start_col + i*7 + 3, value=source_func_id)
                        ws_chains.cell(row=next_row, column=exchange_start_col + i*7 + 4, value=source_func_name)
                        ws_chains.cell(row=next_row, column=exchange_start_col + i*7 + 5, value=target_func_id)
                        ws_chains.cell(row=next_row, column=exchange_start_col + i*7 + 6, value=target_func_name)

                    except Exception as e:
                        print(f"Error processing exchange in {fc_name}: {e}")

                next_row += 1

            except Exception as e:
                print(f"Error processing functional chain {get_name(fc)}: {e}")
                import traceback
                traceback.print_exc()
                continue

        print("Exported Functional Chains successfully")
        return True

    except Exception as e:
        print(f"Error exporting functional chains: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def create_capabilities_headers(ws, max_functional_chains, max_logical_functions, max_involved_components):
    """Create headers for the Capabilities worksheet"""
    # Basic capability info
    ws.cell(row=1, column=1, value="Capability ID")
    ws.cell(row=1, column=2, value="Capability Name")

    # Functional chains headers
    function_start_col = 3
    for i in range(max_functional_chains):
        ws.cell(row=1, column=function_start_col + i*3, value=f"Functional Chain {i+1} ID")
        ws.cell(row=1, column=function_start_col + i*3 + 1, value=f"Functional Chain {i+1} Name")
        ws.cell(row=1, column=function_start_col + i*3 + 2, value=f"Functional Chain {i+1} Involvement ID")

    # Logical functions headers
    function_end_col = function_start_col + max_functional_chains*3
    for i in range(max_logical_functions):
        ws.cell(row=1, column=function_end_col + i*3, value=f"Logical Function {i+1} ID")
        ws.cell(row=1, column=function_end_col + i*3 + 1, value=f"Logical Function {i+1} Name")
        ws.cell(row=1, column=function_end_col + i*3 + 2, value=f"Logical Function {i+1} Involvement ID")

    # Involved components headers
    component_start_col = function_end_col + max_logical_functions*3
    for i in range(max_involved_components):
        ws.cell(row=1, column=component_start_col + i*4, value=f"Involved Component {i+1} ID")
        ws.cell(row=1, column=component_start_col + i*4 + 1, value=f"Involved Component {i+1} Name")
        ws.cell(row=1, column=component_start_col + i*4 + 2, value=f"Involved Component {i+1} Type")
        ws.cell(row=1, column=component_start_col + i*4 + 3, value=f"Involved Component {i+1} Involvement ID")

def get_logical_functions(capability_realization):
    """Get logical functions directly involved with a capability realization"""
    functions = set()

    try:
        # Get the Java object of the capability realization
        java_cr = capability_realization.get_java_object()

        # Get all owned AbstractFunctionAbstractCapabilityInvolvements
        involvements = java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements()

        # Iterate through all involvements to find functions
        for involvement in involvements:
            try:
                # Get the involved element (function)
                involved_element = involvement.getInvolved()
                if involved_element and involved_element.eClass().getName() == "LogicalFunction":
                    functions.add(LogicalFunction(involved_element))
            except Exception as e:
                print(f"Error processing involvement: {e}")
                continue

    except Exception as e:
        print(f"Error getting logical functions: {e}")

    return list(functions)

def get_involved_components_with_ids(capability_realization):
    """Get involved components (actors and components) for a capability realization with involvement IDs"""
    involved_components = []

    try:
        # Get the Java object of the capability realization
        java_cr = capability_realization.get_java_object()

        # Try to use getInvolvedComponents() method
        try:
            # Get all involved components directly
            involved_elements = java_cr.getInvolvedComponents()

            # Get all owned involvements
            involvements = java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements()

            # Create a mapping of element ID to involvement ID
            involvement_map = {}
            for involvement in involvements:
                try:
                    involved_element = involvement.getInvolved()
                    if involved_element:
                        involvement_map[involved_element.getId()] = involvement.getId()
                except Exception as e:
                    print(f"Error processing involvement: {e}")
                    continue

            # Iterate through all involved components
            for element in involved_elements:
                try:
                    element_id = element.getId()
                    involvement_id = involvement_map.get(element_id, "")

                    # Create the appropriate wrapper object based on isActor()
                    if element.isActor():
                        involved_components.append(("actor", LogicalActor(element), involvement_id))
                    else:
                        involved_components.append(("component", LogicalComponent(element), involvement_id))
                except Exception as e:
                    print(f"Error processing involved component: {e}")
                    continue
        except Exception as e:
            print(f"Error using getInvolvedComponents(): {e}")

            # Fallback method: Check CapabilityRealizationInvolvement relationships
            try:
                # Get all owned involvements
                involvements = java_cr.getOwnedAbstractFunctionAbstractCapabilityInvolvements()

                # Iterate through all involvements to find components
                for involvement in involvements:
                    try:
                        involved_element = involvement.getInvolved()
                        if involved_element:
                            involvement_id = involvement.getId()

                            # Create the appropriate wrapper object based on isActor()
                            if involved_element.isActor():
                                involved_components.append(("actor", LogicalActor(involved_element), involvement_id))
                            else:
                                involved_components.append(("component", LogicalComponent(involved_element), involvement_id))
                    except Exception as e:
                        print(f"Error processing involvement: {e}")
                        continue
            except Exception as e:
                print(f"Error using fallback method: {e}")

    except Exception as e:
        print(f"Error getting involved components: {e}")

    return involved_components

def get_functional_chains(capability_realization, se):
    """Get functional chains owned or involved by a capability realization"""
    functional_chains = []

    try:
        # Get the Java object
        java_cr = capability_realization.get_java_object()

        # Get all functional chains in the model
        all_chains = se.get_all_contents_by_type(FunctionalChain)

        for chain in all_chains:
            try:
                # Check if owned by capability realization
                if chain.get_java_object().eContainer() == java_cr:
                    functional_chains.append(chain)
                # Check if capability realization is involved
                else:
                    for involvement in chain.get_java_object().getOwnedFunctionalChainInvolvements():
                        if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                            involved = involvement.getInvolvedElement()
                            if involved and involved.getId() == java_cr.getId():
                                functional_chains.append(chain)
                                break
            except Exception as e:
                print(f"Error checking chain: {e}")
                continue
    except Exception as e:
        print(f"Error getting functional chains: {e}")

    return functional_chains

def get_involvement_id(element, chain):
    """Get the involvement ID for an element in a chain"""
    try:
        for involvement in chain.get_java_object().getOwnedFunctionalChainInvolvements():
            if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                involved = involvement.getInvolvedElement()
                if involved and involved.getId() == element.get_id():
                    return involvement.getId()
            elif involvement.eClass().getName() == "FunctionalChainInvolvementLink":
                involved = involvement.getInvolvedElement()
                if involved and involved.getId() == element.get_id():
                    return involvement.getId()
    except Exception as e:
        print(f"Error getting involvement ID: {e}")
    return ""

def find_max_sizes(capability_realization_pkg, se):
    """Find the maximum number of functional chains, logical functions, and involved components"""
    max_functional_chains = 0
    max_logical_functions = 0
    max_involved_components = 0

    # Temporary analysis to determine maximum sizes
    for cr in capability_realization_pkg.get_owned_capability_realizations():
        try:
            # Count functional chains
            functional_chains = []
            for chain in se.get_all_contents_by_type(FunctionalChain):
                try:
                    # Check if owned by capability realization
                    if chain.get_java_object().eContainer() == cr.get_java_object():
                        functional_chains.append(chain)
                    # Check if capability realization is involved
                    else:
                        for involvement in chain.get_java_object().getOwnedFunctionalChainInvolvements():
                            if involvement.eClass().getName() == "FunctionalChainInvolvementFunction":
                                involved = involvement.getInvolvedElement()
                                if involved and involved.getId() == cr.get_id():
                                    functional_chains.append(chain)
                                    break
                except:
                    continue

            if len(functional_chains) > max_functional_chains:
                max_functional_chains = len(functional_chains)

            # Count logical functions
            functions = get_logical_functions(cr)
            if len(functions) > max_logical_functions:
                max_logical_functions = len(functions)

            # Count involved components
            involved_components = get_involved_components_with_ids(cr)
            if len(involved_components) > max_involved_components:
                max_involved_components = len(involved_components)

        except Exception as e:
            print(f"Error analyzing capability realization: {e}")
            continue

    return max_functional_chains, max_logical_functions, max_involved_components

def export_capabilities(wb):
    """Export all capabilities to the Capabilities worksheet"""
    try:
        # Get model and setup from workbook properties
        model = wb._custom_doc_props["model"]
        se = wb._custom_doc_props["se"]
        la = se.get_logical_architecture()

        # Get the capability realization package
        capability_realization_pkg = la.get_capability_realization_pkg()
        if not capability_realization_pkg:
            print("No Capability Realization Package found in Logical Architecture")
            return False

        # Find max sizes - pass se as parameter
        max_functional_chains, max_logical_functions, max_involved_components = find_max_sizes(capability_realization_pkg, se)

        # Get worksheet
        ws_capabilities = wb["Capabilities"]

        # Create headers
        create_capabilities_headers(ws_capabilities, max_functional_chains, max_logical_functions, max_involved_components)

        # Track the next available row
        next_row = 2

        # Export all capability realizations
        capabilities = capability_realization_pkg.get_owned_capability_realizations()

        for cr in capabilities:
            try:
                # Get basic info
                cr_id = cr.get_id()
                cr_name = cr.get_name()

                # Write basic capability realization info
                ws_capabilities.cell(row=next_row, column=1, value=cr_id)
                ws_capabilities.cell(row=next_row, column=2, value=cr_name)

                # Get and write functional chains - pass se as parameter
                functional_chains = get_functional_chains(cr, se)
                for i, chain in enumerate(functional_chains):
                    try:
                        chain_id = chain.get_id()
                        chain_name = chain.get_name()
                        involvement_id = get_involvement_id(cr, chain)

                        ws_capabilities.cell(row=next_row, column=3 + i*3, value=chain_id)
                        ws_capabilities.cell(row=next_row, column=3 + i*3 + 1, value=chain_name)
                        ws_capabilities.cell(row=next_row, column=3 + i*3 + 2, value=involvement_id)
                    except Exception as e:
                        print(f"Error writing functional chain {i+1}: {e}")

                # Get and write logical functions
                logical_functions = get_logical_functions(cr)
                for i, func in enumerate(logical_functions):
                    try:
                        func_id = func.get_id()  # Use get_id() directly on function
                        func_name = func.get_name()
                        involvement_id = ""

                        # Try to find involvement ID
                        for involvement in cr.get_java_object().getOwnedAbstractFunctionAbstractCapabilityInvolvements():
                            if involvement.getInvolved() and involvement.getInvolved().getId() == func.get_id():
                                involvement_id = involvement.getId()
                                break

                        ws_capabilities.cell(row=next_row, column=3 + max_functional_chains*3 + i*3, value=func_id)
                        ws_capabilities.cell(row=next_row, column=3 + max_functional_chains*3 + i*3 + 1, value=func_name)
                        ws_capabilities.cell(row=next_row, column=3 + max_functional_chains*3 + i*3 + 2, value=involvement_id)
                    except Exception as e:
                        print(f"Error writing logical function {i+1}: {e}")

                # Get and write involved components
                involved_components = get_involved_components_with_ids(cr)
                for i, (comp_type, comp, involvement_id) in enumerate(involved_components):
                    try:
                        comp_id = comp.get_id()  # Use get_id() directly on component
                        comp_name = comp.get_name()

                        ws_capabilities.cell(row=next_row, column=3 + max_functional_chains*3 + max_logical_functions*3 + i*4, value=comp_id)
                        ws_capabilities.cell(row=next_row, column=3 + max_functional_chains*3 + max_logical_functions*3 + i*4 + 1, value=comp_name)
                        ws_capabilities.cell(row=next_row, column=3 + max_functional_chains*3 + max_logical_functions*3 + i*4 + 2, value=comp_type)
                        ws_capabilities.cell(row=next_row, column=3 + max_functional_chains*3 + max_logical_functions*3 + i*4 + 3, value=involvement_id)
                    except Exception as e:
                        print(f"Error writing involved component {i+1}: {e}")

                next_row += 1

            except Exception as e:
                print(f"Error processing capability realization {cr.get_name()}: {e}")
                import traceback
                traceback.print_exc()
                continue
            
        print(f"Exported the Capabilities")
        
        return True
    except Exception as e:
        print(f"Error exporting capabilities: {str(e)}")
        return False
    
def create_requirements(ws):
    """Create the Requirements worksheet with headers"""
    # Set up headers
    ws["A1"] = "Object Type"
    ws["B1"] = 'Req UID'
    ws["C1"] = "Req 'is satisfied by' CapellaObjectID(s)"
    ws["D1"] = 'Parent Req UID(s)'
    ws["E1"] = '(Parent) Chapter UID'
    ws["F1"] = "Requirement Type"
    ws["G1"] = "Req Text"

    # Set column widths for better readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 50

    # Set header style
    for cell in ws[1:1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        

def generate_random_id(length=23):
    """Generate a random ReqIF identifier."""
    prefix = "_"
    characters = string.ascii_letters + string.digits
    random_part = ''.join(random.choices(characters, k=length))
    return prefix + random_part

def get_current_time():
    """Get current time in ReqIF format."""
    import pytz
    from datetime import datetime
    tz = pytz.timezone('Europe/Amsterdam')
    now = datetime.now(tz)
    formatted_time = now.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + now.strftime('%z')
    formatted_time = formatted_time[:-2] + ':' + formatted_time[-2:]
    return formatted_time

def set_ReqIF_identifier(ModelObject, model):
    """Set ReqIF identifier for model objects that don't have one."""
    try:
        ModelObject = ModelObject.get_java_object()
    except:
        pass
    if ModelObject.getReqIFIdentifier():
        return ModelObject.getReqIFIdentifier()
    else:
        ReqIFId = generate_random_id(length=23)
        try:
            model.start_transaction()
            ModelObject.setReqIFIdentifier(ReqIFId)
        except:
            print('Writing to model failed')
            model.rollback_transaction()
        else:
            print('set ReqIF.Id for model object:', ModelObject.getReqIFLongName())
            model.commit_transaction()
        model.save()
    return ReqIFId

XHTML_NS = "http://www.w3.org/1999/xhtml"

def collapse_xhtml_blocks(s, prefix="xhtml"):
    """Collapse XHTML blocks for pretty printing."""
    out = []
    pos = 0
    open_re = re.compile(r'<%s:([A-Za-z0-9_\-]+)\b' % re.escape(prefix))
    while True:
        m = open_re.search(s, pos)
        if not m:
            out.append(s[pos:])
            break
        start = m.start()
        tag = m.group(1)
        out.append(s[pos:start])
        pattern = re.compile(
            r'(</{p}:{t}\s*>)|(<{p}:{t}\b[^>]*?/>)|(<{p}:{t}\b[^>]*>)'.format(p=re.escape(prefix), t=re.escape(tag)),
            flags=re.S,
        )
        depth = 0
        found = False
        for mm in pattern.finditer(s, start):
            tok = mm.group(0)
            if tok.startswith('</'):
                depth -= 1
                if depth == 0:
                    end = mm.end()
                    block = s[start:end]
                    out.append(_collapse_block(block))
                    pos = end
                    found = True
                    break
            elif tok.endswith('/>'):
                if depth == 0:
                    end = mm.end()
                    block = s[start:end]
                    out.append(_collapse_block(block))
                    pos = end
                    found = True
                    break
            else:
                depth += 1
        if not found:
            out.append(s[start:])
            break
    return "".join(out)

def _collapse_block(block):
    """Collapse whitespace in a block."""
    b = re.sub(r'\r?\n', ' ', block)
    b = re.sub(r'\s+', ' ', b)
    b = re.sub(r'>\s+<', '><', b)
    b = re.sub(r'>\s+([^<\s])', r'>\1', b)
    b = re.sub(r'(\S)\s+<', r'\1<', b)
    return b.strip()

def wrap_in_div_p(text: str) -> str:
    """Wrap text in div/p tags."""
    return f'<div><p style="">{html.escape(text)}</p></div>'

class HTMLToXHTMLParser(HTMLParser):
    """Parser for converting HTML to XHTML."""
    def __init__(self):
        super().__init__()
        self.fragments = []
        self.stack = []

    def handle_starttag(self, tag, attrs):
        attrib = dict(attrs)
        if tag == "p" and "style" not in attrib:
            attrib["style"] = ""
        el = ET.Element(f"{{{XHTML_NS}}}{tag}", attrib)
        if self.stack:
            self.stack[-1].append(el)
        else:
            self.fragments.append(el)
        self.stack.append(el)

    def handle_endtag(self, tag):
        if self.stack:
            self.stack.pop()

    def handle_data(self, data):
        if self.stack:
            if self.stack[-1].text:
                self.stack[-1].text += data
            else:
                self.stack[-1].text = data

    def handle_startendtag(self, tag, attrs):
        attrib = dict(attrs)
        el = ET.Element(f"{{{XHTML_NS}}}{tag}", attrib)
        if self.stack:
            self.stack[-1].append(el)
        else:
            self.fragments.append(el)

def convert_html_to_xhtml(html_string):
    """Convert HTML string to XHTML."""
    parser = HTMLToXHTMLParser()
    parser.feed(html_string)
    wrapper = ET.Element(f"{{{XHTML_NS}}}div")
    for frag in parser.fragments:
        if frag.tag == f"{{{XHTML_NS}}}div":
            for child in list(frag):
                wrapper.append(child)
        else:
            wrapper.append(frag)
    return ET.tostring(wrapper, encoding="unicode")

def export_header(wb):
    """Export HEADER worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["HEADER"]

        headers = ["IDENTIFIER", "TITLE", "xmlns=", "version=", "encoding=", "comment", "req-if-tool-id", "source-tool-id", "xmlns:xhtml="]
        for item in se.get_all_contents_by_type(TypesFolder):
            header_id = set_ReqIF_identifier(item, model)
        header_title = se.get_name() + " Requirements Specification"
        values = [header_id, header_title, "http://www.omg.org/spec/ReqIF/20110401/reqif.xsd", "1.0", "UTF-8",
            "Created by: Starion Group", "ReqIF-Export for Capella (https://github.com/STARIONGROUP/...)",
            "Capella Version 7 (https://mbse-capella.org/)", "http://www.w3.org/1999/xhtml"]

        for col, (header, value) in enumerate(zip(headers, values), start=1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=2, column=col, value=value)

        return True
    except Exception as e:
        print(f"Error exporting HEADER: {e}")
        return False

def export_datatypes(wb):
    """Export DATATYPES worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["DATATYPES"]

        headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "DATATYPE-DEFINITION-KIND", "ALTERNATIVE-ID"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(DataTypeDefinition):
            ws[f"A{i}"] = set_ReqIF_identifier(item, model)
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = item.get_java_object().getReqIFLongName()
            ws[f"D{i}"] = 'DATATYPE-DEFINITION-XHTML' if "HTML" in item.get_java_object().getReqIFLongName() else 'DATATYPE-DEFINITION-STRING'
            ws[f"E{i}"] = item.get_java_object().getId()
            i += 1

        for item in se.get_all_contents_by_type(EnumerationDataTypeDefinition):
            ws[f"A{i}"] = set_ReqIF_identifier(item, model)
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = item.get_java_object().getReqIFLongName()
            ws[f"D{i}"] = 'DATATYPE-DEFINITION-ENUMERATION'
            ws[f"E{i}"] = item.get_java_object().getId()
            i += 1

        return True
    except Exception as e:
        print(f"Error exporting DATATYPES: {e}")
        return False

def export_enum_values(wb):
    """Export ENUM-VALUES worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["ENUM-VALUES"]

        headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "KEY", "DATATYPE-DEFINITION-ENUMERATION-REF", "ALTERNATIVE-ID"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(EnumerationDataTypeDefinition):
            for value in item.get_java_object().getSpecifiedValues():
                ws[f"A{i}"] = set_ReqIF_identifier(value, model)
                ws[f"B{i}"] = get_current_time()
                ws[f"C{i}"] = value.getReqIFLongName()
                ws[f"D{i}"] = str(i-1)
                ws[f"E{i}"] = set_ReqIF_identifier(item, model)
                ws[f"F{i}"] = value.getId()
                i += 1

        return True
    except Exception as e:
        print(f"Error exporting ENUM-VALUES: {e}")
        return False

def export_spec_types(wb):
    """Export SPEC-TYPES worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["SPEC-TYPES"]

        headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "KIND", "ALTERNATIVE-ID"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(RequirementType):
            ws[f"A{i}"] = set_ReqIF_identifier(item, model)
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = item.get_java_object().getReqIFLongName()
            ws[f"D{i}"] = 'SPEC-OBJECT-TYPE'
            ws[f"E{i}"] = item.get_java_object().getId()
            i += 1

        for item in se.get_all_contents_by_type(ModuleType):
            ws[f"A{i}"] = set_ReqIF_identifier(item, model)
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = item.get_java_object().getReqIFLongName()
            ws[f"D{i}"] = 'SPECIFICATION-TYPE'
            ws[f"E{i}"] = item.get_java_object().getId()
            i += 1

        for item in se.get_all_contents_by_type(RelationType):
            ws[f"A{i}"] = set_ReqIF_identifier(item, model)
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = item.get_java_object().getReqIFLongName()
            ws[f"D{i}"] = 'SPEC-RELATION-TYPE'
            ws[f"E{i}"] = item.get_java_object().getId()
            i += 1

        return True
    except Exception as e:
        print(f"Error exporting SPEC-TYPES: {e}")
        return False

def export_spec_attributes(wb):
    """Export SPEC-ATTRIBUTES worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["SPEC-ATTRIBUTES"]

        headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "ATTRIBUTE-DEF-KIND", "SPEC-TYPE-REF", "DATATYPE-DEFINITION-STRING-REF", "ALTERNATIVE-ID"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(AttributeDefinition):
            ws[f"A{i}"] = set_ReqIF_identifier(item, model)
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = item.get_java_object().getReqIFLongName()

            def_type = item.get_java_object().getDefinitionType()
            if def_type and def_type.eClass().getName() == "DataTypeDefinition":
                if "HTML" in def_type.getReqIFLongName() or item.get_java_object().getReqIFLongName() == "ReqIF.Text":
                    ws[f"D{i}"] = 'ATTRIBUTE-DEFINITION-XHTML'
                else:
                    ws[f"D{i}"] = 'ATTRIBUTE-DEFINITION-STRING'
            elif def_type and def_type.eClass().getName() == "EnumerationDataTypeDefinition":
                ws[f"D{i}"] = 'ATTRIBUTE-DEFINITION-ENUMERATION'

            ws[f"E{i}"] = item.get_java_object().eContainer().getReqIFIdentifier()
            if item.get_java_object().getReqIFLongName() == "ReqIF.Text":
                for type in se.get_all_contents_by_type(DataTypeDefinition):
                    if type.get_java_object().getReqIFLongName() == "DataTypeDefinitionXHTML":
                        ws[f"F{i}"] = type.get_java_object().getReqIFIdentifier()
            else:
                ws[f"F{i}"] = def_type.getReqIFIdentifier() if def_type else ""
            ws[f"G{i}"] = item.get_java_object().getId()
            i += 1

        return True
    except Exception as e:
        print(f"Error exporting SPEC-ATTRIBUTES: {e}")
        return False

def export_spec_objects(wb):
    """Export SPEC-OBJECTS worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["SPEC-OBJECTS"]

        headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "SPEC-TYPE-REF", "ALTERNATIVE-ID", "Capella BridgeTraces Type"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(Requirement):
            ws[f"A{i}"] = set_ReqIF_identifier(item, model)
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = item.get_java_object().getReqIFLongName()
            ws[f"D{i}"] = item.get_java_object().getRequirementType().getReqIFIdentifier()

            # Set RequirementTypeProxy if needed
            if item.get_java_object().getRequirementTypeProxy() != item.get_java_object().getRequirementType().getReqIFIdentifier():
                try:
                    model.start_transaction()
                    item.get_java_object().setRequirementTypeProxy(item.get_java_object().getRequirementType().getReqIFIdentifier())
                except:
                    print('Writing to model failed')
                    model.rollback_transaction()
                else:
                    print('set RequirementTypeProxy for model object:', item.get_java_object().getReqIFLongName())
                    model.commit_transaction()
                model.save()

            ws[f"E{i}"] = item.get_java_object().getId()
            ws[f"F{i}"] = item.get_java_object().eClass().getName()
            i += 1

        return True
    except Exception as e:
        print(f"Error exporting SPEC-OBJECTS: {e}")
        return False

def export_attribute_values(wb):
    """Export ATTRIBUTE-VALUES worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["ATTRIBUTE-VALUES"]

        headers = ["VALUE / ENUM-VALUE-REF", "ATTRIBUTE-VALUE-KIND", "ATTRIBUTE-DEFINITION-REF", "SPEC-OBJECT-REF", "ALTERNATIVE-ID"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(Requirement):
            attr_list = item.get_java_object().getRequirementType().getOwnedAttributes()
            for attribute in item.get_java_object().getOwnedAttributes():
                try:
                    if attribute.eClass().getName() == "EnumerationValueAttribute":
                        ws[f"A{i}"] = attribute.getValues()[0].getReqIFIdentifier()
                        ws[f"B{i}"] = "ATTRIBUTE-VALUE-ENUMERATION"
                    elif attribute.eClass().getName() == "StringValueAttribute":
                        ws[f"A{i}"] = attribute.getValue()
                        ws[f"B{i}"] = "ATTRIBUTE-VALUE-STRING"
                    ws[f"C{i}"] = attribute.getDefinition().getReqIFIdentifier()
                    ws[f"D{i}"] = attribute.eContainer().getReqIFIdentifier()
                    ws[f"E{i}"] = attribute.getId()
                    i += 1
                except:
                    print(f"Attribute Skipped: attribute value or definition missing for SPEC-OBJECT: {get_name(item)}")

            # Export ReqIF.Name, ReqIF.ChapterName, ReqIF.Prefix, ReqIF.Text
            if item.get_java_object().getReqIFName():
                ws[f"A{i}"] = item.get_java_object().getReqIFName()
                ws[f"B{i}"] = "ATTRIBUTE-VALUE-STRING"
                for attr in attr_list:
                    if attr.getReqIFLongName() == "ReqIF.Name":
                        ws[f"C{i}"] = attr.getReqIFIdentifier()
                ws[f"D{i}"] = set_ReqIF_identifier(item, model)
                i += 1

            if item.get_java_object().getReqIFChapterName():
                ws[f"A{i}"] = item.get_java_object().getReqIFChapterName()
                ws[f"B{i}"] = "ATTRIBUTE-VALUE-STRING"
                for attr in attr_list:
                    if attr.getReqIFLongName() == "ReqIF.ChapterName":
                        ws[f"C{i}"] = attr.getReqIFIdentifier()
                ws[f"D{i}"] = set_ReqIF_identifier(item, model)
                i += 1

            if item.get_java_object().getReqIFPrefix():
                ws[f"A{i}"] = item.get_java_object().getReqIFPrefix()
                ws[f"B{i}"] = "ATTRIBUTE-VALUE-STRING"
                for attr in attr_list:
                    if attr.getReqIFLongName() == "ReqIF.Prefix":
                        ws[f"C{i}"] = attr.getReqIFIdentifier()
                ws[f"D{i}"] = set_ReqIF_identifier(item, model)
                i += 1

            if item.get_java_object().getReqIFText():
                ws[f"A{i}"] = item.get_java_object().getReqIFText()
                ws[f"B{i}"] = "ATTRIBUTE-VALUE-XHTML"
                for attr in attr_list:
                    if attr.getReqIFLongName() == "ReqIF.Text":
                        ws[f"C{i}"] = attr.getReqIFIdentifier()
                ws[f"D{i}"] = set_ReqIF_identifier(item, model)
                i += 1

        return True
    except Exception as e:
        print(f"Error exporting ATTRIBUTE-VALUES: {e}")
        return False

def export_spec_relations(wb):
    """Export SPEC-RELATIONS worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["SPEC-RELATIONS"]

        headers = ["IDENTIFIER", "LAST-CHANGE", "SOURCE-SPEC-OBJECT-REF", "TARGET-SPEC-OBJECT-REF", "SPEC-RELATION-TYPE-REF", "ALTERNATIVE-ID"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(AbstractRelation):
            try:
                ws[f"A{i}"] = set_ReqIF_identifier(item, model)
                ws[f"B{i}"] = get_current_time()
                ws[f"C{i}"] = item.get_java_object().getSource().getReqIFIdentifier()
                ws[f"D{i}"] = item.get_java_object().getTarget().getReqIFIdentifier()
                ws[f"E{i}"] = item.get_java_object().getRelationType().getReqIFIdentifier()
                ws[f"F{i}"] = item.get_java_object().getId()
                i += 1
            except:
                pass

        return True
    except Exception as e:
        print(f"Error exporting SPEC-RELATIONS: {e}")
        return False

def export_specifications(wb):
    """Export SPECIFICATIONS worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["SPECIFICATIONS"]

        headers = ["IDENTIFIER", "LAST-CHANGE", "LONG-NAME", "SPEC-OBJECT-TYPE-REF", "ALTERNATIVE-ID", "Capella BridgeTraces Type"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(CapellaModule):
            ws[f"A{i}"] = set_ReqIF_identifier(item, model)
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = item.get_java_object().getReqIFLongName()
            ws[f"D{i}"] = item.get_java_object().getModuleType().getReqIFIdentifier()
            ws[f"E{i}"] = item.get_java_object().getId()
            ws[f"F{i}"] = "Module"
            i += 1

        return True
    except Exception as e:
        print(f"Error exporting SPECIFICATIONS: {e}")
        return False

def export_spec_hierarchy(wb):
    """Export SPEC-HIERARCHY worksheet for ReqIF"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]
        ws = wb["SPEC-HIERARCHY"]

        headers = ["IDENTIFIER", "LAST-CHANGE", "SPEC-OBJECT-REF", "PARENT-SPEC-OBJECT-REF"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        i = 2
        for item in se.get_all_contents_by_type(Requirement):
            ws[f"A{i}"] = generate_random_id()
            ws[f"B{i}"] = get_current_time()
            ws[f"C{i}"] = set_ReqIF_identifier(item, model)
            ws[f"D{i}"] = item.get_java_object().eContainer().getReqIFIdentifier()
            i += 1

        return True
    except Exception as e:
        print(f"Error exporting SPEC-HIERARCHY: {e}")
        return False
    
def has_requirements(se):
    """Check if the model has requirements (TypesFolder is a prerequisite)"""
    return len(se.get_all_contents_by_type(TypesFolder)) > 0

def export_requirements(wb):
    """Export all ReqIF worksheets for requirements, or skip if no requirements exist"""
    try:
        se = wb._custom_doc_props["se"]
        model = wb._custom_doc_props["model"]

        # Check if model has requirements (TypesFolder is prerequisite)
        has_reqs = has_requirements(se)
        wb._custom_doc_props["has_requirements"] = has_reqs  # Store flag for later

        if not has_reqs:
            # Delete ALL requirement-related worksheets
            req_worksheets = [
                "HEADER", "DATATYPES", "ENUM-VALUES", "SPEC-TYPES",
                "SPEC-ATTRIBUTES", "SPEC-OBJECTS", "ATTRIBUTE-VALUES",
                "SPEC-RELATIONS", "SPECIFICATIONS", "SPEC-HIERARCHY"
            ]
            for ws_name in req_worksheets:
                if ws_name in wb.sheetnames:
                    del wb[ws_name]
            return True  # Not a failure - just no requirements

        # Model has requirements - proceed with normal export
        success = True
        success &= export_header(wb)
        success &= export_datatypes(wb)
        success &= export_enum_values(wb)
        success &= export_spec_types(wb)
        success &= export_spec_attributes(wb)
        success &= export_spec_objects(wb)
        success &= export_attribute_values(wb)
        success &= export_spec_relations(wb)
        success &= export_specifications(wb)
        success &= export_spec_hierarchy(wb)
        
        print("Exported the Requirements")

        return success

    except Exception as e:
        print(f"Error in export_requirements: {e}")
        import traceback
        traceback.print_exc()
        return False        

def full_export(aird_path):
    """
    Perform a full export of all model elements to a single Excel workbook
    with multiple worksheets.
    """
    try:
        # Setup workbook and get model information
        wb = setup_workbook(aird_path)
        if not wb:
            print("Failed to setup workbook")
            return False

        print("Starting the export")

        # Export all elements
        success = True
        success &= export_systems(wb)
        success &= export_functions(wb)
        success &= export_links(wb)
        success &= export_functional_exchange_sheet(wb)
        success &= export_component_exchanges(wb)
        success &= export_link_exchanges(wb)
        success &= export_functional_chains(wb)
        success &= export_capabilities(wb)
        success &= export_requirements(wb)

        if not success:
            print("Some exports failed. Check error messages above.")
            return False      

        # Save and refresh with better error handling
        try:
            se = wb._custom_doc_props["se"]
            folder = wb._custom_doc_props["folder"]
            project_name = wb._custom_doc_props["project_name"]

            xlsx_file_name = CapellaPlatform.getAbsolutePath(folder) + "/" + project_name + "_complete_export.xlsx"

            # Check if file exists and is locked
            if os.path.exists(xlsx_file_name):
                try:
                    # Try to open the file in exclusive mode to check if it's locked
                    with open(xlsx_file_name, 'r+b') as f:
                        pass
                except IOError:
                    print(f"Error: The file '{xlsx_file_name}' is already open in another application.")
                    print("Please close the file and try again.")
                    return False

            print("Writing " + xlsx_file_name)
            wb.save(filename=xlsx_file_name)
            print("File saved successfully.")
            CapellaPlatform.refresh(wb._custom_doc_props["folder"])
            
            has_reqs = wb._custom_doc_props.get("has_requirements", True)
            
            if not has_reqs:
                print("Model successfully exported, model did not contain any requirements")
            else:
                print("Full export completed successfully!")
            return True

        except PermissionError:
            print(f"Permission denied: Could not save file to {xlsx_file_name}.")
            print("The file is likely open in another application. Please close it and try again.")
            return False
        except Exception as e:
            print(f"An error occurred while saving: {e}")
            return False

    except Exception as e:
        print(f"Error in full export: {str(e)}")
        import traceback
        traceback.print_exc()
        return False