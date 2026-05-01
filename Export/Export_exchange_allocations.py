#!/usr/bin/env python3
"""
export_exchange_allocations.py

Exports exchange allocations from the generated SysML files to an Excel file.
Reads Parts_generated.sysml for component exchange connection defs and interface
defs (with 'flow of' lines), and Functions_generated.sysml for functional
exchange connection defs, action defs, and item defs.

Output columns match DVS_Exchange_Allocations.xlsx:
  Component Exchange ID, Component Exchange Name,
  Component Source Port ID, Component Source Port Name,
  Source Component ID, Source Component Name,
  Component Target Port ID, Component Target Port Name,
  Target Component ID, Target Component Name,
  Functional Exchange ID, Functional Exchange Name,
  Functional Source Port ID, Functional Source Port Name,
  Functional Target Port ID, Functional Target Port Name,
  Allocation ID (empty - not stored in SysML),
  Source Function ID, Target Function ID

Usage:
  python export_exchange_allocations.py [parts.sysml] [functions.sysml] [output.xlsx]
"""

import pathlib
import re
import sys
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import PatternFill


def from_pascal_case(name: str) -> str:
    """Convert PascalCase/ID-suffixed name to space-separated words."""
    if not name:
        return name

    name_without_id = re.sub(r'_[0-9a-f]{4}$', '', name)

    if name_without_id.startswith('CODE'):
        rest_part = name_without_id[4:]
        processed_rest = []
        for i, char in enumerate(rest_part):
            if char.isupper() and i > 0:
                processed_rest.append(' ')
            processed_rest.append(char)
        return f"CODE {''.join(processed_rest)}".strip()

    result = [name_without_id[0]]
    i = 1
    n = len(name_without_id)
    while i < n:
        char = name_without_id[i]
        prev_char = name_without_id[i - 1]
        if char.isupper():
            if prev_char.islower():
                result.append(' ')
                result.append(char)
                i += 1
            else:
                j = i
                while j < n and name_without_id[j].isupper():
                    j += 1
                if j < n and name_without_id[j].islower():
                    result.append(name_without_id[i:j - 1])
                    result.append(' ')
                    result.append(name_without_id[j - 1])
                    i = j
                else:
                    result.append(char)
                    i += 1
        else:
            result.append(char)
            i += 1
    return ''.join(result)


def format_port_name(port_usage: str) -> str:
    """Convert port usage like CP1_9b89 or FOP1_04b2 to CP 1 or FOP 1."""
    base = port_usage.split('_')[0]
    base = re.sub(r'([a-zA-Z])([0-9])', r'\1 \2', base)
    base = re.sub(r'([0-9])([a-zA-Z])', r'\1 \2', base)
    return base


def format_exchange_name(ce_sysml_name: str) -> str:
    """Convert CE SysML name like Groundstation_Antenna_UHF to Groundstation - Antenna - UHF."""
    return ce_sysml_name.replace('_', ' - ')


def extract_id(text: str) -> str:
    """Extract UUID from a /* ID: uuid */ doc comment."""
    m = re.search(r'/\*\s*ID:\s*([0-9a-f-]+)\s*\*/', text)
    return m.group(1) if m else ""


def get_block(text: str, start_pos: int) -> Tuple[str, int]:
    """Return (block_content, end_pos) for the { } block starting at start_pos."""
    depth = 1
    pos = start_pos
    while pos < len(text) and depth > 0:
        if text[pos] == '{':
            depth += 1
        elif text[pos] == '}':
            depth -= 1
        pos += 1
    return (text[start_pos:pos - 1] if depth == 0 else ""), pos


def parse_parts_file(file_path: pathlib.Path):
    """Parse Parts_generated.sysml.

    Returns:
      part_defs          — {type_name: {'id': str, 'ports': {usage: {'id': str}}}}
      ce_connections     — {ce_name: {'id', 'from_part_type', 'to_part_type',
                                       'from_port_usage', 'to_port_usage'}}
      interface_defs     — {iface_name: {'flows': [fe_type, ...]}}
    """
    content = file_path.read_text(encoding='utf-8')

    part_defs: Dict = {}
    for m in re.finditer(r'part def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        ports = {}
        for pm in re.finditer(r'port (\w+)\s*:\s*\w+\s*\{', body):
            port_name = pm.group(1)
            port_body, _ = get_block(body, pm.end())
            ports[port_name] = {'id': extract_id(port_body)}
        part_defs[name] = {'id': extract_id(body), 'ports': ports}

    ce_connections: Dict = {}
    for m in re.finditer(r'connection def (\w+)\s*\{', content):
        ce_name = m.group(1)
        body, _ = get_block(content, m.end())
        ce_id = extract_id(body)
        end_parts = re.findall(r'end part (\w+)\s*:\s*(\w+);', body)
        if len(end_parts) < 2:
            continue
        iface_m = re.search(
            r'interface\s*:\s*\w+\s+connect\s+(\w+)\.(\w+)\s+to\s+(\w+)\.(\w+)',
            body,
        )
        if not iface_m:
            continue
        from_part_usage, from_port_usage, to_part_usage, to_port_usage = iface_m.groups()
        usage_to_type = {u: t for u, t in end_parts}
        ce_connections[ce_name] = {
            'id': ce_id,
            'from_part_type': usage_to_type.get(from_part_usage, ''),
            'to_part_type': usage_to_type.get(to_part_usage, ''),
            'from_port_usage': from_port_usage,
            'to_port_usage': to_port_usage,
        }

    interface_defs: Dict = {}
    for m in re.finditer(r'interface def (\w+)\s*\{', content):
        iface_name = m.group(1)
        body, _ = get_block(content, m.end())
        flows = re.findall(
            r'flow of (\w+)\s+from\s+\w+\.\w+\s+to\s+\w+\.\w+\s*;', body
        )
        interface_defs[iface_name] = {'flows': flows}

    return part_defs, ce_connections, interface_defs


def parse_functions_file(file_path: pathlib.Path):
    """Parse Functions_generated.sysml.

    Returns:
      action_defs          — {type_name: {'id': str, 'ports': {usage: {'id': str}}}}
      item_defs            — {item_name: {'id': str}}
      fe_conns_by_id       — {fe_id: {'from_action_type', 'to_action_type',
                                       'from_port_usage', 'to_port_usage'}}
    """
    content = file_path.read_text(encoding='utf-8')

    action_defs: Dict = {}
    for m in re.finditer(r'action def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        ports = {}
        for pm in re.finditer(r'port (\w+)\s*:\s*\w+\s*\{', body):
            port_name = pm.group(1)
            port_body, _ = get_block(body, pm.end())
            ports[port_name] = {'id': extract_id(port_body)}
        action_defs[name] = {'id': extract_id(body), 'ports': ports}

    item_defs: Dict = {}
    for m in re.finditer(r'item def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        item_defs[name] = {'id': extract_id(body)}

    fe_conns_by_id: Dict = {}
    for m in re.finditer(r'connection def (\w+)\s*\{', content):
        body, _ = get_block(content, m.end())
        conn_id = extract_id(body)
        if not conn_id:
            continue
        end_actions = re.findall(r'end action (\w+)\s*:\s*(\w+);', body)
        if len(end_actions) < 2:
            continue
        iface_m = re.search(
            r'interface\s*:\s*\w+\s+connect\s+(\w+)\.(\w+)\s+to\s+(\w+)\.(\w+)',
            body,
        )
        if not iface_m:
            continue
        # FE connect order is reversed: TARGET.to_port to SOURCE.from_port
        to_action_usage, to_port_usage, from_action_usage, from_port_usage = iface_m.groups()
        usage_to_type = {u: t for u, t in end_actions}
        fe_conns_by_id[conn_id] = {
            'from_action_type': usage_to_type.get(from_action_usage, ''),
            'to_action_type': usage_to_type.get(to_action_usage, ''),
            'from_port_usage': from_port_usage,
            'to_port_usage': to_port_usage,
        }

    return action_defs, item_defs, fe_conns_by_id


def build_allocations(
    part_defs, ce_connections, interface_defs,
    action_defs, item_defs, fe_conns_by_id,
) -> List[Dict]:
    """Cross-reference parsed data and produce one dict per exchange allocation row."""
    rows = []

    for iface_name, iface in interface_defs.items():
        if not iface['flows']:
            continue
        if not iface_name.endswith('_Interface'):
            continue

        ce_name = iface_name[: -len('_Interface')]
        ce = ce_connections.get(ce_name)
        if not ce:
            continue

        ce_id = ce['id']
        ce_display_name = format_exchange_name(ce_name)

        from_part = part_defs.get(ce['from_part_type'], {})
        to_part = part_defs.get(ce['to_part_type'], {})

        src_component_id = from_part.get('id', '')
        src_component_name = from_pascal_case(ce['from_part_type'])
        src_port_id = from_part.get('ports', {}).get(ce['from_port_usage'], {}).get('id', '')
        src_port_name = format_port_name(ce['from_port_usage'])

        tgt_component_id = to_part.get('id', '')
        tgt_component_name = from_pascal_case(ce['to_part_type'])
        tgt_port_id = to_part.get('ports', {}).get(ce['to_port_usage'], {}).get('id', '')
        tgt_port_name = format_port_name(ce['to_port_usage'])

        for fe_type in iface['flows']:
            fe_id = item_defs.get(fe_type, {}).get('id', '')
            fe_name = from_pascal_case(fe_type)

            fe_conn = fe_conns_by_id.get(fe_id, {})
            from_action_type = fe_conn.get('from_action_type', '')
            to_action_type = fe_conn.get('to_action_type', '')
            fe_from_port_usage = fe_conn.get('from_port_usage', '')
            fe_to_port_usage = fe_conn.get('to_port_usage', '')

            from_action = action_defs.get(from_action_type, {})
            to_action = action_defs.get(to_action_type, {})

            src_func_id = from_action.get('id', '')
            tgt_func_id = to_action.get('id', '')
            fe_src_port_id = from_action.get('ports', {}).get(fe_from_port_usage, {}).get('id', '')
            fe_src_port_name = format_port_name(fe_from_port_usage) if fe_from_port_usage else ''
            fe_tgt_port_id = to_action.get('ports', {}).get(fe_to_port_usage, {}).get('id', '')
            fe_tgt_port_name = format_port_name(fe_to_port_usage) if fe_to_port_usage else ''

            rows.append({
                'ce_id': ce_id,
                'ce_name': ce_display_name,
                'src_port_id': src_port_id,
                'src_port_name': src_port_name,
                'src_component_id': src_component_id,
                'src_component_name': src_component_name,
                'tgt_port_id': tgt_port_id,
                'tgt_port_name': tgt_port_name,
                'tgt_component_id': tgt_component_id,
                'tgt_component_name': tgt_component_name,
                'fe_id': fe_id,
                'fe_name': fe_name,
                'fe_src_port_id': fe_src_port_id,
                'fe_src_port_name': fe_src_port_name,
                'fe_tgt_port_id': fe_tgt_port_id,
                'fe_tgt_port_name': fe_tgt_port_name,
                'src_func_id': src_func_id,
                'tgt_func_id': tgt_func_id,
            })

    return rows


def export_to_excel(rows: List[Dict], output_path: pathlib.Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    if wb.active.title == "Sheet":
        wb.remove(wb.active)
    ws = wb.create_sheet("Exchange Allocations")

    headers = [
        "Component Exchange ID", "Component Exchange Name",
        "Component Source Port ID", "Component Source Port Name",
        "Source Component ID", "Source Component Name",
        "Component Target Port ID", "Component Target Port Name",
        "Target Component ID", "Target Component Name",
        "Functional Exchange ID", "Functional Exchange Name",
        "Functional Source Port ID", "Functional Source Port Name",
        "Functional Target Port ID", "Functional Target Port Name",
        "Allocation ID",
        "Source Function ID", "Target Function ID",
    ]

    gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = gray

    for row in rows:
        r = ws.max_row + 1
        ws.cell(r, 1, row['ce_id'])
        ws.cell(r, 2, row['ce_name'])
        ws.cell(r, 3, row['src_port_id'])
        ws.cell(r, 4, row['src_port_name'])
        ws.cell(r, 5, row['src_component_id'])
        ws.cell(r, 6, row['src_component_name'])
        ws.cell(r, 7, row['tgt_port_id'])
        ws.cell(r, 8, row['tgt_port_name'])
        ws.cell(r, 9, row['tgt_component_id'])
        ws.cell(r, 10, row['tgt_component_name'])
        ws.cell(r, 11, row['fe_id'])
        ws.cell(r, 12, row['fe_name'])
        ws.cell(r, 13, row['fe_src_port_id'])
        ws.cell(r, 14, row['fe_src_port_name'])
        ws.cell(r, 15, row['fe_tgt_port_id'])
        ws.cell(r, 16, row['fe_tgt_port_name'])
        ws.cell(r, 17, "")  # Allocation ID not stored in SysML
        ws.cell(r, 18, row['src_func_id'])
        ws.cell(r, 19, row['tgt_func_id'])

    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=0,
        )
        ws.column_dimensions[col[0].column_letter].width = (max_len + 2) * 1.2

    wb.save(output_path)
    print(f"Exported {len(rows)} exchange allocations to: {output_path}")


def main() -> None:
    _DVS_DIR = pathlib.Path(__file__).parent.parent
    DEFAULT_PARTS = _DVS_DIR / "Parts_generated.sysml"
    DEFAULT_FUNCTIONS = _DVS_DIR / "Functions_generated.sysml"
    DEFAULT_OUTPUT = _DVS_DIR / "results" / "DVS_Exchange_Allocations_export.xlsx"

    args = sys.argv[1:]
    if args and args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    parts_path = pathlib.Path(args[0]) if len(args) >= 1 else DEFAULT_PARTS
    functions_path = pathlib.Path(args[1]) if len(args) >= 2 else DEFAULT_FUNCTIONS
    output_path = pathlib.Path(args[2]) if len(args) >= 3 else DEFAULT_OUTPUT

    for path in (parts_path, functions_path):
        if not path.exists():
            print(f"Error: file not found: {path}", file=sys.stderr)
            sys.exit(1)

    part_defs, ce_connections, interface_defs = parse_parts_file(parts_path)
    action_defs, item_defs, fe_conns_by_id = parse_functions_file(functions_path)

    rows = build_allocations(
        part_defs, ce_connections, interface_defs,
        action_defs, item_defs, fe_conns_by_id,
    )

    if not rows:
        print("Warning: no exchange allocations found.", file=sys.stderr)

    export_to_excel(rows, output_path)


if __name__ == "__main__":
    main()
