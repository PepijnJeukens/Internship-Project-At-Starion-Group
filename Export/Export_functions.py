#!/usr/bin/env python3
"""
export_functions_to_excel.py
Exports functions from a SysML v2 file to an Excel file with hierarchical structure.
Uses text parsing to avoid syside import issues.
Each function is exported only once at its correct hierarchy level.
Now removes ID suffix and properly handles spacing for all function names.
"""

import pathlib
import re
import sys
from typing import Dict, List

import openpyxl
from openpyxl.styles import PatternFill

def from_pascal_case(name: str) -> str:
    """
    Convert PascalCase to space-separated words with proper handling of acronyms.
    Rules:
    1. First remove ID suffix (underscore and 4 hex chars) if present
    2. For CODE-prefixed names: split each capital letter with spaces
    3. For all other names: add space before capital letters that follow lowercase letters
    """
    if not name:
        return name

    # Remove ID suffix if present (underscore followed by 4 hex characters)
    name_without_id = re.sub(r'_[0-9a-f]{4}$', '', name)

    # Special handling for CODE prefix
    if name_without_id.startswith('CODE'):
        # Split CODE and the rest
        code_part = 'CODE'
        rest_part = name_without_id[4:]

        # Process the rest part by adding space before each uppercase letter
        processed_rest = []
        for i, char in enumerate(rest_part):
            if char.isupper() and i > 0:
                processed_rest.append(' ')
            processed_rest.append(char)
        return f"{code_part} {''.join(processed_rest)}".strip()

    # General case for all other names
    result = [name_without_id[0]]

    for i in range(1, len(name_without_id)):
        char = name_without_id[i]
        prev_char = name_without_id[i-1]

        # Add space before uppercase letters that follow lowercase letters
        if char.isupper() and prev_char.islower():
            result.append(' ')
        result.append(char)

    return ''.join(result)

def extract_id_from_text(text: str) -> str:
    """Extract ID from text using the pattern: /* ID: uuid */"""
    match = re.search(r"/\*\s*ID:\s*([0-9a-f-]+)\s*\*/", text)
    return match.group(1) if match else ""

def get_block_content(text: str, start_pos: int) -> tuple:
    """Get content of a block starting with {, handling nested braces."""
    brace_count = 1
    end_pos = start_pos
    while end_pos < len(text) and brace_count > 0:
        if text[end_pos] == '{':
            brace_count += 1
        elif text[end_pos] == '}':
            brace_count -= 1
        end_pos += 1
    return text[start_pos:end_pos-1] if brace_count == 0 else "", end_pos

def parse_functions_file(file_path: pathlib.Path) -> List[Dict]:
    """
    Parse the SysML file to extract function hierarchy.
    Returns a list of dictionaries with function information and hierarchy level.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the FunctionsGenerated package content
    functions_package_match = re.search(r'package FunctionsGenerated\s*\{', content)
    if not functions_package_match:
        print("Error: FunctionsGenerated package not found in file", file=sys.stderr)
        return []

    start_pos = functions_package_match.end()
    functions_content, end_pos = get_block_content(content, start_pos)

    # First, find all action definitions
    action_defs = {}
    for action_def_match in re.finditer(r'action def (\w+)\s*\{', functions_content):
        action_name = action_def_match.group(1)
        start = action_def_match.end()
        content_block, end = get_block_content(functions_content, start)
        action_id = extract_id_from_text(content_block)

        # Find all action usages in this action
        usages = []
        for usage_match in re.finditer(r'action (\w+)\s*:\s*(\w+);', content_block):
            usage_name, usage_type = usage_match.groups()
            usages.append(usage_type)

        action_defs[action_name] = {
            'id': action_id,
            'usages': usages,
            'content': content_block
        }

    # Build hierarchy
    functions = []

    # Find top-level functions (not used by any other function)
    all_used_types = set()
    for action_name, action_data in action_defs.items():
        all_used_types.update(action_data['usages'])

    top_level_functions = [name for name in action_defs if name not in all_used_types]

    # Process each top-level function and its hierarchy
    for top_func_name in top_level_functions:
        stack = [(top_func_name, 0)]
        while stack:
            func_name, level = stack.pop()
            func_data = action_defs[func_name]

            # Add the function to our list
            functions.append({
                'name': func_name,
                'id': func_data['id'],
                'level': level
            })

            # Add children to stack in reverse order (so they're processed in order)
            for child_type in reversed(func_data['usages']):
                if child_type in action_defs:
                    stack.append((child_type, level + 1))

    return functions

def remove_id_suffix(name: str) -> str:
    """
    Remove the ID suffix (underscore and first 4 ID characters) from a name if present.
    Example: "PrepareEducationalSpaceExperiment_83c0" -> "PrepareEducationalSpaceExperiment"
    """
    return re.sub(r'_[0-9a-f]{4}$', '', name)

def export_functions_to_excel(functions: List[Dict], output_path: pathlib.Path) -> None:
    """Export all functions to an Excel file with hierarchical structure.
    Now exports names without the ID suffix and with proper spacing.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    if wb.active.title == "Sheet":
        wb.remove(wb.active)

    ws = wb.create_sheet("Functions")

    # Determine the maximum hierarchy level
    max_level = max([f['level'] for f in functions]) if functions else 0

    # Create headers dynamically based on max_level
    headers = []
    for level in range(max_level + 1):
        prefix = "Sub" * level if level > 0 else ""
        headers.extend([f"{prefix}Function ID", f"{prefix}Function Name", f"{prefix}Function Kind"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Export functions in hierarchy order (already in correct order from parse_functions_file)
    for func in functions:
        new_row = ws.max_row + 1

        # Format the name without ID suffix for export
        display_name = from_pascal_case(remove_id_suffix(func['name']))

        # Fill in all columns for this function's level and above
        for level in range(func['level'] + 1):
            col = 1 + (level * 3)
            if level == func['level']:
                # This is our current level - fill ID and name
                ws.cell(row=new_row, column=col, value=func['id'])
                ws.cell(row=new_row, column=col+1, value=display_name)
                # Leave kind column empty as requested
            else:
                # Higher levels - leave empty (will be filled by parent)
                pass

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Exported {len(functions)} functions to: {output_path}")

def main() -> None:
    _DVS_DIR = pathlib.Path(__file__).parent.parent
    DEFAULT_INPUT = _DVS_DIR / "Functions_generated.sysml"
    DEFAULT_OUTPUT = _DVS_DIR / "results" / "DVS_Functions_export.xlsx"

    args = sys.argv[1:]
    if args and args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    input_path = pathlib.Path(args[0]) if args else DEFAULT_INPUT
    if not input_path.exists():
        print(f"Error: file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = DEFAULT_OUTPUT if len(args) < 2 else pathlib.Path(args[1])

    print(f"Reading: {input_path}")
    functions = parse_functions_file(input_path)

    if not functions:
        print("Warning: No functions found in the file!")
    else:
        print(f"Found {len(functions)} functions in the file")
        # Print hierarchy for verification
        for func in functions:
            indent = "  " * func['level']
            print(f"{indent}{from_pascal_case(remove_id_suffix(func['name']))} (Level {func['level']}, ID: {func['id']})")

    export_functions_to_excel(functions, output_path)

if __name__ == "__main__":
    main()