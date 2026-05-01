#!/usr/bin/env python3
"""
export_functional_chains.py

Exports functional chains from Functions_generated.sysml to an Excel file.

Output columns match DVS_Functional_Chains.xlsx (dynamic width):
  Functional Chain ID, Functional Chain Name,
  Start Function ID, Start Function Name,
  End Function ID, End Function Name,
  Function 1 ID, Function 1 Name, Function 1 Involvement ID,
  ...  (one group per function up to max chain length)
  Exchange 1 ID, Exchange 1 Name, Exchange 1 Involvement ID,
  ...  (one group per exchange up to max exchanges in any chain)

Involvement ID columns are left empty (not stored in SysML).

Usage:
  python export_functional_chains.py [functions.sysml] [output.xlsx]
"""

import pathlib
import re
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import PatternFill


# ---------------------------------------------------------------------------
# Name helpers  (identical to the other export scripts)
# ---------------------------------------------------------------------------

def from_pascal_case(name: str) -> str:
    """Convert PascalCase/ID-suffixed name to space-separated words."""
    if not name:
        return name

    name_without_id = re.sub(r'_[0-9a-f]{4}$', '', name)

    if name_without_id.startswith('CODE'):
        rest_part = name_without_id[4:]
        processed_rest = []
        for i, char in enumerate(rest_part):
            if char.isupper() and i > 0:
                processed_rest.append(' ')
            processed_rest.append(char)
        return f"CODE {''.join(processed_rest)}".strip()

    result = [name_without_id[0]]
    i = 1
    n = len(name_without_id)
    while i < n:
        char = name_without_id[i]
        prev_char = name_without_id[i - 1]
        if char.isupper():
            if prev_char.islower():
                result.append(' ')
                result.append(char)
                i += 1
            else:
                j = i
                while j < n and name_without_id[j].isupper():
                    j += 1
                if j < n and name_without_id[j].islower():
                    result.append(name_without_id[i:j - 1])
                    result.append(' ')
                    result.append(name_without_id[j - 1])
                    i = j
                else:
                    result.append(char)
                    i += 1
        else:
            result.append(char)
            i += 1
    return ''.join(result)


def extract_id(text: str) -> str:
    """Extract UUID from a /* ID: uuid */ doc comment."""
    m = re.search(r'/\*\s*ID:\s*([0-9a-f-]+)\s*\*/', text)
    return m.group(1) if m else ""


def get_block(text: str, start_pos: int) -> Tuple[str, int]:
    """Return (block_content, end_pos) for the { } block starting at start_pos."""
    depth = 1
    pos = start_pos
    while pos < len(text) and depth > 0:
        if text[pos] == '{':
            depth += 1
        elif text[pos] == '}':
            depth -= 1
        pos += 1
    return (text[start_pos:pos - 1] if depth == 0 else ""), pos


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class FunctionalChain:
    chain_id: str
    chain_type: str
    function_types: List[str] = field(default_factory=list)
    exchange_types: List[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# SysML parsing
# ---------------------------------------------------------------------------

def parse_functions_file(file_path: pathlib.Path):
    """Parse Functions_generated.sysml.

    Returns:
      chains       — list of FunctionalChain (action defs with 'first start;')
      action_ids   — {type_name: id}  for all regular action defs
      item_ids     — {type_name: id}  for all item defs
    """
    content = file_path.read_text(encoding='utf-8')

    action_ids: Dict[str, str] = {}
    chains: List[FunctionalChain] = []

    for m in re.finditer(r'action def (\w+)\s*\{', content):
        type_name = m.group(1)
        body, _ = get_block(content, m.end())
        block_id = extract_id(body)

        if 'first start;' in body:
            # Functional chain: extract exchange type names and function sequence
            exchange_m = re.search(r'/\*\s*Exchanges:\s*([^*]+)\*/', body)
            exchange_types = []
            if exchange_m:
                exchange_types = [
                    t.strip() for t in exchange_m.group(1).split(',') if t.strip()
                ]
            function_types = re.findall(
                r'then action \w+\s*:\s*(\w+)\s*;', body
            )
            chains.append(FunctionalChain(
                chain_id=block_id,
                chain_type=type_name,
                function_types=function_types,
                exchange_types=exchange_types,
            ))
        else:
            # Regular action def: store id for cross-reference
            if block_id:
                action_ids[type_name] = block_id

    item_ids: Dict[str, str] = {}
    for m in re.finditer(r'item def (\w+)\s*\{', content):
        type_name = m.group(1)
        body, _ = get_block(content, m.end())
        block_id = extract_id(body)
        if block_id:
            item_ids[type_name] = block_id

    return chains, action_ids, item_ids


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_to_excel(
    chains: List[FunctionalChain],
    action_ids: Dict[str, str],
    item_ids: Dict[str, str],
    output_path: pathlib.Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    max_functions = max((len(c.function_types) for c in chains), default=0)
    max_exchanges = max((len(c.exchange_types) for c in chains), default=0)

    wb = openpyxl.Workbook()
    if wb.active.title == "Sheet":
        wb.remove(wb.active)
    ws = wb.create_sheet("Functional Chains")

    headers = [
        "Functional Chain ID", "Functional Chain Name",
        "Start Function ID", "Start Function Name",
        "End Function ID", "End Function Name",
    ]
    for n in range(1, max_functions + 1):
        headers += [f"Function {n} ID", f"Function {n} Name", f"Function {n} Involvement ID"]
    for n in range(1, max_exchanges + 1):
        headers += [f"Exchange {n} ID", f"Exchange {n} Name", f"Exchange {n} Involvement ID"]

    gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = gray

    for chain in chains:
        chain_name = from_pascal_case(chain.chain_type)

        func_ids = [action_ids.get(t, '') for t in chain.function_types]
        func_names = [from_pascal_case(t) for t in chain.function_types]

        exch_ids = [item_ids.get(t, '') for t in chain.exchange_types]
        exch_names = [from_pascal_case(t) for t in chain.exchange_types]

        start_func_id = func_ids[0] if func_ids else ''
        start_func_name = func_names[0] if func_names else ''
        end_func_id = func_ids[-1] if func_ids else ''
        end_func_name = func_names[-1] if func_names else ''

        row_values = [
            chain.chain_id, chain_name,
            start_func_id, start_func_name,
            end_func_id, end_func_name,
        ]

        for i in range(max_functions):
            row_values.append(func_ids[i] if i < len(func_ids) else '')
            row_values.append(func_names[i] if i < len(func_names) else '')
            row_values.append('')  # Involvement ID not stored in SysML

        for i in range(max_exchanges):
            row_values.append(exch_ids[i] if i < len(exch_ids) else '')
            row_values.append(exch_names[i] if i < len(exch_names) else '')
            row_values.append('')  # Involvement ID not stored in SysML

        r = ws.max_row + 1
        for col, value in enumerate(row_values, 1):
            ws.cell(r, col, value)

    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None and c.value != ''),
            default=0,
        )
        ws.column_dimensions[col[0].column_letter].width = (max_len + 2) * 1.2

    wb.save(output_path)
    print(f"Exported {len(chains)} functional chains to: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    _DVS_DIR = pathlib.Path(__file__).parent.parent
    DEFAULT_INPUT = _DVS_DIR / "Functions_generated.sysml"
    DEFAULT_OUTPUT = _DVS_DIR / "results" / "DVS_Functional_Chains_export.xlsx"

    args = sys.argv[1:]
    if args and args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    input_path = pathlib.Path(args[0]) if args else DEFAULT_INPUT
    if not input_path.exists():
        print(f"Error: file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = pathlib.Path(args[1]) if len(args) >= 2 else DEFAULT_OUTPUT

    chains, action_ids, item_ids = parse_functions_file(input_path)

    if not chains:
        print("Warning: no functional chains found.", file=sys.stderr)

    export_to_excel(chains, action_ids, item_ids, output_path)


if __name__ == "__main__":
    main()
