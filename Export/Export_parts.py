#!/usr/bin/env python3
"""
export_parts_to_excel.py
Exports parts from a SysML v2 file to an Excel file with proper hierarchical structure.
Each system is immediately followed by its subsystems in the Excel file.
Now excludes the ID suffix (underscore and first 4 ID characters) from exported names.
Improved name conversion to match import behavior:
- "HighSchoolTeacher" -> "High School Teacher"
- "AIVTeam" -> "AIV Team"
- "EPS" -> "EPS"
- "CODETest1" -> "CODE Test 1"
"""

import pathlib
import re
import sys
from typing import Dict, List

import openpyxl
from openpyxl.styles import PatternFill

def from_pascal_case(name: str) -> str:
    """
    Convert PascalCase to space-separated words with proper handling of acronyms.
    Rules:
    1. First letter is always kept as start of first word
    2. For subsequent uppercase letters:
       - If followed by lowercase: add space before the letter (e.g., "School" in "HighSchool")
       - If in a sequence of uppercase letters: add space before the last letter (e.g., "AIV" in "AIVTeam")
    3. Remove ID suffix if present (underscore followed by 4 hex characters)
    """
    if not name:
        return name

    # Remove ID suffix if present (underscore followed by 4 hex characters)
    name_without_id = re.sub(r'_[0-9a-f]{4}$', '', name)

    # If the name is all uppercase, return as-is
    if name_without_id.isupper():
        return name_without_id

    # Handle special case: single word with no uppercase after first letter
    if len(name_without_id) == 1:
        return name_without_id

    result = [name_without_id[0].upper()]

    i = 1
    n = len(name_without_id)

    while i < n:
        # Check if current character is uppercase
        if name_without_id[i].isupper():
            # Look ahead to see if next character is lowercase
            if i + 1 < n and name_without_id[i+1].islower():
                # This is the start of a new word (e.g., "School" in "HighSchool")
                result.append(' ')
                result.append(name_without_id[i].upper())
                i += 1
            else:
                # This is part of an acronym sequence (e.g., "AIV" in "AIVTeam")
                # Find the end of the uppercase sequence
                j = i
                while j < n and name_without_id[j].isupper():
                    j += 1

                # Add space before the last uppercase letter if followed by lowercase
                if j < n and name_without_id[j].islower():
                    # Add all but last uppercase letter
                    result.append(name_without_id[i:j-1])
                    result.append(' ')
                    result.append(name_without_id[j-1].upper())
                else:
                    # All uppercase at end, add as-is
                    result.append(name_without_id[i:j])

                i = j
        else:
            # Lowercase letter - just add it
            result.append(name_without_id[i])
            i += 1

    # Join all parts and clean up
    return ''.join(result).replace('  ', ' ').strip()

def extract_id_from_text(text: str) -> str:
    """Extract ID from text using the pattern: /* ID: uuid */"""
    match = re.search(r"/\*\s*ID:\s*([0-9a-f-]+)\s*\*/", text)
    return match.group(1) if match else ""

def get_block_content(text: str, start_pos: int) -> tuple:
    """Get content of a block starting with {, handling nested braces."""
    brace_count = 1
    end_pos = start_pos
    while end_pos < len(text) and brace_count > 0:
        if text[end_pos] == '{':
            brace_count += 1
        elif text[end_pos] == '}':
            brace_count -= 1
        end_pos += 1
    return text[start_pos:end_pos-1] if brace_count == 0 else "", end_pos

def parse_parts_file(file_path: pathlib.Path) -> tuple:
    """
    Parse the SysML file to extract part hierarchy.
    Returns a tuple of (parts, part_defs) where:
    - parts: list of dictionaries with part information and hierarchy level
    - part_defs: dictionary of all part definitions with their usages
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the PartsGenerated package content
    parts_package_match = re.search(r'package PartsGenerated\s*\{', content)
    if not parts_package_match:
        print("Error: PartsGenerated package not found in file", file=sys.stderr)
        return [], {}

    start_pos = parts_package_match.end()
    parts_content, end_pos = get_block_content(content, start_pos)

    # First, find all part definitions
    part_defs = {}
    for part_def_match in re.finditer(r'part def (\w+)\s*\{', parts_content):
        part_name = part_def_match.group(1)
        start = part_def_match.end()
        content_block, end = get_block_content(parts_content, start)
        part_id = extract_id_from_text(content_block)

        # Find all part usages in this part
        usages = []
        for usage_match in re.finditer(r'part (\w+)\s*:\s*(\w+);', content_block):
            usage_name, usage_type = usage_match.groups()
            usages.append(usage_type)

        part_defs[part_name] = {
            'id': part_id,
            'usages': usages,
            'content': content_block
        }

    # Build hierarchy - find top-level parts (not used by any other part)
    all_used_types = set()
    for part_name, part_data in part_defs.items():
        all_used_types.update(part_data['usages'])

    top_level_parts = [name for name in part_defs if name not in all_used_types]

    # If LogicalSystem exists (with or without ID suffix), make it the first top-level part
    logical_system_variants = [name for name in top_level_parts if name.startswith('LogicalSystem')]
    if logical_system_variants:
        top_level_parts.remove(logical_system_variants[0])
        top_level_parts.insert(0, logical_system_variants[0])

    # Build the complete hierarchy
    parts = []
    for top_part_name in top_level_parts:
        stack = [(top_part_name, 0)]
        while stack:
            part_name, level = stack.pop()
            part_data = part_defs[part_name]

            # Add the part to our list
            parts.append({
                'name': part_name,
                'id': part_data['id'],
                'level': level
            })

            # Add children to stack in reverse order (so they're processed in order)
            for child_type in reversed(part_data['usages']):
                if child_type in part_defs:
                    stack.append((child_type, level + 1))

    return parts, part_defs

def remove_id_suffix(name: str) -> str:
    """
    Remove the ID suffix (underscore and first 4 ID characters) from a name if present.
    Example: "EPS_cc31" -> "EPS"
             "LogicalSystem_e9b6" -> "LogicalSystem"
             "StoreEnergyInTheBatteries_baf4" -> "StoreEnergyInTheBatteries"
    """
    return re.sub(r'_[0-9a-f]{4}$', '', name)

def export_parts_to_excel(parts: List[Dict], part_defs: Dict, output_path: pathlib.Path) -> None:
    """Export all parts to an Excel file with hierarchical structure.
    Now exports names without the ID suffix and with proper spacing.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    if wb.active.title == "Sheet":
        wb.remove(wb.active)

    ws = wb.create_sheet("Parts")

    # Determine the maximum hierarchy level
    max_level = max([p['level'] for p in parts]) if parts else 0

    # Create headers dynamically based on max_level
    headers = []
    for level in range(max_level + 1):
        prefix = "Sub" * level if level > 0 else ""
        headers.extend([f"{prefix}System ID", f"{prefix}System Name", f"{prefix}System Type"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Export parts in hierarchy order (already in correct order from parse_parts_file)
    for part in parts:
        new_row = ws.max_row + 1

        # Format the name without ID suffix for export
        display_name = from_pascal_case(remove_id_suffix(part['name']))

        # Fill in all columns for this part's level and above
        for level in range(part['level'] + 1):
            col = 1 + (level * 3)
            if level == part['level']:
                # This is our current level - fill ID and name (without ID suffix)
                ws.cell(row=new_row, column=col, value=part['id'])
                ws.cell(row=new_row, column=col+1, value=display_name)
            else:
                # Higher levels - leave empty (will be filled by parent)
                pass

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)

def main() -> None:
    _DVS_DIR = pathlib.Path(__file__).parent.parent
    DEFAULT_INPUT = _DVS_DIR / "Parts_generated.sysml"
    DEFAULT_OUTPUT = _DVS_DIR / "results" / "DVS_Parts_export.xlsx"

    args = sys.argv[1:]
    if args and args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    input_path = pathlib.Path(args[0]) if args else DEFAULT_INPUT
    if not input_path.exists():
        print(f"Error: file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = DEFAULT_OUTPUT if len(args) < 2 else pathlib.Path(args[1])

    print(f"Reading: {input_path}")
    parts, part_defs = parse_parts_file(input_path)

    if not parts:
        print("Warning: No parts found in the file!")
    else:
        print(f"Found {len(parts)} parts in the file")
        # Print hierarchy for verification
        for part in parts:
            indent = "  " * part['level']
            print(f"{indent}{from_pascal_case(remove_id_suffix(part['name']))} (Level {part['level']}, ID: {part['id']})")

    export_parts_to_excel(parts, part_defs, output_path)

if __name__ == "__main__":
    main()