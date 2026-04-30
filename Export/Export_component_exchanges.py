#!/usr/bin/env python3
"""
export_component_exchanges_to_excel.py
Exports component exchanges from a SysML v2 file to an Excel file.
Properly extracts port IDs and applies naming conventions:
- Port names: CP1_9b89 -> CP 1 (exclude after _, add space between letters/numbers)
- Exchange names: EW_HST_HTML -> EW - HST - HTML (underscores to hyphen-space)
"""

import pathlib
import re
import sys
from typing import Dict, List

import openpyxl
from openpyxl.styles import PatternFill

def from_pascal_case(name: str) -> str:
    """Convert PascalCase to space-separated words with proper handling of acronyms."""
    if not name:
        return name
    if name.isupper():
        return name
    result = []
    i = 0
    n = len(name)
    while i < n:
        if name[i].isupper():
            j = i
            while j < n and name[j].isupper():
                j += 1
            if j == n:
                result.append(name[i:j])
                i = j
            elif j < n and name[j].islower():
                result.append(name[i:j-1])
                result.append(" " + name[j-1])
                i = j
            else:
                result.append(name[i:j])
                i = j
        else:
            result.append(name[i])
            i += 1
    return "".join(result).replace("  ", " ").strip()

def format_port_name(port_name: str) -> str:
    """Convert port name like CP1_9b89 to CP 1 (exclude after _, add space between letters/numbers)."""
    # Take only the part before the first underscore
    base_name = port_name.split('_')[0]
    # Insert space between letters and numbers
    result = re.sub(r'([a-zA-Z])([0-9])', r'\1 \2', base_name)
    result = re.sub(r'([0-9])([a-zA-Z])', r'\1 \2', result)
    return result

def format_exchange_name(exchange_name: str) -> str:
    """Convert exchange name like EW_HST_HTML to EW - HST - HTML."""
    return exchange_name.replace('_', ' - ')

def extract_id_from_text(text: str) -> str:
    """Extract ID from text using the pattern: ID: uuid"""
    match = re.search(r"/\*\s*ID:\s*([0-9a-f-]+)\s*\*/", text)
    return match.group(1) if match else ""

def get_block_content(text: str, start_pos: int) -> tuple:
    """Get content of a block starting with {, handling nested braces."""
    brace_count = 1
    end_pos = start_pos
    while end_pos < len(text) and brace_count > 0:
        if text[end_pos] == '{':
            brace_count += 1
        elif text[end_pos] == '}':
            brace_count -= 1
        end_pos += 1
    return text[start_pos:end_pos-1] if brace_count == 0 else "", end_pos

def parse_sysml_file(file_path: pathlib.Path) -> List[Dict]:
    """
    Parse the SysML file to extract connection definitions.
    Returns a list of dictionaries with connection information.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # First, create a dictionary of all part definitions with their ports and IDs
    part_port_map = {}
    for part_match in re.finditer(r'part def (\w+)\s*\{', content):
        part_name = part_match.group(1)
        start = part_match.end()
        part_content, end = get_block_content(content, start)

        part_port_map[part_name] = {}
        part_id = extract_id_from_text(part_content)

        # Find all port definitions in this part
        for port_match in re.finditer(r'port (\w+)\s*:\s*\w+\s*\{', part_content):
            port_name = port_match.group(1)
            port_start = port_match.end()
            port_content, port_end = get_block_content(part_content, port_start)
            port_id = extract_id_from_text(port_content)
            part_port_map[part_name][port_name] = {
                'id': port_id,
                'part_id': part_id
            }

    # Find all connection definitions
    connections = []
    for conn_match in re.finditer(r'connection def (\w+)\s*\{', content):
        conn_name = conn_match.group(1)
        start = conn_match.end()
        conn_content, end = get_block_content(content, start)

        # Extract connection ID
        conn_id = extract_id_from_text(conn_content)

        # Extract end parts
        end_part_matches = re.findall(r'end part (\w+)\s*:\s*(\w+);', conn_content)
        if len(end_part_matches) < 2:
            continue

        from_part_usage, from_part_type = end_part_matches[0]
        to_part_usage, to_part_type = end_part_matches[1]

        # Extract interface and connection details
        interface_match = re.search(
            r'interface : \w+\s+connect\s+(\w+)\.(\w+)\s+to\s+(\w+)\.(\w+)',
            conn_content
        )
        if not interface_match:
            continue

        from_part_name_in_conn, from_port_usage, to_part_name_in_conn, to_port_usage = interface_match.groups()

        # Get from part/port info
        from_info = part_port_map.get(from_part_type, {}).get(from_port_usage, {})
        from_part_id = from_info.get('part_id', "")
        from_port_id = from_info.get('id', "")

        # Get to part/port info
        to_info = part_port_map.get(to_part_type, {}).get(to_port_usage, {})
        to_part_id = to_info.get('part_id', "")
        to_port_id = to_info.get('id', "")

        connections.append({
            "conn_id": conn_id,
            "conn_name": conn_name,
            "from_part_id": from_part_id,
            "from_part_name": from_part_type,
            "from_port_id": from_port_id,
            "from_port_name": from_port_usage,  # Keep original for ID lookup
            "to_part_id": to_part_id,
            "to_part_name": to_part_type,
            "to_port_id": to_port_id,
            "to_port_name": to_port_usage  # Keep original for ID lookup
        })

    return connections

def create_workbook() -> openpyxl.Workbook:
    """Create a new workbook with headers."""
    wb = openpyxl.Workbook()
    if wb.active.title == "Sheet":
        wb.remove(wb.active)
    ws = wb.create_sheet("Component Exchanges")

    headers = [
        "Component From ID", "Component From Name", "Component From Port ID", "Component From Port Name",
        "Component From Port Direction", "Component From Port Kind",
        "Component Exchange ID", "Component Exchange Name", "Component Exchange Kind",
        "Component To ID", "Component To Name", "Component To Port ID", "Component To Port Name",
        "Component To Port Direction", "Component To Port Kind"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    return wb

def export_to_excel(connections: List[Dict], output_path: pathlib.Path) -> None:
    """Export connections to an Excel file with formatted names."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = create_workbook()
    ws = wb["Component Exchanges"]

    for conn in connections:
        new_row = ws.max_row + 1

        # Format names for export only (keep original names for ID lookup)
        formatted_from_port_name = format_port_name(conn["from_port_name"])
        formatted_to_port_name = format_port_name(conn["to_port_name"])
        formatted_exchange_name = format_exchange_name(conn["conn_name"])

        # From part and port
        ws.cell(row=new_row, column=1, value=conn["from_part_id"])
        ws.cell(row=new_row, column=2, value=from_pascal_case(conn["from_part_name"]))
        ws.cell(row=new_row, column=3, value=conn["from_port_id"])
        ws.cell(row=new_row, column=4, value=formatted_from_port_name)  # Formatted port name
        ws.cell(row=new_row, column=5, value="")  # Direction - leave empty
        ws.cell(row=new_row, column=6, value="")  # Kind - leave empty

        # Connection info
        ws.cell(row=new_row, column=7, value=conn["conn_id"])
        ws.cell(row=new_row, column=8, value=formatted_exchange_name)  # Formatted exchange name
        ws.cell(row=new_row, column=9, value="")  # Exchange Kind - leave empty

        # To part and port
        ws.cell(row=new_row, column=10, value=conn["to_part_id"])
        ws.cell(row=new_row, column=11, value=from_pascal_case(conn["to_part_name"]))
        ws.cell(row=new_row, column=12, value=conn["to_port_id"])
        ws.cell(row=new_row, column=13, value=formatted_to_port_name)  # Formatted port name
        ws.cell(row=new_row, column=14, value="")  # Direction - leave empty
        ws.cell(row=new_row, column=15, value="")  # Kind - leave empty

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Exported {len(connections)} component exchanges to: {output_path}")

def main() -> None:
    _DVS_DIR = pathlib.Path(__file__).parent.parent
    DEFAULT_INPUT = _DVS_DIR / "Parts_generated.sysml"
    DEFAULT_OUTPUT = _DVS_DIR / "results" / "DVS_Component_Exchanges_export.xlsx"

    args = sys.argv[1:]
    if args and args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    input_path = pathlib.Path(args[0]) if args else DEFAULT_INPUT
    if not input_path.exists():
        print(f"Error: file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = DEFAULT_OUTPUT if len(args) < 2 else pathlib.Path(args[1])

    print(f"Reading: {input_path}")
    connections = parse_sysml_file(input_path)

    if not connections:
        print("Warning: No connections found in the file!")
    else:
        print(f"Found {len(connections)} connections in the file")
        # Print debug info for verification
        for i, conn in enumerate(connections, 1):
            print(f"\nConnection {i}: {format_exchange_name(conn['conn_name'])}")
            print(f"  From: {conn['from_part_name']}.{format_port_name(conn['from_port_name'])} (ID: {conn['from_part_id']}, Port ID: {conn['from_port_id']})")
            print(f"  To:   {conn['to_part_name']}.{format_port_name(conn['to_port_name'])} (ID: {conn['to_part_id']}, Port ID: {conn['to_port_id']})")

    export_to_excel(connections, output_path)

if __name__ == "__main__":
    main()