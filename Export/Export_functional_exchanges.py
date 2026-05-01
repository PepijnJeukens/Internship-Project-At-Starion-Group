#!/usr/bin/env python3
"""
export_functional_exchanges_to_excel.py
Exports functional exchanges from a SysML v2 file to an Excel file.
Formats names for export by:
- Removing ID suffixes (underscore and 4 hex chars)
- Converting PascalCase to space-separated words with proper handling of acronyms
- Formatting port names by removing ID suffix and adding spaces between letters/numbers only
"""

import pathlib
import re
import sys
from typing import Dict, List

import openpyxl
from openpyxl.styles import PatternFill

def from_pascal_case(name: str) -> str:
    """
    Convert PascalCase to space-separated words with proper handling of acronyms.
    Rules:
    1. First remove ID suffix (underscore and 4 hex chars) if present
    2. For CODE-prefixed names: split each capital letter with spaces
    3. For uppercase sequences followed by lowercase: add space between last two uppercase letters
       - DVSTeam -> DVS Team
       - AIVTeam -> AIV Team
    4. For lowercase followed by uppercase: add space before uppercase letter
       - GroundStation -> Ground Station
    """
    if not name:
        return name

    # Remove ID suffix if present (underscore followed by 4 hex characters)
    name_without_id = re.sub(r'_[0-9a-f]{4}$', '', name)

    # Special handling for CODE prefix
    if name_without_id.startswith('CODE'):
        # Split CODE and the rest
        code_part = 'CODE'
        rest_part = name_without_id[4:]

        # Process the rest part by adding space before each uppercase letter
        processed_rest = []
        for i, char in enumerate(rest_part):
            if char.isupper() and i > 0:
                processed_rest.append(' ')
            processed_rest.append(char)
        return f"{code_part} {''.join(processed_rest)}".strip()

    # General case for all other names
    result = [name_without_id[0]]

    i = 1
    n = len(name_without_id)

    while i < n:
        char = name_without_id[i]
        prev_char = name_without_id[i-1]

        # Check if current character is uppercase
        if char.isupper():
            # Check if previous character is lowercase
            if prev_char.islower():
                # Add space before this uppercase letter
                result.append(' ')
                result.append(char)
                i += 1
            else:
                # Previous character is uppercase - we're in an uppercase sequence
                # Look ahead to see if this sequence is followed by lowercase
                j = i
                while j < n and name_without_id[j].isupper():
                    j += 1

                if j < n and name_without_id[j].islower():
                    # This is an uppercase sequence followed by lowercase
                    # Add all but the last uppercase letter
                    result.append(name_without_id[i:j-1])
                    # Add space and the last uppercase letter
                    result.append(' ')
                    result.append(name_without_id[j-1])
                    i = j
                else:
                    # This is an uppercase sequence at the end or followed by another uppercase
                    result.append(char)
                    i += 1
        else:
            # Lowercase letter - just add it
            result.append(char)
            i += 1

    return ''.join(result)

def format_port_name_for_export(port_usage_name: str) -> str:
    """Convert port usage name for export by removing ID suffix and formatting.
    Example: 'FOP1_1974' -> 'FOP 1', 'FIP3_2128' -> 'FIP 3'
    Only adds spaces between letters and numbers, not between letters.
    """
    # Remove everything after and including the first underscore
    base_name = port_usage_name.split('_')[0]

    # Insert space between letters and numbers only
    formatted = re.sub(r'([a-zA-Z])([0-9])', r'\1 \2', base_name)
    formatted = re.sub(r'([0-9])([a-zA-Z])', r'\1 \2', formatted)

    return formatted

def extract_id_from_text(text: str) -> str:
    """Extract ID from text using the pattern: /* ID: uuid */"""
    match = re.search(r"/\*\s*ID:\s*([0-9a-f-]+)\s*\*/", text)
    return match.group(1) if match else ""

def get_block_content(text: str, start_pos: int) -> tuple:
    """Get content of a block starting with {, handling nested braces."""
    brace_count = 1
    end_pos = start_pos
    while end_pos < len(text) and brace_count > 0:
        if text[end_pos] == '{':
            brace_count += 1
        elif text[end_pos] == '}':
            brace_count -= 1
        end_pos += 1
    return text[start_pos:end_pos-1] if brace_count == 0 else "", end_pos

def parse_sysml_file(file_path: pathlib.Path) -> List[Dict]:
    """
    Parse the SysML file to extract functional exchange definitions.
    Returns a list of dictionaries with functional exchange information.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # First, create a dictionary of all action definitions with their ports and IDs
    action_info = {}
    for action_match in re.finditer(r'action def (\w+)\s*\{', content):
        action_name = action_match.group(1)
        start = action_match.end()
        action_content, end = get_block_content(content, start)

        # Extract action ID
        action_id = extract_id_from_text(action_content)

        # Find all port definitions in this action
        action_info[action_name] = {
            'id': action_id,
            'ports': {}
        }

        for port_match in re.finditer(r'port (\w+)\s*:\s*\w+\s*\{', action_content):
            port_usage_name = port_match.group(1)
            port_start = port_match.end()
            port_content, port_end = get_block_content(action_content, port_start)

            # Extract port ID
            port_id = extract_id_from_text(port_content)

            action_info[action_name]['ports'][port_usage_name] = {
                'id': port_id,
                'name': port_usage_name  # Store the exact port usage name for ID lookup
            }

    # Find all connection definitions (functional exchanges)
    functional_exchanges = []
    for conn_match in re.finditer(r'connection def (\w+)\s*\{', content):
        conn_name = conn_match.group(1)
        start = conn_match.end()
        conn_content, end = get_block_content(content, start)

        # Extract connection ID
        conn_id = extract_id_from_text(conn_content)

        # Extract end actions
        end_action_matches = re.findall(r'end action (\w+)\s*:\s*(\w+);', conn_content)
        if len(end_action_matches) < 2:
            continue

        from_action_usage, from_action_type = end_action_matches[0]
        to_action_usage, to_action_type = end_action_matches[1]

        # Extract interface and connection details
        interface_match = re.search(
            r'interface : \w+\s+connect\s+(\w+)\.(\w+)\s+to\s+(\w+)\.(\w+)',
            conn_content
        )
        if not interface_match:
            continue

        # The connection is from to_port to from_port (order in the connect statement)
        to_action_name_in_conn, to_port_usage, from_action_name_in_conn, from_port_usage = interface_match.groups()

        # Get from action info
        from_action_data = action_info.get(from_action_type, {})
        from_action_id = from_action_data.get('id', "")
        from_port_data = from_action_data.get('ports', {}).get(from_port_usage, {})
        from_port_id = from_port_data.get('id', "")

        # Get to action info
        to_action_data = action_info.get(to_action_type, {})
        to_action_id = to_action_data.get('id', "")
        to_port_data = to_action_data.get('ports', {}).get(to_port_usage, {})
        to_port_id = to_port_data.get('id', "")

        functional_exchanges.append({
            "exchange_id": conn_id,
            "exchange_name": conn_name,
            "from_action_id": from_action_id,
            "from_action_name": from_action_type,
            "from_port_id": from_port_id,
            "from_port_name": from_port_usage,  # Full name for ID lookup (e.g., FIP1_24b3)
            "to_action_id": to_action_id,
            "to_action_name": to_action_type,
            "to_port_id": to_port_id,
            "to_port_name": to_port_usage  # Full name for ID lookup (e.g., FOP1_ac6c)
        })

    return functional_exchanges

def create_workbook() -> openpyxl.Workbook:
    """Create a new workbook with headers."""
    wb = openpyxl.Workbook()

    if wb.active.title == "Sheet":
        wb.remove(wb.active)

    ws = wb.create_sheet("Functional Exchanges")

    headers = [
        "Function From ID", "Function From Name", "Function From Port ID", "Function From Port Name",
        "Functional Exchange ID", "Functional Exchange Name",
        "Function To ID", "Function To Name", "Function To Port ID", "Function To Port Name"
    ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    return wb

def export_to_excel(functional_exchanges: List[Dict], output_path: pathlib.Path) -> None:
    """Export functional exchanges to an Excel file with formatted names."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = create_workbook()
    ws = wb["Functional Exchanges"]

    for exchange in functional_exchanges:
        new_row = ws.max_row + 1

        # Format names for export
        formatted_from_action_name = from_pascal_case(exchange["from_action_name"])
        formatted_to_action_name = from_pascal_case(exchange["to_action_name"])
        formatted_exchange_name = from_pascal_case(exchange["exchange_name"])
        formatted_from_port_name = format_port_name_for_export(exchange["from_port_name"])
        formatted_to_port_name = format_port_name_for_export(exchange["to_port_name"])

        # From function and port
        ws.cell(row=new_row, column=1, value=exchange["from_action_id"])
        ws.cell(row=new_row, column=2, value=formatted_from_action_name)
        ws.cell(row=new_row, column=3, value=exchange["from_port_id"])
        ws.cell(row=new_row, column=4, value=formatted_from_port_name)  # Formatted port name

        # Functional exchange info
        ws.cell(row=new_row, column=5, value=exchange["exchange_id"])
        ws.cell(row=new_row, column=6, value=formatted_exchange_name)

        # To function and port
        ws.cell(row=new_row, column=7, value=exchange["to_action_id"])
        ws.cell(row=new_row, column=8, value=formatted_to_action_name)
        ws.cell(row=new_row, column=9, value=exchange["to_port_id"])
        ws.cell(row=new_row, column=10, value=formatted_to_port_name)  # Formatted port name

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Exported {len(functional_exchanges)} functional exchanges to: {output_path}")

def main() -> None:
    _DVS_DIR = pathlib.Path(__file__).parent.parent
    DEFAULT_INPUT = _DVS_DIR / "Functions_generated.sysml"
    DEFAULT_OUTPUT = _DVS_DIR / "results" / "DVS_Functional_Exchanges_export.xlsx"

    args = sys.argv[1:]
    if args and args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    input_path = pathlib.Path(args[0]) if args else DEFAULT_INPUT
    if not input_path.exists():
        print(f"Error: file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_path = DEFAULT_OUTPUT if len(args) < 2 else pathlib.Path(args[1])

    print(f"Reading: {input_path}")
    functional_exchanges = parse_sysml_file(input_path)

    if not functional_exchanges:
        print("Warning: No functional exchanges found in the file!")
    else:
        print(f"Found {len(functional_exchanges)} functional exchanges in the file")
        # Print debug info for the first few exchanges
        for i, exchange in enumerate(functional_exchanges[:3], 1):
            # Format names for display
            display_from_action = from_pascal_case(exchange["from_action_name"])
            display_to_action = from_pascal_case(exchange["to_action_name"])
            display_exchange = from_pascal_case(exchange["exchange_name"])
            display_from_port = format_port_name_for_export(exchange["from_port_name"])
            display_to_port = format_port_name_for_export(exchange["to_port_name"])

            print(f"\nSample exchange {i}:")
            print(f"  From: {display_from_action}.{display_from_port} (ID: {exchange['from_action_id']}, Port ID: {exchange['from_port_id']})")
            print(f"  To:   {display_to_action}.{display_to_port} (ID: {exchange['to_action_id']}, Port ID: {exchange['to_port_id']})")
            print(f"  Exchange: {display_exchange} (ID: {exchange['exchange_id']})")

    export_to_excel(functional_exchanges, output_path)

if __name__ == "__main__":
    main()