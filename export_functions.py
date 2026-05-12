#!/usr/bin/env python3
"""
export_functions.py

Contains all unique export functions consolidated from the individual export scripts:
  export_parts_to_excel.py, export_functions_from_excel.py, export_allocations.py,
  export_functional_exchanges.py, export_component_exchanges.py,
  export_exchange_allocations.py, export_functional_chains.py

Public API
----------
  export_systems(ws, parts_path)
  export_functions(ws, functions_path)
  export_allocations(ws, parts_path, functions_path)
  export_functional_exchanges(ws, functions_path)
  export_component_exchanges(ws, parts_path)
  export_exchange_allocations(ws, parts_path, functions_path)
  export_functional_chains(ws, functions_path)
  full_export(parts_filename, functions_filename, excel_name)
"""

import pathlib
import re
import sys
import uuid as _uuid
from dataclasses import dataclass, field
from typing import Dict, List, Tuple

# Fixed namespace for deterministic fallback ID generation (matches uuid.NAMESPACE_URL)
_FALLBACK_NS = _uuid.UUID('6ba7b810-9dad-11d1-80b4-00c04fd430c8')

import openpyxl
from openpyxl.styles import PatternFill


# ============================================================================
# Shared utility functions
# ============================================================================

def from_pascal_case(name: str) -> str:
    """Convert PascalCase / ID-suffixed name to space-separated words.

    Rules:
      - Removes trailing ID suffix (_xxxx where xxxx is 4 hex chars).
      - Names starting with 'CODE': split each following capital with a space.
      - Uppercase sequence followed by lowercase (DVSTeam -> DVS Team).
      - Lowercase followed by uppercase (GroundStation -> Ground Station).
    """
    if not name:
        return name

    name_without_id = re.sub(r'_[0-9a-f]{4}$', '', name)

    if name_without_id.startswith('CODE'):
        rest_part = name_without_id[4:]
        processed_rest = []
        for i, char in enumerate(rest_part):
            if char.isupper() and i > 0:
                processed_rest.append(' ')
            processed_rest.append(char)
        return f"CODE {''.join(processed_rest)}".strip()

    result = [name_without_id[0]]
    i = 1
    n = len(name_without_id)

    while i < n:
        char = name_without_id[i]
        prev_char = name_without_id[i - 1]

        if char.isupper():
            if prev_char.islower():
                result.append(' ')
                result.append(char)
                i += 1
            else:
                j = i
                while j < n and name_without_id[j].isupper():
                    j += 1
                if j < n and name_without_id[j].islower():
                    result.append(name_without_id[i:j - 1])
                    result.append(' ')
                    result.append(name_without_id[j - 1])
                    i = j
                else:
                    result.append(char)
                    i += 1
        else:
            result.append(char)
            i += 1

    return ''.join(result)


def remove_id_suffix(name: str) -> str:
    """Remove the trailing ID suffix (_xxxx where xxxx is 4 hex chars)."""
    return re.sub(r'_[0-9a-f]{4}$', '', name)


def _has_id_comment(text: str) -> bool:
    """Return True if *text* contains an explicit /* ID: uuid */ doc comment."""
    return bool(re.search(r'/\*\s*ID:\s*[0-9a-f-]+\s*\*/', text))


def extract_id(text: str, element=None) -> str:
    """Extract UUID for a SysML element using three strategies in order:

    1. Parse a ``/* ID: <uuid> */`` doc comment from *text*.
    2. Read ``element.element_id`` from a syside Element (if provided).
    3. Generate a deterministic UUID from *text* so the same element
       always receives the same ID regardless of which sheet is being exported.
    """
    # 1. Explicit ID comment
    m = re.search(r'/\*\s*ID:\s*([0-9a-f-]+)\s*\*/', text)
    if m:
        return m.group(1)

    # 2. syside element_id
    if element is not None:
        try:
            eid = element.element_id
            if eid:
                return str(eid)
        except Exception:
            pass

    # 3. Deterministic fallback — hash the block text so the same element
    #    always gets the same ID regardless of which sheet is being exported.
    return str(_uuid.uuid5(_FALLBACK_NS, text))


def get_block(text: str, start_pos: int) -> Tuple[str, int]:
    """Return (block_content, end_pos) for the { } block starting at start_pos."""
    depth = 1
    pos = start_pos
    while pos < len(text) and depth > 0:
        if text[pos] == '{':
            depth += 1
        elif text[pos] == '}':
            depth -= 1
        pos += 1
    return (text[start_pos:pos - 1] if depth == 0 else ""), pos


def format_port_name(port_usage: str) -> str:
    """Convert port usage name to display form.

    Examples: CP1_9b89 -> CP 1,  FOP1_04b2 -> FOP 1
    Strips everything from the first underscore, then inserts a space between
    letters and digits.
    """
    base = port_usage.split('_')[0]
    base = re.sub(r'([a-zA-Z])([0-9])', r'\1 \2', base)
    base = re.sub(r'([0-9])([a-zA-Z])', r'\1 \2', base)
    return base


def format_exchange_name(name: str) -> str:
    """Convert component-exchange SysML name to display form.

    Example: EW_HST_HTML -> EW - HST - HTML
    """
    return name.replace('_', ' - ')


def _apply_header_style(ws) -> None:
    """Fill the first row of *ws* with a light-gray background."""
    gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.fill = gray


def _auto_adjust_columns(ws) -> None:
    """Set each column width to fit its widest cell value."""
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col if c.value is not None and c.value != ''),
            default=0,
        )
        ws.column_dimensions[col[0].column_letter].width = (max_len + 2) * 1.2


def _get_direct_allocated_functions(body: str) -> List[Dict]:
    """Return allocated functions declared directly in *body*, skipping nested blocks.

    Only ``perform action`` usages at the immediate level of *body* are returned.
    Any perform actions inside a nested ``{ }`` child block are skipped so that a
    parent usage block does not absorb its children's allocations.
    """
    result: List[Dict] = []
    i = 0
    n = len(body)

    while i < n:
        m_action = re.search(r'\bperform action (\w+)\s*:\s*(\w+)\s*([{;])', body[i:])
        if not m_action:
            break

        abs_action_start = i + m_action.start()
        next_brace = body.find('{', i)

        if next_brace != -1 and next_brace < abs_action_start:
            # A nested block opens before this perform action — skip the block.
            _, abs_end = get_block(body, next_brace + 1)
            i = abs_end
        else:
            # The perform action is at the current level — capture it.
            usage_name = m_action.group(1)
            func_type  = m_action.group(2)
            terminator = m_action.group(3)
            if terminator == '{':
                func_body, abs_end = get_block(body, i + m_action.end())
            else:
                func_body = ''
                abs_end   = i + m_action.end()
            result.append({
                'name':       func_type,
                'id':         extract_id(func_body) if func_body else extract_id(f'{usage_name}:{func_type}'),
                'usage_name': usage_name,
            })
            i = abs_end

    return result


def _get_direct_child_types(body: str) -> List[Tuple[str, int]]:
    """Return (type_name, count) of direct child part/item usages from a block body.

    Handles multiplicities in both positions:
      part [N] name : Type   (multiplicity before usage name)
      part name [N] : Type   (multiplicity after usage name)
    Skips nested blocks so only direct children are returned, not descendants.
    """
    children: List[Tuple[str, int]] = []
    i = 0
    while i < len(body):
        m = re.search(
            r'(?:part|item)\s+(?:\[(\d+)\]\s+)?(\w+)\s*(?:\[(\d+)\])?\s*:\s*(\w+)\s*([{;])',
            body[i:]
        )
        if not m:
            break
        mult_before, _usage, mult_after, type_name, terminator = m.groups()
        count = int(mult_before or mult_after or 1)
        children.append((type_name, count))
        abs_end = i + m.end()
        if terminator == '{':
            _, after_block = get_block(body, abs_end)
            i = after_block
        else:
            i = abs_end
    return children


def _build_usage_to_type_map(content: str) -> Dict[str, str]:
    """Return {usage_name: type_name} from all part/item usage declarations in *content*."""
    result: Dict[str, str] = {}
    for m in re.finditer(r'(?:part|item)\s+(\w+)\s*:\s*(\w+)\s*[{;]', content):
        result.setdefault(m.group(1), m.group(2))
    return result


def _get_package_content(content: str, file_path: pathlib.Path) -> Tuple[str, str]:
    """Find the first top-level package declaration in *content*.

    Returns (package_name, package_body).
    Raises ValueError if no package declaration is found.
    """
    m = re.search(r'package\s+(\w+)\s*\{', content)
    if not m:
        raise ValueError(f"No package declaration found in {file_path}")
    package_name = m.group(1)
    body, _ = get_block(content, m.end())
    return package_name, body


# ============================================================================
# Systems export  (Parts_generated.sysml -> "Systems" worksheet)
# ============================================================================

def _collect_exchange_item_names(content: str) -> set:
    """Return the set of item def type names that are exchange items (not actors).

    An item def is classified as an exchange item if its name appears as a
    flowing item type in any port def, interface def, or action def within *content*:
      - port def body:      'in/out/inout item _ : ItemType'
      - interface def body: 'flow of ItemType from ... to ...'
      - action def body:    'in/out/inout item _ : ItemType'

    Item defs whose names do not appear in any of these contexts are treated as actors.
    """
    exchange_items: set = set()

    for m in re.finditer(r'\bport def \w+\s*\{', content):
        body, _ = get_block(content, m.end())
        for pm in re.finditer(r'\b(?:in|out|inout)\s+item\s+\w+\s*:\s*(\w+)', body):
            exchange_items.add(pm.group(1))

    for m in re.finditer(r'\binterface def \w+\s*\{', content):
        body, _ = get_block(content, m.end())
        for fm in re.finditer(r'\bflow\s+of\s+(\w+)\b', body):
            exchange_items.add(fm.group(1))

    for m in re.finditer(r'\baction def \w+\s*\{', content):
        body, _ = get_block(content, m.end())
        for pm in re.finditer(r'\b(?:in|out|inout)\s+item\s+\w+\s*:\s*(\w+)', body):
            exchange_items.add(pm.group(1))

    return exchange_items

def _parse_parts_hierarchy(file_path: pathlib.Path) -> Tuple[List[Dict], Dict]:
    """Parse a parts SysML file and return (parts_list, part_defs).

    Supports both old-style (hierarchy inside def bodies) and new-style
    (bare defs with hierarchy expressed through nested usage blocks).
    Bare semicolon defs (part def X; / item def X;) are also recognised.
    Multiplicities (part name [N] : Type or part [N] name : Type) cause the
    child to appear N times in the output, each with a freshly generated ID.

    parts_list entries: {'name': str, 'id': str, 'id_from_comment': bool, 'level': int, 'node_type': str}
    part_defs entries:  {'id': str, 'id_from_comment': bool,
                         'usages': [(type_name, inst_id), ...], 'content': str, 'node_type': str}
                        inst_id is '' to inherit the type's own id, or a uuid4 string for
                        multiplicity instances that need distinct ids.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    try:
        _, parts_content = _get_package_content(content, file_path)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return [], {}

    # Identify exchange items so they can be excluded from the Systems worksheet.
    # Exchange items are item defs whose type names appear as flowing data in
    # port defs or interface defs; all other item defs are actors.
    exchange_item_names = _collect_exchange_item_names(parts_content)

    part_defs: Dict = {}

    # --- defs with bodies ---
    for m in re.finditer(r'(part|item) def (\w+)\s*\{', parts_content):
        def_keyword = m.group(1)
        part_name = m.group(2)
        body, _ = get_block(parts_content, m.end())
        if def_keyword == 'item' and part_name in exchange_item_names:
            continue  # Exchange items belong in Functional Exchanges, not Systems
        # Old-style: child usages appear as bare semicolon usages inside def body
        usages = [(t, '') for _, t in re.findall(r'(?:part|item) (\w+)\s*:\s*(\w+);', body)]
        node_type = "Actor" if def_keyword == "item" else "Component"
        part_defs[part_name] = {
            'id': extract_id(body),
            'id_from_comment': _has_id_comment(body),
            'usages': usages,
            'content': body,
            'node_type': node_type,
        }

    # --- bare semicolon defs (part def X; / item def X;) ---
    for m in re.finditer(r'(part|item) def (\w+)\s*;', parts_content):
        def_keyword = m.group(1)
        part_name = m.group(2)
        if part_name not in part_defs:
            if def_keyword == 'item' and part_name in exchange_item_names:
                continue  # Exchange items belong in Functional Exchanges, not Systems
            node_type = "Actor" if def_keyword == "item" else "Component"
            part_defs[part_name] = {
                'id': extract_id(part_name),
                'id_from_comment': False,
                'usages': [],
                'content': '',
                'node_type': node_type,
            }

    # New-style detection: if no def body contains child usages, the hierarchy
    # is expressed through nested usage blocks instead.
    if not any(d['usages'] for d in part_defs.values()):
        for m in re.finditer(
            r'(?:part|item)\s+(?:\[\d+\]\s+)?\w+\s*(?:\[\d+\])?\s*:\s*(\w+)\s*\{',
            parts_content
        ):
            parent_type = m.group(1)
            if parent_type not in part_defs:
                continue
            body, _ = get_block(parts_content, m.end())
            raw_children = [(t, n) for t, n in _get_direct_child_types(body) if t in part_defs]
            if raw_children and not part_defs[parent_type]['usages']:
                expanded: List[Tuple[str, str]] = []
                for child_type, count in raw_children:
                    if count == 1:
                        expanded.append((child_type, ''))
                    else:
                        for _ in range(count):
                            expanded.append((child_type, str(_uuid.uuid4())))
                part_defs[parent_type]['usages'] = expanded

    all_used = {t for d in part_defs.values() for t, _ in d['usages']}
    top_level = [n for n in part_defs if n not in all_used]

    logical = [n for n in top_level if n.startswith('LogicalSystem')]
    if logical:
        top_level.remove(logical[0])
        top_level.insert(0, logical[0])

    parts: List[Dict] = []
    for root in top_level:
        stack = [(root, 0, '')]
        while stack:
            name, level, inst_id = stack.pop()
            actual_id = inst_id if inst_id else part_defs[name]['id']
            parts.append({
                'name': name,
                'id': actual_id,
                'id_from_comment': part_defs[name]['id_from_comment'],
                'level': level,
                'node_type': part_defs[name]['node_type'],
            })
            for child_type, child_inst_id in reversed(part_defs[name]['usages']):
                if child_type in part_defs:
                    stack.append((child_type, level + 1, child_inst_id))

    return parts, part_defs


def export_systems(ws, parts_path: pathlib.Path) -> None:
    """Populate *ws* (Systems worksheet) from Parts_generated.sysml."""
    parts, _ = _parse_parts_hierarchy(parts_path)

    max_level = max((p['level'] for p in parts), default=0)
    headers = []
    for level in range(max_level + 1):
        prefix = "Sub" * level
        headers.extend([f"{prefix}System ID", f"{prefix}System Name", f"{prefix}System Type"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws)

    for part in parts:
        row = ws.max_row + 1
        col = 1 + part['level'] * 3
        display_name = remove_id_suffix(part['name']) if part['id_from_comment'] else part['name']
        ws.cell(row=row, column=col,     value=part['id'])
        ws.cell(row=row, column=col + 1, value=from_pascal_case(display_name))
        ws.cell(row=row, column=col + 2, value=part['node_type'])

    _auto_adjust_columns(ws)
    print(f"  -> {len(parts)} systems exported.")


# ============================================================================
# Functions export  (Functions_generated.sysml -> "Functions" worksheet)
# ============================================================================

def _parse_actions_hierarchy(file_path: pathlib.Path) -> List[Dict]:
    """Parse Functions_generated.sysml and return a flat list of action defs.

    Entries: {'name': str, 'id': str, 'id_from_comment': bool, 'level': int}
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    try:
        _, funcs_content = _get_package_content(content, file_path)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return []

    action_defs: Dict = {}
    for m in re.finditer(r'action def (\w+)\s*\{', funcs_content):
        name = m.group(1)
        body, _ = get_block(funcs_content, m.end())
        # Skip functional chains — they contain 'then action' which is not a regular function
        if re.search(r'\bthen action\b', body):
            continue
        usages = [t for _, t in re.findall(r'action (\w+)\s*:\s*(\w+);', body)]
        action_defs[name] = {'id': extract_id(body), 'id_from_comment': _has_id_comment(body), 'usages': usages}

    all_used = {t for d in action_defs.values() for t in d['usages']}
    top_level = [n for n in action_defs if n not in all_used]

    functions: List[Dict] = []
    for root in top_level:
        stack = [(root, 0)]
        while stack:
            name, level = stack.pop()
            functions.append({'name': name, 'id': action_defs[name]['id'], 'id_from_comment': action_defs[name]['id_from_comment'], 'level': level})
            for child in reversed(action_defs[name]['usages']):
                if child in action_defs:
                    stack.append((child, level + 1))

    return functions


def export_functions(ws, functions_path: pathlib.Path) -> None:
    """Populate *ws* (Functions worksheet) from Functions_generated.sysml."""
    functions = _parse_actions_hierarchy(functions_path)

    max_level = max((f['level'] for f in functions), default=0)
    headers = []
    for level in range(max_level + 1):
        prefix = "Sub" * level
        headers.extend([f"{prefix}Function ID", f"{prefix}Function Name", f"{prefix}Function Kind"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws)

    for func in functions:
        row = ws.max_row + 1
        col = 1 + func['level'] * 3
        display_name = remove_id_suffix(func['name']) if func['id_from_comment'] else func['name']
        ws.cell(row=row, column=col,     value=func['id'])
        ws.cell(row=row, column=col + 1, value=from_pascal_case(display_name))
        ws.cell(row=row, column=col + 2, value="FUNCTION") # Hard coded as function, to be specified by the user when diagramming in Capella

    _auto_adjust_columns(ws)
    print(f"  -> {len(functions)} functions exported.")


# ============================================================================
# Allocations export  (Parts -> "Link Systems and Functions" worksheet)
# ============================================================================

def _parse_parts_with_allocated_functions(file_path: pathlib.Path) -> Dict:
    """Parse a parts SysML file and return part_defs including perform actions.

    Supports both old-style (perform actions in def bodies) and new-style
    (bare defs with perform actions in usage blocks).
    Bare semicolon defs (part def X; / item def X;) are also recognised.
    Multiplicities in usages are expanded, each instance receiving a fresh ID.

    part_defs entries: {
      'id': str,
      'usages': [(type_name, inst_id), ...],   inst_id '' = inherit type's own id
      'functions': [{'name': str, 'id': str, 'usage_name': str}, ...],
      'content': str,
    }
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    try:
        _, parts_content = _get_package_content(content, file_path)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return {}

    # Identify exchange items so they are excluded from allocations (same logic as Systems).
    exchange_item_names = _collect_exchange_item_names(parts_content)

    part_defs: Dict = {}

    # --- defs with bodies ---
    for m in re.finditer(r'(part|item) def (\w+)\s*\{', parts_content):
        def_keyword = m.group(1)
        part_name = m.group(2)
        body, _ = get_block(parts_content, m.end())
        if def_keyword == 'item' and part_name in exchange_item_names:
            continue  # Exchange items belong in Functional Exchanges, not allocations
        # Old-style: child usages and perform actions inside def body
        usages = [(t, '') for _, t in re.findall(r'(?:part|item) (\w+)\s*:\s*(\w+);', body)]
        allocated_functions = []
        for pm in re.finditer(r'perform action (\w+)\s*:\s*(\w+)\s*([{;])', body):
            usage_name, func_type, terminator = pm.groups()
            if terminator == '{':
                func_body, _ = get_block(body, pm.end())
            else:
                func_body = ''
            func_id = extract_id(func_body) if func_body else extract_id(f'{usage_name}:{func_type}')
            allocated_functions.append({'name': func_type, 'id': func_id, 'usage_name': usage_name})
        part_defs[part_name] = {
            'id': extract_id(body),
            'usages': usages,
            'functions': allocated_functions,
            'content': body,
        }

    # --- bare semicolon defs (part def X; / item def X;) ---
    for m in re.finditer(r'(part|item) def (\w+)\s*;', parts_content):
        def_keyword = m.group(1)
        part_name = m.group(2)
        if part_name not in part_defs:
            if def_keyword == 'item' and part_name in exchange_item_names:
                continue  # Exchange items belong in Functional Exchanges, not allocations
            part_defs[part_name] = {
                'id': extract_id(part_name),
                'usages': [],
                'functions': [],
                'content': '',
            }

    # Always scan part/item usage blocks for both hierarchy and perform-action
    # allocations, regardless of what was found in def bodies.  Results are merged
    # with any data already captured from def bodies (deduplication by type name).
    # This handles:
    #   - New-style: bare defs where hierarchy and allocations live in usage blocks.
    #   - Mixed-style: some allocations in def bodies, others in usage blocks.
    #   - Old-style: everything in def bodies (usage scan finds nothing new).
    # _get_direct_allocated_functions prevents a parent usage block from absorbing
    # its children's allocations by skipping nested { } blocks.
    for m in re.finditer(
        r'(?:part|item)\s+(?:\[\d+\]\s+)?\w+\s*(?:\[\d+\])?\s*:\s*(\w+)\s*\{',
        parts_content
    ):
        parent_type = m.group(1)
        if parent_type not in part_defs:
            continue
        body, _ = get_block(parts_content, m.end())

        # Hierarchy: add child types not already present from def bodies.
        raw_children = [(t, n) for t, n in _get_direct_child_types(body) if t in part_defs]
        if raw_children:
            existing_child_types = {t for t, _ in part_defs[parent_type]['usages']}
            for child_type, count in raw_children:
                if child_type not in existing_child_types:
                    if count == 1:
                        part_defs[parent_type]['usages'].append((child_type, ''))
                    else:
                        for _ in range(count):
                            part_defs[parent_type]['usages'].append((child_type, str(_uuid.uuid4())))
                    existing_child_types.add(child_type)

        # Functions: merge usage-block allocations with any already captured from
        # def bodies, using the function type name as the deduplication key.
        existing_func_types = {f['name'] for f in part_defs[parent_type]['functions']}
        for func_entry in _get_direct_allocated_functions(body):
            if func_entry['name'] not in existing_func_types:
                part_defs[parent_type]['functions'].append(func_entry)
                existing_func_types.add(func_entry['name'])

    return part_defs


def _parts_with_subtree_functions(part_defs: Dict) -> set:
    """Return the set of part names that have at least one function in their subtree.

    A part is included if it has direct function allocations OR any descendant does.
    """
    has_func: set = set()

    def _check(name: str, visiting: set) -> bool:
        if name in has_func:
            return True
        if name in visiting:
            return False
        visiting.add(name)
        data = part_defs.get(name)
        if data is None:
            return False
        if data['functions']:
            has_func.add(name)
            return True
        for child_type, _ in data['usages']:
            if _check(child_type, visiting):
                has_func.add(name)
                return True
        return False

    for name in part_defs:
        _check(name, set())

    return has_func


def _build_allocation_hierarchy(part_defs: Dict) -> List[Dict]:
    """Return a flat list of part and function items in hierarchy order.

    Parts with no functions anywhere in their subtree are omitted entirely.
    """
    all_used = {t for d in part_defs.values() for t, _ in d['usages']}
    top_level = [n for n in part_defs if n not in all_used]

    logical = [n for n in top_level if n.startswith('LogicalSystem')]
    if logical:
        top_level.remove(logical[0])
        top_level.insert(0, logical[0])

    # Pre-compute which parts have at least one function somewhere in their subtree.
    parts_with_funcs = _parts_with_subtree_functions(part_defs)

    result: List[Dict] = []
    for root in top_level:
        if root not in parts_with_funcs:
            continue
        stack = [(root, 0, '')]
        while stack:
            name, level, inst_id = stack.pop()
            data = part_defs[name]
            actual_id = inst_id if inst_id else data['id']
            result.append({'type': 'part', 'name': name, 'id': actual_id, 'level': level})
            for func in data['functions']:
                result.append({'type': 'function', 'name': func['name'], 'id': func['id'],
                                'level': level, 'parent': name})
            for child_type, child_inst_id in reversed(data['usages']):
                if child_type in part_defs and child_type in parts_with_funcs:
                    stack.append((child_type, level + 1, child_inst_id))

    return result


def export_allocations(ws, parts_path: pathlib.Path, functions_path: pathlib.Path) -> None:
    """Populate *ws* (Link Systems and Functions worksheet) from Parts_generated.sysml."""
    part_defs = _parse_parts_with_allocated_functions(parts_path)
    hierarchy = _build_allocation_hierarchy(part_defs)

    max_level = max((item['level'] for item in hierarchy), default=0)
    headers = []
    for level in range(max_level + 1):
        prefix = "Sub" * level
        headers.extend([f"{prefix}System ID", f"{prefix}System Name"])
    headers.extend(["Function ID", "Function Name"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws)

    func_id_col   = len(headers) - 1
    func_name_col = len(headers)

    for item in hierarchy:
        row = ws.max_row + 1
        if item['type'] == 'part':
            col = 1 + item['level'] * 2
            ws.cell(row=row, column=col,     value=item['id'])
            ws.cell(row=row, column=col + 1, value=from_pascal_case(item['name']))
        else:
            ws.cell(row=row, column=func_id_col,   value=item['id'])
            ws.cell(row=row, column=func_name_col, value=from_pascal_case(item['name']))

    _auto_adjust_columns(ws)
    print(f"  -> {len(hierarchy)} allocation items exported.")


# ============================================================================
# Functional Exchanges export  (Functions_generated.sysml -> worksheet)
# ============================================================================

def _parse_functional_exchanges(file_path: pathlib.Path) -> List[Dict]:
    """Parse a functions SysML file and return a list of functional exchange dicts."""
    with open(file_path, 'r', encoding='utf-8') as f:
        raw = f.read()

    try:
        _, content = _get_package_content(raw, file_path)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return []

    action_info: Dict = {}
    for m in re.finditer(r'action def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        ports: Dict = {}
        for pm in re.finditer(r'port (\w+)\s*:\s*\w+\s*\{', body):
            port_name = pm.group(1)
            port_body, _ = get_block(body, pm.end())
            ports[port_name] = {'id': extract_id(port_body), 'name': port_name}
        action_info[name] = {'id': extract_id(body), 'ports': ports}

    # Map item def ID -> item type name so we can resolve the flowing item name
    # (each connection def's ID equals the corresponding item def's ID)
    item_id_to_name: Dict[str, str] = {}
    for m in re.finditer(r'item def (\w+)\s*\{', content):
        item_name = m.group(1)
        body, _ = get_block(content, m.end())
        item_id = extract_id(body)
        if item_id:
            item_id_to_name[item_id] = item_name

    exchanges: List[Dict] = []
    for m in re.finditer(r'connection def (\w+)\s*\{', content):
        conn_name = m.group(1)
        body, _ = get_block(content, m.end())
        conn_id = extract_id(body)

        end_actions = re.findall(r'end action (\w+)\s*:\s*(\w+);', body)
        if len(end_actions) < 2:
            continue

        from_action_type = end_actions[0][1]
        to_action_type   = end_actions[1][1]

        iface_m = re.search(
            r'interface : \w+\s+connect\s+(\w+)\.(\w+)\s+to\s+(\w+)\.(\w+)', body
        )
        if not iface_m:
            continue

        # FE connect order: TARGET.to_port to SOURCE.from_port
        _, to_port_usage, _, from_port_usage = iface_m.groups()

        from_data = action_info.get(from_action_type, {})
        to_data   = action_info.get(to_action_type,   {})

        # The connection def ID equals the item def ID — use the item name as the exchange name
        exchange_name = item_id_to_name.get(conn_id, conn_name)

        exchanges.append({
            'exchange_id':      conn_id,
            'exchange_name':    exchange_name,
            'from_action_id':   from_data.get('id', ''),
            'from_action_name': from_action_type,
            'from_port_id':     from_data.get('ports', {}).get(from_port_usage, {}).get('id', ''),
            'from_port_name':   from_port_usage,
            'to_action_id':     to_data.get('id', ''),
            'to_action_name':   to_action_type,
            'to_port_id':       to_data.get('ports', {}).get(to_port_usage, {}).get('id', ''),
            'to_port_name':     to_port_usage,
        })

    return exchanges


def export_functional_exchanges(ws, functions_path: pathlib.Path) -> None:
    """Populate *ws* (Functional Exchanges worksheet) from Functions_generated.sysml."""
    exchanges = _parse_functional_exchanges(functions_path)

    headers = [
        "Function From ID", "Function From Name",
        "Function From Port ID", "Function From Port Name",
        "Functional Exchange ID", "Functional Exchange Name",
        "Function To ID", "Function To Name",
        "Function To Port ID", "Function To Port Name",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws)

    for ex in exchanges:
        row = ws.max_row + 1
        ws.cell(row=row, column=1,  value=ex['from_action_id'])
        ws.cell(row=row, column=2,  value=from_pascal_case(ex['from_action_name']))
        ws.cell(row=row, column=3,  value=ex['from_port_id'])
        ws.cell(row=row, column=4,  value=format_port_name(ex['from_port_name']))
        ws.cell(row=row, column=5,  value=ex['exchange_id'])
        ws.cell(row=row, column=6,  value=from_pascal_case(ex['exchange_name']))
        ws.cell(row=row, column=7,  value=ex['to_action_id'])
        ws.cell(row=row, column=8,  value=from_pascal_case(ex['to_action_name']))
        ws.cell(row=row, column=9,  value=ex['to_port_id'])
        ws.cell(row=row, column=10, value=format_port_name(ex['to_port_name']))

    _auto_adjust_columns(ws)
    print(f"  -> {len(exchanges)} functional exchanges exported.")


# ============================================================================
# Component Exchanges export  (Parts_generated.sysml -> worksheet)
# ============================================================================

def _determine_port_direction(port_def_body: str) -> str:
    """Return IN, OUT, or INOUT based on item directions declared in a port def body."""
    has_in  = bool(re.search(r'\bin\s+item\b',  port_def_body))
    has_out = bool(re.search(r'\bout\s+item\b', port_def_body))
    if has_in and has_out:
        return 'INOUT'
    if has_in:
        return 'IN'
    if has_out:
        return 'OUT'
    return ''


_OPPOSITE_DIRECTION = {'IN': 'OUT', 'OUT': 'IN', 'INOUT': 'INOUT'}


def _fill_missing_port_direction(conn: Dict) -> None:
    """Infer a missing port direction from the opposite side (delegation ports in Capella).

    If only one side has a direction, the other is set to its mirror:
      IN <-> OUT, INOUT <-> INOUT.
    Mutates *conn* in place.
    """
    frm = conn['from_port_direction']
    to  = conn['to_port_direction']
    if frm and not to:
        conn['to_port_direction']   = _OPPOSITE_DIRECTION.get(frm, '')
    elif to and not frm:
        conn['from_port_direction'] = _OPPOSITE_DIRECTION.get(to, '')


def _parse_component_exchanges(file_path: pathlib.Path) -> List[Dict]:
    """Parse a parts SysML file and return a list of component exchange dicts.

    Supports both old-style (ports and interface connects inside def bodies) and
    new-style (bare defs, ports in usage blocks, interface usages at package/container level).
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        raw = f.read()

    try:
        _, content = _get_package_content(raw, file_path)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return []

    # Build port def name -> direction mapping (same for both styles)
    port_def_directions: Dict[str, str] = {}
    for m in re.finditer(r'port def (\w+)\s*\{', content):
        port_type_name = m.group(1)
        body, _ = get_block(content, m.end())
        port_def_directions[port_type_name] = _determine_port_direction(body)

    # Collect part/item def IDs (needed in both styles for fallback)
    part_def_ids: Dict[str, str] = {}
    for m in re.finditer(r'(?:part|item) def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        part_def_ids[name] = extract_id(body)

    # --- Build part_port_map: type_name -> {port_usage: {id, part_id, direction}} ---
    # Old-style: ports live inside part/item def bodies
    part_port_map: Dict = {name: {} for name in part_def_ids}
    for m in re.finditer(r'(?:part|item) def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        part_id = part_def_ids[name]
        for pm in re.finditer(r'port (\w+)\s*:\s*(\w+)\s*\{', body):
            port_name = pm.group(1)
            port_type = pm.group(2)
            port_body, _ = get_block(body, pm.end())
            part_port_map[name][port_name] = {
                'id':        extract_id(port_body),
                'part_id':   part_id,
                'direction': port_def_directions.get(port_type, ''),
            }

    # New-style: if no ports found in defs, scan usage blocks instead
    if not any(ports for ports in part_port_map.values()):
        for m in re.finditer(r'(?:part|item)\s+\w+\s*:\s*(\w+)\s*\{', content):
            type_name = m.group(1)
            if type_name not in part_def_ids:
                continue
            body, _ = get_block(content, m.end())
            part_id = part_def_ids[type_name]
            for pm in re.finditer(r'port\s+(\w+)\s*:\s*(\w+)\s*\{', body):
                port_name = pm.group(1)
                port_type = pm.group(2)
                port_body, _ = get_block(body, pm.end())
                if port_name not in part_port_map[type_name]:
                    part_port_map[type_name][port_name] = {
                        'id':        extract_id(port_body),
                        'part_id':   part_id,
                        'direction': port_def_directions.get(port_type, ''),
                    }

    # Build a global usage-name → type-name map for resolving interface connect paths
    usage_to_type = _build_usage_to_type_map(content)

    connections: List[Dict] = []
    for m in re.finditer(r'connection def (\w+)\s*\{', content):
        conn_name = m.group(1)
        body, _ = get_block(content, m.end())
        conn_id = extract_id(body)

        end_parts = re.findall(r'end (?:part|item) (\w+)\s*:\s*(\w+);', body)
        if len(end_parts) < 2:
            continue

        from_part_type = end_parts[0][1]
        to_part_type   = end_parts[1][1]

        # Old-style: interface connect is embedded inside the connection def body
        iface_m = re.search(
            r'interface : \w+\s+connect\s+(\w+)\.(\w+)\s+to\s+(\w+)\.(\w+)', body
        )
        if iface_m:
            _, from_port_usage, _, to_port_usage = iface_m.groups()
            # from/to part types stay as declared in end_parts
        else:
            # New-style: find the separate interface usage statement for this connection
            iface_name = f"{conn_name}_Interface"
            iface_pat = (
                rf'interface\s*:\s*{re.escape(iface_name)}'
                rf'\s+connect\s+(\S+?)\s+to\s+(\S+?)\s*;'
            )
            iu_m = re.search(iface_pat, content)
            if not iu_m:
                continue
            from_path_parts = iu_m.group(1).split('.')
            to_path_parts   = iu_m.group(2).split('.')
            from_port_usage = from_path_parts[-1]
            to_port_usage   = to_path_parts[-1]
            # The part usage is the second-to-last segment of the path
            from_uname = from_path_parts[-2] if len(from_path_parts) >= 2 else from_path_parts[0]
            to_uname   = to_path_parts[-2]   if len(to_path_parts)   >= 2 else to_path_parts[0]
            from_part_type = usage_to_type.get(from_uname, from_part_type)
            to_part_type   = usage_to_type.get(to_uname,   to_part_type)

        from_info = part_port_map.get(from_part_type, {}).get(from_port_usage, {})
        to_info   = part_port_map.get(to_part_type,   {}).get(to_port_usage,   {})

        connections.append({
            'conn_id':             conn_id,
            'conn_name':           conn_name,
            'from_part_id':        from_info.get('part_id', part_def_ids.get(from_part_type, '')),
            'from_part_name':      from_part_type,
            'from_port_id':        from_info.get('id', ''),
            'from_port_name':      from_port_usage,
            'from_port_direction': from_info.get('direction', ''),
            'to_part_id':          to_info.get('part_id', part_def_ids.get(to_part_type, '')),
            'to_part_name':        to_part_type,
            'to_port_id':          to_info.get('id', ''),
            'to_port_name':        to_port_usage,
            'to_port_direction':   to_info.get('direction', ''),
        })

    for conn in connections:
        _fill_missing_port_direction(conn)

    return connections


def export_component_exchanges(ws, parts_path: pathlib.Path) -> None:
    """Populate *ws* (Component Exchanges worksheet) from Parts_generated.sysml."""
    connections = _parse_component_exchanges(parts_path)

    headers = [
        "Component From ID", "Component From Name",
        "Component From Port ID", "Component From Port Name",
        "Component From Port Direction", "Component From Port Kind",
        "Component Exchange ID", "Component Exchange Name", "Component Exchange Kind",
        "Component To ID", "Component To Name",
        "Component To Port ID", "Component To Port Name",
        "Component To Port Direction", "Component To Port Kind",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws)

    for conn in connections:
        row = ws.max_row + 1
        ws.cell(row=row, column=1,  value=conn['from_part_id'])
        ws.cell(row=row, column=2,  value=from_pascal_case(conn['from_part_name']))
        ws.cell(row=row, column=3,  value=conn['from_port_id'])
        ws.cell(row=row, column=4,  value=format_port_name(conn['from_port_name']))
        ws.cell(row=row, column=5,  value=conn['from_port_direction'])
        ws.cell(row=row, column=6,  value='FLOW')  # Kind — hard coded as FLOW as it is the standard in Capella
        ws.cell(row=row, column=7,  value=conn['conn_id'])
        ws.cell(row=row, column=8,  value=format_exchange_name(conn['conn_name']))
        ws.cell(row=row, column=9,  value='')  # Exchange Kind — not stored in SysML
        ws.cell(row=row, column=10, value=conn['to_part_id'])
        ws.cell(row=row, column=11, value=from_pascal_case(conn['to_part_name']))
        ws.cell(row=row, column=12, value=conn['to_port_id'])
        ws.cell(row=row, column=13, value=format_port_name(conn['to_port_name']))
        ws.cell(row=row, column=14, value=conn['to_port_direction'])
        ws.cell(row=row, column=15, value='FLOW')  # Kind — hard coded as FLOW as it is the standard in Capella

    _auto_adjust_columns(ws)
    print(f"  -> {len(connections)} component exchanges exported.")


# ============================================================================
# Exchange Allocations export  (Parts + Functions -> "Link Exchanges" worksheet)
# ============================================================================

def _parse_parts_for_exchange_allocations(file_path: pathlib.Path):
    """Parse a parts SysML file for exchange-allocation cross-referencing.

    Supports both old-style (ports in def bodies, interface connect in connection def) and
    new-style (bare defs, ports in usage blocks, interface connects as separate usages).

    Returns (part_defs, ce_connections, interface_defs).
    """
    raw = file_path.read_text(encoding='utf-8')
    try:
        _, content = _get_package_content(raw, file_path)
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    # --- Part defs: collect IDs and ports ---
    # Old-style: ports live inside def bodies
    part_defs: Dict = {}
    for m in re.finditer(r'(?:part|item) def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        ports: Dict = {}
        for pm in re.finditer(r'port (\w+)\s*:\s*\w+\s*\{', body):
            port_name = pm.group(1)
            port_body, _ = get_block(body, pm.end())
            ports[port_name] = {'id': extract_id(port_body)}
        part_defs[name] = {'id': extract_id(body), 'ports': ports}

    # New-style: if no ports found in defs, scan usage blocks for port IDs
    if not any(pd['ports'] for pd in part_defs.values()):
        for m in re.finditer(r'(?:part|item)\s+\w+\s*:\s*(\w+)\s*\{', content):
            type_name = m.group(1)
            if type_name not in part_defs:
                continue
            body, _ = get_block(content, m.end())
            for pm in re.finditer(r'port\s+(\w+)\s*:\s*\w+\s*\{', body):
                port_name = pm.group(1)
                port_body, _ = get_block(body, pm.end())
                if port_name not in part_defs[type_name]['ports']:
                    part_defs[type_name]['ports'][port_name] = {'id': extract_id(port_body)}

    # Global usage-name → type-name map for resolving interface connect paths
    usage_to_type = _build_usage_to_type_map(content)

    # --- Connection defs ---
    ce_connections: Dict = {}
    for m in re.finditer(r'connection def (\w+)\s*\{', content):
        ce_name = m.group(1)
        body, _ = get_block(content, m.end())
        ce_id = extract_id(body)
        end_parts = re.findall(r'end (?:part|item) (\w+)\s*:\s*(\w+);', body)
        if len(end_parts) < 2:
            continue

        # Old-style: interface connect embedded in the connection def body
        iface_m = re.search(
            r'interface\s*:\s*\w+\s+connect\s+(\w+)\.(\w+)\s+to\s+(\w+)\.(\w+)', body
        )
        if iface_m:
            from_part_usage, from_port_usage, to_part_usage, to_port_usage = iface_m.groups()
            local_map = {u: t for u, t in end_parts}
            ce_connections[ce_name] = {
                'id':              ce_id,
                'from_part_type':  local_map.get(from_part_usage, ''),
                'to_part_type':    local_map.get(to_part_usage, ''),
                'from_port_usage': from_port_usage,
                'to_port_usage':   to_port_usage,
            }
        else:
            # New-style: find the separate interface usage statement
            iface_name = f"{ce_name}_Interface"
            iface_pat = (
                rf'interface\s*:\s*{re.escape(iface_name)}'
                rf'\s+connect\s+(\S+?)\s+to\s+(\S+?)\s*;'
            )
            iu_m = re.search(iface_pat, content)
            if not iu_m:
                continue
            from_path_parts = iu_m.group(1).split('.')
            to_path_parts   = iu_m.group(2).split('.')
            from_port_usage = from_path_parts[-1]
            to_port_usage   = to_path_parts[-1]
            from_uname = from_path_parts[-2] if len(from_path_parts) >= 2 else from_path_parts[0]
            to_uname   = to_path_parts[-2]   if len(to_path_parts)   >= 2 else to_path_parts[0]
            ce_connections[ce_name] = {
                'id':              ce_id,
                'from_part_type':  usage_to_type.get(from_uname, end_parts[0][1]),
                'to_part_type':    usage_to_type.get(to_uname,   end_parts[1][1]),
                'from_port_usage': from_port_usage,
                'to_port_usage':   to_port_usage,
            }

    # --- Interface defs: collect flow-of items and allocation IDs ---
    interface_defs: Dict = {}
    _flow_re = re.compile(
        r'flow\s+of\s+(\w+)\s+from\s+\w+\.\w+\s+to\s+\w+\.\w+\s*(?:;|\{([^}]*)\})',
        re.DOTALL,
    )
    for m in re.finditer(r'interface def (\w+)\s*\{', content):
        iface_name = m.group(1)
        body, _ = get_block(content, m.end())
        flows = []
        for fm in _flow_re.finditer(body):
            fe_type = fm.group(1)
            block_content = fm.group(2) or ''
            flows.append({'type': fe_type, 'allocation_id': extract_id(block_content)})
        interface_defs[iface_name] = {'flows': flows}

    return part_defs, ce_connections, interface_defs


def _parse_functions_for_exchange_allocations(file_path: pathlib.Path):
    """Parse a functions SysML file for exchange-allocation cross-referencing.

    Returns (action_defs, item_defs, fe_conns_by_id).
    """
    raw = file_path.read_text(encoding='utf-8')
    try:
        _, content = _get_package_content(raw, file_path)
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    action_defs: Dict = {}
    for m in re.finditer(r'action def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        ports: Dict = {}
        for pm in re.finditer(r'port (\w+)\s*:\s*\w+\s*\{', body):
            port_name = pm.group(1)
            port_body, _ = get_block(body, pm.end())
            ports[port_name] = {'id': extract_id(port_body)}
        action_defs[name] = {'id': extract_id(body), 'ports': ports}

    item_defs: Dict = {}
    for m in re.finditer(r'item def (\w+)\s*\{', content):
        name = m.group(1)
        body, _ = get_block(content, m.end())
        item_defs[name] = {'id': extract_id(body)}

    fe_conns_by_id: Dict = {}
    for m in re.finditer(r'connection def (\w+)\s*\{', content):
        body, _ = get_block(content, m.end())
        conn_id = extract_id(body)
        if not conn_id:
            continue
        end_actions = re.findall(r'end action (\w+)\s*:\s*(\w+);', body)
        if len(end_actions) < 2:
            continue
        iface_m = re.search(
            r'interface\s*:\s*\w+\s+connect\s+(\w+)\.(\w+)\s+to\s+(\w+)\.(\w+)', body
        )
        if not iface_m:
            continue
        # FE connect order: TARGET.to_port to SOURCE.from_port
        to_action_usage, to_port_usage, from_action_usage, from_port_usage = iface_m.groups()
        usage_to_type = {u: t for u, t in end_actions}
        fe_conns_by_id[conn_id] = {
            'from_action_type': usage_to_type.get(from_action_usage, ''),
            'to_action_type':   usage_to_type.get(to_action_usage,   ''),
            'from_port_usage':  from_port_usage,
            'to_port_usage':    to_port_usage,
        }

    return action_defs, item_defs, fe_conns_by_id


def _build_exchange_allocation_rows(
    part_defs, ce_connections, interface_defs,
    action_defs, item_defs, fe_conns_by_id,
) -> List[Dict]:
    """Cross-reference parsed data and return one dict per exchange allocation row."""
    rows: List[Dict] = []

    for iface_name, iface in interface_defs.items():
        if not iface['flows'] or not iface_name.endswith('_Interface'):
            continue

        ce_name = iface_name[:-len('_Interface')]
        ce = ce_connections.get(ce_name)
        if not ce:
            continue

        from_part = part_defs.get(ce['from_part_type'], {})
        to_part   = part_defs.get(ce['to_part_type'],   {})

        src_component_id   = from_part.get('id', '')
        src_component_name = from_pascal_case(ce['from_part_type'])
        src_port_id        = from_part.get('ports', {}).get(ce['from_port_usage'], {}).get('id', '')
        src_port_name      = format_port_name(ce['from_port_usage'])

        tgt_component_id   = to_part.get('id', '')
        tgt_component_name = from_pascal_case(ce['to_part_type'])
        tgt_port_id        = to_part.get('ports', {}).get(ce['to_port_usage'], {}).get('id', '')
        tgt_port_name      = format_port_name(ce['to_port_usage'])

        for flow in iface['flows']:
            fe_type       = flow['type']
            allocation_id = flow['allocation_id']
            fe_id   = item_defs.get(fe_type, {}).get('id', '')
            fe_name = from_pascal_case(fe_type)

            fe_conn          = fe_conns_by_id.get(fe_id, {})
            from_action_type = fe_conn.get('from_action_type', '')
            to_action_type   = fe_conn.get('to_action_type',   '')
            fe_from_port     = fe_conn.get('from_port_usage',  '')
            fe_to_port       = fe_conn.get('to_port_usage',    '')

            from_action = action_defs.get(from_action_type, {})
            to_action   = action_defs.get(to_action_type,   {})

            rows.append({
                'ce_id':             ce['id'],
                'ce_name':           format_exchange_name(ce_name),
                'src_port_id':       src_port_id,
                'src_port_name':     src_port_name,
                'src_component_id':  src_component_id,
                'src_component_name': src_component_name,
                'tgt_port_id':       tgt_port_id,
                'tgt_port_name':     tgt_port_name,
                'tgt_component_id':  tgt_component_id,
                'tgt_component_name': tgt_component_name,
                'fe_id':             fe_id,
                'fe_name':           fe_name,
                'fe_src_port_id':    from_action.get('ports', {}).get(fe_from_port, {}).get('id', ''),
                'fe_src_port_name':  format_port_name(fe_from_port) if fe_from_port else '',
                'fe_tgt_port_id':    to_action.get('ports', {}).get(fe_to_port, {}).get('id', ''),
                'fe_tgt_port_name':  format_port_name(fe_to_port) if fe_to_port else '',
                'allocation_id':     allocation_id,
                'src_func_id':       from_action.get('id', ''),
                'tgt_func_id':       to_action.get('id', ''),
            })

    return rows


def export_exchange_allocations(ws, parts_path: pathlib.Path, functions_path: pathlib.Path) -> None:
    """Populate *ws* (Link Exchanges worksheet) from Parts + Functions SysML files."""
    part_defs, ce_connections, interface_defs = _parse_parts_for_exchange_allocations(parts_path)
    action_defs, item_defs, fe_conns_by_id   = _parse_functions_for_exchange_allocations(functions_path)
    rows = _build_exchange_allocation_rows(
        part_defs, ce_connections, interface_defs,
        action_defs, item_defs, fe_conns_by_id,
    )

    headers = [
        "Component Exchange ID", "Component Exchange Name",
        "Component Source Port ID", "Component Source Port Name",
        "Source Component ID", "Source Component Name",
        "Component Target Port ID", "Component Target Port Name",
        "Target Component ID", "Target Component Name",
        "Functional Exchange ID", "Functional Exchange Name",
        "Functional Source Port ID", "Functional Source Port Name",
        "Functional Target Port ID", "Functional Target Port Name",
        "Allocation ID",
        "Source Function ID", "Target Function ID",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws)

    for row_data in rows:
        r = ws.max_row + 1
        ws.cell(r,  1, row_data['ce_id'])
        ws.cell(r,  2, row_data['ce_name'])
        ws.cell(r,  3, row_data['src_port_id'])
        ws.cell(r,  4, row_data['src_port_name'])
        ws.cell(r,  5, row_data['src_component_id'])
        ws.cell(r,  6, row_data['src_component_name'])
        ws.cell(r,  7, row_data['tgt_port_id'])
        ws.cell(r,  8, row_data['tgt_port_name'])
        ws.cell(r,  9, row_data['tgt_component_id'])
        ws.cell(r, 10, row_data['tgt_component_name'])
        ws.cell(r, 11, row_data['fe_id'])
        ws.cell(r, 12, row_data['fe_name'])
        ws.cell(r, 13, row_data['fe_src_port_id'])
        ws.cell(r, 14, row_data['fe_src_port_name'])
        ws.cell(r, 15, row_data['fe_tgt_port_id'])
        ws.cell(r, 16, row_data['fe_tgt_port_name'])
        ws.cell(r, 17, row_data['allocation_id'])
        ws.cell(r, 18, row_data['src_func_id'])
        ws.cell(r, 19, row_data['tgt_func_id'])

    _auto_adjust_columns(ws)
    print(f"  -> {len(rows)} exchange allocations exported.")


# ============================================================================
# Functional Chains export  (Functions_generated.sysml -> worksheet)
# ============================================================================

@dataclass
class FunctionalChain:
    chain_id: str
    chain_type: str
    function_types: List[str] = field(default_factory=list)
    exchanges: List[Dict] = field(default_factory=list)


def _parse_chain_actions_from_body(body: str) -> List[Dict]:
    """Parse a chain action def body into ordered action items.

    Returns list of {'type': action_type_name, 'ctx': 'main'|'if'|'else'}.
    Handles the generated if/else/fork pattern.
    """
    items: List[Dict] = []
    state = 'main'  # 'main', 'if', 'pending_else', 'else'

    for raw_line in body.split('\n'):
        s = raw_line.strip()
        if not s:
            continue

        if re.match(r'^then if .+\{$', s):
            state = 'if'
        elif s == '}':
            if state == 'if':
                state = 'pending_else'
            elif state == 'else':
                state = 'main'
        elif re.match(r'^else\s*\{', s) and state == 'pending_else':
            state = 'else'
        else:
            m = re.match(r'then action \w+\s*:\s*(\w+)\s*;', s)
            if m:
                ctx = 'if' if state == 'if' else ('else' if state == 'else' else 'main')
                items.append({'type': m.group(1), 'ctx': ctx})

    return items


def _build_chain_edges(items: List[Dict]) -> List[Tuple[str, str]]:
    """Build (source_type, target_type) pairs from an ordered action item list.

    Handles linear sequences and branches (if/else groups).
    The "active tails" — the last action(s) before the next segment — are tracked
    so that branch convergences produce edges from both branches to the join node.
    """
    edges: List[Tuple[str, str]] = []
    last_mains: List[str] = []
    i = 0
    n = len(items)

    while i < n:
        ctx = items[i]['ctx']

        if ctx == 'main':
            for last in last_mains:
                edges.append((last, items[i]['type']))
            last_mains = [items[i]['type']]
            i += 1

        elif ctx == 'if':
            # Collect the full if-branch then the full else-branch
            if_items: List[str] = []
            while i < n and items[i]['ctx'] == 'if':
                if_items.append(items[i]['type'])
                i += 1
            else_items: List[str] = []
            while i < n and items[i]['ctx'] == 'else':
                else_items.append(items[i]['type'])
                i += 1

            # Edges from each active tail to the first item of each branch
            for last in last_mains:
                if if_items:
                    edges.append((last, if_items[0]))
                if else_items:
                    edges.append((last, else_items[0]))

            # Edges within each branch (for multi-step branches)
            for k in range(len(if_items) - 1):
                edges.append((if_items[k], if_items[k + 1]))
            for k in range(len(else_items) - 1):
                edges.append((else_items[k], else_items[k + 1]))

            # New active tails are the last item of each branch
            last_mains = []
            if if_items:
                last_mains.append(if_items[-1])
            if else_items:
                last_mains.append(else_items[-1])

        else:
            i += 1

    return edges


def _parse_functional_chains(file_path: pathlib.Path):
    """Parse chains from a functions SysML file and return (chains, action_ids).

    For each chain, the exchanges list is built by:
      1. Parsing the chain body for the ordered action sequence (incl. branches).
      2. Deriving source→target edges from that sequence.
      3. Looking up the connection def for each edge to get the exchange ID/name.
    """
    raw = file_path.read_text(encoding='utf-8')
    try:
        _, content = _get_package_content(raw, file_path)
    except ValueError as exc:
        raise ValueError(str(exc)) from exc

    # action_ids: type_name → id  (regular, non-chain action defs)
    action_ids: Dict[str, str] = {}
    chain_data: List[Tuple[str, str, str]] = []  # (type_name, chain_id, body)

    for m in re.finditer(r'action def (\w+)\s*\{', content):
        type_name = m.group(1)
        body, _ = get_block(content, m.end())
        block_id = extract_id(body)
        if 'first start;' in body:
            chain_data.append((type_name, block_id, body))
        else:
            if block_id:
                action_ids[type_name] = block_id

    # item_id_to_name: exchange_id → human-readable exchange name
    item_id_to_name: Dict[str, str] = {}
    for m in re.finditer(r'item def (\w+)\s*\{', content):
        type_name = m.group(1)
        body, _ = get_block(content, m.end())
        block_id = extract_id(body)
        if block_id:
            item_id_to_name[block_id] = from_pascal_case(type_name)

    # conn_lookup: (from_action_type, to_action_type) → exchange_id
    # The connection def's /* ID: ... */ equals the item def ID for that exchange.
    conn_lookup: Dict[Tuple[str, str], str] = {}
    for m in re.finditer(r'connection def (\w+)\s*\{', content):
        body, _ = get_block(content, m.end())
        conn_id = extract_id(body)
        if not conn_id:
            continue
        end_actions = re.findall(r'end action (\w+)\s*:\s*(\w+);', body)
        if len(end_actions) < 2:
            continue
        # Generated order: first end action is the source (Out), second is target (In)
        from_type = end_actions[0][1]
        to_type = end_actions[1][1]
        conn_lookup[(from_type, to_type)] = conn_id

    # Build FunctionalChain objects
    chains: List[FunctionalChain] = []
    for type_name, chain_id, body in chain_data:
        items = _parse_chain_actions_from_body(body)
        function_types = [item['type'] for item in items]
        edges = _build_chain_edges(items)

        exchange_list: List[Dict] = []
        for src_type, tgt_type in edges:
            ex_id = conn_lookup.get((src_type, tgt_type), '')
            exchange_list.append({
                'exchange_id':          ex_id,
                'exchange_name':        item_id_to_name.get(ex_id, '') if ex_id else '',
                'involvement_id':       '',  # not stored in SysML
                'source_function_id':   action_ids.get(src_type, ''),
                'source_function_name': from_pascal_case(src_type),
                'target_function_id':   action_ids.get(tgt_type, ''),
                'target_function_name': from_pascal_case(tgt_type),
            })

        chains.append(FunctionalChain(
            chain_id=chain_id,
            chain_type=type_name,
            function_types=function_types,
            exchanges=exchange_list,
        ))

    return chains, action_ids


def export_functional_chains(ws, functions_path: pathlib.Path) -> None:
    """Populate *ws* (Functional Chains worksheet) from Functions_generated.sysml."""
    chains, action_ids = _parse_functional_chains(functions_path)

    max_functions = max((len(c.function_types) for c in chains), default=0)
    max_exchanges = max((len(c.exchanges)       for c in chains), default=0)

    headers = ["Functional Chain ID", "Functional Chain Name"]
    for n in range(1, max_functions + 1):
        headers += [f"Function {n} ID", f"Function {n} Name", f"Function {n} Involvement ID"]
    for n in range(1, max_exchanges + 1):
        headers += [
            f"Exchange {n} ID",
            f"Exchange {n} Name",
            f"Exchange {n} Involvement ID",
            f"Exchange {n} Source Function ID",
            f"Exchange {n} Source Function Name",
            f"Exchange {n} Target Function ID",
            f"Exchange {n} Target Function Name",
        ]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws)

    for chain in chains:
        func_ids   = [action_ids.get(t, '') for t in chain.function_types]
        func_names = [from_pascal_case(t)   for t in chain.function_types]

        row_values = [chain.chain_id, from_pascal_case(chain.chain_type)]

        for i in range(max_functions):
            row_values += [
                func_ids[i]   if i < len(func_ids)   else '',
                func_names[i] if i < len(func_names) else '',
                '',  # Involvement ID — not stored in SysML
            ]

        for i in range(max_exchanges):
            if i < len(chain.exchanges):
                ex = chain.exchanges[i]
                row_values += [
                    ex['exchange_id'],
                    ex['exchange_name'],
                    ex['involvement_id'],
                    ex['source_function_id'],
                    ex['source_function_name'],
                    ex['target_function_id'],
                    ex['target_function_name'],
                ]
            else:
                row_values += ['', '', '', '', '', '', '']

        r = ws.max_row + 1
        for col, value in enumerate(row_values, 1):
            ws.cell(r, col, value)

    _auto_adjust_columns(ws)
    print(f"  -> {len(chains)} functional chains exported.")


# ============================================================================
# full_export  — creates the complete multi-sheet Excel file
# ============================================================================

def full_export(parts_filename: str, functions_filename: str, excel_name: str) -> None:
    """Create a complete Excel export with all worksheets.

    Args:
        parts_filename:     Filename of the Parts SysML file.  Relative paths are
                            resolved against the DVS directory (parent of this script).
        functions_filename: Filename of the Functions SysML file (same resolution).
        excel_name:         Base name for the output file.  The workbook is saved as
                            ``<excel_name>_complete_export.xlsx`` inside DVS/results/.

    Worksheet order
    ---------------
      Systems                    <- export_systems
      Functions                  <- export_functions
      Link Systems and Functions <- export_allocations
      Functional Exchanges       <- export_functional_exchanges
      Component Exchanges        <- export_component_exchanges
      Link Exchanges             <- export_exchange_allocations
      Functional Chains          <- export_functional_chains
      Capabilities               <- empty worksheet

    Error handling
    --------------
    If any step raises an exception the error is printed, the exception is
    re-raised, and the Excel file is NOT saved.
    """
    _FOLDER_DIR    = pathlib.Path(__file__).parent.parent
    results_dir = _FOLDER_DIR / "results"

    def _resolve(filename: str) -> pathlib.Path:
        p = pathlib.Path(filename)
        return p if p.is_absolute() else _FOLDER_DIR / filename

    if not parts_filename and not functions_filename:
        print("Error: No file paths provided. At least one of parts_filename or functions_filename must be specified.", file=sys.stderr)
        raise ValueError("No file paths provided for export.")

    if not parts_filename:
        print(f"Note: No parts file specified; using functions file '{functions_filename}' for all sheets.")
        parts_path = functions_path = _resolve(functions_filename)
    elif not functions_filename:
        print(f"Note: No functions file specified; using parts file '{parts_filename}' for all sheets.")
        parts_path = functions_path = _resolve(parts_filename)
    else:
        parts_path     = _resolve(parts_filename)
        functions_path = _resolve(functions_filename)

    output_path    = results_dir / f"{excel_name}_complete_export.xlsx"

    paths_to_check = [(parts_path, "Parts"), (functions_path, "Functions")]
    if parts_path == functions_path:
        paths_to_check = [(parts_path, "Input")]
    for path, label in paths_to_check:
        if not path.exists():
            print(f"Error: {label} file not found: {path}", file=sys.stderr)
            raise FileNotFoundError(f"{label} file not found: {path}")

    results_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    if wb.active.title == "Sheet":
        wb.remove(wb.active)

    # (sheet name, export function or None for empty, positional args after ws)
    worksheet_steps = [
        ("Systems",                    export_systems,              (parts_path,)),
        ("Functions",                  export_functions,            (functions_path,)),
        ("Link Systems and Functions", export_allocations,          (parts_path, functions_path)),
        ("Functional Exchanges",       export_functional_exchanges, (functions_path,)),
        ("Component Exchanges",        export_component_exchanges,  (parts_path,)),
        ("Link Exchanges",             export_exchange_allocations, (parts_path, functions_path)),
        ("Functional Chains",          export_functional_chains,    (functions_path,)),
        ("Capabilities",               None,                        ()),
    ]

    for sheet_name, fn, args in worksheet_steps:
        ws = wb.create_sheet(sheet_name)
        if fn is None:
            print(f"  '{sheet_name}' — empty worksheet created.")
            continue
        try:
            print(f"Exporting '{sheet_name}'...")
            fn(ws, *args)
        except Exception as exc:
            print(
                f"\nError while exporting worksheet '{sheet_name}': {exc}\n"
                f"Export aborted. The Excel file has NOT been saved.",
                file=sys.stderr,
            )
            raise

    wb.save(output_path)
    print(f"\nComplete export saved to: {output_path}")
