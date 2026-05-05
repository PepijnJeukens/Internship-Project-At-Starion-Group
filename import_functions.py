#!/usr/bin/env python3
"""
import_functions.py
Combined import library for all DVS Excel -> SysML v2 import operations.

Public API
----------
import_systems(rows, output_path, package_name="")
import_functions(rows, output_path, package_name="")
import_allocations(rows, parts_path, functions_path, functions_pkg="")
import_functional_exchanges(rows, functions_path)
import_component_exchanges(rows, parts_path)
import_exchange_allocations(rows, parts_path)
import_functional_chains(rows, functions_path)
full_import(excel_file, parts_filename, functions_filename,
            parts_package="", functions_package="")

All import_* functions accept 'rows' as a list[tuple] from a worksheet
(first row is the header).  full_import opens a single multi-sheet workbook
and dispatches each sheet to the appropriate function.
"""

"""
Import 
"""


import csv
import pathlib
import re
import subprocess
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install it with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

INDENT = "    "
ALLOC_MARKER = "/* exchange allocation items */"

# ---------------------------------------------------------------------------
# Shared name-conversion helpers
# ---------------------------------------------------------------------------

def to_type_name(name: str, item_id: str = "") -> str:
    """PascalCase SysML type identifier.  ALL-CAPS abbreviations are preserved.
    Appends first 4 characters of item_id when provided to ensure uniqueness.
    """
    words = re.split(r"[\s/\-_]+", name.strip())
    result = []
    for word in words:
        cleaned = re.sub(r"[^A-Za-z0-9]", "", word)
        if not cleaned:
            continue
        if cleaned.isupper() and len(cleaned) > 1:
            result.append(cleaned)
        else:
            result.append(cleaned[0].upper() + cleaned[1:])
    type_name = "".join(result)
    if item_id:
        return f"{type_name}_{item_id[:4]}"
    return type_name


def to_usage_name(name: str, item_id: str = "") -> str:
    """camelCase SysML usage identifier.
    Initial ALL-CAPS abbreviations are lowercased as a group.
    Appends first 4 characters of item_id when provided to ensure uniqueness.
    """
    type_name = to_type_name(name, item_id)
    if not type_name:
        return name

    underscore_pos = type_name.find("_")
    if underscore_pos != -1:
        name_part = type_name[:underscore_pos]
        id_part = type_name[underscore_pos:]
    else:
        name_part = type_name
        id_part = ""

    n = len(name_part)
    end_prefix = 0
    for i in range(n):
        ch = name_part[i]
        if ch.isupper():
            next_is_lower = i + 1 < n and name_part[i + 1].islower()
            if next_is_lower and i > 0:
                end_prefix = i
                break
            else:
                end_prefix = i + 1
        else:
            break

    if end_prefix == 0:
        end_prefix = 1

    return name_part[:end_prefix].lower() + name_part[end_prefix:] + id_part


def to_exchange_name(raw_name: str) -> str:
    """Convert a component exchange name to a SysML identifier.
    Replaces '-' with '_', removes spaces, collapses consecutive '_'.
    """
    name = raw_name.replace("-", "_").replace(" ", "")
    return re.sub(r"_+", "_", name).strip("_")


def to_port_usage_name(port_name: str, port_id: str) -> str:
    """Port usage name: stripped port name + '_' + first 4 chars of ID."""
    return f"{port_name.replace(' ', '')}_{port_id[:4]}"


def to_component_suffix(raw_name: str) -> str:
    """PascalCase suffix derived from a raw component name (e.g. 'Ground station' -> 'GroundStation')."""
    words = re.split(r"[\s/\-_]+", raw_name.strip())
    result = []
    for word in words:
        cleaned = re.sub(r"[^A-Za-z0-9]", "", word)
        if not cleaned:
            continue
        if cleaned.isupper() and len(cleaned) > 1:
            result.append(cleaned)
        else:
            result.append(cleaned[0].upper() + cleaned[1:])
    return "".join(result)


def extract_id_prefix_from_name(name: str) -> str:
    """Return the 4-char hex ID prefix appended to a name (e.g. 'EPS_cc31' -> 'cc31'), or ''."""
    pos = name.rfind("_")
    if pos != -1:
        suffix = name[pos + 1:]
        if len(suffix) == 4 and all(c in "0123456789abcdef" for c in suffix.lower()):
            return suffix
    return ""


def _stem_to_package(path) -> str:
    """Derive a SysML package name from a file stem."""
    return to_type_name(pathlib.Path(path).stem)


# ---------------------------------------------------------------------------
# Shared SysML-manipulation helpers
# ---------------------------------------------------------------------------

def _find_block_end(content: str, open_brace_pos: int) -> int:
    """Index of the '}' that closes the '{' at open_brace_pos."""
    depth = 0
    for i in range(open_brace_pos, len(content)):
        if content[i] == "{":
            depth += 1
        elif content[i] == "}":
            depth -= 1
            if depth == 0:
                return i
    return -1


def validate(path: pathlib.Path, check_dir: bool = False) -> bool:
    """Run syside check on path (or its parent directory when check_dir=True)."""
    target = str(path.parent if check_dir else path)
    try:
        proc = subprocess.run(
            ["syside", "check", target],
            capture_output=True, text=True,
        )
        output = proc.stdout + proc.stderr
        if proc.returncode != 0 or "error" in output.lower():
            print("Validation errors:")
            print(output)
            return False
        print("Validation passed.")
        return True
    except FileNotFoundError:
        print("syside not found — skipping validation.")
        return True


# ---------------------------------------------------------------------------
# Shared header-detection helper
# ---------------------------------------------------------------------------

def _detect_id_name_pairs(header: tuple) -> List[Tuple[int, int]]:
    """Scan a header row and return (id_col, name_col) index pairs.
    Works regardless of extra columns between each ID/Name pair.
    """
    h = [str(c).lower().strip() if c else "" for c in header]
    pairs: List[Tuple[int, int]] = []
    for id_col, v in enumerate(h):
        if "id" in v and v:
            for name_col in range(id_col + 1, len(h)):
                if "name" in h[name_col]:
                    pairs.append((id_col, name_col))
                    break
    return pairs


# ---------------------------------------------------------------------------
# Shared action/part look-up helpers
# ---------------------------------------------------------------------------

def get_action_id_to_name_map(content: str) -> Dict[str, str]:
    """Return {action_id: type_name} from SysML text by scanning action def bodies."""
    id_to_name: Dict[str, str] = {}
    for m in re.finditer(r"action def (\w+)\s*\{([^}]*)\}", content, re.DOTALL):
        id_match = re.search(r"/\*\s*ID:\s*([\w\-]+)\s*\*/", m.group(2))
        if id_match:
            id_to_name[id_match.group(1)] = m.group(1)
    return id_to_name


def get_all_part_defs(content: str) -> Dict[str, str]:
    """Return {type_name: full_definition} for all part defs in SysML text."""
    return {
        m.group(1): m.group(0)
        for m in re.finditer(r"part def (\w+)\s*\{([^}]*)\}", content, re.DOTALL)
    }


def get_part_id_to_name_map(content: str) -> Dict[str, str]:
    """Return {part_id: type_name} from SysML text by scanning part def bodies."""
    id_to_name: Dict[str, str] = {}
    for m in re.finditer(r"part def (\w+)\s*\{([^}]*)\}", content, re.DOTALL):
        id_match = re.search(r"/\*\s*ID:\s*([\w\-]+)\s*\*/", m.group(2))
        if id_match:
            id_to_name[id_match.group(1)] = m.group(1)
    return id_to_name


def parse_id_to_def_name(sysml_text: str) -> Dict[str, str]:
    """Return {action_id: def_name} using the tighter inline-ID regex pattern."""
    id_to_name: Dict[str, str] = {}
    for m in re.finditer(
        r"\baction def\s+(\w+)\s*\{[^{]*?/\* ID: ([\w\-]+) \*/",
        sysml_text,
        re.DOTALL,
    ):
        id_to_name[m.group(2)] = m.group(1)
    return id_to_name


def match_action_name(action_name: str, action_id: str, id_to_name_map: Dict[str, str]) -> str:
    """Find the actual SysML action def name for an Excel action entry.
    Tries exact match, then ID lookup, then base-name match, then computed name.
    """
    if action_name in id_to_name_map.values():
        return action_name
    if action_id in id_to_name_map:
        return id_to_name_map[action_id]

    base = action_name
    pos = action_name.rfind("_")
    if pos != -1:
        base = action_name[:pos]
    for tname in id_to_name_map.values():
        tbase = tname[:tname.rfind("_")] if tname.rfind("_") != -1 else tname
        if tbase == base:
            return tname

    return to_type_name(action_name, action_id)


def match_part_name(part_name: str, part_id: str, id_to_name_map: Dict[str, str]) -> str:
    """Find the actual SysML part def name for an Excel part entry.
    Tries exact match, then ID lookup, then base-name match, then computed name.
    """
    if part_name in id_to_name_map.values():
        return part_name
    if part_id in id_to_name_map:
        return id_to_name_map[part_id]

    base = part_name
    pos = part_name.rfind("_")
    if pos != -1:
        base = part_name[:pos]
    for tname in id_to_name_map.values():
        tbase = tname[:tname.rfind("_")] if tname.rfind("_") != -1 else tname
        if tbase == base:
            return tname

    return to_type_name(part_name, part_id)


# ===========================================================================
# SECTION: Systems / Parts  (worksheet "Systems")
# ===========================================================================

@dataclass
class PartNode:
    name: str
    id: str = ""
    children: list = field(default_factory=list)  # list[PartNode]


def _parse_part_rows(rows: List[tuple]) -> List[PartNode]:
    """Build a PartNode tree from worksheet rows (first row is header)."""
    if not rows:
        return []

    level_cols = _detect_id_name_pairs(rows[0])
    if not level_cols:
        print("Warning: could not detect ID/Name column pairs from header.", file=sys.stderr)
        print(f"  Header: {rows[0]}", file=sys.stderr)
        return []

    roots: List[PartNode] = []
    current_at_level: Dict[int, PartNode] = {}

    for row in rows[1:]:
        row = list(row) + [None] * (len(level_cols) * 3)
        for level, (col_id, col_name) in enumerate(level_cols):
            cell_id = row[col_id]
            cell_name = row[col_name]
            if cell_id and cell_name:
                node = PartNode(name=str(cell_name).strip(), id=str(cell_id).strip())
                current_at_level[level] = node
                for deeper in [k for k in current_at_level if k > level]:
                    del current_at_level[deeper]
                if level == 0:
                    roots.append(node)
                elif (level - 1) in current_at_level:
                    current_at_level[level - 1].children.append(node)
                break

    return roots


def _read_excel_parts(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[PartNode]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return _parse_part_rows(list(ws.iter_rows(values_only=True)))


def _read_csv_parts(path: pathlib.Path) -> List[PartNode]:
    sample = path.read_text(encoding="utf-8-sig")[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    with path.open(encoding="utf-8-sig") as f:
        rows = [tuple(row) for row in csv.reader(f, delimiter=delimiter)]
    return _parse_part_rows(rows)


def load_parts(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[PartNode]:
    """Load parts from an Excel or CSV file."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return _read_excel_parts(path, sheet_name)
    if suffix in (".csv", ".tsv"):
        return _read_csv_parts(path)
    try:
        return _read_excel_parts(path, sheet_name)
    except Exception:
        return _read_csv_parts(path)


def _collect_canonical_parts(parts: List[PartNode]) -> Dict[str, PartNode]:
    """Walk the tree; return type_name -> canonical PartNode (prefer the richer version)."""
    canonical: Dict[str, PartNode] = {}

    def walk(node: PartNode) -> None:
        tname = to_type_name(node.name, node.id)
        if tname not in canonical:
            canonical[tname] = node
        elif node.children and not canonical[tname].children:
            canonical[tname] = node
        for child in node.children:
            walk(child)

    for root in parts:
        walk(root)
    return canonical


def _collect_all_part_ids(parts: List[PartNode]) -> Dict[str, List[str]]:
    """Return type_name -> list of all IDs sharing that type name."""
    all_ids: Dict[str, List[str]] = {}

    def walk(node: PartNode) -> None:
        tname = to_type_name(node.name, node.id)
        if node.id and node.id not in all_ids.get(tname, []):
            all_ids.setdefault(tname, []).append(node.id)
        for child in node.children:
            walk(child)

    for root in parts:
        walk(root)
    return all_ids


def generate_parts_sysml(parts: List[PartNode], package_name: str = "Parts") -> str:
    """Produce a complete SysML v2 package string for part definitions."""
    canonical = _collect_canonical_parts(parts)
    all_ids = _collect_all_part_ids(parts)
    emitted: Set[str] = set()

    lines: List[str] = [f"package {package_name} {{", ""]

    def emit(node: PartNode) -> None:
        tname = to_type_name(node.name, node.id)
        if tname in emitted:
            return
        emitted.add(tname)
        cn = canonical[tname]
        lines.append(f"{INDENT}part def {tname} {{")
        for id_val in all_ids.get(tname, []):
            lines.append(f"{INDENT * 2}doc")
            lines.append(f"{INDENT * 2}/* ID: {id_val} */")
        for child in cn.children:
            child_type = to_type_name(child.name, child.id)
            child_usage = to_usage_name(child.name, child.id)
            lines.append(f"{INDENT * 2}part {child_usage} : {child_type};")
        lines.append(f"{INDENT}}}")
        lines.append("")
        for child in cn.children:
            emit(child)

    for root in parts:
        emit(root)

    lines.append("}")
    return "\n".join(lines)


def import_systems(rows: List[tuple], output_path, package_name: str = "") -> None:
    """Generate a parts/systems SysML file from worksheet rows."""
    output_path = pathlib.Path(output_path)
    pkg = package_name or _stem_to_package(output_path)

    parts = _parse_part_rows(rows)
    if not parts:
        print("Warning: no systems/parts found in the input rows.", file=sys.stderr)
        return

    total = len(_collect_canonical_parts(parts))
    print(f"Systems:  {len(parts)} top-level, {total} unique part types")

    sysml = generate_parts_sysml(parts, pkg)
    output_path.write_text(sysml, encoding="utf-8")
    print(f"Written:  {output_path}")
    validate(output_path)


# ===========================================================================
# SECTION: Functions  (worksheet "Functions")
# ===========================================================================

@dataclass
class FunctionNode:
    name: str
    id: str = ""
    kind: str = "FUNCTION"
    children: list = field(default_factory=list)  # list[FunctionNode]


def _detect_kind_col(header: tuple, id_col: int) -> int:
    """Return the column index of the Kind/Type cell for a given level, or -1."""
    h = [str(c).lower().strip() if c else "" for c in header]
    for col in range(id_col + 1, min(id_col + 4, len(h))):
        if "kind" in h[col] or "type" in h[col]:
            return col
    return -1


def _parse_func_rows(rows: List[tuple]) -> List[FunctionNode]:
    """Build a FunctionNode tree from worksheet rows (first row is header)."""
    if not rows:
        return []

    header = rows[0]
    level_cols = _detect_id_name_pairs(header)
    if not level_cols:
        print("Warning: could not detect ID/Name column pairs from header.", file=sys.stderr)
        print(f"  Header: {header}", file=sys.stderr)
        return []

    kind_cols = [_detect_kind_col(header, id_col) for id_col, _ in level_cols]

    roots: List[FunctionNode] = []
    current_at_level: Dict[int, FunctionNode] = {}

    for row in rows[1:]:
        row = list(row) + [None] * (len(level_cols) * 4)
        for level, (col_id, col_name) in enumerate(level_cols):
            cell_id = row[col_id]
            cell_name = row[col_name]
            if cell_id and cell_name:
                kind_col = kind_cols[level]
                kind = str(row[kind_col]).upper() if kind_col >= 0 and row[kind_col] else "FUNCTION"
                node = FunctionNode(
                    name=str(cell_name).strip(),
                    id=str(cell_id).strip(),
                    kind=kind,
                )
                current_at_level[level] = node
                for deeper in [k for k in current_at_level if k > level]:
                    del current_at_level[deeper]
                if level == 0:
                    roots.append(node)
                elif (level - 1) in current_at_level:
                    current_at_level[level - 1].children.append(node)
                break

    return roots


def _read_excel_functions(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[FunctionNode]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return _parse_func_rows(list(ws.iter_rows(values_only=True)))


def _read_csv_functions(path: pathlib.Path) -> List[FunctionNode]:
    sample = path.read_text(encoding="utf-8-sig")[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    with path.open(encoding="utf-8-sig") as f:
        rows = [tuple(row) for row in csv.reader(f, delimiter=delimiter)]
    return _parse_func_rows(rows)


def load_functions(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[FunctionNode]:
    """Load functions from an Excel or CSV file."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return _read_excel_functions(path, sheet_name)
    if suffix in (".csv", ".tsv"):
        return _read_csv_functions(path)
    try:
        return _read_excel_functions(path, sheet_name)
    except Exception:
        return _read_csv_functions(path)


def _collect_canonical_functions(nodes: List[FunctionNode]) -> Dict[str, FunctionNode]:
    """Return type_name -> canonical FunctionNode (prefer richer/with-children version)."""
    canonical: Dict[str, FunctionNode] = {}

    def walk(node: FunctionNode) -> None:
        tname = to_type_name(node.name, node.id)
        if tname not in canonical:
            canonical[tname] = node
        elif node.children and not canonical[tname].children:
            canonical[tname] = node
        for child in node.children:
            walk(child)

    for root in nodes:
        walk(root)
    return canonical


def generate_functions_sysml(functions: List[FunctionNode], package_name: str = "Functions") -> str:
    """Produce a SysML v2 package with action defs for all functions."""
    canonical = _collect_canonical_functions(functions)
    emitted: Set[str] = set()
    lines: List[str] = [f"package {package_name} {{", ""]

    def emit(node: FunctionNode) -> None:
        tname = to_type_name(node.name, node.id)
        if tname in emitted:
            return
        emitted.add(tname)
        cn = canonical[tname]
        lines.append(f"{INDENT}action def {tname} {{")
        if cn.id:
            lines.append(f"{INDENT * 2}doc")
            lines.append(f"{INDENT * 2}/* ID: {cn.id} */")
        for child in cn.children:
            child_type = to_type_name(child.name, child.id)
            child_usage = to_usage_name(child.name, child.id)
            lines.append(f"{INDENT * 2}action {child_usage} : {child_type};")
        lines.append(f"{INDENT}}}")
        lines.append("")
        for child in cn.children:
            emit(child)

    for root in functions:
        emit(root)

    lines.append("}")
    return "\n".join(lines)


def import_functions(rows: List[tuple], output_path, package_name: str = "") -> None:
    """Generate a functions SysML file from worksheet rows."""
    output_path = pathlib.Path(output_path)
    pkg = package_name or _stem_to_package(output_path)

    functions = _parse_func_rows(rows)
    if not functions:
        print("Warning: no functions found in the input rows.", file=sys.stderr)
        return

    canonical = _collect_canonical_functions(functions)
    parents = sum(1 for n in canonical.values() if n.children)
    print(f"Functions: {len(functions)} top-level, {len(canonical)} unique action defs ({parents} with subfunctions)")

    sysml = generate_functions_sysml(functions, pkg)
    output_path.write_text(sysml, encoding="utf-8")
    print(f"Written:  {output_path}")
    validate(output_path)


# ===========================================================================
# SECTION: Function-System Allocations  (worksheet "Link Systems and Functions")
# ===========================================================================

@dataclass
class Allocation:
    target_system_path: List[str]
    target_system_id: str
    function_name: str
    function_id: str


def _detect_alloc_columns(header: tuple) -> Tuple[List[Tuple[int, int]], List[Tuple[int, int]]]:
    """Return (system_pairs, function_pairs) from an allocation header row."""
    h = [str(c).lower().strip() if c else "" for c in header]
    system_pairs: List[Tuple[int, int]] = []
    function_pairs: List[Tuple[int, int]] = []

    for id_col, id_header in enumerate(h):
        if "id" not in id_header or not id_header:
            continue
        for name_col in range(id_col + 1, len(h)):
            if "name" in h[name_col]:
                is_function = "function" in id_header or "function" in h[name_col]
                (function_pairs if is_function else system_pairs).append((id_col, name_col))
                break

    return system_pairs, function_pairs


def _parse_allocation_rows(rows: List[tuple]) -> List[Allocation]:
    """Build Allocation list from worksheet rows (first row is header)."""
    if not rows:
        return []

    header = rows[0]
    system_pairs, function_pairs = _detect_alloc_columns(header)

    if not system_pairs:
        print("Warning: no system ID/Name column pairs found in header.", file=sys.stderr)
        return []
    if not function_pairs:
        print("Warning: no function ID/Name column pairs found in header.", file=sys.stderr)
        return []

    allocations: List[Allocation] = []
    current_system_path: List[str] = []
    current_system_id: str = ""

    for row in rows[1:]:
        row = list(row) + [None] * max(len(header), 20)

        for level, (id_col, name_col) in enumerate(system_pairs):
            cell_id = row[id_col]
            cell_name = row[name_col]
            if cell_id and cell_name:
                current_system_path = current_system_path[:level]
                current_system_path.append(str(cell_name).strip())
                current_system_id = str(cell_id).strip()
                break

        function_name = ""
        function_id = ""
        for id_col, name_col in reversed(function_pairs):
            val = row[name_col]
            if val and str(val).strip():
                function_name = str(val).strip()
                function_id = str(row[id_col]).strip() if row[id_col] else ""
                break

        if function_name:
            if current_system_path:
                allocations.append(Allocation(
                    target_system_path=current_system_path.copy(),
                    target_system_id=current_system_id,
                    function_name=function_name,
                    function_id=function_id,
                ))
            else:
                print(f"Warning: function '{function_name}' has no target system.", file=sys.stderr)

    return allocations


def load_function_allocations(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[Allocation]:
    """Load function-system allocations from an Excel file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return _parse_allocation_rows(list(ws.iter_rows(values_only=True)))


def _strip_perform_actions(content: str) -> str:
    """Remove all previously injected 'perform action ...' blocks (idempotency)."""
    lines = content.split("\n")
    result: List[str] = []
    skipping = False
    depth = 0

    for line in lines:
        if not skipping:
            if re.match(r"^\s+perform\s+action\b", line):
                skipping = True
                depth = line.count("{") - line.count("}")
                if depth <= 0:
                    skipping = False
            else:
                result.append(line)
        else:
            depth += line.count("{") - line.count("}")
            if depth <= 0:
                skipping = False

    return "\n".join(result)


def _ensure_import(content: str, functions_package: str) -> str:
    """Insert 'private import <pkg>::*;' after the package opening brace if absent."""
    import_stmt = f"private import {functions_package}::*;"
    if import_stmt in content:
        return content
    m = re.search(r"\bpackage\s+\w[\w.]*\s*\{", content)
    if not m:
        return content
    return content[:m.end()] + f"\n\n{INDENT}{import_stmt}" + content[m.end():]


def _build_perform_block(action_name: str, action_type: str, function_id: str) -> str:
    i1, i2 = INDENT * 2, INDENT * 3
    return "\n".join([
        f"{i1}perform action {action_name} : {action_type} {{",
        f"{i2}doc",
        f"{i2}/* ID: {function_id} */",
        f"{i1}}}",
    ])


def _insert_into_part_def(
    content: str, type_name: str, perform_blocks: List[str], part_defs: Dict[str, str]
) -> str:
    """Insert perform_blocks before the closing '}' of the named part def."""
    matched_name = None
    # Try exact match, then prefix match
    if type_name in part_defs:
        matched_name = type_name
    else:
        base = type_name[:type_name.rfind("_")] if type_name.rfind("_") != -1 else type_name
        for pname in part_defs:
            if pname == base or pname.startswith(base + "_"):
                matched_name = pname
                break

    if not matched_name:
        print(f"Warning: 'part def {type_name}' not found; allocation skipped.", file=sys.stderr)
        return content

    pattern = rf"part def {re.escape(matched_name)}\s*\{{"
    m = re.search(pattern, content)
    if not m:
        print(f"Warning: 'part def {matched_name}' not found; allocation skipped.", file=sys.stderr)
        return content

    brace_pos = m.end() - 1
    end_pos = _find_block_end(content, brace_pos)
    if end_pos == -1:
        print(f"Warning: unmatched '{{' for 'part def {matched_name}'.", file=sys.stderr)
        return content

    line_start = content.rfind("\n", 0, end_pos) + 1
    return content[:line_start] + "\n".join(perform_blocks) + "\n" + content[line_start:]


def merge_allocations_into_parts(
    parts_content: str,
    allocations: List[Allocation],
    functions_content: str,
    functions_package: str = "FunctionsGenerated",
) -> str:
    """Inject perform action statements into the parts SysML content."""
    content = _strip_perform_actions(parts_content)
    content = _ensure_import(content, functions_package)

    part_defs = get_all_part_defs(content)
    function_id_to_name = get_action_id_to_name_map(functions_content)

    by_type: Dict[str, List[Allocation]] = {}
    for alloc in allocations:
        tname = to_type_name(alloc.target_system_path[-1], alloc.target_system_id)
        by_type.setdefault(tname, []).append(alloc)

    for type_name, allocs in by_type.items():
        perform_blocks = []
        for a in allocs:
            matched_fn = match_action_name(a.function_name, a.function_id, function_id_to_name)
            perform_blocks.append(
                _build_perform_block(
                    to_usage_name(a.function_name, a.function_id),
                    matched_fn,
                    a.function_id,
                )
            )
        content = _insert_into_part_def(content, type_name, perform_blocks, part_defs)

    return content


def import_allocations(
    rows: List[tuple],
    parts_path,
    functions_path,
    functions_pkg: str = "",
) -> None:
    """Inject function-system perform actions into the parts SysML file."""
    parts_path = pathlib.Path(parts_path)
    functions_path = pathlib.Path(functions_path)

    if not parts_path.exists():
        print(f"Error: parts file not found: {parts_path}", file=sys.stderr)
        return
    if not functions_path.exists():
        print(f"Error: functions file not found: {functions_path}", file=sys.stderr)
        return

    pkg = functions_pkg or _stem_to_package(functions_path)

    allocations = _parse_allocation_rows(rows)
    if not allocations:
        print("Warning: no allocations found in the input rows.", file=sys.stderr)
        return
    print(f"Allocations: {len(allocations)}")

    parts_content = parts_path.read_text(encoding="utf-8")
    functions_content = functions_path.read_text(encoding="utf-8")
    updated = merge_allocations_into_parts(parts_content, allocations, functions_content, pkg)

    parts_path.write_text(updated, encoding="utf-8")
    print(f"Written:  {parts_path}")
    validate(parts_path, check_dir=True)


# ===========================================================================
# SECTION: Functional Exchanges  (worksheet "Functional Exchanges")
# ===========================================================================

@dataclass
class FunctionalExchange:
    from_action_id: str
    from_action_name: str
    from_port_id: str
    from_port_name: str
    exchange_id: str
    exchange_name: str
    to_action_id: str
    to_action_name: str
    to_port_id: str
    to_port_name: str


def load_functional_exchanges(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[FunctionalExchange]:
    """Load functional exchanges from an Excel file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))

    exchanges: List[FunctionalExchange] = []
    for row in rows[1:]:
        if not any(row):
            continue
        (from_id, from_name, from_port_id, from_port_name,
         exch_id, exch_name,
         to_id, to_name, to_port_id, to_port_name) = row[:10]
        if not exch_id:
            continue
        exchanges.append(FunctionalExchange(
            from_action_id=str(from_id).strip(),
            from_action_name=str(from_name).strip(),
            from_port_id=str(from_port_id).strip(),
            from_port_name=str(from_port_name).strip(),
            exchange_id=str(exch_id).strip(),
            exchange_name=str(exch_name).strip(),
            to_action_id=str(to_id).strip(),
            to_action_name=str(to_name).strip(),
            to_port_id=str(to_port_id).strip(),
            to_port_name=str(to_port_name).strip(),
        ))
    return exchanges


def _parse_functional_exchange_rows(rows: List[tuple]) -> List[FunctionalExchange]:
    """Parse functional exchanges from worksheet rows (first row is header, skipped)."""
    exchanges: List[FunctionalExchange] = []
    for row in rows[1:]:
        if not any(row):
            continue
        (from_id, from_name, from_port_id, from_port_name,
         exch_id, exch_name,
         to_id, to_name, to_port_id, to_port_name) = (list(row) + [None] * 10)[:10]
        if not exch_id:
            continue
        exchanges.append(FunctionalExchange(
            from_action_id=str(from_id).strip(),
            from_action_name=str(from_name).strip(),
            from_port_id=str(from_port_id).strip() if from_port_id else "",
            from_port_name=str(from_port_name).strip() if from_port_name else "",
            exchange_id=str(exch_id).strip(),
            exchange_name=str(exch_name).strip(),
            to_action_id=str(to_id).strip(),
            to_action_name=str(to_name).strip(),
            to_port_id=str(to_port_id).strip() if to_port_id else "",
            to_port_name=str(to_port_name).strip() if to_port_name else "",
        ))
    return exchanges


def parse_existing_def_names(sysml_text: str) -> Set[str]:
    """Return all defined names (action, part, port, item, connection, interface, attribute defs)."""
    names: Set[str] = set()
    for pattern in [
        r"\baction def\s+(\w+)", r"\bpart def\s+(\w+)", r"\bport def\s+(\w+)",
        r"\bitem def\s+(\w+)", r"\bconnection def\s+(\w+)", r"\binterface def\s+(\w+)",
        r"\battribute def\s+(\w+)",
    ]:
        for m in re.finditer(pattern, sysml_text):
            names.add(m.group(1))
    return names


def parse_non_item_def_names(sysml_text: str) -> Set[str]:
    """Return defined names excluding item defs (for conflict resolution)."""
    names: Set[str] = set()
    for pattern in [
        r"\baction def\s+(\w+)", r"\bpart def\s+(\w+)", r"\bport def\s+(\w+)",
        r"\bconnection def\s+(\w+)", r"\binterface def\s+(\w+)", r"\battribute def\s+(\w+)",
    ]:
        for m in re.finditer(pattern, sysml_text):
            names.add(m.group(1))
    return names


def build_item_name_registry(
    exchanges: List[FunctionalExchange], existing_names: Set[str]
) -> Dict[str, str]:
    """Return exchange_name -> safe_item_def_name, appending 'Item' on conflict."""
    registry: Dict[str, str] = {}
    reserved = set(existing_names)

    for exch in exchanges:
        raw = to_type_name(exch.exchange_name)
        if raw in registry.values():
            continue
        name = f"{raw}Item" if raw in reserved else raw
        registry[exch.exchange_name] = name
        reserved.add(name)

    return registry


def _funcout_def_name(exch: FunctionalExchange, registry: Dict[str, str]) -> str:
    return f"{registry.get(exch.exchange_name, to_type_name(exch.exchange_name))}FuncOut"


def _funcin_def_name(exch: FunctionalExchange, registry: Dict[str, str]) -> str:
    return f"{registry.get(exch.exchange_name, to_type_name(exch.exchange_name))}FuncIn"


def _item_type_name(exch: FunctionalExchange, registry: Dict[str, str]) -> str:
    return registry.get(exch.exchange_name, to_type_name(exch.exchange_name))


def _item_usage_name(exch: FunctionalExchange, registry: Dict[str, str]) -> str:
    return to_usage_name(_item_type_name(exch, registry))


def _assign_fe_conn_iface_names(
    exchanges: List[FunctionalExchange], id_to_def_name: Dict[str, str]
) -> List[Tuple[str, str]]:
    """Return (conn_name, iface_name) per exchange, with numeric suffixes on collisions."""
    seen: Dict[str, int] = {}
    result: List[Tuple[str, str]] = []
    for exch in exchanges:
        from_type = match_action_name(exch.from_action_name, exch.from_action_id, id_to_def_name)
        to_type = match_action_name(exch.to_action_name, exch.to_action_id, id_to_def_name)
        base = f"{from_type}To{to_type}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        suffix = str(count) if count > 1 else ""
        result.append((f"{base}Connection{suffix}", f"{base}Interface{suffix}"))
    return result


def _build_fe_action_ports(
    exchanges: List[FunctionalExchange],
    registry: Dict[str, str],
    id_to_def_name: Dict[str, str],
) -> Dict[str, List[Tuple[str, str, str]]]:
    """Return action_id -> [(port_usage_name, port_def_name, port_id), ...]."""
    action_ports: Dict[str, List[Tuple[str, str, str]]] = {}

    for exch in exchanges:
        from_action_name = match_action_name(exch.from_action_name, exch.from_action_id, id_to_def_name)
        to_action_name = match_action_name(exch.to_action_name, exch.to_action_id, id_to_def_name)

        entries = [
            (exch.from_action_id, from_action_name, to_port_usage_name(exch.from_port_name, exch.from_port_id), _funcout_def_name(exch, registry), exch.from_port_id),
            (exch.to_action_id, to_action_name, to_port_usage_name(exch.to_port_name, exch.to_port_id), _funcin_def_name(exch, registry), exch.to_port_id),
        ]
        for action_id, matched_name, usage, def_name, port_id in entries:
            for pid, pname in id_to_def_name.items():
                if pname == matched_name and pid == action_id:
                    action_ports.setdefault(action_id, [])
                    if not any(p[0] == usage for p in action_ports[action_id]):
                        action_ports[action_id].append((usage, def_name, port_id))
                    break

    return action_ports


def _scan_existing_action_ports(sysml_text: str) -> Set[Tuple[str, str]]:
    """Return (action_id, port_usage_name) pairs already present in the SysML text."""
    pairs: Set[Tuple[str, str]] = set()
    current_id: Optional[str] = None
    action_body_depth: Optional[int] = None
    depth = 0
    action_def_re = re.compile(r"\baction def\s+\w+")

    for line in sysml_text.splitlines():
        if action_def_re.search(line):
            action_body_depth = depth + 1
            current_id = None

        id_m = re.search(r"/\* ID: ([\w\-]+) \*/", line)
        if id_m and action_body_depth is not None and depth == action_body_depth:
            current_id = id_m.group(1)

        port_m = re.search(r"\bport\s+(\w+)\s*:", line)
        if port_m and current_id and action_body_depth is not None and depth == action_body_depth:
            pairs.add((current_id, port_m.group(1)))

        depth += line.count("{") - line.count("}")
        if action_body_depth is not None and depth < action_body_depth:
            current_id = None
            action_body_depth = None

    return pairs


def inject_functional_exchanges(sysml_text: str, exchanges: List[FunctionalExchange]) -> str:
    """Return SysML text augmented with port usages, item/port/interface/connection defs."""
    id_to_def_name = parse_id_to_def_name(sysml_text)
    non_item_names = parse_non_item_def_names(sysml_text)
    registry = build_item_name_registry(exchanges, non_item_names)
    action_ports = _build_fe_action_ports(exchanges, registry, id_to_def_name)
    conn_iface_names = _assign_fe_conn_iface_names(exchanges, id_to_def_name)

    injected_pairs: Set[Tuple[str, str]] = _scan_existing_action_ports(sysml_text)
    existing_item_defs: Set[str] = set(re.findall(r"\bitem def\s+(\w+)", sysml_text))
    existing_port_defs: Set[str] = set(re.findall(r"\bport def\s+(\w+)", sysml_text))
    existing_iface_defs: Set[str] = set(re.findall(r"\binterface def\s+(\w+)", sysml_text))
    existing_conn_defs: Set[str] = set(re.findall(r"\bconnection def\s+(\w+)", sysml_text))

    lines = sysml_text.split("\n")
    result: List[str] = []

    for line in lines:
        result.append(line)
        id_match = re.search(r"/\* ID: ([\w\-]+) \*/", line)
        if id_match:
            action_id = id_match.group(1)
            if action_id in action_ports:
                base_indent = " " * (len(line) - len(line.lstrip()))
                for usage, def_name, port_id in action_ports[action_id]:
                    if (action_id, usage) in injected_pairs:
                        continue
                    injected_pairs.add((action_id, usage))
                    result.append(f"{base_indent}port {usage} : {def_name} {{")
                    result.append(f"{base_indent}{INDENT}doc")
                    result.append(f"{base_indent}{INDENT}/* ID: {port_id} */")
                    result.append(f"{base_indent}}}")

    append_lines: List[str] = []

    for exch in exchanges:
        item_name = _item_type_name(exch, registry)
        if item_name not in existing_item_defs:
            existing_item_defs.add(item_name)
            item_usage = _item_usage_name(exch, registry)
            append_lines += [
                f"{INDENT}item def {item_name} {{",
                f"{INDENT * 2}doc",
                f"{INDENT * 2}/* ID: {exch.exchange_id} */",
                f"{INDENT}}}", "",
            ]

    for exch in exchanges:
        funcout = _funcout_def_name(exch, registry)
        funcin = _funcin_def_name(exch, registry)
        if funcout not in existing_port_defs:
            existing_port_defs.add(funcout)
            existing_port_defs.add(funcin)
            item_type = _item_type_name(exch, registry)
            item_usage = _item_usage_name(exch, registry)
            append_lines += [
                f"{INDENT}port def {funcout} {{",
                f"{INDENT * 2}doc",
                f"{INDENT * 2}/* ID: {exch.from_port_id} */",
                f"{INDENT * 2}out item {item_usage} : {item_type};",
                f"{INDENT}}}", "",
                f"{INDENT}port def {funcin} {{",
                f"{INDENT * 2}doc",
                f"{INDENT * 2}/* ID: {exch.to_port_id} */",
                f"{INDENT * 2}in item {item_usage} : {item_type};",
                f"{INDENT}}}", "",
            ]

    for exch, (conn_name, iface_name) in zip(exchanges, conn_iface_names):
        if iface_name not in existing_iface_defs:
            existing_iface_defs.add(iface_name)
            item_type = _item_type_name(exch, registry)
            item_usage = _item_usage_name(exch, registry)
            append_lines += [
                f"{INDENT}interface def {iface_name} {{",
                f"{INDENT * 2}end port outPort : {_funcout_def_name(exch, registry)};",
                f"{INDENT * 2}end port inPort : {_funcin_def_name(exch, registry)};",
                f"{INDENT * 2}flow of {item_type} from outPort.{item_usage} to inPort.{item_usage};",
                f"{INDENT}}}", "",
            ]

    for exch, (conn_name, iface_name) in zip(exchanges, conn_iface_names):
        if conn_name not in existing_conn_defs:
            existing_conn_defs.add(conn_name)
            from_type = match_action_name(exch.from_action_name, exch.from_action_id, id_to_def_name)
            to_type = match_action_name(exch.to_action_name, exch.to_action_id, id_to_def_name)
            item_usage = _item_usage_name(exch, registry)
            from_end = f"{item_usage}Out"
            to_end = f"{item_usage}In"
            funcout_usage = to_port_usage_name(exch.from_port_name, exch.from_port_id)
            funcin_usage = to_port_usage_name(exch.to_port_name, exch.to_port_id)
            append_lines += [
                f"{INDENT}connection def {conn_name} {{",
                f"{INDENT * 2}doc",
                f"{INDENT * 2}/* ID: {exch.exchange_id} */",
                f"{INDENT * 2}end action {from_end} : {from_type};",
                f"{INDENT * 2}end action {to_end} : {to_type};",
                f"{INDENT * 2}interface : {iface_name}"
                f" connect {to_end}.{funcin_usage}"
                f" to {from_end}.{funcout_usage};",
                f"{INDENT}}}", "",
            ]

    closing_idx: Optional[int] = None
    for i in range(len(result) - 1, -1, -1):
        if result[i].strip() == "}":
            closing_idx = i
            break

    if closing_idx is None:
        result.extend(append_lines)
        result.append("}")
    else:
        result = result[:closing_idx] + append_lines + result[closing_idx:]

    return "\n".join(result)


def import_functional_exchanges(rows: List[tuple], functions_path) -> None:
    """Inject functional exchanges into the functions SysML file."""
    functions_path = pathlib.Path(functions_path)
    if not functions_path.exists():
        print(f"Error: functions file not found: {functions_path}", file=sys.stderr)
        return

    exchanges = _parse_functional_exchange_rows(rows)
    if not exchanges:
        print("Warning: no functional exchanges found in the input rows.", file=sys.stderr)
        return
    print(f"Functional exchanges: {len(exchanges)}")

    sysml_text = functions_path.read_text(encoding="utf-8")
    augmented = inject_functional_exchanges(sysml_text, exchanges)
    functions_path.write_text(augmented, encoding="utf-8")
    print(f"Written:  {functions_path}")
    validate(functions_path)


# ===========================================================================
# SECTION: Component Exchanges  (worksheet "Component Exchanges")
# ===========================================================================

@dataclass
class ComponentExchange:
    exchange_id: str
    exchange_name: str
    exchange_kind: str
    from_part_id: str
    from_part_name: str
    from_port_id: str
    from_port_name: str
    from_port_direction: str
    to_part_id: str
    to_part_name: str
    to_port_id: str
    to_port_name: str
    to_port_direction: str


def _parse_component_exchange_rows(rows: List[tuple]) -> List[ComponentExchange]:
    """Parse component exchanges from worksheet rows (first row is header, skipped)."""
    exchanges: List[ComponentExchange] = []
    for row in rows[1:]:
        if not any(row):
            continue
        padded = list(row) + [None] * 15
        (from_id, from_name, from_port_id, from_port_name, from_port_dir, _from_port_kind,
         exch_id, exch_name, exch_kind,
         to_id, to_name, to_port_id, to_port_name, to_port_dir, _to_port_kind) = padded[:15]
        if not exch_id:
            continue
        exchanges.append(ComponentExchange(
            exchange_id=str(exch_id).strip(),
            exchange_name=str(exch_name).strip(),
            exchange_kind=str(exch_kind).strip() if exch_kind else "",
            from_part_id=str(from_id).strip(),
            from_part_name=str(from_name).strip(),
            from_port_id=str(from_port_id).strip(),
            from_port_name=str(from_port_name).strip(),
            from_port_direction=str(from_port_dir).strip() if from_port_dir else "",
            to_part_id=str(to_id).strip(),
            to_part_name=str(to_name).strip(),
            to_port_id=str(to_port_id).strip(),
            to_port_name=str(to_port_name).strip(),
            to_port_direction=str(to_port_dir).strip() if to_port_dir else "",
        ))
    return exchanges


def load_component_exchanges(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[ComponentExchange]:
    """Load component exchanges from an Excel file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return _parse_component_exchange_rows(list(ws.iter_rows(values_only=True)))


def _build_ce_part_ports(
    exchanges: List[ComponentExchange], part_id_to_name: Dict[str, str]
) -> Dict[str, List[Tuple[str, str, str]]]:
    """Return part_name -> [(port_usage_name, port_def_name, port_id), ...]."""
    part_ports: Dict[str, List[Tuple[str, str, str]]] = {}

    for exch in exchanges:
        exch_name = to_exchange_name(exch.exchange_name)
        from_suffix = to_component_suffix(exch.from_part_name)
        to_suffix = to_component_suffix(exch.to_part_name)
        from_port_def = f"{exch_name}ConnectionPoint_{from_suffix}"
        to_port_def = f"{exch_name}ConnectionPoint_{to_suffix}"

        from_part_name = match_part_name(exch.from_part_name, exch.from_part_id, part_id_to_name)
        to_part_name = match_part_name(exch.to_part_name, exch.to_part_id, part_id_to_name)

        for part_name, port_name, port_id, port_def in [
            (from_part_name, exch.from_port_name, exch.from_port_id, from_port_def),
            (to_part_name,   exch.to_port_name,   exch.to_port_id,   to_port_def),
        ]:
            usage = to_port_usage_name(port_name, port_id)
            part_ports.setdefault(part_name, [])
            if not any(p[0] == usage for p in part_ports[part_name]):
                part_ports[part_name].append((usage, port_def, port_id))

    return part_ports


def inject_component_exchanges(sysml_text: str, exchanges: List[ComponentExchange]) -> str:
    """Return SysML text augmented with port usages, port defs, interface defs, and connection defs."""
    part_id_to_name = get_part_id_to_name_map(sysml_text)
    part_ports = _build_ce_part_ports(exchanges, part_id_to_name)

    existing_port_usages: Set[str] = set(re.findall(r"\bport\s+(\w+)\s*:", sysml_text))
    existing_port_defs: Set[str] = set(re.findall(r"\bport def\s+(\w+)", sysml_text))
    existing_iface_defs: Set[str] = set(re.findall(r"\binterface def\s+(\w+)", sysml_text))
    existing_conn_defs: Set[str] = set(re.findall(r"\bconnection def\s+(\w+)", sysml_text))

    lines = sysml_text.split("\n")
    result: List[str] = []

    for line in lines:
        result.append(line)
        id_match = re.search(r"/\* ID: ([\w\-]+) \*/", line)
        if id_match:
            part_id = id_match.group(1)
            matched_part_name = part_id_to_name.get(part_id)
            if matched_part_name and matched_part_name in part_ports:
                base_indent = " " * (len(line) - len(line.lstrip()))
                for usage, port_def, p_id in part_ports[matched_part_name]:
                    if usage in existing_port_usages:
                        continue
                    existing_port_usages.add(usage)
                    result.append(f"{base_indent}port {usage} : {port_def} {{")
                    result.append(f"{base_indent}{INDENT}doc")
                    result.append(f"{base_indent}{INDENT}/* ID: {p_id} */")
                    result.append(f"{base_indent}}}")

    append_lines: List[str] = []

    for exch in exchanges:
        exch_name = to_exchange_name(exch.exchange_name)
        from_suffix = to_component_suffix(exch.from_part_name)
        to_suffix = to_component_suffix(exch.to_part_name)
        for pd in [f"{exch_name}ConnectionPoint_{from_suffix}", f"{exch_name}ConnectionPoint_{to_suffix}"]:
            if pd not in existing_port_defs:
                existing_port_defs.add(pd)
                append_lines += [f"{INDENT}port def {pd};", ""]

    for exch in exchanges:
        iface_name = f"{to_exchange_name(exch.exchange_name)}_Interface"
        if iface_name not in existing_iface_defs:
            existing_iface_defs.add(iface_name)
            exch_name = to_exchange_name(exch.exchange_name)
            from_suffix = to_component_suffix(exch.from_part_name)
            to_suffix = to_component_suffix(exch.to_part_name)
            from_port_def = f"{exch_name}ConnectionPoint_{from_suffix}"
            to_port_def = f"{exch_name}ConnectionPoint_{to_suffix}"
            from_part_name = match_part_name(exch.from_part_name, exch.from_part_id, part_id_to_name)
            to_part_name = match_part_name(exch.to_part_name, exch.to_part_id, part_id_to_name)
            from_port_end = f"{to_usage_name(from_part_name)}Port"
            to_port_end = f"{to_usage_name(to_part_name)}Port"
            append_lines += [
                f"{INDENT}interface def {iface_name} {{",
                f"{INDENT * 2}end port {from_port_end} : {from_port_def};",
                f"{INDENT * 2}end port {to_port_end} : {to_port_def};",
                f"{INDENT}}}", "",
            ]

    for exch in exchanges:
        exch_name = to_exchange_name(exch.exchange_name)
        if exch_name not in existing_conn_defs:
            existing_conn_defs.add(exch_name)
            from_part_name = match_part_name(exch.from_part_name, exch.from_part_id, part_id_to_name)
            to_part_name = match_part_name(exch.to_part_name, exch.to_part_id, part_id_to_name)
            from_port = to_port_usage_name(exch.from_port_name, exch.from_port_id)
            to_port = to_port_usage_name(exch.to_port_name, exch.to_port_id)
            from_end = to_usage_name(from_part_name)
            to_end = to_usage_name(to_part_name)
            append_lines += [
                f"{INDENT}connection def {exch_name} {{",
                f"{INDENT * 2}doc",
                f"{INDENT * 2}/* ID: {exch.exchange_id} */",
                f"{INDENT * 2}end part {from_end} : {from_part_name};",
                f"{INDENT * 2}end part {to_end} : {to_part_name};",
                f"{INDENT * 2}interface : {exch_name} connect {from_end}.{from_port} to {to_end}.{to_port};",
                f"{INDENT}}}", "",
            ]

    closing_idx: Optional[int] = None
    for i in range(len(result) - 1, -1, -1):
        if result[i].strip() == "}":
            closing_idx = i
            break

    if closing_idx is None:
        result.extend(append_lines)
        result.append("}")
    else:
        result = result[:closing_idx] + append_lines + result[closing_idx:]

    return "\n".join(result)


def import_component_exchanges(rows: List[tuple], parts_path) -> None:
    """Inject component exchanges into the parts SysML file."""
    parts_path = pathlib.Path(parts_path)
    if not parts_path.exists():
        print(f"Error: parts file not found: {parts_path}", file=sys.stderr)
        return

    exchanges = _parse_component_exchange_rows(rows)
    if not exchanges:
        print("Warning: no component exchanges found in the input rows.", file=sys.stderr)
        return
    print(f"Component exchanges: {len(exchanges)}")

    sysml_text = parts_path.read_text(encoding="utf-8")
    augmented = inject_component_exchanges(sysml_text, exchanges)
    parts_path.write_text(augmented, encoding="utf-8")
    print(f"Written:  {parts_path}")
    validate(parts_path, check_dir=True)


# ===========================================================================
# SECTION: Exchange Allocations  (worksheet "Link Exchanges")
# ===========================================================================

@dataclass(frozen=True)
class ExchangeAllocation:
    component_exchange_name: str
    functional_exchange_name: str
    allocation_id: str


def _parse_exchange_allocation_rows(rows: List[tuple]) -> List[ExchangeAllocation]:
    """Parse exchange allocations from worksheet rows (first row is header)."""
    if not rows:
        return []

    header = [str(c).strip() if c else "" for c in rows[0]]
    try:
        ce_col = header.index("Component Exchange Name")
        fe_col = header.index("Functional Exchange Name")
        id_col = header.index("Allocation ID")
    except ValueError as exc:
        print(f"Error: required column not found in header: {exc}", file=sys.stderr)
        return []

    seen: set = set()
    allocations: List[ExchangeAllocation] = []
    for row in rows[1:]:
        row = list(row) + [None] * 20
        ce = str(row[ce_col]).strip() if row[ce_col] else ""
        fe = str(row[fe_col]).strip() if row[fe_col] else ""
        aid = str(row[id_col]).strip() if row[id_col] else ""
        if not (ce and fe and aid):
            continue
        key = (ce, fe, aid)
        if key not in seen:
            seen.add(key)
            allocations.append(ExchangeAllocation(ce, fe, aid))

    return allocations


def load_exchange_allocations(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[ExchangeAllocation]:
    """Load exchange allocations from an Excel file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return _parse_exchange_allocation_rows(list(ws.iter_rows(values_only=True)))


def _strip_flow_lines(content: str) -> str:
    lines = content.split("\n")
    return "\n".join(line for line in lines if not re.match(r"^\s+flow\s+of\b", line))


def _strip_allocation_port_items(content: str) -> str:
    """Revert injected port def bodies (marked with ALLOC_MARKER) back to bare form."""
    pattern = re.compile(
        r"([ \t]+)port\s+def\s+(\w+)\s*\{"
        r"[^{}]*?" + re.escape(ALLOC_MARKER) + r"[^{}]*?\}",
        re.DOTALL,
    )
    return pattern.sub(lambda m: f"{m.group(1)}port def {m.group(2)};", content)


def _find_interface_port_info(content: str, iface_name: str) -> List[Tuple[str, str, bool]]:
    """Return (port_name, port_type, is_conjugated) for each end port in an interface def."""
    m = re.search(rf"\binterface\s+def\s+{re.escape(iface_name)}\s*\{{", content)
    if not m:
        return []
    close_pos = _find_block_end(content, m.end() - 1)
    if close_pos == -1:
        return []
    block = content[m.end() - 1: close_pos + 1]
    raw = re.findall(r"\bend\s+port\s+(\w+)\s*:\s*(~?)(\w+)\s*;", block)
    return [(name, ptype, tilde == "~") for name, tilde, ptype in raw]


def _ensure_port_def_has_item(
    content: str, port_def_name: str, fe_type: str, fe_usage: str, direction: str = "out"
) -> str:
    """Add 'direction item feUsage : feType;' to a port def body."""
    item_line = f"{INDENT * 2}{direction} item {fe_usage} : {fe_type};"

    pattern_body = re.compile(rf"\bport\s+def\s+{re.escape(port_def_name)}\s*\{{")
    m = pattern_body.search(content)
    if m:
        close_pos = _find_block_end(content, m.end() - 1)
        if close_pos == -1:
            return content
        block = content[m.end() - 1: close_pos + 1]
        if re.search(rf"\bitem\s+{re.escape(fe_usage)}\s*:", block):
            return content
        line_start = content.rfind("\n", 0, close_pos) + 1
        return content[:line_start] + item_line + "\n" + content[line_start:]

    pattern_semi = re.compile(rf"([ \t]+)port\s+def\s+{re.escape(port_def_name)}\s*;")
    m = pattern_semi.search(content)
    if not m:
        print(f"Warning: 'port def {port_def_name}' not found; item '{fe_usage}' not added.", file=sys.stderr)
        return content

    indent = m.group(1)
    new_def = (
        f"{indent}port def {port_def_name} {{\n"
        f"{indent}{INDENT}{ALLOC_MARKER}\n"
        f"{item_line}\n"
        f"{indent}}}"
    )
    return content[: m.start()] + new_def + content[m.end():]


def _insert_into_interface_def(content: str, iface_name: str, flow_lines: List[str]) -> str:
    """Insert flow_lines before the closing '}' of 'interface def iface_name { ... }'."""
    m = re.search(rf"\binterface\s+def\s+{re.escape(iface_name)}\s*\{{", content)
    if not m:
        print(f"Warning: 'interface def {iface_name}' not found; flows skipped.", file=sys.stderr)
        return content
    close_pos = _find_block_end(content, m.end() - 1)
    if close_pos == -1:
        print(f"Warning: unmatched '{{' for 'interface def {iface_name}'.", file=sys.stderr)
        return content
    line_start = content.rfind("\n", 0, close_pos) + 1
    return content[:line_start] + "\n".join(flow_lines) + "\n" + content[line_start:]


def inject_exchange_allocations(parts_content: str, allocations: List[ExchangeAllocation]) -> str:
    """Inject 'flow of' lines (and required port-def items) for each allocation."""
    content = _strip_flow_lines(parts_content)
    content = _strip_allocation_port_items(content)

    by_ce: Dict[str, List[ExchangeAllocation]] = {}
    for alloc in allocations:
        by_ce.setdefault(to_exchange_name(alloc.component_exchange_name), []).append(alloc)

    for ce_sysml_name, allocs in by_ce.items():
        iface_name = f"{ce_sysml_name}_Interface"
        port_info = _find_interface_port_info(content, iface_name)

        if len(port_info) < 2:
            print(f"Warning: could not resolve two end ports for '{iface_name}'; flows skipped.", file=sys.stderr)
            continue

        from_port_name, from_port_type, from_conj = port_info[0]
        to_port_name, to_port_type, to_conj = port_info[1]

        flow_lines: List[str] = []
        for alloc in allocs:
            fe_type = to_type_name(alloc.functional_exchange_name)
            fe_usage = to_usage_name(alloc.functional_exchange_name)

            if not from_conj:
                content = _ensure_port_def_has_item(content, from_port_type, fe_type, fe_usage, "out")
            if not to_conj:
                content = _ensure_port_def_has_item(content, to_port_type, fe_type, fe_usage, "in")

            flow_lines.append(
                f"{INDENT * 2}flow of {fe_type} from {from_port_name}.{fe_usage} to {to_port_name}.{fe_usage};"
            )

        content = _insert_into_interface_def(content, iface_name, flow_lines)

    return content


def import_exchange_allocations(rows: List[tuple], parts_path) -> None:
    """Inject exchange allocation flow lines into the parts SysML file."""
    parts_path = pathlib.Path(parts_path)
    if not parts_path.exists():
        print(f"Error: parts file not found: {parts_path}", file=sys.stderr)
        return

    allocations = _parse_exchange_allocation_rows(rows)
    if not allocations:
        print("Warning: no exchange allocations found in the input rows.", file=sys.stderr)
        return
    print(f"Exchange allocations: {len(allocations)}")

    parts_content = parts_path.read_text(encoding="utf-8")
    updated = inject_exchange_allocations(parts_content, allocations)
    parts_path.write_text(updated, encoding="utf-8")
    print(f"Written:  {parts_path}")
    validate(parts_path, check_dir=True)


# ===========================================================================
# SECTION: Functional Chains  (worksheet "Functional Chains")
# ===========================================================================

@dataclass
class FunctionalChain:
    chain_id: str
    chain_name: str
    start_function_id: str
    start_function_name: str
    end_function_id: str
    end_function_name: str
    function_ids: List[str] = field(default_factory=list)
    function_names: List[str] = field(default_factory=list)
    exchange_ids: List[str] = field(default_factory=list)
    exchange_names: List[str] = field(default_factory=list)


def _detect_chain_column_groups(header: tuple) -> Tuple[List[Tuple[int, int]], List[Tuple[int, int]]]:
    """Detect function and exchange column groups from a functional-chains header row."""
    function_cols: List[Tuple[int, int]] = []
    exchange_cols: List[Tuple[int, int]] = []

    n = 1
    while True:
        try:
            function_cols.append((header.index(f"Function {n} ID"), header.index(f"Function {n} Name")))
            n += 1
        except ValueError:
            break

    n = 1
    while True:
        try:
            exchange_cols.append((header.index(f"Exchange {n} ID"), header.index(f"Exchange {n} Name")))
            n += 1
        except ValueError:
            break

    return function_cols, exchange_cols


def _parse_chain_rows(rows: List[tuple]) -> List[FunctionalChain]:
    """Build a FunctionalChain list from worksheet rows (first row is header)."""
    if not rows:
        return []

    header = rows[0]
    function_cols, exchange_cols = _detect_chain_column_groups(header)

    if not function_cols:
        print("Warning: no 'Function N ID/Name' columns found in header.", file=sys.stderr)
    if not exchange_cols:
        print("Warning: no 'Exchange N ID/Name' columns found in header.", file=sys.stderr)

    print(f"Detected: {len(function_cols)} function column group(s), {len(exchange_cols)} exchange column group(s)")

    try:
        chain_id_col = header.index("Functional Chain ID")
        chain_name_col = header.index("Functional Chain Name")
        start_fn_id_col = header.index("Start Function ID")
        start_fn_name_col = header.index("Start Function Name")
        end_fn_id_col = header.index("End Function ID")
        end_fn_name_col = header.index("End Function Name")
    except ValueError as exc:
        print(f"Error: required column not found in header: {exc}", file=sys.stderr)
        return []

    chains: List[FunctionalChain] = []
    for row in rows[1:]:
        if not any(row):
            continue
        chain_id = row[chain_id_col]
        chain_name = row[chain_name_col]
        if not chain_id or not chain_name:
            continue

        function_ids = [str(row[ic]).strip() for ic, nc in function_cols if row[ic]]
        function_names = [str(row[nc]).strip() if row[nc] else "" for ic, nc in function_cols if row[ic]]
        exchange_ids = [str(row[ic]).strip() for ic, nc in exchange_cols if row[ic]]
        exchange_names = [str(row[nc]).strip() if row[nc] else "" for ic, nc in exchange_cols if row[ic]]

        chains.append(FunctionalChain(
            chain_id=str(chain_id).strip(),
            chain_name=str(chain_name).strip(),
            start_function_id=str(row[start_fn_id_col]).strip() if row[start_fn_id_col] else "",
            start_function_name=str(row[start_fn_name_col]).strip() if row[start_fn_name_col] else "",
            end_function_id=str(row[end_fn_id_col]).strip() if row[end_fn_id_col] else "",
            end_function_name=str(row[end_fn_name_col]).strip() if row[end_fn_name_col] else "",
            function_ids=function_ids,
            function_names=function_names,
            exchange_ids=exchange_ids,
            exchange_names=exchange_names,
        ))

    return chains


def load_chains(path: pathlib.Path, sheet_name: Optional[str] = None) -> List[FunctionalChain]:
    """Load functional chains from an Excel file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return _parse_chain_rows(list(ws.iter_rows(values_only=True)))


def inject_chains(sysml_text: str, chains: List[FunctionalChain]) -> str:
    """Return SysML text augmented with functional chain action defs."""
    id_to_def_name = parse_id_to_def_name(sysml_text)
    existing_names: Set[str] = set(re.findall(r"\baction def\s+(\w+)", sysml_text))

    append_lines: List[str] = []

    for chain in chains:
        chain_type = to_type_name(chain.chain_name, chain.chain_id)
        if chain_type in existing_names:
            print(f"  Skipping (already exists): {chain_type}")
            continue
        existing_names.add(chain_type)

        lines = [
            f"{INDENT}action def {chain_type} {{",
            f"{INDENT * 2}doc",
            f"{INDENT * 2}/* ID: {chain.chain_id} */",
        ]
        if chain.exchange_names:
            exchange_type_names = [to_type_name(n) for n in chain.exchange_names if n]
            lines.append(f"{INDENT * 2}/* Exchanges: {', '.join(exchange_type_names)} */")
        lines.append(f"{INDENT * 2}first start;")
        for func_id, func_name in zip(chain.function_ids, chain.function_names):
            action_type = id_to_def_name.get(func_id, to_type_name(func_name, func_id))
            action_usage = to_usage_name(func_name, func_id)
            lines.append(f"{INDENT * 2}then action {action_usage} : {action_type};")
        lines.append(f"{INDENT}}}")
        lines.append("")
        append_lines.extend(lines)

    if not append_lines:
        return sysml_text

    result = sysml_text.split("\n")
    closing_idx: Optional[int] = None
    for i in range(len(result) - 1, -1, -1):
        if result[i].strip() == "}":
            closing_idx = i
            break

    if closing_idx is None:
        result.extend(append_lines)
        result.append("}")
    else:
        result = result[:closing_idx] + append_lines + result[closing_idx:]

    return "\n".join(result)


def import_functional_chains(rows: List[tuple], functions_path) -> None:
    """Inject functional chain action defs into the functions SysML file."""
    functions_path = pathlib.Path(functions_path)
    if not functions_path.exists():
        print(f"Error: functions file not found: {functions_path}", file=sys.stderr)
        return

    chains = _parse_chain_rows(rows)
    if not chains:
        print("Warning: no functional chains found in the input rows.", file=sys.stderr)
        return
    print(f"Functional chains: {len(chains)}")

    sysml_text = functions_path.read_text(encoding="utf-8")
    augmented = inject_chains(sysml_text, chains)
    functions_path.write_text(augmented, encoding="utf-8")
    print(f"Written:  {functions_path}")
    validate(functions_path)


# ===========================================================================
# SECTION: full_import orchestrator
# ===========================================================================

# Worksheet name -> function mapping
_SHEET_MAP = {
    "Systems":                      "Systems",
    "Functions":                    "Functions",
    "Link Systems and Functions":   "Link Systems and Functions",
    "Functional Exchanges":         "Functional Exchanges",
    "Component Exchanges":          "Component Exchanges",
    "Link Exchanges":               "Link Exchanges",
    "Functional Chains":            "Functional Chains",
    # "Capabilities" is skipped
}


def _get_sheet_rows(wb, sheet_name: str) -> Optional[List[tuple]]:
    """Return rows from a named worksheet, or None if the sheet doesn't exist."""
    if sheet_name not in wb.sheetnames:
        print(f"Warning: worksheet '{sheet_name}' not found — skipping.", file=sys.stderr)
        return None
    return list(wb[sheet_name].iter_rows(values_only=True))


def full_import(
    excel_file,
    parts_filename,
    functions_filename,
    parts_package: str = "",
    functions_package: str = "",
) -> None:
    """Run the complete import pipeline from a single multi-sheet Excel workbook.

    Worksheet-to-function mapping:
      Systems                     -> import_systems       -> parts_filename
      Functions                   -> import_functions     -> functions_filename
      Link Systems and Functions  -> import_allocations   -> modifies parts_filename
      Functional Exchanges        -> import_functional_exchanges -> modifies functions_filename
      Component Exchanges         -> import_component_exchanges -> modifies parts_filename
      Link Exchanges              -> import_exchange_allocations -> modifies parts_filename
      Functional Chains           -> import_functional_chains   -> modifies functions_filename
      Capabilities                -> skipped
    """
    excel_file = pathlib.Path(excel_file)
    parts_path = pathlib.Path(parts_filename)
    functions_path = pathlib.Path(functions_filename)

    if not excel_file.exists():
        print(f"Error: Excel file not found: {excel_file}", file=sys.stderr)
        sys.exit(1)

    print(f"Opening: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    # 1. Systems -> Parts SysML
    rows = _get_sheet_rows(wb, "Systems")
    if rows:
        print("\n--- Systems ---")
        import_systems(rows, parts_path, parts_package)

    # 2. Functions -> Functions SysML
    rows = _get_sheet_rows(wb, "Functions")
    if rows:
        print("\n--- Functions ---")
        import_functions(rows, functions_path, functions_package)

    # 3. Link Systems and Functions -> augment Parts SysML
    rows = _get_sheet_rows(wb, "Link Systems and Functions")
    if rows:
        print("\n--- Link Systems and Functions ---")
        import_allocations(rows, parts_path, functions_path, functions_pkg=functions_package)

    # 4. Functional Exchanges -> augment Functions SysML
    rows = _get_sheet_rows(wb, "Functional Exchanges")
    if rows:
        print("\n--- Functional Exchanges ---")
        import_functional_exchanges(rows, functions_path)

    # 5. Component Exchanges -> augment Parts SysML
    rows = _get_sheet_rows(wb, "Component Exchanges")
    if rows:
        print("\n--- Component Exchanges ---")
        import_component_exchanges(rows, parts_path)

    # 6. Link Exchanges -> augment Parts SysML
    rows = _get_sheet_rows(wb, "Link Exchanges")
    if rows:
        print("\n--- Link Exchanges ---")
        import_exchange_allocations(rows, parts_path)

    # 7. Functional Chains -> augment Functions SysML
    rows = _get_sheet_rows(wb, "Functional Chains")
    if rows:
        print("\n--- Functional Chains ---")
        import_functional_chains(rows, functions_path)

    print("\nDone.")
