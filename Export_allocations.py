#!/usr/bin/env python3
"""
export_allocations_to_excel.py
Exports function allocations from SysML v2 files to an Excel file with proper hierarchical structure.
Each system is followed by its allocated functions in the Excel file.
"""

import pathlib
import re
import sys
from typing import Dict, List

import openpyxl
from openpyxl.styles import PatternFill

def from_pascal_case(name: str) -> str:
    """Convert PascalCase to space-separated words with proper handling of acronyms."""
    if not name:
        return name
    if name.isupper():
        return name
    result = []
    i = 0
    n = len(name)
    while i < n:
        if name[i].isupper():
            j = i
            while j < n and name[j].isupper():
                j += 1
            if j == n:
                result.append(name[i:j])
                i = j
            elif j < n and name[j].islower():
                result.append(name[i:j-1])
                result.append(" " + name[j-1])
                i = j
            else:
                result.append(name[i:j])
                i = j
        else:
            result.append(name[i])
            i += 1
    return "".join(result).replace("  ", " ").strip()

def extract_id_from_text(text: str) -> str:
    """Extract ID from text using the pattern: /* ID: uuid */"""
    match = re.search(r"/\*\s*ID:\s*([0-9a-f-]+)\s*\*/", text)
    return match.group(1) if match else ""

def get_block_content(text: str, start_pos: int) -> tuple:
    """Get content of a block starting with {, handling nested braces."""
    brace_count = 1
    end_pos = start_pos
    while end_pos < len(text) and brace_count > 0:
        if text[end_pos] == '{':
            brace_count += 1
        elif text[end_pos] == '}':
            brace_count -= 1
        end_pos += 1
    return text[start_pos:end_pos-1] if brace_count == 0 else "", end_pos

def parse_parts_file(file_path: pathlib.Path) -> Dict[str, Dict]:
    """
    Parse the Parts_generated.sysml file to extract part hierarchy and their allocated functions.
    Returns a dictionary of parts with their IDs, hierarchy, and allocated functions.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the PartsGenerated package content
    parts_package_match = re.search(r'package PartsGenerated\s*\{', content)
    if not parts_package_match:
        print("Error: PartsGenerated package not found in file", file=sys.stderr)
        return {}

    start_pos = parts_package_match.end()
    parts_content, end_pos = get_block_content(content, start_pos)

    # First, find all part definitions
    part_defs = {}
    for part_def_match in re.finditer(r'part def (\w+)\s*\{', parts_content):
        part_name = part_def_match.group(1)
        start = part_def_match.end()
        content_block, end = get_block_content(parts_content, start)
        part_id = extract_id_from_text(content_block)

        # Find all part usages in this part
        usages = []
        for usage_match in re.finditer(r'part (\w+)\s*:\s*(\w+);', content_block):
            usage_name, usage_type = usage_match.groups()
            usages.append(usage_type)

        # Find all allocated functions (perform action)
        allocated_functions = []
        for perform_match in re.finditer(r'perform action (\w+)\s*:\s*(\w+)\s*\{', content_block):
            usage_name, function_type = perform_match.groups()
            function_id = extract_id_from_text(content_block[perform_match.end():])
            allocated_functions.append({
                'name': function_type,
                'id': function_id,
                'usage_name': usage_name
            })

        part_defs[part_name] = {
            'id': part_id,
            'usages': usages,
            'functions': allocated_functions,
            'content': content_block
        }

    return part_defs

def parse_functions_file(file_path: pathlib.Path) -> Dict[str, Dict]:
    """
    Parse the Functions_generated.sysml file to extract function definitions and their IDs.
    Returns a dictionary of functions with their IDs.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Find the FunctionsGenerated package content
    functions_package_match = re.search(r'package FunctionsGenerated\s*\{', content)
    if not functions_package_match:
        print("Error: FunctionsGenerated package not found in file", file=sys.stderr)
        return {}

    start_pos = functions_package_match.end()
    functions_content, end_pos = get_block_content(content, start_pos)

    # Find all action definitions
    function_defs = {}
    for action_def_match in re.finditer(r'action def (\w+)\s*\{', functions_content):
        function_name = action_def_match.group(1)
        start = action_def_match.end()
        content_block, end = get_block_content(functions_content, start)
        function_id = extract_id_from_text(content_block)

        function_defs[function_name] = {
            'id': function_id,
            'content': content_block
        }

    return function_defs

def build_allocation_hierarchy(part_defs: Dict[str, Dict]) -> List[Dict]:
    """
    Build a flat list of parts and their allocated functions with hierarchy levels.
    Each part is followed by its allocated functions.
    """
    # Find top-level parts (not used by any other part)
    all_used_types = set()
    for part_name, part_data in part_defs.items():
        all_used_types.update(part_data['usages'])

    top_level_parts = [name for name in part_defs if name not in all_used_types]

    # If LogicalSystem exists, make it the first top-level part
    if 'LogicalSystem' in top_level_parts:
        top_level_parts.remove('LogicalSystem')
        top_level_parts.insert(0, 'LogicalSystem')

    result = []
    for top_part_name in top_level_parts:
        stack = [(top_part_name, 0)]
        while stack:
            part_name, level = stack.pop()
            part_data = part_defs[part_name]

            # Add the part to our list
            result.append({
                'type': 'part',
                'name': part_name,
                'id': part_data['id'],
                'level': level
            })

            # Add its allocated functions (if any)
            for func in part_data['functions']:
                result.append({
                    'type': 'function',
                    'name': func['name'],
                    'id': func['id'],
                    'level': level,
                    'parent': part_name
                })

            # Add children to stack in reverse order (so they're processed in order)
            for child_type in reversed(part_data['usages']):
                if child_type in part_defs:
                    stack.append((child_type, level + 1))

    return result

def export_allocations_to_excel(allocation_hierarchy: List[Dict], output_path: pathlib.Path) -> None:
    """Export allocations to an Excel file with hierarchical structure."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    if wb.active.title == "Sheet":
        wb.remove(wb.active)

    ws = wb.create_sheet("Allocations")

    # Determine the maximum hierarchy level
    max_level = max([item['level'] for item in allocation_hierarchy]) if allocation_hierarchy else 0

    # Create headers dynamically based on max_level
    headers = []
    for level in range(max_level + 1):
        prefix = "Sub" * level if level > 0 else ""
        headers.extend([f"{prefix}System ID", f"{prefix}System Name"])
    headers.extend(["Function ID", "Function Name"])

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Export items in hierarchy order
    for item in allocation_hierarchy:
        new_row = ws.max_row + 1

        if item['type'] == 'part':
            # This is a system/part - fill system columns
            system_col = 1 + (item['level'] * 2)
            ws.cell(row=new_row, column=system_col, value=item['id'])
            ws.cell(row=new_row, column=system_col+1, value=from_pascal_case(item['name']))
        elif item['type'] == 'function':
            # This is a function - fill function columns
            function_id_col = len(headers) - 1
            function_name_col = len(headers)
            ws.cell(row=new_row, column=function_id_col, value=item['id'])
            ws.cell(row=new_row, column=function_name_col, value=from_pascal_case(item['name']))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Exported allocations to: {output_path}")

def main() -> None:
    _DVS_DIR = pathlib.Path(__file__).parent.parent
    DEFAULT_PARTS_INPUT = _DVS_DIR / "Parts_generated.sysml"
    DEFAULT_FUNCTIONS_INPUT = _DVS_DIR / "Functions_generated.sysml"
    DEFAULT_OUTPUT = _DVS_DIR / "results" / "DVS_Allocations_export.xlsx"

    args = sys.argv[1:]
    if args and args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    parts_input = DEFAULT_PARTS_INPUT
    functions_input = DEFAULT_FUNCTIONS_INPUT
    output_path = DEFAULT_OUTPUT

    if len(args) >= 1:
        parts_input = pathlib.Path(args[0])
    if len(args) >= 2:
        functions_input = pathlib.Path(args[1])
    if len(args) >= 3:
        output_path = pathlib.Path(args[2])

    if not parts_input.exists():
        print(f"Error: Parts file not found: {parts_input}", file=sys.stderr)
        sys.exit(1)
    if not functions_input.exists():
        print(f"Error: Functions file not found: {functions_input}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading parts from: {parts_input}")
    print(f"Reading functions from: {functions_input}")

    # Parse files
    part_defs = parse_parts_file(parts_input)
    function_defs = parse_functions_file(functions_input)

    # Build allocation hierarchy
    allocation_hierarchy = build_allocation_hierarchy(part_defs)

    if not allocation_hierarchy:
        print("Warning: No allocations found in the file!")
    else:
        print(f"Found {len(allocation_hierarchy)} items (parts and functions) in the hierarchy")
        # Print hierarchy for verification
        for item in allocation_hierarchy:
            indent = "  " * item['level']
            if item['type'] == 'part':
                print(f"{indent}{item['name']} (System, Level {item['level']}, ID: {item['id']})")
            else:
                print(f"{indent}{item['name']} (Function, Level {item['level']}, ID: {item['id']})")

    export_allocations_to_excel(allocation_hierarchy, output_path)

if __name__ == "__main__":
    main()