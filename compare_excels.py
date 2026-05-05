#!/usr/bin/env python3
"""
compare_excel_files.py
Compares two Excel files (all worksheets) and creates a new file with the content
of the first file, coloring cells green for matches (case-insensitive) and red for differences.
"""

import pathlib
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

def compare_excel_files(file1_path: pathlib.Path, file2_path: pathlib.Path, output_path: pathlib.Path) -> None:
    """
    Compare two Excel files (all worksheets) and create a new file with the first file's content,
    coloring cells based on whether they match (green) or differ (red) from the second file.
    Comparison is case-insensitive.
    """
    # Load both workbooks
    wb1 = load_workbook(filename=str(file1_path), data_only=True)
    wb2 = load_workbook(filename=str(file2_path), data_only=True)

    # Create a new output workbook
    output_wb = Workbook()
    # Remove the default sheet created by Workbook()
    if 'Sheet' in output_wb.sheetnames:
        del output_wb['Sheet']

    # Define fill styles
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")    # Light red

    # Get all sheet names from the first workbook
    sheet_names = wb1.sheetnames

    for sheet_name in sheet_names:
        # Get sheets from both workbooks
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name] if sheet_name in wb2.sheetnames else None

        # Create a new sheet in the output workbook
        output_sheet = output_wb.create_sheet(title=sheet_name)

        # Copy all cell values and formatting from sheet1 to output_sheet
        for row in sheet1.iter_rows():
            for cell in row:
                output_cell = output_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                # Copy formatting if needed
                if cell.has_style:
                    output_cell.font = cell.font.copy()
                    output_cell.border = cell.border.copy()
                    output_cell.fill = cell.fill.copy()
                    output_cell.number_format = cell.number_format
                    output_cell.protection = cell.protection.copy()
                    output_cell.alignment = cell.alignment.copy()

        # Now color the cells based on comparison with sheet2
        if sheet2 is not None:
            for row in sheet1.iter_rows():
                for cell in row:
                    col_letter = cell.column_letter
                    row_num = cell.row

                    # Check if the cell exists in the second sheet
                    if row_num <= sheet2.max_row and cell.column <= sheet2.max_column:
                        cell2 = sheet2[f"{col_letter}{row_num}"]

                        # Get cell values as strings, case-insensitive comparison
                        val1 = str(cell.value).strip().lower() if cell.value is not None else ""
                        val2 = str(cell2.value).strip().lower() if cell2.value is not None else ""

                        output_cell = output_sheet[f"{col_letter}{row_num}"]

                        if val1 == val2:
                            # Values match (case-insensitive)
                            output_cell.fill = green_fill
                        else:
                            # Values differ
                            output_cell.fill = red_fill
                    else:
                        # Cell doesn't exist in second sheet - mark as different
                        output_cell = output_sheet[f"{col_letter}{row_num}"]
                        output_cell.fill = red_fill

            # Mark cells that exist in sheet2 but not in sheet1 as red
            for row in sheet2.iter_rows():
                for cell in row:
                    col_letter = cell.column_letter
                    row_num = cell.row

                    if row_num > sheet1.max_row or cell.column > sheet1.max_column:
                        output_cell = output_sheet[f"{col_letter}{row_num}"]
                        output_cell.value = cell.value  # Copy the value
                        output_cell.fill = red_fill

    # Save the output workbook
    output_wb.save(filename=str(output_path))
    print(f"Comparison result saved to: {output_path}")

def main() -> None:
    # Get the directory where this script is located
    script_dir = pathlib.Path(__file__).parent

    # Define your file paths directly here
    data_dir = script_dir / "Data"  # Assuming your files are in a Data subdirectory

    file1_path = data_dir / "TestExport_complete_export.xlsx"
    file2_path = data_dir / "DVS_complete_export.xlsx"
    output_path = data_dir / "Comparison_Result.xlsx"  # Output file path

    # Verify files exist
    if not file1_path.exists():
        print(f"Error: First file not found at: {file1_path}", file=sys.stderr)
        print(f"Current directory: {script_dir}", file=sys.stderr)
        sys.exit(1)
    if not file2_path.exists():
        print(f"Error: Second file not found at: {file2_path}", file=sys.stderr)
        print(f"Current directory: {script_dir}", file=sys.stderr)
        sys.exit(1)

    print(f"Comparing: {file1_path} with {file2_path}")
    compare_excel_files(file1_path, file2_path, output_path)

if __name__ == "__main__":
    main()